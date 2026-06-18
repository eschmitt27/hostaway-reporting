"""
lot4_saisie_template.py
Lot 4 — Creation SAISIE_ReservationsHorsHostaway.xlsx
- Onglet SAISIE (30 colonnes + data validation + MFC)
- Onglet REF_LOCALE (11 listes + lookup taux commission)
- Onglet CONTROLES_SAISIE (13 controles automatiques)
- Onglet README (instructions)

Decisions integrees :
  QM-L4-01 : canal_id obligatoire (REF_Canaux_Reservation)
  QM-L4-02 : taux_commission VLOOKUP REF_PROPRIETAIRE, fallback A_CONTROLER
  QM-L4-03 : VRBO Unknown dans ce fichier, meme canal_id=CANAL_003
  QM-L4-04 : separation reservation/charge, lien via paye_avec_montant_recupere (Lot 3)
  QM-L4-05 : PK RESHH-AAAA-MM-NNN, NNN reset 001 chaque mois, valeur figee
  QM-L4-06 : chemin 01_SOURCES_BRUTES/ReservationsHH/
"""

import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import FormulaRule

BASE     = "C:/Users/Ewan/OneDrive/Documents/Conciergerie/Pilotage_Conciergerie"
REF_PATH = f"{BASE}/01_SOURCES_BRUTES/REF_Setup/REF_Setup.xlsm"
OUT_DIR  = f"{BASE}/01_SOURCES_BRUTES/ReservationsHH"
OUT_PATH = f"{OUT_DIR}/SAISIE_ReservationsHorsHostaway.xlsx"

os.makedirs(OUT_DIR, exist_ok=True)

# ── 1. Lire REF_Setup ─────────────────────────────────────────────────────────
ref_wb = openpyxl.load_workbook(REF_PATH, read_only=True, keep_vba=True)

# REF_Canaux_Reservation : col0=canal_id, col5=actif
canaux = [
    r[0] for r in ref_wb['REF_Canaux_Reservation'].iter_rows(min_row=2, values_only=True)
    if r[0] and r[5] == 'OUI'
]

# REF_Proprietaires : col0=proprietaire_id, col7=taux_commission, col8=actif
proprietaires_raw = [
    (r[0], r[7]) for r in ref_wb['REF_Proprietaires'].iter_rows(min_row=2, values_only=True)
    if r[0] and r[8] == 'OUI'
]
proprietaires = [p[0] for p in proprietaires_raw]
taux_prop     = [p[1] for p in proprietaires_raw]

# REF_Logements : col0=logement_id, col12=actif
logements = [
    r[0] for r in ref_wb['REF_Logements'].iter_rows(min_row=2, values_only=True)
    if r[0] and r[12] == 'OUI'
]

# REF_Associes : col0=associe_id, col3=actif
associes = [
    r[0] for r in ref_wb['REF_Associes'].iter_rows(min_row=2, values_only=True)
    if r[0] and r[3] == 'OUI'
]

# REF_Modes_Paiement : col0=mode_paiement_id, col5=actif
modes = [
    r[0] for r in ref_wb['REF_Modes_Paiement'].iter_rows(min_row=2, values_only=True)
    if r[0] and r[5] == 'OUI'
]

ref_wb.close()
print(f"  REF lu — canaux:{len(canaux)}, proprios:{len(proprietaires)}, "
      f"logements:{len(logements)}, associes:{len(associes)}, modes:{len(modes)}")

# Listes statiques
SOURCE_FINANCIERE      = ['SAISIE_MANUELLE', 'HOSTAWAY_REFERENCE', 'VRBO_UNKNOWN',
                          'DIRECT_HA_PAYANT', 'A_CONTROLER']
TAUX_COMMISSION_SOURCE = ['REF_PROPRIETAIRE', 'REF_LOGEMENT', 'SAISIE_MANUELLE', 'A_CONTROLER']
CODES_IMPACT           = ['HC', 'IC', 'HR']
COMPTABILISATION       = ['OUI', 'NON']
STATUTS_CONTROLE       = ['VALIDE', 'A_CONTROLER', 'EXCLU_RESULTAT', 'A_VENTILER']
NIVEAUX_ANOMALIE       = ['INFO', 'A_CONTROLER', 'BLOQUANT']

# ── 2. Workbook ───────────────────────────────────────────────────────────────
wb        = Workbook()
ws_saisie = wb.active
ws_saisie.title = 'SAISIE'
ws_ref    = wb.create_sheet('REF_LOCALE')
ws_ctrl   = wb.create_sheet('CONTROLES_SAISIE')
ws_readme = wb.create_sheet('README')

H_FONT = Font(bold=True, color='FFFFFF', size=9)
H_FILL = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')

# ── 3. REF_LOCALE — listes simples ────────────────────────────────────────────
simple_lists = [
    ('lst_Canaux',               canaux),
    ('lst_Proprietaires',        proprietaires),
    ('lst_Logements',            logements),
    ('lst_Associes',             associes),
    ('lst_ModesPaiement',        modes),
    ('lst_Source_Financiere',    SOURCE_FINANCIERE),
    ('lst_TauxCommissionSource', TAUX_COMMISSION_SOURCE),
    ('lst_Codes_Impact',         CODES_IMPACT),
    ('lst_Comptabilisation',     COMPTABILISATION),
    ('lst_Statuts_Controle',     STATUTS_CONTROLE),
    ('lst_Niveau_Anomalie',      NIVEAUX_ANOMALIE),
]

for col_i, (lst_name, lst_vals) in enumerate(simple_lists, 1):
    col_l = get_column_letter(col_i)
    c = ws_ref.cell(row=1, column=col_i, value=lst_name)
    c.font = H_FONT
    c.fill = H_FILL
    ws_ref.column_dimensions[col_l].width = (
        max(len(lst_name), max((len(str(v)) for v in lst_vals), default=8)) + 2
    )
    for row_i, val in enumerate(lst_vals, 2):
        ws_ref.cell(row=row_i, column=col_i, value=val)

# Lookup taux_commission (2 colonnes): prop_id | taux — pour VLOOKUP col N
TAUX_COL_START = len(simple_lists) + 2  # une colonne vide de separation
taux_col_id    = TAUX_COL_START
taux_col_val   = TAUX_COL_START + 1
dark_fill = PatternFill(start_color='404040', end_color='404040', fill_type='solid')

c_id  = ws_ref.cell(row=1, column=taux_col_id,  value='prop_taux_id')
c_val = ws_ref.cell(row=1, column=taux_col_val, value='prop_taux_val')
for c in (c_id, c_val):
    c.font = H_FONT
    c.fill = dark_fill

for row_i, (pid, taux) in enumerate(proprietaires_raw, 2):
    ws_ref.cell(row=row_i, column=taux_col_id,  value=pid)
    ws_ref.cell(row=row_i, column=taux_col_val, value=taux)

ws_ref.freeze_panes = 'A2'
print(f"[OK] REF_LOCALE — {len(simple_lists)} listes + lookup taux {len(proprietaires_raw)} proprios")

# ── 4. Named Ranges ───────────────────────────────────────────────────────────
for col_i, (lst_name, lst_vals) in enumerate(simple_lists, 1):
    col_l = get_column_letter(col_i)
    n = len(lst_vals)
    ref_str = f"'REF_LOCALE'!${col_l}$2:${col_l}${n + 1}"
    wb.defined_names[lst_name] = DefinedName(name=lst_name, attr_text=ref_str)

# Named range 2-col pour VLOOKUP taux
tl_id  = get_column_letter(taux_col_id)
tl_val = get_column_letter(taux_col_val)
n_t    = len(proprietaires_raw)
ref_taux = f"'REF_LOCALE'!${tl_id}$2:${tl_val}${n_t + 1}"
wb.defined_names['lst_PropTauxLookup'] = DefinedName(
    name='lst_PropTauxLookup', attr_text=ref_taux
)
print(f"[OK] Named Ranges — {len(simple_lists)} listes + lst_PropTauxLookup")

# ── 5. SAISIE — 30 colonnes ───────────────────────────────────────────────────
# (nom_colonne, couleur_groupe_hex)
SAISIE_COLS = [
    # Groupe 1 — Identification (A-B)
    ('reservation_hh_id',            '1F497D'),   # A  col 1
    ('ROW_HASH',                     '1F497D'),   # B  col 2
    # Groupe 2 — Rattachement (C-H)
    ('mois',                         '2E75B6'),   # C  col 3
    ('canal_id',                     '2E75B6'),   # D  col 4
    ('source_financiere',            '2E75B6'),   # E  col 5
    ('proprietaire_id',              '2E75B6'),   # F  col 6
    ('logement_id',                  '2E75B6'),   # G  col 7
    ('reservation_id_hostaway',      '2E75B6'),   # H  col 8
    # Groupe 3 — Sejour (I-K)
    ('date_arrivee',                 '375623'),   # I  col 9
    ('date_depart',                  '375623'),   # J  col 10
    ('nuits',                        '375623'),   # K  col 11
    # Groupe 4 — Financier encaissement (L-Q)
    ('total_percu',                  '7030A0'),   # L  col 12
    ('menage',                       '7030A0'),   # M  col 13
    ('taux_commission',              '7030A0'),   # N  col 14
    ('taux_commission_source',       '7030A0'),   # O  col 15
    ('commentaire_taux_commission',  '7030A0'),   # P  col 16
    ('commission',                   '7030A0'),   # Q  col 17
    # Groupe 5 — Associe recuperateur (R-S)
    ('montant_recupere',             '974706'),   # R  col 18
    ('associe_id_recuperateur',      '974706'),   # S  col 19
    # Groupe 6 — Reversement proprietaire (T-U)
    ('montant_reverse_proprietaire', 'C55A11'),   # T  col 20
    ('mode_paiement_id',             'C55A11'),   # U  col 21
    # Groupe 7 — Acompte (V)
    ('acompte_facture',              '833C00'),   # V  col 22
    # Groupe 8 — Statut / impact (W-AC)
    ('code_impact',                  'C00000'),   # W  col 23
    ('comptabilisation',             'C00000'),   # X  col 24
    ('impact_resultat_reel',         'C00000'),   # Y  col 25
    ('impact_resultat_comptable',    'C00000'),   # Z  col 26
    ('statut_controle',              '7B0000'),   # AA col 27
    ('niveau_anomalie',              '7B0000'),   # AB col 28
    ('code_anomalie',                '7B0000'),   # AC col 29
    # Groupe 9 — Note (AD)
    ('commentaire',                  '595959'),   # AD col 30
]

assert len(SAISIE_COLS) == 30, f"Attendu 30 colonnes, got {len(SAISIE_COLS)}"

COL_WIDTHS = [
    28, 46,          # A-B  : reservation_hh_id, ROW_HASH
    12, 14, 22,      # C-E  : mois, canal_id, source_financiere
    14, 22, 22,      # F-H  : proprietaire_id, logement_id, reservation_id_hostaway
    14, 14, 8,       # I-K  : date_arrivee, date_depart, nuits
    14, 12, 16,      # L-N  : total_percu, menage, taux_commission
    24, 30, 14,      # O-Q  : taux_commission_source, commentaire_taux, commission
    14, 22,          # R-S  : montant_recupere, associe_id_recuperateur
    26, 18,          # T-U  : montant_reverse_proprietaire, mode_paiement_id
    14,              # V    : acompte_facture
    12, 16,          # W-X  : code_impact, comptabilisation
    22, 24,          # Y-Z  : impact_resultat_reel, impact_resultat_comptable
    18, 18, 24,      # AA-AC: statut_controle, niveau_anomalie, code_anomalie
    30,              # AD   : commentaire
]

for col_i, (col_name, color) in enumerate(SAISIE_COLS, 1):
    c = ws_saisie.cell(row=1, column=col_i, value=col_name)
    c.font = Font(bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws_saisie.column_dimensions[get_column_letter(col_i)].width = COL_WIDTHS[col_i - 1]

ws_saisie.row_dimensions[1].height = 42
ws_saisie.freeze_panes = 'A2'

# ── 6. Formules calculees (500 lignes) ────────────────────────────────────────
N_FORMULA = 500
for row in range(2, N_FORMULA + 2):
    r = row
    # C3 — mois (depuis date_arrivee col I)
    ws_saisie.cell(r, 3).value = (
        f'=IF(I{r}="","",TEXT(I{r},"YYYY-MM"))'
    )
    # K11 — nuits
    ws_saisie.cell(r, 11).value = (
        f'=IF(OR(I{r}="",J{r}=""),"",J{r}-I{r})'
    )
    # N14 — taux_commission : VLOOKUP depuis proprietaire_id (col F)
    ws_saisie.cell(r, 14).value = (
        f'=IF(F{r}="","",IFERROR(VLOOKUP(F{r},lst_PropTauxLookup,2,0),""))'
    )
    # O15 — taux_commission_source
    ws_saisie.cell(r, 15).value = (
        f'=IF(F{r}="","",'
        f'IF(ISNUMBER(IFERROR(VLOOKUP(F{r},lst_PropTauxLookup,2,0),"")),'
        f'"REF_PROPRIETAIRE","A_CONTROLER"))'
    )
    # Q17 — commission : (total_percu - menage) * taux
    ws_saisie.cell(r, 17).value = (
        f'=IF(OR(L{r}="",N{r}=""),"",IFERROR((L{r}-IF(M{r}="",0,M{r}))*N{r},0))'
    )
    # V22 — acompte_facture
    ws_saisie.cell(r, 22).value = (
        f'=IF(L{r}="","",L{r}'
        f'-IF(M{r}="",0,M{r})'
        f'-IF(Q{r}="",0,Q{r})'
        f'-IF(T{r}="",0,T{r}))'
    )
    # Y25 — impact_resultat_reel (depuis code_impact col W)
    ws_saisie.cell(r, 25).value = (
        f'=IF(W{r}="IC","OUI",'
        f'IF(W{r}="HC","OUI",'
        f'IF(W{r}="HR","NON","A_CONTROLER")))'
    )
    # Z26 — impact_resultat_comptable
    ws_saisie.cell(r, 26).value = (
        f'=IF(W{r}="IC","OUI",'
        f'IF(W{r}="HC","NON",'
        f'IF(W{r}="HR","NON","A_CONTROLER")))'
    )
    # B2 — ROW_HASH
    ws_saisie.cell(r, 2).value = (
        f'=IF(A{r}="","",A{r}&"|"&D{r}&"|"&F{r}&"|"&G{r}'
        f'&"|"&TEXT(I{r},"YYYYMMDD")&"|"&TEXT(L{r},"0.00"))'
    )

# ── 7. Data Validations ───────────────────────────────────────────────────────
# 11 DVs : 10 du plan + niveau_anomalie (coherence Lot 3 — DV informationnel)
DV_MAP = [
    ('D',  'lst_Canaux'),
    ('E',  'lst_Source_Financiere'),
    ('F',  'lst_Proprietaires'),
    ('G',  'lst_Logements'),
    ('O',  'lst_TauxCommissionSource'),
    ('S',  'lst_Associes'),
    ('U',  'lst_ModesPaiement'),
    ('W',  'lst_Codes_Impact'),
    ('X',  'lst_Comptabilisation'),
    ('AA', 'lst_Statuts_Controle'),
    ('AB', 'lst_Niveau_Anomalie'),
]

for col_l, lst_name in DV_MAP:
    dv = DataValidation(
        type="list",
        formula1=lst_name,
        allow_blank=True,
        showErrorMessage=True,
        error='Valeur hors liste. Utiliser la liste deroulante.',
        errorTitle='Valeur invalide',
    )
    ws_saisie.add_data_validation(dv)
    dv.sqref = f'{col_l}2:{col_l}501'

# ── 8. MFC ────────────────────────────────────────────────────────────────────
red_fill    = PatternFill(start_color='FFAAAA', end_color='FFAAAA', fill_type='solid')
orange_fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
ws_saisie.conditional_formatting.add(
    'A2:AD501', FormulaRule(formula=['$AB2="BLOQUANT"'],    fill=red_fill)
)
ws_saisie.conditional_formatting.add(
    'A2:AD501', FormulaRule(formula=['$AB2="A_CONTROLER"'], fill=orange_fill)
)
print("[OK] SAISIE — 30 colonnes + 11 DV + formules + MFC")

# ── 9. CONTROLES_SAISIE — 13 controles ───────────────────────────────────────
CTRL_HEADERS = ['Code', 'Libelle', 'Valeur', 'Seuil', 'Note']
for col_i, h in enumerate(CTRL_HEADERS, 1):
    c = ws_ctrl.cell(row=1, column=col_i, value=h)
    c.font = H_FONT
    c.fill = H_FILL

CONTROLS = [
    ('CTR-L4-01', 'Lignes saisies (reservation_hh_id renseigne)',
     '=MAX(0,COUNTA(SAISIE!A:A)-1)',
     'INFO', ''),
    ('CTR-L4-02', 'Lignes VALIDE (statut_controle)',
     '=COUNTIF(SAISIE!AA:AA,"VALIDE")',
     'INFO', ''),
    ('CTR-L4-03', 'Lignes A_CONTROLER (statut_controle)',
     '=COUNTIF(SAISIE!AA:AA,"A_CONTROLER")',
     'ATTENTION si > 0', ''),
    ('CTR-L4-04', 'Lignes EXCLU_RESULTAT',
     '=COUNTIF(SAISIE!AA:AA,"EXCLU_RESULTAT")',
     'INFO', ''),
    ('CTR-L4-05', 'Lignes A_VENTILER',
     '=COUNTIF(SAISIE!AA:AA,"A_VENTILER")',
     'ATTENTION si > 0', ''),
    ('CTR-L4-06', 'Lignes BLOQUANT (niveau_anomalie)',
     '=COUNTIF(SAISIE!AB:AB,"BLOQUANT")',
     'BLOQUANT si > 0', 'Doit etre 0 avant livraison vers MASTER'),
    ('CTR-L4-07', 'reservation_hh_id dupliques',
     '=SUMPRODUCT((COUNTIF(SAISIE!A2:A501,SAISIE!A2:A501)-1)*(SAISIE!A2:A501<>""))',
     'BLOQUANT si > 0', 'PK unique obligatoire — RESH_DOUBLON_HH_ID'),
    ('CTR-L4-08', 'Proprietaire manquant (ligne renseignee)',
     '=COUNTIFS(SAISIE!A2:A501,"<>",SAISIE!F2:F501,"")',
     'BLOQUANT si > 0', 'RESH_PROPRIETAIRE_MANQUANT'),
    ('CTR-L4-09', 'Logement manquant (ligne renseignee)',
     '=COUNTIFS(SAISIE!A2:A501,"<>",SAISIE!G2:G501,"")',
     'BLOQUANT si > 0', 'RESH_LOGEMENT_MANQUANT'),
    ('CTR-L4-10', 'Canal manquant (ligne renseignee)',
     '=COUNTIFS(SAISIE!A2:A501,"<>",SAISIE!D2:D501,"")',
     'BLOQUANT si > 0', 'RESH_CANAL_MANQUANT'),
    ('CTR-L4-11', 'Source financiere manquante (ligne renseignee)',
     '=COUNTIFS(SAISIE!A2:A501,"<>",SAISIE!E2:E501,"")',
     'BLOQUANT si > 0', 'RESH_SOURCE_FINANCIERE_MANQUANTE'),
    ('CTR-L4-12', 'Acompte facture negatif',
     '=COUNTIF(SAISIE!V2:V501,"<0")',
     'BLOQUANT si > 0', 'RESH_ACOMPTE_NEGATIF — verifier la formule'),
    ('CTR-L4-13', 'VRBO sans montant (VRBO_UNKNOWN + total_percu vide)',
     '=COUNTIFS(SAISIE!E2:E501,"VRBO_UNKNOWN",SAISIE!L2:L501,"")',
     'A_CONTROLER si > 0', 'ANO-004 — saisir le montant VRBO reel'),
]

for row_i, ctrl in enumerate(CONTROLS, 2):
    for col_i, val in enumerate(ctrl, 1):
        ws_ctrl.cell(row=row_i, column=col_i, value=val)

ws_ctrl.column_dimensions['A'].width = 14
ws_ctrl.column_dimensions['B'].width = 55
ws_ctrl.column_dimensions['C'].width = 65
ws_ctrl.column_dimensions['D'].width = 22
ws_ctrl.column_dimensions['E'].width = 55
ws_ctrl.freeze_panes = 'A2'
print("[OK] CONTROLES_SAISIE — 13 controles")

# ── 10. README ────────────────────────────────────────────────────────────────
README_LINES = [
    ('SAISIE_ReservationsHorsHostaway.xlsx — Guide de saisie (Lot 4)', True),
    ('', False),
    ('PERIMETRE', True),
    ('Reservations non gerees financierement par Hostaway.', False),
    ('  - 29 VRBO paymentStatus=Unknown (ANO-004) : source_financiere=VRBO_UNKNOWN', False),
    ('  - 13 Direct Hostaway totalPrice>0 dont la verite financiere est ici : source_financiere=DIRECT_HA_PAYANT', False),
    ('  - Reservations absentes de Hostaway : source_financiere=SAISIE_MANUELLE', False),
    ('EXCLUT : ownerStay (montant 0). IK/avantages (Lot 7). Acomptes proprietaires (Lot 5).', False),
    ('', False),
    ('REGLES OBLIGATOIRES', True),
    ('1. reservation_hh_id : PK unique. Format : RESHH-AAAA-MM-NNN (NNN reset a 001 chaque mois).', False),
    ('   Exemple : RESHH-2026-05-001 / RESHH-2026-06-001.', False),
    ('   SAISIR comme valeur fixe (taper en dur ou Coller valeur seule). NE PAS utiliser de formule.', False),
    ('   La cle ne doit jamais etre regeneree automatiquement apres saisie initiale.', False),
    ('2. canal_id : obligatoire. VRBO=CANAL_003 / Direct=CANAL_004 / Autre=CANAL_005.', False),
    ('3. source_financiere : obligatoire (voir tableau ci-dessous).', False),
    ('4. reservation_id_hostaway : obligatoire si source_financiere IN', False),
    ('   {VRBO_UNKNOWN, DIRECT_HA_PAYANT, HOSTAWAY_REFERENCE}.', False),
    ('5. taux_commission : pre-rempli via VLOOKUP depuis REF_Proprietaires (col N).', False),
    ('   Si la formule renvoie vide → taux absent du referentiel.', False),
    ('   Dans ce cas : saisir le taux manuellement ET changer taux_commission_source a SAISIE_MANUELLE.', False),
    ('6. acompte_facture = total_percu - menage - commission - montant_reverse_proprietaire (calcule, col V).', False),
    ('   Si negatif → BLOQUANT. Verifier coherence des montants saisis.', False),
    ('7. Si une charge est payee avec le montant recupere par un associe :', False),
    ('   → saisir la charge dans SAISIE_Charges_Flux.xlsx (Lot 3),', False),
    ('     champ paye_avec_montant_recupere = reservation_hh_id de cette ligne.', False),
    ('   → NE PAS compenser dans ce fichier.', False),
    ('8. Passer statut_controle a VALIDE avant integration dans MASTER (D029).', False),
    ('', False),
    ('TABLE source_financiere', True),
    ('  VRBO_UNKNOWN        : VRBO dans Hostaway, paymentStatus=Unknown — montant saisi ici.', False),
    ('  DIRECT_HA_PAYANT    : Direct Hostaway avec totalPrice>0, verite financiere = Lot 4.', False),
    ('  SAISIE_MANUELLE     : Reservation absente de Hostaway, saisie integrale.', False),
    ('  HOSTAWAY_REFERENCE  : Hostaway sert de reference non financiere uniquement.', False),
    ('  A_CONTROLER         : Source non clairement identifiable.', False),
    ('', False),
    ('DOUBLON ANTI-COMPTAGE (Lot 4bis)', True),
    ('Les lignes DIRECT_HA_PAYANT seront reconciliees vs Hostaway direct via reservation_id_hostaway.', False),
    ('Chaque reservation n alimente MASTER_CALC_Flux qu une seule fois (TC2 — REGLES_METIER §10).', False),
    ('', False),
    ('ONGLETS', True),
    ('  SAISIE             : 30 colonnes, 11 DV, MFC rouge=BLOQUANT / orange=A_CONTROLER', False),
    ('  REF_LOCALE         : 11 listes + lookup taux commission (ne pas modifier manuellement)', False),
    ('  CONTROLES_SAISIE   : 13 controles automatiques — verifier avant chaque livraison', False),
    ('  README             : ce fichier', False),
    ('', False),
    ('CONTROLES OBLIGATOIRES AVANT LIVRAISON (D029 — JOURNAL_CONTROLES)', True),
    ('CTR-L4-06 = 0  (aucun BLOQUANT dans niveau_anomalie)', False),
    ('CTR-L4-07 = 0  (aucun reservation_hh_id duplique)', False),
    ('CTR-L4-08 = 0  (aucun proprietaire manquant)', False),
    ('CTR-L4-09 = 0  (aucun logement manquant)', False),
    ('CTR-L4-10 = 0  (aucun canal manquant)', False),
    ('CTR-L4-11 = 0  (aucune source_financiere manquante)', False),
    ('CTR-L4-12 = 0  (aucun acompte_facture negatif)', False),
    ('CTR-L4-13 : noter le nombre de VRBO sans montant restants (ANO-004)', False),
]

title_font  = Font(bold=True, size=12, color='1F497D')
normal_font = Font(size=10)
ws_readme.column_dimensions['A'].width = 95

for row_i, (line, bold) in enumerate(README_LINES, 1):
    c = ws_readme.cell(row=row_i, column=1, value=line)
    c.font = title_font if bold else normal_font

print("[OK] README")

# ── 11. Save ──────────────────────────────────────────────────────────────────
wb.save(OUT_PATH)
print(f"[SAVED] {OUT_PATH}")
print("\n=== SAISIE_ReservationsHorsHostaway.xlsx cree avec succes ===")
