"""
lot4bis_master_calc_reservations.py
Lot 4bis — Creation MASTER_CALC_Reservations.xlsx (squelette + M-code Power Query)
- Onglet MASTER    : 24 colonnes (22 base + niveau_anomalie + code_anomalie)
- Onglet VUE_FLUX  : memes 24 colonnes — filtre statut_controle=VALIDE + impact_reel=OUI + montant non nul
- Onglet POWER_QUERY_CODE : M-code a copier dans Excel (7 requetes)

Anti-double-comptage :
  Direct HA + HH liee => 1 ligne HH (HOSTAWAY_DIRECT_HH) — ligne HA exclue
  VRBO HA + HH liee   => 1 ligne HH (HOSTAWAY_VRBO_HH)   — ligne HA exclue
  VRBO HA sans HH     => 1 ligne HA A_CONTROLER (HOSTAWAY_VRBO_A_CONTROLER)
  Airbnb / Booking HA => 1 ligne HA (HOSTAWAY_AIRBNB / HOSTAWAY_BOOKING)
  OwnerStay HA        => 1 ligne HA EXCLU_RESULTAT (OWNERSTAY_EXCLU)
  HH pure             => 1 ligne HH (MANUEL_HORS_HOSTAWAY)
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE     = "C:/Users/Ewan/OneDrive/Documents/Conciergerie/Pilotage_Conciergerie"
OUT_DIR  = f"{BASE}/02_TRAVAIL/Lot4bis_TableCommune"
OUT_PATH = f"{OUT_DIR}/MASTER_CALC_Reservations.xlsx"

os.makedirs(OUT_DIR, exist_ok=True)

# ── 24 colonnes MASTER ────────────────────────────────────────────────────────
MASTER_COLS = [
    # Bloc identification
    ('reservation_calc_id',       '1F497D'),
    ('ROW_HASH',                  '1F497D'),
    # Bloc rattachement
    ('source',                    '2E75B6'),
    ('reservation_id_hostaway',   '2E75B6'),
    ('reservation_hh_id',         '2E75B6'),
    ('mois',                      '2E75B6'),
    ('logement_id',               '2E75B6'),
    ('proprietaire_id',           '2E75B6'),
    # Bloc sejour
    ('date_arrivee',              '375623'),
    ('date_depart',               '375623'),
    ('nuits',                     '375623'),
    # Bloc financier
    ('montant_retenu',            '7030A0'),
    ('source_montant',            '7030A0'),
    # Bloc impact
    ('code_impact',               'C00000'),
    ('impact_resultat_reel',      'C00000'),
    ('impact_resultat_comptable', 'C00000'),
    # Bloc statut
    ('statut_controle',           '7B0000'),
    ('niveau_anomalie',           '7B0000'),
    ('code_anomalie',             '7B0000'),
    ('commentaire',               '595959'),
    # Bloc systeme PQ
    ('source_module',             '404040'),
    ('source_table',              '404040'),
    ('source_pk',                 '404040'),
    ('date_integration',          '404040'),
]

assert len(MASTER_COLS) == 24, f"Attendu 24 colonnes, got {len(MASTER_COLS)}"

# ── M-code Power Query ────────────────────────────────────────────────────────

MCODE_HA_RES = r"""
// Requete : HA_Reservations_Source
// Lit MASTER_FACT_HA_Reservations.xlsx onglet MASTER
// Mode : Connexion uniquement

let
    BASE = "C:\Users\Ewan\OneDrive\Documents\Conciergerie\Pilotage_Conciergerie\",
    Source = Excel.Workbook(
        File.Contents(BASE & "02_TRAVAIL\Lot1_Hostaway\MASTER_FACT_HA_Reservations.xlsx"),
        null, true
    ),
    Sheet = Source{[Item="MASTER", Kind="Sheet"]}[Data],
    Promoted = Table.PromoteHeaders(Sheet, [PromoteAllScalars=true]),
    Typed = Table.TransformColumnTypes(Promoted, {
        {"reservation_id",    type text},
        {"listingMapId",      type text},
        {"source",            type text},
        {"channel_type",      type text},
        {"source_financiere", type text},
        {"status",            type text},
        {"paymentStatus",     type text},
        {"checkInDate",       type date},
        {"checkOutDate",      type date},
        {"nights",            type number},
        {"guestCount",        type number},
        {"guestName",         type text},
        {"totalPrice",        type number},
        {"cleaningFee_res",   type number},
        {"channelCommission", type number},
        {"airbnbExpectedPayout", type number},
        {"is_ownerStay",      type text},
        {"inclure_resultat",  type text},
        {"updatedOn",         type text},
        {"createdOn",         type text},
        {"extrait_le",        type text},
        {"ROW_HASH",          type text}
    }),
    Filtered = Table.SelectRows(Typed, each
        [reservation_id] <> null and [reservation_id] <> ""
    )
in
    Filtered
""".strip()

MCODE_HA_PAY = r"""
// Requete : HA_Payout_Source
// Lit MASTER_CALC_HA_Payout.xlsx onglet MASTER
// Mode : Connexion uniquement

let
    BASE = "C:\Users\Ewan\OneDrive\Documents\Conciergerie\Pilotage_Conciergerie\",
    Source = Excel.Workbook(
        File.Contents(BASE & "02_TRAVAIL\Lot1_Hostaway\MASTER_CALC_HA_Payout.xlsx"),
        null, true
    ),
    Sheet = Source{[Item="MASTER", Kind="Sheet"]}[Data],
    Promoted = Table.PromoteHeaders(Sheet, [PromoteAllScalars=true]),
    Typed = Table.TransformColumnTypes(Promoted, {
        {"reservation_id",         type text},
        {"listingMapId",           type text},
        {"source",                 type text},
        {"channel_type",           type text},
        {"statut_calcul_payout",   type text},
        {"payout_calcule",         type number},
        {"source_payout",          type text},
        {"menage_retenu",          type number},
        {"assiette_commission",    type number},
        {"inclure_resultat_auto",  type text},
        {"extrait_le",             type text},
        {"ROW_HASH",               type text}
    }),
    Filtered = Table.SelectRows(Typed, each
        [reservation_id] <> null and [reservation_id] <> ""
    )
in
    Filtered
""".strip()

MCODE_HH = r"""
// Requete : HH_Source
// Lit SAISIE_ReservationsHorsHostaway.xlsx directement (sans dependance refresh Lot 4)
// Filtre : statut_controle = VALIDE uniquement
// Mode : Connexion uniquement

let
    BASE = "C:\Users\Ewan\OneDrive\Documents\Conciergerie\Pilotage_Conciergerie\",
    Source = Excel.Workbook(
        File.Contents(BASE & "01_SOURCES_BRUTES\ReservationsHH\SAISIE_ReservationsHorsHostaway.xlsx"),
        null, true
    ),
    Sheet = Source{[Item="SAISIE", Kind="Sheet"]}[Data],
    Promoted = Table.PromoteHeaders(Sheet, [PromoteAllScalars=true]),
    Typed = Table.TransformColumnTypes(Promoted, {
        {"reservation_hh_id",         type text},
        {"mois",                      type text},
        {"canal_id",                  type text},
        {"source_financiere",         type text},
        {"proprietaire_id",           type text},
        {"logement_id",               type text},
        {"reservation_id_hostaway",   type text},
        {"date_arrivee",              type date},
        {"date_depart",               type date},
        {"nuits",                     type number},
        {"total_percu",               type number},
        {"code_impact",               type text},
        {"impact_resultat_reel",      type text},
        {"impact_resultat_comptable", type text},
        {"statut_controle",           type text},
        {"commentaire",               type text}
    }),
    Filtered = Table.SelectRows(Typed, each
        [reservation_hh_id] <> null and [reservation_hh_id] <> ""
        and [statut_controle] = "VALIDE"
    ),
    // Colonnes utiles seulement pour Lot 4bis
    Selected = Table.SelectColumns(Filtered, {
        "reservation_hh_id", "mois", "source_financiere",
        "proprietaire_id", "logement_id", "reservation_id_hostaway",
        "date_arrivee", "date_depart", "nuits",
        "total_percu", "code_impact",
        "impact_resultat_reel", "impact_resultat_comptable",
        "statut_controle", "commentaire"
    })
in
    Selected
""".strip()

MCODE_REF_MAP = r"""
// Requete : REF_Mapping_Source
// Lit REF_Mapping_Logements depuis REF_Setup.xlsm
// Mode : Connexion uniquement

let
    BASE = "C:\Users\Ewan\OneDrive\Documents\Conciergerie\Pilotage_Conciergerie\",
    Source = Excel.Workbook(
        File.Contents(BASE & "01_SOURCES_BRUTES\REF_Setup\REF_Setup.xlsm"),
        null, true
    ),
    Sheet = Source{[Item="REF_Mapping_Logements", Kind="Sheet"]}[Data],
    Promoted = Table.PromoteHeaders(Sheet, [PromoteAllScalars=true]),
    Typed = Table.TransformColumnTypes(Promoted, {
        {"mapping_id",    type text},
        {"listingMapId",  type text},
        {"logement_id",   type text},
        {"actif",         type text}
    }),
    Actif = Table.SelectRows(Typed, each [actif] = "OUI")
in
    Actif
""".strip()

MCODE_REF_LOG = r"""
// Requete : REF_Logements_Source
// Lit REF_Logements depuis REF_Setup.xlsm
// Mode : Connexion uniquement

let
    BASE = "C:\Users\Ewan\OneDrive\Documents\Conciergerie\Pilotage_Conciergerie\",
    Source = Excel.Workbook(
        File.Contents(BASE & "01_SOURCES_BRUTES\REF_Setup\REF_Setup.xlsm"),
        null, true
    ),
    Sheet = Source{[Item="REF_Logements", Kind="Sheet"]}[Data],
    Promoted = Table.PromoteHeaders(Sheet, [PromoteAllScalars=true]),
    Typed = Table.TransformColumnTypes(Promoted, {
        {"logement_id",    type text},
        {"proprietaire_id",type text},
        {"actif",          type text}
    }),
    Actif = Table.SelectRows(Typed, each [actif] = "OUI")
in
    Actif
""".strip()

MCODE_MASTER = r"""
// Requete : MASTER_CALC_Reservations
// Table commune 24 colonnes — anti-double-comptage HA + HH
// Charger dans : onglet MASTER (Clic droit > Charger dans > Table > feuille MASTER)
//
// channel_type HA : AIRBNB / BOOKING / VRBO / DIRECT / UNKNOWN
// is_ownerStay HA : "OUI" / "NON"
// source_financiere HH : DIRECT_HA_PAYANT / VRBO_UNKNOWN / SAISIE_MANUELLE / HOSTAWAY_REFERENCE

let
    // ── 1. Sources ────────────────────────────────────────────────────────────
    HA_Res  = HA_Reservations_Source,
    HA_Pay  = HA_Payout_Source,
    HH      = HH_Source,
    RefMap  = REF_Mapping_Source,
    RefLog  = REF_Logements_Source,

    // ── 2. Comptage mappings par listingMapId (detection mapping multiple) ────
    RefMap_Count = Table.Group(
        RefMap, {"listingMapId"},
        {{"map_count", each Table.RowCount(_), type number}}
    ),

    // ── 3. IDs Hostaway lies a une ligne HH (anti-double-comptage) ───────────
    HH_Linked_IDs = List.Distinct(
        List.Select(
            Table.Column(HH, "reservation_id_hostaway"),
            each _ <> null and _ <> ""
        )
    ),

    // ── 4. JOIN HA Reservations + Payout (LEFT) ───────────────────────────────
    HA_Pay_Join = Table.NestedJoin(
        HA_Res, {"reservation_id"},
        HA_Pay, {"reservation_id"},
        "Pay", JoinKind.LeftOuter
    ),
    HA_Pay_Exp = Table.ExpandTableColumn(
        HA_Pay_Join, "Pay",
        {"payout_calcule", "statut_calcul_payout"},
        {"payout_calcule", "statut_calcul_payout"}
    ),

    // ── 5. JOIN HA + REF_Mapping (listingMapId -> logement_id) ───────────────
    HA_Map_Join = Table.NestedJoin(
        HA_Pay_Exp, {"listingMapId"},
        RefMap, {"listingMapId"},
        "Map", JoinKind.LeftOuter
    ),
    HA_Map_Exp = Table.ExpandTableColumn(
        HA_Map_Join, "Map",
        {"logement_id"},
        {"logement_id_m"}
    ),

    // ── 6. JOIN HA + RefMap_Count (detection mapping multiple) ────────────────
    HA_Cnt_Join = Table.NestedJoin(
        HA_Map_Exp, {"listingMapId"},
        RefMap_Count, {"listingMapId"},
        "Cnt", JoinKind.LeftOuter
    ),
    HA_Cnt_Exp = Table.ExpandTableColumn(
        HA_Cnt_Join, "Cnt",
        {"map_count"},
        {"map_count"}
    ),

    // ── 7. JOIN HA + REF_Logements (logement_id -> proprietaire_id) ──────────
    HA_Log_Join = Table.NestedJoin(
        HA_Cnt_Exp, {"logement_id_m"},
        RefLog, {"logement_id"},
        "Log", JoinKind.LeftOuter
    ),
    HA_Log_Exp = Table.ExpandTableColumn(
        HA_Log_Join, "Log",
        {"proprietaire_id"},
        {"proprietaire_id_m"}
    ),

    // ── 8. Filtre anti-double-comptage : retirer lignes HA liees a HH ─────────
    // Logique : si reservation_id HA existe dans reservation_id_hostaway HH → HH est la verite
    HA_Filtered = Table.SelectRows(HA_Log_Exp, each
        not List.Contains(HH_Linked_IDs, [reservation_id])
    ),

    // ── 9. Colonnes derivees branche HA ──────────────────────────────────────
    HA_Mois = Table.AddColumn(HA_Filtered, "_mois", each
        if [checkInDate] <> null
        then Date.ToText([checkInDate], "yyyy-MM")
        else "9999-99"
    , type text),

    HA_Src = Table.AddColumn(HA_Mois, "_source", each
        if [is_ownerStay] = "OUI"          then "OWNERSTAY_EXCLU"
        else if [channel_type] = "AIRBNB"  then "HOSTAWAY_AIRBNB"
        else if [channel_type] = "BOOKING" then "HOSTAWAY_BOOKING"
        else if [channel_type] = "VRBO"    then "HOSTAWAY_VRBO_A_CONTROLER"
        else if [channel_type] = "DIRECT"  then "HOSTAWAY_DIRECT_HH"
        else "A_CONTROLER"
    , type text),

    HA_CI = Table.AddColumn(HA_Src, "_code_impact", each
        if [is_ownerStay] = "OUI"                                      then "HR"
        else if [channel_type] = "AIRBNB" or [channel_type] = "BOOKING" then "IC"
        else "HC"
    , type text),

    HA_Mnt = Table.AddColumn(HA_CI, "_montant", each
        if [is_ownerStay] = "OUI" then 0.0
        else [payout_calcule]
    , type number),

    HA_SM = Table.AddColumn(HA_Mnt, "_source_montant", each
        if [is_ownerStay] = "OUI"                                      then "NON_CONCERNE"
        else if [channel_type] = "AIRBNB" or [channel_type] = "BOOKING" then "HOSTAWAY_PAYOUT"
        else "A_CONTROLER"
    , type text),

    HA_IR = Table.AddColumn(HA_SM, "_impact_reel", each
        if [_code_impact] = "IC" or [_code_impact] = "HC" then "OUI"
        else if [_code_impact] = "HR"                     then "NON"
        else "A_CONTROLER"
    , type text),

    HA_IC = Table.AddColumn(HA_IR, "_impact_compta", each
        if [_code_impact] = "IC"                               then "OUI"
        else if [_code_impact] = "HC" or [_code_impact] = "HR" then "NON"
        else "A_CONTROLER"
    , type text),

    // Controles anomalies HA (tous A_CONTROLER — BLOQUANT ajoute post-combine)
    HA_Ano = Table.AddColumn(HA_IC, "_ano_code", each
        if [map_count] = null
        then "RESERVATION_LOGEMENT_NON_MAPPE"
        else if [map_count] > 1
        then "RESERVATION_MAPPING_MULTIPLE"
        else if [channel_type] = "DIRECT"
             and [totalPrice] <> null and [totalPrice] > 0
        then "RESERVATION_HOSTAWAY_DIRECT_AVEC_MONTANT_SANS_HH"
        else if ([channel_type] = "AIRBNB" or [channel_type] = "BOOKING")
             and [is_ownerStay] = "NON"
             and ([payout_calcule] = null or [payout_calcule] = 0)
        then "RESERVATION_PAYOUT_MANQUANT"
        else null
    , type text),

    HA_Niv = Table.AddColumn(HA_Ano, "_niveau_ano", each
        if [_ano_code] = null then null else "A_CONTROLER"
    , type text),

    HA_Stat = Table.AddColumn(HA_Niv, "_statut", each
        if [is_ownerStay] = "OUI"   then "EXCLU_RESULTAT"
        else if [_ano_code] <> null then "A_CONTROLER"
        else "VALIDE"
    , type text),

    // ── 10. Indexer HA par mois (PK counter reset /mois) ─────────────────────
    HA_Sort = Table.Sort(HA_Stat, {{"_mois", Order.Ascending}, {"reservation_id", Order.Ascending}}),
    HA_Grp  = Table.Group(HA_Sort, {"_mois"}, {
        {"rows", each Table.AddIndexColumn(_, "_idx", 1, 1), type table}
    }),
    HA_Ung  = Table.Combine(HA_Grp[rows]),
    HA_PK   = Table.AddColumn(HA_Ung, "_calc_id",
        each "RES-" & [_mois] & "-HA-" & Text.PadStart(Text.From([_idx]), 3, "0")
    , type text),

    // ── 11. Selection + renommage colonnes HA → 24 standard ──────────────────
    HA_Sel  = Table.SelectColumns(HA_PK, {
        "_calc_id", "_source",
        "reservation_id", "_mois",
        "logement_id_m", "proprietaire_id_m",
        "checkInDate", "checkOutDate", "nights",
        "_montant", "_source_montant",
        "_code_impact", "_impact_reel", "_impact_compta",
        "_statut", "_niveau_ano", "_ano_code"
    }),
    HA_Ren  = Table.RenameColumns(HA_Sel, {
        {"_calc_id",          "reservation_calc_id"},
        {"_source",           "source"},
        {"reservation_id",    "reservation_id_hostaway"},
        {"_mois",             "mois"},
        {"logement_id_m",     "logement_id"},
        {"proprietaire_id_m", "proprietaire_id"},
        {"checkInDate",       "date_arrivee"},
        {"checkOutDate",      "date_depart"},
        {"nights",            "nuits"},
        {"_montant",          "montant_retenu"},
        {"_source_montant",   "source_montant"},
        {"_code_impact",      "code_impact"},
        {"_impact_reel",      "impact_resultat_reel"},
        {"_impact_compta",    "impact_resultat_comptable"},
        {"_statut",           "statut_controle"},
        {"_niveau_ano",       "niveau_anomalie"},
        {"_ano_code",         "code_anomalie"}
    }),
    HA_A1  = Table.AddColumn(HA_Ren,  "reservation_hh_id",   each null,                           type text),
    HA_A2  = Table.AddColumn(HA_A1,   "ROW_HASH",            each null,                           type text),
    HA_A3  = Table.AddColumn(HA_A2,   "commentaire",         each null,                           type text),
    HA_A4  = Table.AddColumn(HA_A3,   "source_module",       each "LOT4BIS_CALC_RESERVATIONS",    type text),
    HA_A5  = Table.AddColumn(HA_A4,   "source_table",        each "MASTER_FACT_HA_Reservations",  type text),
    HA_A6  = Table.AddColumn(HA_A5,   "source_pk",           each [reservation_id_hostaway],      type text),
    HA_A7  = Table.AddColumn(HA_A6,   "date_integration",    each DateTime.LocalNow(),            type datetime),

    // ── 12. Colonnes derivees branche HH ─────────────────────────────────────
    HH_Src  = Table.AddColumn(HH, "_source_hh", each
        if [source_financiere] = "DIRECT_HA_PAYANT" then "HOSTAWAY_DIRECT_HH"
        else if [source_financiere] = "VRBO_UNKNOWN" then "HOSTAWAY_VRBO_HH"
        else "MANUEL_HORS_HOSTAWAY"
    , type text),

    HH_SM   = Table.AddColumn(HH_Src, "_sm_hh", each
        if [source_financiere] = "VRBO_UNKNOWN" then "MANUEL_VRBO"
        else "MANUEL_HH"
    , type text),

    // Controle RESERVATION_HH_NON_VALIDE : ne devrait pas arriver (HH_Source filtre VALIDE)
    // conserve pour securite en cas de lecture depuis MASTER non filtre
    HH_Ano  = Table.AddColumn(HH_SM, "_ano_hh", each
        if [statut_controle] <> "VALIDE" then "RESERVATION_HH_NON_VALIDE" else null
    , type text),

    HH_Niv  = Table.AddColumn(HH_Ano, "_niv_hh", each
        if [_ano_hh] <> null then "A_CONTROLER" else null
    , type text),

    // ── 13. Indexer HH par mois ───────────────────────────────────────────────
    HH_Sort = Table.Sort(HH_Niv, {{"mois", Order.Ascending}, {"reservation_hh_id", Order.Ascending}}),
    HH_Grp  = Table.Group(HH_Sort, {"mois"}, {
        {"rows", each Table.AddIndexColumn(_, "_idx", 1, 1), type table}
    }),
    HH_Ung  = Table.Combine(HH_Grp[rows]),
    HH_PK   = Table.AddColumn(HH_Ung, "_calc_id_hh",
        each "RES-" & [mois] & "-HH-" & Text.PadStart(Text.From([_idx]), 3, "0")
    , type text),

    // ── 14. Selection + renommage colonnes HH → 24 standard ──────────────────
    HH_Sel  = Table.SelectColumns(HH_PK, {
        "_calc_id_hh", "_source_hh",
        "reservation_id_hostaway", "reservation_hh_id",
        "mois", "logement_id", "proprietaire_id",
        "date_arrivee", "date_depart", "nuits",
        "total_percu", "_sm_hh",
        "code_impact", "impact_resultat_reel", "impact_resultat_comptable",
        "statut_controle", "_niv_hh", "_ano_hh",
        "commentaire"
    }),
    HH_Ren  = Table.RenameColumns(HH_Sel, {
        {"_calc_id_hh",  "reservation_calc_id"},
        {"_source_hh",   "source"},
        {"total_percu",  "montant_retenu"},
        {"_sm_hh",       "source_montant"},
        {"_niv_hh",      "niveau_anomalie"},
        {"_ano_hh",      "code_anomalie"}
    }),
    HH_A1   = Table.AddColumn(HH_Ren, "ROW_HASH",         each null,                                       type text),
    HH_A2   = Table.AddColumn(HH_A1,  "source_module",    each "LOT4BIS_CALC_RESERVATIONS",                type text),
    HH_A3   = Table.AddColumn(HH_A2,  "source_table",     each "MASTER_FACT_MAN_ReservationsHorsHostaway", type text),
    HH_A4   = Table.AddColumn(HH_A3,  "source_pk",        each [reservation_hh_id],                        type text),
    HH_A5   = Table.AddColumn(HH_A4,  "date_integration", each DateTime.LocalNow(),                        type datetime),

    // ── 15. Empilement HA + HH ────────────────────────────────────────────────
    COL_STD = {
        "reservation_calc_id", "ROW_HASH", "source",
        "reservation_id_hostaway", "reservation_hh_id", "mois",
        "logement_id", "proprietaire_id",
        "date_arrivee", "date_depart", "nuits",
        "montant_retenu", "source_montant",
        "code_impact", "impact_resultat_reel", "impact_resultat_comptable",
        "statut_controle", "niveau_anomalie", "code_anomalie", "commentaire",
        "source_module", "source_table", "source_pk", "date_integration"
    },
    HA_Std    = Table.ReorderColumns(Table.SelectColumns(HA_A7, COL_STD), COL_STD),
    HH_Std    = Table.ReorderColumns(Table.SelectColumns(HH_A5, COL_STD), COL_STD),
    Combined  = Table.Combine({HA_Std, HH_Std}),

    // ── 16. Controle BLOQUANT : RESERVATION_DOUBLON_HOSTAWAY_HH ──────────────
    // reservation_id_hostaway non vide rattache a 2+ lignes actives (hors EXCLU_RESULTAT)
    Active_Lines = Table.SelectRows(Combined, each
        [statut_controle] <> "EXCLU_RESULTAT"
        and [reservation_id_hostaway] <> null
        and [reservation_id_hostaway] <> ""
    ),
    Doublon_Cnt = Table.Group(Active_Lines, {"reservation_id_hostaway"}, {
        {"cnt", each Table.RowCount(_), type number}
    }),
    Doublon_IDs = List.Transform(
        Table.Column(
            Table.SelectRows(Doublon_Cnt, each [cnt] >= 2),
            "reservation_id_hostaway"
        ),
        each _
    ),

    // ── 17. Controle BLOQUANT : RESERVATION_CALC_ID_DUPLIQUE ─────────────────
    CalcID_Cnt  = Table.Group(Combined, {"reservation_calc_id"}, {
        {"cnt2", each Table.RowCount(_), type number}
    }),
    DupCalcIDs  = List.Transform(
        Table.Column(
            Table.SelectRows(CalcID_Cnt, each [cnt2] >= 2),
            "reservation_calc_id"
        ),
        each _
    ),

    // ── 18. Flags temporaires ─────────────────────────────────────────────────
    F1 = Table.AddColumn(Combined, "_is_ha_doublon", each
        [reservation_id_hostaway] <> null
        and [reservation_id_hostaway] <> ""
        and List.Contains(Doublon_IDs, [reservation_id_hostaway])
    , type logical),
    F2 = Table.AddColumn(F1, "_is_id_dup", each
        List.Contains(DupCalcIDs, [reservation_calc_id])
    , type logical),

    // ── 19. Anomalies finales (priorite : ID_DUP > HA_DOUBLON > existant) ────
    F3  = Table.AddColumn(F2, "_ano_final", each
        if [_is_id_dup]       then "RESERVATION_CALC_ID_DUPLIQUE"
        else if [_is_ha_doublon] then "RESERVATION_DOUBLON_HOSTAWAY_HH"
        else [code_anomalie]
    , type text),
    F4  = Table.AddColumn(F3, "_niv_final", each
        if [_is_id_dup] or [_is_ha_doublon] then "BLOQUANT"
        else [niveau_anomalie]
    , type text),
    F5  = Table.AddColumn(F4, "_stat_final", each
        if [_is_id_dup] or [_is_ha_doublon] then "A_CONTROLER"
        else [statut_controle]
    , type text),

    // ── 20. Nettoyage et ordre final 24 colonnes ──────────────────────────────
    Cleaned = Table.RemoveColumns(F5, {
        "statut_controle", "niveau_anomalie", "code_anomalie",
        "_is_ha_doublon", "_is_id_dup"
    }),
    Renamed = Table.RenameColumns(Cleaned, {
        {"_ano_final",  "code_anomalie"},
        {"_niv_final",  "niveau_anomalie"},
        {"_stat_final", "statut_controle"}
    }),
    Final   = Table.ReorderColumns(Renamed, COL_STD)
in
    Final
""".strip()

MCODE_VUE = r"""
// Requete : VUE_FLUX_Reservations
// Filtre MASTER_CALC_Reservations :
//   statut_controle = VALIDE
//   impact_resultat_reel = OUI
//   montant_retenu non nul et non zero
// Exclut : EXCLU_RESULTAT / A_CONTROLER / ownerStay / VRBO non renseigne
// Alimente : Lot 9 (MASTER_CALC_Flux)
// Charger dans : onglet VUE_FLUX (Clic droit > Charger dans > Table > feuille VUE_FLUX)

let
    Source = MASTER_CALC_Reservations,
    Filtre = Table.SelectRows(Source, each
        [statut_controle]      = "VALIDE"
        and [impact_resultat_reel] = "OUI"
        and [montant_retenu]   <> null
        and [montant_retenu]   <> 0
    )
in
    Filtre
""".strip()

INSTRUCTIONS = """
INSTRUCTIONS POWER QUERY — MASTER_CALC_Reservations.xlsx
==========================================================
24 colonnes — Anti-double-comptage HA + HH

PREREQUIS :
1. MASTER_FACT_HA_Reservations.xlsx ouvert et onglet MASTER refresh (Lot 1).
2. MASTER_CALC_HA_Payout.xlsx ouvert et onglet MASTER refresh (Lot 1).
3. SAISIE_ReservationsHorsHostaway.xlsx saisie et validee (Lot 4).
4. REF_Setup.xlsm ouvert (REF_Mapping_Logements + REF_Logements actifs).

CREER 7 REQUETES dans cet ordre exact :

Req 1 : HA_Reservations_Source        — Mode : Connexion uniquement
Req 2 : HA_Payout_Source              — Mode : Connexion uniquement
Req 3 : HH_Source                     — Mode : Connexion uniquement
Req 4 : REF_Mapping_Source            — Mode : Connexion uniquement
Req 5 : REF_Logements_Source          — Mode : Connexion uniquement
Req 6 : MASTER_CALC_Reservations      — Charger dans > Table > feuille MASTER
Req 7 : VUE_FLUX_Reservations         — Charger dans > Table > feuille VUE_FLUX

Pour chaque requete :
  Donnees > Obtenir des donnees > Editeur Power Query
  Accueil > Nouvelle source > Requete vide
  Coller le M-code > OK > Fermer et charger selon mode indique

NOTES :
- Adapter BASE dans chaque requete si changement de lecteur ou utilisateur.
- channel_type HA attendus : AIRBNB / BOOKING / VRBO / DIRECT / UNKNOWN.
- Si valeurs differentes dans vos donnees HA, adapter les blocs if/else dans MASTER_CALC_Reservations.
- BLOQUANTS detectes post-empilement : RESERVATION_DOUBLON_HOSTAWAY_HH + RESERVATION_CALC_ID_DUPLIQUE.
- VUE_FLUX : uniquement lignes exploitables pour Lot 9 (VALIDE + impact_reel=OUI + montant non nul).
""".strip()

# ── Workbook ──────────────────────────────────────────────────────────────────
wb          = Workbook()
ws_master   = wb.active
ws_master.title = 'MASTER'
ws_vue      = wb.create_sheet('VUE_FLUX')
ws_pq       = wb.create_sheet('POWER_QUERY_CODE')

H_FONT = Font(bold=True, color='FFFFFF', size=10)

def write_headers(ws, cols):
    for col_i, (col_name, color) in enumerate(cols, 1):
        c = ws.cell(row=1, column=col_i, value=col_name)
        c.font = H_FONT
        c.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(col_i)].width = max(len(col_name) + 2, 14)
    ws.row_dimensions[1].height = 42
    ws.freeze_panes = 'A2'

# MASTER — 24 colonnes
write_headers(ws_master, MASTER_COLS)
ws_master.cell(row=2, column=1,
    value='[Charge par Power Query — requete MASTER_CALC_Reservations]'
).font = Font(italic=True, color='808080')
print(f"[OK] MASTER — {len(MASTER_COLS)} colonnes")

# VUE_FLUX — memes 24 colonnes
write_headers(ws_vue, MASTER_COLS)
ws_vue.cell(row=2, column=1,
    value='[Vue par Power Query — requete VUE_FLUX_Reservations — VALIDE + impact_reel=OUI + montant<>0]'
).font = Font(italic=True, color='808080')
print(f"[OK] VUE_FLUX — {len(MASTER_COLS)} colonnes")

# POWER_QUERY_CODE
ws_pq.column_dimensions['A'].width = 40
ws_pq.column_dimensions['C'].width = 120

title_font = Font(bold=True, size=12, color='1F497D')
code_font  = Font(name='Courier New', size=9, color='262626')
note_font  = Font(italic=True, size=10, color='595959')

ws_pq.cell(row=1, column=1,
    value='POWER QUERY M-CODE — MASTER_CALC_Reservations').font = title_font
ws_pq.cell(row=2, column=1,
    value='Ne pas modifier — reference uniquement').font = note_font

row_off = 4
for line in INSTRUCTIONS.split('\n'):
    ws_pq.cell(row=row_off, column=1, value=line).font = Font(size=9)
    row_off += 1
row_off += 2

for label, mcode in [
    ('REQUETE 1 : HA_Reservations_Source',   MCODE_HA_RES),
    ('REQUETE 2 : HA_Payout_Source',         MCODE_HA_PAY),
    ('REQUETE 3 : HH_Source',                MCODE_HH),
    ('REQUETE 4 : REF_Mapping_Source',       MCODE_REF_MAP),
    ('REQUETE 5 : REF_Logements_Source',     MCODE_REF_LOG),
    ('REQUETE 6 : MASTER_CALC_Reservations', MCODE_MASTER),
    ('REQUETE 7 : VUE_FLUX_Reservations',    MCODE_VUE),
]:
    ws_pq.cell(row=row_off, column=1, value=label).font = Font(bold=True, size=10, color='7030A0')
    row_off += 1
    for line in mcode.split('\n'):
        ws_pq.cell(row=row_off, column=3, value=line).font = code_font
        ws_pq.row_dimensions[row_off].height = 14
        row_off += 1
    row_off += 2

print("[OK] POWER_QUERY_CODE — 7 requetes M-code")

wb.save(OUT_PATH)
print(f"[SAVED] {OUT_PATH}")
print(f"\n=== MASTER_CALC_Reservations.xlsx cree avec succes ({len(MASTER_COLS)} colonnes) ===")
