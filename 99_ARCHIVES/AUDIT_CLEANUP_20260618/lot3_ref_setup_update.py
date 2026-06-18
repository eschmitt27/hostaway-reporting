"""
lot3_ref_setup_update.py
Lot 3 — Mise à jour REF_Setup.xlsm
- Backup obligatoire
- REF_Categories_Charges : ajout colonne filtre_vue_menage + CHG_021/022/023
- REF_Types_Flux : ajout TYPE_FLUX_009 à 012
- REF_Types_Affectation : ajout AFF_GLOBAL + AFF_NON_AFFECTABLE
- REF_Statuts : désactivation STAT_022 + ajout STAT_024-029
- REF_Charges_Recurrentes : nouvel onglet + 2 lignes REC_001/002
"""

import openpyxl
import shutil
import os
from datetime import datetime, date

BASE = "C:/Users/Ewan/OneDrive/Documents/Conciergerie/Pilotage_Conciergerie"
REF_PATH = f"{BASE}/01_SOURCES_BRUTES/REF_Setup/REF_Setup.xlsm"
ARCHIVE_DIR = f"{BASE}/99_ARCHIVES/LOT3_Charges"
TS = datetime.now().strftime("%Y%m%d_%H%M%S")
BACKUP = f"{ARCHIVE_DIR}/REF_Setup_BACKUP_{TS}.xlsm"

# ── 1. BACKUP ──────────────────────────────────────────────────────────────────
os.makedirs(ARCHIVE_DIR, exist_ok=True)
shutil.copy2(REF_PATH, BACKUP)
print(f"[BACKUP] {BACKUP}")

# ── 2. LOAD ────────────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(REF_PATH, keep_vba=True)

# ── 3. REF_Categories_Charges ─────────────────────────────────────────────────
ws = wb['REF_Categories_Charges']

# Ajout en-tête colonne 9
ws.cell(row=1, column=9, value='filtre_vue_menage')

# filtre_vue_menage par ID (OUI uniquement pour CHG_003/004/018/023)
VUE = {'CHG_003': 'OUI', 'CHG_004': 'OUI', 'CHG_018': 'OUI'}

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    cid = row[0].value
    if cid:
        ws.cell(row=row[0].row, column=9, value=VUE.get(str(cid), 'NON'))

# 3 nouvelles catégories
new_cats = [
    ('CHG_021', 'Incident', 'Incident voyageur',
     "Charge liée à un incident voyageur (dégradation, réclamation, compensation, "
     "intervention urgente). reservation_id obligatoire (D041).",
     'OUI', 'A_DEFINIR', 'NON', 'OUI', 'NON'),
    ('CHG_022', 'AirCover', 'Prestation AirCover refacturée',
     "Prestation conciergerie dans le cadre d'un sinistre AirCover, facturée au "
     "propriétaire. Alimente charges_exceptionnelles_refacturees (D042).",
     'OUI', 'OUI', 'NON', 'OUI', 'NON'),
    ('CHG_023', 'Ménage', 'Forfait local cave',
     "Loyer mensuel de la cave, à répartir dans le coût des ménages internes selon "
     "la clé NOMBRE_MENAGES. Alimente le coût complet ménage interne via "
     "VUE_ACHATS_MENAGE_VALIDES.",
     'OUI', 'NON', 'OUI', 'OUI', 'OUI'),
]
r = ws.max_row + 1
for cat in new_cats:
    for c, v in enumerate(cat, 1):
        ws.cell(row=r, column=c, value=v)
    r += 1
print("[OK] REF_Categories_Charges — filtre_vue_menage + CHG_021/022/023")

# ── 4. REF_Types_Flux ─────────────────────────────────────────────────────────
ws = wb['REF_Types_Flux']
# Colonnes: type_flux_id, type_flux, description, code_impact_defaut,
#           avantage_brut_defaut, deduit_avantage_defaut, comptabilisable_defaut,
#           actif, commentaire
new_types = [
    ('TYPE_FLUX_009', 'ACHAT_MENAGE',
     "Achat lié à l'activité ménage (linge, produits, consommables, matériel)",
     'HC', 'NON', 'NON', 'NON', 'OUI',
     "Éligible VUE_ACHATS_MENAGE_VALIDES si filtre_vue_menage=OUI"),
    ('TYPE_FLUX_010', 'FRAIS_LOCAL',
     "Charge locale interne récurrente (loyer cave, accès local) répartie par clé ménage",
     'HC', 'NON', 'NON', 'NON', 'OUI',
     "Coût interne HC, non refacturable, alimente coût complet ménage interne"),
    ('TYPE_FLUX_011', 'CHARGE_EXCEPTIONNELLE_REFACTURABLE',
     "Charge exceptionnelle et ponctuelle refacturée au propriétaire",
     'IC', 'NON', 'NON', 'OUI', 'OUI',
     "Impact montant_du_conciergerie uniquement — jamais revenu_net_exploitation (EP7)"),
    ('TYPE_FLUX_012', 'CHARGE_RECURRENTE_REFACTURABLE',
     "Charge récurrente et prévisible refacturée au propriétaire "
     "(forfait mensuel logiciel, consommables)",
     'IC', 'NON', 'NON', 'OUI', 'OUI',
     "Distinct de TYPE_FLUX_011 (exceptionnel/ponctuel). Portée par REF_Charges_Recurrentes."),
]
r = ws.max_row + 1
for t in new_types:
    for c, v in enumerate(t, 1):
        ws.cell(row=r, column=c, value=v)
    r += 1
print("[OK] REF_Types_Flux — TYPE_FLUX_009 à 012")

# ── 5. REF_Types_Affectation ──────────────────────────────────────────────────
ws = wb['REF_Types_Affectation']
# Colonnes: affectation_id, type_affectation, description,
#           logement_obligatoire, proprietaire_obligatoire, actif, commentaire
new_aff = [
    ('AFF_GLOBAL', 'GLOBAL',
     "Charge non rattachée à un logement ou propriétaire précis — activité globale",
     'NON', 'NON', 'OUI', ''),
    ('AFF_NON_AFFECTABLE', 'NON_AFFECTABLE',
     "Charge non affectable (personnelle, mixte, provisoire)",
     'NON', 'NON', 'OUI', ''),
]
r = ws.max_row + 1
for a in new_aff:
    for c, v in enumerate(a, 1):
        ws.cell(row=r, column=c, value=v)
    r += 1
print("[OK] REF_Types_Affectation — AFF_GLOBAL + AFF_NON_AFFECTABLE")

# ── 6. REF_Statuts ────────────────────────────────────────────────────────────
ws = wb['REF_Statuts']
# Colonnes: statut_id, famille_statut, statut, ordre_affichage, actif, commentaire

# Désactivation STAT_022
for row in ws.iter_rows(min_row=2):
    if row[0].value == 'STAT_022':
        ws.cell(row=row[0].row, column=5, value='NON')
        ws.cell(row=row[0].row, column=6,
                value='Déplacé vers niveau_anomalie — voir STAT_029 (D044)')
        print(f"  STAT_022 désactivé (ligne {row[0].row})")
        break

new_stats = [
    ('STAT_024', 'statut_controle', 'A_CONTROLER', 2, 'OUI',
     "Ligne en attente de vérification — non incluse dans le résultat final"),
    ('STAT_025', 'statut_controle', 'EXCLU_RESULTAT', 6, 'OUI',
     "Ligne exclue du résultat avec justification (équivalent IGNORE_JUSTIFIE, "
     "Lot 3+) — D044"),
    ('STAT_026', 'statut_controle', 'A_VENTILER', 7, 'OUI',
     "Ligne à ventiler entre logements ou périodes avant intégration"),
    ('STAT_027', 'niveau_anomalie', 'INFO', 1, 'OUI',
     "Anomalie informative — aucun impact calcul"),
    ('STAT_028', 'niveau_anomalie', 'A_CONTROLER', 2, 'OUI',
     "Anomalie à vérifier — calcul provisoire possible"),
    ('STAT_029', 'niveau_anomalie', 'BLOQUANT', 3, 'OUI',
     "Anomalie bloquante — ligne non intégrable en l'état"),
]
r = ws.max_row + 1
for s in new_stats:
    for c, v in enumerate(s, 1):
        ws.cell(row=r, column=c, value=v)
    r += 1
print("[OK] REF_Statuts — STAT_022 désactivé + STAT_024-029")

# ── 7. REF_Charges_Recurrentes (nouveau onglet) ───────────────────────────────
ws_rec = wb.create_sheet('REF_Charges_Recurrentes')

headers_rec = [
    'charge_recurrente_id', 'libelle_charge_recurrente', 'categorie_charge_id',
    'type_flux_id', 'montant_ttc', 'devise', 'periodicite',
    'date_debut_validite', 'date_fin_validite', 'affectation_type',
    'logement_id', 'proprietaire_id', 'cle_repartition',
    'code_impact_defaut', 'prise_en_compta_defaut', 'refacturable_defaut',
    'filtre_vue_menage', 'actif', 'commentaire',
]
for c, h in enumerate(headers_rec, 1):
    ws_rec.cell(row=1, column=c, value=h)

rec_data = [
    ('REC_001', 'Forfait client logiciel et consommables', 'CHG_016',
     'TYPE_FLUX_012', None, 'EUR', 'MENSUEL',
     date(2026, 1, 1), None, 'LOGEMENT',
     None, None, 'AUCUNE',
     'IC', 'OUI', 'OUI', 'NON', 'OUI',
     'Facturé mensuellement au propriétaire — logiciels et consommables refacturés'),
    ('REC_002', 'Forfait local cave', 'CHG_023',
     'TYPE_FLUX_010', 50.0, 'EUR', 'MENSUEL',
     date(2026, 1, 1), None, 'GLOBAL',
     None, None, 'NOMBRE_MENAGES',
     'HC', 'NON', 'NON', 'OUI', 'OUI',
     'Loyer cave réparti par nombre de ménages — coût complet ménage interne HC'),
]
for ri, row_data in enumerate(rec_data, 2):
    for ci, v in enumerate(row_data, 1):
        ws_rec.cell(row=ri, column=ci, value=v)
print("[OK] REF_Charges_Recurrentes — onglet créé + REC_001/002")

# ── 8. SAVE ────────────────────────────────────────────────────────────────────
wb.save(REF_PATH)
print(f"[SAVED] {REF_PATH}")
print("\n=== REF_Setup update terminé ===")
