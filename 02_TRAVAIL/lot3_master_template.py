"""
lot3_master_template.py
Lot 3 — Creation MASTER_FACT_MAN_Charges.xlsx (squelette + M-code Power Query)
- Onglet MASTER          : 37 colonnes (31 SAISIE + 6 colonnes PQ)
- Onglet VUE_MENAGE      : memes 37 colonnes — filtre statut_controle=VALIDE ET filtre_vue_menage=OUI
- Onglet POWER_QUERY_CODE : M-code a copier dans Excel

Colonnes PQ ajoutees (6) :
  sens              — derive de sens_flux (DEPENSE→CHARGE, RECUPERATION→PRODUIT, etc.)
  filtre_vue_menage — lookup REF_Categories_Charges
  source_module     — constante LOT3_CHARGES
  source_table      — constante SAISIE_Charges_Flux
  source_pk         — = charge_id
  date_integration  — DateTime.LocalNow() a chaque refresh PQ
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = "C:/Users/Ewan/OneDrive/Documents/Conciergerie/Pilotage_Conciergerie"
OUT_PATH = f"{BASE}/02_TRAVAIL/Lot3_Charges/MASTER_FACT_MAN_Charges.xlsx"

# ── 37 colonnes MASTER ────────────────────────────────────────────────────────
# (nom, couleur_groupe)
# Marquage [PQ] = colonne ajoutee par Power Query (pas dans SAISIE)
MASTER_COLS = [
    # Groupe 1 — Identification
    ('charge_id',                  '1F497D'),
    ('date_charge',                '1F497D'),
    ('mois',                       '1F497D'),
    # Groupe 2 — Montant + sens
    ('montant',                    '375623'),
    ('sens_flux',                  '375623'),
    ('sens',                       '375623'),   # [PQ] DEPENSE→CHARGE, RECUPERATION→PRODUIT, etc.
    # Groupe 3 — Classification
    ('categorie_charge_id',        '7030A0'),
    ('filtre_vue_menage',          '7030A0'),   # [PQ] lookup REF_Categories_Charges
    ('type_flux_id',               '7030A0'),
    # Groupe 4 — Impact resultat
    ('code_impact',                'C00000'),
    ('impact_resultat_reel',       'C00000'),
    ('impact_resultat_comptable',  'C00000'),
    ('prise_en_compta',            'C00000'),
    # Groupe 5 — Qui paie
    ('associe_id',                 '974706'),
    ('mode_paiement_id',           '974706'),
    ('carte_id',                   '974706'),
    # Groupe 6 — Affectation
    ('affectation_type',           '1F3864'),
    ('logement_id',                '1F3864'),
    ('proprietaire_id',            '1F3864'),
    # Groupe 7 — Conditions speciales
    ('reservation_id',             '833C00'),
    ('refacturable',               '833C00'),
    # Groupe 8 — Tracabilite saisie
    ('source_flux',                '595959'),
    ('methode_traitement',         '595959'),
    ('paye_avec_montant_recupere', '595959'),
    ('lien_virement_banque',       '595959'),
    # Groupe 9 — Controle
    ('statut_controle',            '7B0000'),
    ('niveau_anomalie',            '7B0000'),
    ('code_anomalie',              '7B0000'),
    ('statut_rapprochement',       '7B0000'),
    # Groupe 10 — Documents
    ('justificatif',               '595959'),
    ('commentaire',                '595959'),
    # Groupe 11 — Systeme saisie
    ('ROW_HASH',                   '262626'),
    ('date_saisie',                '262626'),
    # Groupe 12 — Systeme integration PQ [PQ]
    ('source_module',              '404040'),   # [PQ] = LOT3_CHARGES
    ('source_table',               '404040'),   # [PQ] = SAISIE_Charges_Flux
    ('source_pk',                  '404040'),   # [PQ] = charge_id
    ('date_integration',           '404040'),   # [PQ] = DateTime.LocalNow()
]

assert len(MASTER_COLS) == 37, f"Attendu 37 colonnes, got {len(MASTER_COLS)}"

# ── M-code Power Query ────────────────────────────────────────────────────────

MCODE_SAISIE = r"""
// Requete : SAISIE_Source
// Lit SAISIE_Charges_Flux.xlsx, typage fort, filtre lignes vides

let
    BASE_PATH = "C:\Users\Ewan\OneDrive\Documents\Conciergerie\Pilotage_Conciergerie\",
    Source = Excel.Workbook(
        File.Contents(BASE_PATH & "01_SOURCES_BRUTES\Charges\SAISIE_Charges_Flux.xlsx"),
        null, true
    ),
    SAISIE_Sheet = Source{[Item="SAISIE",Kind="Sheet"]}[Data],
    Promoted = Table.PromoteHeaders(SAISIE_Sheet, [PromoteAllScalars=true]),
    Typed = Table.TransformColumnTypes(Promoted, {
        {"charge_id",                  type text},
        {"date_charge",                type date},
        {"mois",                       type text},
        {"montant",                    type number},
        {"sens_flux",                  type text},
        {"categorie_charge_id",        type text},
        {"type_flux_id",               type text},
        {"code_impact",                type text},
        {"impact_resultat_reel",       type text},
        {"impact_resultat_comptable",  type text},
        {"prise_en_compta",            type text},
        {"associe_id",                 type text},
        {"mode_paiement_id",           type text},
        {"carte_id",                   type text},
        {"affectation_type",           type text},
        {"logement_id",                type text},
        {"proprietaire_id",            type text},
        {"reservation_id",             type text},
        {"refacturable",               type text},
        {"source_flux",                type text},
        {"methode_traitement",         type text},
        {"paye_avec_montant_recupere", type text},
        {"lien_virement_banque",       type text},
        {"statut_controle",            type text},
        {"niveau_anomalie",            type text},
        {"code_anomalie",              type text},
        {"statut_rapprochement",       type text},
        {"justificatif",               type text},
        {"commentaire",                type text},
        {"ROW_HASH",                   type text},
        {"date_saisie",                type date}
    }),
    Filtered = Table.SelectRows(Typed, each
        [charge_id] <> null and [charge_id] <> ""
    )
in
    Filtered
""".strip()

MCODE_REF_CATS = r"""
// Requete : REF_Categories_Source
// Lit REF_Categories_Charges depuis REF_Setup.xlsm
// Mode : Connexion uniquement (ne pas charger dans une feuille)

let
    BASE_PATH = "C:\Users\Ewan\OneDrive\Documents\Conciergerie\Pilotage_Conciergerie\",
    Source = Excel.Workbook(
        File.Contents(BASE_PATH & "01_SOURCES_BRUTES\REF_Setup\REF_Setup.xlsm"),
        null, true
    ),
    Sheet = Source{[Item="REF_Categories_Charges",Kind="Sheet"]}[Data],
    Promoted = Table.PromoteHeaders(Sheet, [PromoteAllScalars=true]),
    Selected = Table.SelectColumns(Promoted, {"categorie_charge_id", "filtre_vue_menage"})
in
    Selected
""".strip()

MCODE_MASTER = r"""
// Requete : MASTER_FACT_MAN_Charges
// 37 colonnes = 31 SAISIE + 6 colonnes PQ
// Charger dans : onglet MASTER (Load To > Table > feuille existante MASTER)
// Mode : Connexion uniquement pour SAISIE_Source et REF_Categories_Source

let
    Source = SAISIE_Source,

    // (1) LEFT JOIN pour filtre_vue_menage (REF_Categories_Charges)
    Joined = Table.NestedJoin(
        Source, {"categorie_charge_id"},
        REF_Categories_Source, {"categorie_charge_id"},
        "_REF", JoinKind.LeftOuter
    ),
    Expanded = Table.ExpandTableColumn(Joined, "_REF", {"filtre_vue_menage"}),

    // (2) Colonne 'sens' derivee de 'sens_flux'
    //     DEPENSE        -> CHARGE
    //     RECUPERATION   -> PRODUIT
    //     REMBOURSEMENT  -> NEUTRALISATION
    //     REFACTURATION  -> CHARGE  (charge avancee puis recuperee sur proprietaire)
    //     NEUTRE         -> NEUTRALISATION
    AddSens = Table.AddColumn(Expanded, "sens", each
        if      [sens_flux] = "DEPENSE"       then "CHARGE"
        else if [sens_flux] = "RECUPERATION"  then "PRODUIT"
        else if [sens_flux] = "REMBOURSEMENT" then "NEUTRALISATION"
        else if [sens_flux] = "REFACTURATION" then "CHARGE"
        else if [sens_flux] = "NEUTRE"        then "NEUTRALISATION"
        else null,
        type text
    ),

    // (3) Colonnes systeme d'integration
    AddSourceModule = Table.AddColumn(AddSens,        "source_module",    each "LOT3_CHARGES",       type text),
    AddSourceTable  = Table.AddColumn(AddSourceModule, "source_table",     each "SAISIE_Charges_Flux", type text),
    AddSourcePK     = Table.AddColumn(AddSourceTable,  "source_pk",        each [charge_id],           type text),
    AddDateInteg    = Table.AddColumn(AddSourcePK,     "date_integration", each DateTime.LocalNow(),   type datetime),

    // (4) Reordonnement final — 37 colonnes
    Reordered = Table.ReorderColumns(AddDateInteg, {
        "charge_id", "date_charge", "mois",
        "montant", "sens_flux", "sens",
        "categorie_charge_id", "filtre_vue_menage", "type_flux_id",
        "code_impact", "impact_resultat_reel", "impact_resultat_comptable",
        "prise_en_compta",
        "associe_id", "mode_paiement_id", "carte_id",
        "affectation_type", "logement_id", "proprietaire_id",
        "reservation_id", "refacturable",
        "source_flux", "methode_traitement",
        "paye_avec_montant_recupere", "lien_virement_banque",
        "statut_controle", "niveau_anomalie", "code_anomalie", "statut_rapprochement",
        "justificatif", "commentaire",
        "ROW_HASH", "date_saisie",
        "source_module", "source_table", "source_pk", "date_integration"
    })
in
    Reordered
""".strip()

MCODE_VUE = r"""
// Requete : VUE_ACHATS_MENAGE_VALIDES
// Filtre MASTER : filtre_vue_menage=OUI ET statut_controle=VALIDE
// 37 colonnes identiques a MASTER (pas de suppression de colonnes)
// Charger dans : onglet VUE_MENAGE (Load To > Table > feuille existante VUE_MENAGE)
// Usage : cout complet menage interne — agregation par logement/mois (D028)
// Anti-double-comptage : ces lignes NE figurent pas dans M04 (D027/D038)

let
    Source = MASTER_FACT_MAN_Charges,
    Filtre = Table.SelectRows(Source, each
        [filtre_vue_menage] = "OUI" and
        [statut_controle]   = "VALIDE"
    )
in
    Filtre
""".strip()

INSTRUCTIONS = """
INSTRUCTIONS POWER QUERY — MASTER_FACT_MAN_Charges.xlsx
==========================================================
37 colonnes = 31 colonnes SAISIE_Charges_Flux + 6 colonnes PQ :
  sens, filtre_vue_menage, source_module, source_table, source_pk, date_integration

1. Ouvrir MASTER_FACT_MAN_Charges.xlsx dans Excel
2. Onglet Donnees > Obtenir des donnees > Editeur Power Query

CREER 4 REQUETES dans cet ordre exact :

Req 1 : SAISIE_Source
  Donnees > Requete vide > coller le M-code ci-dessous
  Mode : Connexion uniquement (ne pas charger dans une feuille)

Req 2 : REF_Categories_Source
  Donnees > Requete vide > coller le M-code ci-dessous
  Mode : Connexion uniquement

Req 3 : MASTER_FACT_MAN_Charges
  Donnees > Requete vide > coller le M-code ci-dessous
  Puis : Clic droit > Charger dans > Table > Feuille MASTER

Req 4 : VUE_ACHATS_MENAGE_VALIDES
  Donnees > Requete vide > coller le M-code ci-dessous
  Puis : Clic droit > Charger dans > Table > Feuille VUE_MENAGE

NOTE : adapter BASE_PATH si le projet change de lecteur ou d'utilisateur.
""".strip()

# ── Workbook ───────────────────────────────────────────────────────────────────
wb = Workbook()
ws_master = wb.active
ws_master.title = 'MASTER'
ws_vue    = wb.create_sheet('VUE_MENAGE')
ws_pq     = wb.create_sheet('POWER_QUERY_CODE')

H_FONT = Font(bold=True, color='FFFFFF', size=10)

def write_headers(ws, cols):
    for col_i, (col_name, color) in enumerate(cols, 1):
        c = ws.cell(row=1, column=col_i, value=col_name)
        c.font = H_FONT
        c.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(col_i)].width = max(len(col_name) + 2, 14)
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = 'A2'

# MASTER — 37 colonnes
write_headers(ws_master, MASTER_COLS)
ws_master.cell(row=2, column=1,
    value='[Charge par Power Query — requete MASTER_FACT_MAN_Charges]'
).font = Font(italic=True, color='808080')
print(f"[OK] MASTER — {len(MASTER_COLS)} colonnes")

# VUE_MENAGE — memes 37 colonnes
write_headers(ws_vue, MASTER_COLS)
ws_vue.cell(row=2, column=1,
    value='[Vue par Power Query — requete VUE_ACHATS_MENAGE_VALIDES — filtre_vue_menage=OUI ET statut_controle=VALIDE]'
).font = Font(italic=True, color='808080')
print(f"[OK] VUE_MENAGE — {len(MASTER_COLS)} colonnes")

# POWER_QUERY_CODE
ws_pq.column_dimensions['A'].width = 32
ws_pq.column_dimensions['C'].width = 100

title_font = Font(bold=True, size=12, color='1F497D')
code_font  = Font(name='Courier New', size=9, color='262626')
note_font  = Font(italic=True, size=10, color='595959')

ws_pq.cell(row=1, column=1, value='POWER QUERY M-CODE — MASTER_FACT_MAN_Charges').font = title_font
ws_pq.cell(row=2, column=1, value='Ne pas modifier — reference uniquement').font = note_font

row_off = 4
for line in INSTRUCTIONS.split('\n'):
    ws_pq.cell(row=row_off, column=1, value=line).font = Font(size=9)
    row_off += 1
row_off += 2

for label, mcode in [
    ('REQUETE 1 : SAISIE_Source',             MCODE_SAISIE),
    ('REQUETE 2 : REF_Categories_Source',     MCODE_REF_CATS),
    ('REQUETE 3 : MASTER_FACT_MAN_Charges',   MCODE_MASTER),
    ('REQUETE 4 : VUE_ACHATS_MENAGE_VALIDES', MCODE_VUE),
]:
    ws_pq.cell(row=row_off, column=1, value=label).font = Font(bold=True, size=10, color='7030A0')
    row_off += 1
    for line in mcode.split('\n'):
        ws_pq.cell(row=row_off, column=3, value=line).font = code_font
        ws_pq.row_dimensions[row_off].height = 14
        row_off += 1
    row_off += 2

print("[OK] POWER_QUERY_CODE — 4 requetes M-code (37 colonnes)")

wb.save(OUT_PATH)
print(f"[SAVED] {OUT_PATH}")
print(f"\n=== MASTER_FACT_MAN_Charges.xlsx cree avec succes ({len(MASTER_COLS)} colonnes) ===")
