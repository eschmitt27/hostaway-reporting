"""
lot6b_m04_menages_internes.py — Alimentation DÉTERMINISTE de M04 (sans Power Query)
================================================================================
Remplace la dépendance Power Query : le MASTER M04 est reconstruit DIRECTEMENT
par ce script à chaque run, depuis la Google Sheet "Suivi ménage".

D027 (M04 = MO interne HC, TYPE_FLUX_013) conservée. D106 (refonte ménages).

URL CSV : lue depuis REF_Setup.xlsm > REF_Sources_Systeme
          (nom_source = GOOGLE_SHEET_M04_DECLARATIONS, actif=OUI). PAS d'URL en dur.

Sorties écrites automatiquement :
  - 02_TRAVAIL/Lot6b_DeclarationsInternes/MASTER_NORM_Declarations_Internes.xlsx
  - 02_DONNEES_NORMALISEES/menages/M04_MENAGES_PowerQuery.xlsx :
        SOURCE_RAW (traçabilité), MASTER (calculé), VUE_ACTIVE (VALIDE)
  POWER_QUERY_CODE conservé en documentation/archive, non utilisé.

Contrôles BLOQUANTS : URL absente / inaccessible / structure Google Sheet inattendue.
Ne touche pas : banque, Hostaway, factures, résultats aval (lot9-12).
"""

import sys, os, io, csv, subprocess, shutil, hashlib, datetime, collections, unicodedata, warnings
warnings.filterwarnings("ignore")
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import openpyxl
from openpyxl.styles import Font, PatternFill

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REF  = os.path.join(ROOT, "01_SOURCES_BRUTES", "REF_Setup", "REF_Setup.xlsm")
M04  = os.path.join(ROOT, "02_DONNEES_NORMALISEES", "menages", "M04_MENAGES_PowerQuery.xlsx")
NORM_DIR = os.path.join(ROOT, "02_TRAVAIL", "Lot6b_DeclarationsInternes")
NORM_OUT = os.path.join(NORM_DIR, "MASTER_NORM_Declarations_Internes.xlsx")
NOW = datetime.datetime.now().isoformat(timespec="seconds")

MOIS = {"janvier":"01","fevrier":"02","mars":"03","avril":"04","mai":"05","juin":"06","juillet":"07","aout":"08","septembre":"09","octobre":"10","novembre":"11","decembre":"12"}
INTMAP = {"imene": ("INT_0001","Imène"), "kira": ("INT_0002","Kheira"), "kheira": ("INT_0002","Kheira")}
REQUIRED_COLS = ["Prénom", "Mois des ménages", "Année des ménages", "Appartement"]

def abort(msg):
    print(f"\n[BLOQUANT lot6b] {msg}\nM04 NON modifié.")
    sys.exit(1)

def norm(s):
    s = str(s or "").strip().lower()
    return " ".join("".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn").split())

def rh(*v): return hashlib.sha256("|".join("" if x is None else str(x) for x in v).encode()).hexdigest()[:16]
def fnum(x):
    try: return float(x)
    except (TypeError, ValueError): return None
def to_d(v):
    try: return datetime.date.fromisoformat(str(v)[:10])
    except (ValueError, TypeError): return None

def sh(p, s):
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True); ws = wb[s]
    rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]; wb.close()
    return [dict(zip([str(c) for c in rows[0]], r)) for r in rows[1:]]

# ── URL depuis REF (bloquant) ────────────────────────────────────────────────
url = None
for d in sh(REF, "REF_Sources_Systeme"):
    if str(d.get("nom_source")) == "GOOGLE_SHEET_M04_DECLARATIONS" and str(d.get("actif")) == "OUI":
        url = str(d.get("dossier_source") or "").strip()
if not url or not url.startswith("http"):
    abort("URL GOOGLE_SHEET_M04_DECLARATIONS absente/invalide dans REF_Sources_Systeme (SRC_011).")

# ── Fetch CSV (bloquant) ─────────────────────────────────────────────────────
res = subprocess.run(["curl", "-sL", "-A", "Mozilla/5.0", url], capture_output=True, text=True, encoding="utf-8")
txt = res.stdout or ""
if res.returncode != 0 or len(txt) < 50:
    abort(f"Google Sheet inaccessible (rc={res.returncode}, taille={len(txt)}).")
rows = list(csv.reader(io.StringIO(txt)))
if not rows: abort("CSV vide.")
hdr = rows[0]
missing = [c for c in REQUIRED_COLS if c not in hdr]
if missing: abort(f"Structure Google Sheet inattendue, colonnes manquantes : {missing}")

# ── Référentiels ─────────────────────────────────────────────────────────────
lmap, lognom, logtype, logprop, loghaid = {}, {}, {}, {}, {}
for d in sh(REF, "REF_Mapping_Logements"):
    if d.get("valeur_source"): lmap[norm(d["valeur_source"])] = d.get("logement_id")
for d in sh(REF, "REF_Logements"):
    lid = d.get("logement_id")
    if lid and str(lid) != "logement_id":
        lognom[lid] = d.get("nom_logement_officiel"); logtype[lid] = d.get("type_logement_id")
        logprop[lid] = d.get("proprietaire_id"); loghaid[lid] = d.get("hostaway_listing_id")
std_ref = sh(REF, "REF_Couts_Standards_Menage")
taux = next((fnum(d.get("valeur")) for d in sh(REF, "REF_Parametres_Generaux") if d.get("nom_parametre") == "TAUX_HORAIRE_MENAGE_INTERNE"), None)
def std_unit(type_id, dref):
    best = None
    for d in std_ref:
        if d.get("type_logement_id") != type_id or str(d.get("actif")) != "OUI": continue
        deb, fin = to_d(d.get("date_debut_validite")), to_d(d.get("date_fin_validite"))
        if deb and dref < deb: continue
        if fin and dref > fin: continue
        best = fnum(d.get("cout_standard_menage"))
    return best

i_pre = hdr.index("Prénom"); i_mois = hdr.index("Mois des ménages"); i_an = hdr.index("Année des ménages")
appcols = [i for i, h in enumerate(hdr) if h.strip() == "Appartement"]
i_lav_na = next((i for i, h in enumerate(hdr) if h.strip().startswith("Coûts de lavage du linge (hors")), None)

# ── Dépivot ──────────────────────────────────────────────────────────────────
norm_rows = []
for r in rows[1:]:
    if not any(c.strip() for c in r): continue
    pre = r[i_pre].strip(); mm = MOIS.get(norm(r[i_mois])); yr = r[i_an].strip()
    miso = f"{yr}-{mm}" if (mm and yr) else None
    iid, inom = INTMAP.get(norm(pre), (None, pre))
    lav_na = fnum(r[i_lav_na]) if (i_lav_na is not None and i_lav_na < len(r) and r[i_lav_na].strip()) else 0
    for ci in appcols:
        app = r[ci].strip() if ci < len(r) else ""
        if not app: continue
        nb = int(fnum(r[ci+1]) or 0) if (ci+1 < len(r) and r[ci+1].strip()) else 0
        nh = fnum(r[ci+2]) if (ci+2 < len(r) and r[ci+2].strip()) else None
        lav = fnum(r[ci+3]) if (ci+3 < len(r) and r[ci+3].strip()) else 0
        lid = lmap.get(norm(app))
        statut, code = ("VALIDE", "")
        if lid is None: statut, code = "A_CONTROLER", "LOGEMENT_NON_MAPPE"
        elif iid is None: statut, code = "A_CONTROLER", "INTERVENANT_NON_MAPPE"
        norm_rows.append({"mois": miso, "annee": yr, "mois_saisie": r[i_mois].strip(),
            "appartement_source": app, "nom_appartement": lognom.get(lid), "logement_id": lid,
            "intervenant_source": pre, "intervenant_id": iid, "nom_intervenant": inom, "type_intervenant": "INTERNE",
            "nb_menages": nb, "nb_heures": nh, "cout_lavage_attribue": lav, "lavage_non_attribuable_mois": lav_na,
            "statut_controle": statut, "code_controle": code, "source_url": url, "date_extraction": NOW,
            "ROW_HASH": rh(miso, lid, iid, nb)})

# ── 1) MASTER_NORM ───────────────────────────────────────────────────────────
os.makedirs(NORM_DIR, exist_ok=True)
NCOLS = ["mois","annee","mois_saisie","appartement_source","nom_appartement","logement_id",
    "intervenant_source","intervenant_id","nom_intervenant","type_intervenant","nb_menages","nb_heures",
    "cout_lavage_attribue","lavage_non_attribuable_mois","statut_controle","code_controle","source_url","date_extraction","ROW_HASH"]
wbn = openpyxl.Workbook(); wsn = wbn.active; wsn.title = "MASTER_NORMALISE"; wsn.append(NCOLS)
for c in wsn[1]: c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="DDDDDD")
for d in norm_rows: wsn.append([d.get(c) for c in NCOLS])
wbn.save(NORM_OUT)

# ── 2) M04 SOURCE_RAW + MASTER + VUE_ACTIVE (autres onglets préservés) ────────
backup = M04 + ".BAK_" + datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
shutil.copy(M04, backup)
wb = openpyxl.load_workbook(M04)
ws = wb["SOURCE_RAW"]
if ws.max_row > 1: ws.delete_rows(2, ws.max_row - 1)
for d in norm_rows:
    ws.append([d["mois"], d["appartement_source"], d["intervenant_source"], "MENAGE_STANDARD",
               d["nb_menages"], d["nb_heures"], f'lavage={d["cout_lavage_attribue"] or 0}; import lot6b'])
if "tbl_SOURCE_RAW" in ws.tables:
    ws.tables["tbl_SOURCE_RAW"].ref = f"A1:G{1+len(norm_rows)}"

MASTER_HEADERS = [c.value for c in wb["MASTER"][1]]
master_rows = []; cnt = collections.Counter()
for d in norm_rows:
    miso = d["mois"]; lid = d["logement_id"]; type_id = logtype.get(lid)
    dref = to_d((miso + "-01")) if miso else datetime.date.today()
    nb = d["nb_menages"]; nh = d["nb_heures"]
    su = std_unit(type_id, dref or datetime.date.today())
    cet = round((nh or 0) * taux, 2) if (nh is not None and taux is not None) else None
    ceu = round(cet / nb, 2) if (cet is not None and nb) else None
    cst = round((su or 0) * nb, 2) if su is not None else None
    ec = round(cst - cet, 2) if (cst is not None and cet is not None) else None
    cnt[miso] += 1
    mid = f"MEN-{miso or '0000-00'}-{cnt[miso]:03d}"
    master_rows.append({"menage_calc_id": mid, "ROW_HASH": d["ROW_HASH"], "mois": miso, "annee": d["annee"],
        "mois_num": (miso[5:7] if miso else None), "logement_id": lid, "proprietaire_id": logprop.get(lid),
        "hostaway_listing_id": loghaid.get(lid), "appartement_source": d["appartement_source"],
        "intervenant_id": d["intervenant_id"], "nom_intervenant": d["nom_intervenant"], "type_intervenant": "INTERNE",
        "type_menage": "MENAGE_STANDARD", "nb_menages": nb, "nb_heures": nh, "taux_horaire_intervenant": taux,
        "cout_execution_total": cet, "cout_execution_unitaire": ceu, "cout_standard": su,
        "cout_standard_total_ligne": cst, "ecart_main_oeuvre_vs_standard": ec, "total_execution": cet,
        "type_flux_id": "TYPE_FLUX_013", "sens": "CHARGE", "code_impact": "HC",
        "impact_resultat_reel": "OUI", "impact_resultat_comptable": "NON",
        "statut_controle": d["statut_controle"], "niveau_anomalie": ("INFO" if d["statut_controle"] == "VALIDE" else "A_CONTROLER"),
        "code_anomalie": d["code_controle"] or None, "source_module": "lot6b", "source_table": "SOURCE_RAW",
        "source_pk": mid, "date_integration": NOW})
for sheetname, only_valide in [("MASTER", False), ("VUE_ACTIVE", True)]:
    wsm = wb[sheetname]
    if wsm.max_row > 1: wsm.delete_rows(2, wsm.max_row - 1)
    for r in master_rows:
        if only_valide and r["statut_controle"] != "VALIDE": continue
        wsm.append([r.get(h) for h in MASTER_HEADERS])
wb.save(M04); wb.close()

# ── Rapport ──────────────────────────────────────────────────────────────────
mai = [d for d in norm_rows if d["mois"] == "2026-05"]
ag = collections.Counter()
for d in mai: ag[(d["intervenant_id"], d["nom_intervenant"])] += d["nb_menages"]
nmap = sum(1 for d in norm_rows if d["code_controle"])
print(f"[lot6b] URL REF OK (SRC_011) | CSV {len(rows)-1} lignes | normalisées {len(norm_rows)} | M04 MASTER {len(master_rows)} lignes")
print(f"[lot6b] MASTER_NORM -> {NORM_OUT}")
print(f"[lot6b] M04 SOURCE_RAW+MASTER+VUE_ACTIVE reconstruit SANS Power Query | backup {os.path.basename(backup)}")
print(f"[lot6b] mai 2026 par intervenant : {dict(ag)}")
print(f"[lot6b] lignes A_CONTROLER (mapping) : {nmap}")
