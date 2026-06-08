"""
lot3_saisie_template.py
Lot 3 — Creation SAISIE_Charges_Flux.xlsx
- Onglet SAISIE (31 colonnes + data validation + MFC)
- Onglet REF_LOCALE (18 listes depuis REF_Setup.xlsm)
- Onglet CONTROLES_SAISIE (13 controles automatiques)
- Onglet README (instructions)
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import FormulaRule

BASE = "C:/Users/Ewan/OneDrive/Documents/Conciergerie/Pilotage_Conciergerie"
REF_PATH = f"{BASE}/01_SOURCES_BRUTES/REF_Setup/REF_Setup.xlsm"
OUT_PATH = f"{BASE}/01_SOURCES_BRUTES/Charges/SAISIE_Charges_Flux.xlsx"

# ── 1. Lire REF_Setup ──────────────────────────────────────────────────────────
ref_wb = openpyxl.load_workbook(REF_PATH, read_only=True, keep_vba=True)

# IDs actifs par feuille
cats = [
    r[0] for r in ref_wb['REF_Categories_Charges'].iter_rows(min_row=2, values_only=True)
    if r[0] and r[7] == 'OUI'
]

EXCLUDED_TF = {'TYPE_FLUX_001', 'TYPE_FLUX_003', 'TYPE_FLUX_005', 'TYPE_FLUX_006', 'TYPE_FLUX_007'}
types_flux = [
    r[0] for r in ref_wb['REF_Types_Flux'].iter_rows(min_row=2, values_only=True)
    if r[0] and r[0] not in EXCLUDED_TF and r[7] == 'OUI'
]

modes = [
    r[0] for r in ref_wb['REF_Modes_Paiement'].iter_rows(min_row=2, values_only=True)
    if r[0] and r[5] == 'OUI'
]

cartes = [
    r[0] for r in ref_wb['REF_Cartes_Paiement'].iter_rows(min_row=2, values_only=True)
    if r[0] and r[8] == 'OUI'
]

associes = [
    r[0] for r in ref_wb['REF_Associes'].iter_rows(min_row=2, values_only=True)
    if r[0] and r[3] == 'OUI'
]

# Logements actifs (inclut APPARTEMENT_DIVERS / LOGEMENT_DIVERS — codes techniques)
logements = [
    r[0] for r in ref_wb['REF_Logements'].iter_rows(min_row=2, values_only=True)
    if r[0] and r[12] == 'OUI'
]

proprietaires = [
    r[0] for r in ref_wb['REF_Proprietaires'].iter_rows(min_row=2, values_only=True)
    if r[0] and r[8] == 'OUI'
]

# type_affectation (libelles) — col 1, tous les enregistrements
affectations = [
    r[1] for r in ref_wb['REF_Types_Affectation'].iter_rows(min_row=2, values_only=True)
    if r[0] and r[1]
]

ref_wb.close()
print(f"  REF lu — cats:{len(cats)}, types_flux:{len(types_flux)}, logements:{len(logements)}, proprios:{len(proprietaires)}")

# Listes statiques
SENS_FLUX              = ['DEPENSE', 'RECUPERATION', 'REMBOURSEMENT', 'REFACTURATION', 'NEUTRE']
CODES_IMPACT           = ['IC', 'HC', 'HR']
STATUTS_CONTROLE       = ['VALIDE', 'A_CONTROLER', 'EXCLU_RESULTAT', 'A_VENTILER']
NIVEAUX_ANOMALIE       = ['INFO', 'A_CONTROLER', 'BLOQUANT']
SOURCES_FLUX           = ['SAISIE_MANUELLE', 'FACTURE_PDF', 'RAPPROCHEMENT_BANCAIRE', 'RESERVATION_HH', 'AIRCOVER']
METHODES               = ['DIRECT', 'VIA_VUE_MENAGE', 'REFACTURATION_PROPRIETAIRE', 'NEUTRALISATION', 'A_CONFIRMER']
STATUTS_RAPPROCHEMENT  = ['NON_RAPPROCHE', 'RAPPROCHE_BANQUE', 'SANS_RAPPROCHEMENT', 'ECART_A_ANALYSER']
PRISE_COMPTA           = ['OUI', 'NON']
REFACTURABLE           = ['OUI', 'NON']
JUSTIFICATIF           = ['OUI', 'NON', 'LIEN']

# ── 2. Workbook ────────────────────────────────────────────────────────────────
wb = Workbook()
ws_saisie = wb.active
ws_saisie.title = 'SAISIE'
ws_ref   = wb.create_sheet('REF_LOCALE')
ws_ctrl  = wb.create_sheet('CONTROLES_SAISIE')
ws_readme = wb.create_sheet('README')

H_FONT  = Font(bold=True, color='FFFFFF', size=9)
H_FILL  = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')

# ── 3. REF_LOCALE ─────────────────────────────────────────────────────────────
ref_lists = [
    ('lst_Categories',            cats),
    ('lst_TypesFlux_Lot3',        types_flux),
    ('lst_Codes_Impact',          CODES_IMPACT),
    ('lst_Sens_Flux',             SENS_FLUX),
    ('lst_Associes',              associes),
    ('lst_ModesPaiement',         modes),
    ('lst_Cartes',                cartes),
    ('lst_Affectation',           affectations),
    ('lst_Logements',             logements),
    ('lst_Proprietaires',         proprietaires),
    ('lst_Statuts_Controle',      STATUTS_CONTROLE),
    ('lst_Niveau_Anomalie',       NIVEAUX_ANOMALIE),
    ('lst_Source_Flux',           SOURCES_FLUX),
    ('lst_Methode',               METHODES),
    ('lst_Statut_Rapprochement',  STATUTS_RAPPROCHEMENT),
    ('lst_Prise_Compta',          PRISE_COMPTA),
    ('lst_Refacturable',          REFACTURABLE),
    ('lst_Justificatif',          JUSTIFICATIF),
]

for col_i, (lst_name, lst_vals) in enumerate(ref_lists, 1):
    col_l = get_column_letter(col_i)
    c = ws_ref.cell(row=1, column=col_i, value=lst_name)
    c.font = H_FONT
    c.fill = H_FILL
    ws_ref.column_dimensions[col_l].width = max(len(lst_name), max((len(str(v)) for v in lst_vals), default=8)) + 2
    for row_i, val in enumerate(lst_vals, 2):
        ws_ref.cell(row=row_i, column=col_i, value=val)

ws_ref.freeze_panes = 'A2'
print(f"[OK] REF_LOCALE — {len(ref_lists)} listes")

# ── 4. Named Ranges ───────────────────────────────────────────────────────────
for col_i, (lst_name, lst_vals) in enumerate(ref_lists, 1):
    col_l = get_column_letter(col_i)
    n = len(lst_vals)
    ref_str = f"'REF_LOCALE'!${col_l}$2:${col_l}${n + 1}"
    dn = DefinedName(name=lst_name, attr_text=ref_str)
    wb.defined_names[lst_name] = dn
print(f"[OK] Named Ranges — {len(ref_lists)}")

# ── 5. SAISIE — 31 colonnes ───────────────────────────────────────────────────
# (nom, groupe_num, couleur_hexa)
SAISIE_COLS = [
    # Groupe 1 — Identification
    ('charge_id',                  1, '1F497D'),   # A col 1
    ('date_charge',                1, '1F497D'),   # B col 2
    ('mois',                       1, '1F497D'),   # C col 3
    # Groupe 2 — Montant + sens
    ('montant',                    2, '375623'),   # D col 4
    ('sens_flux',                  2, '375623'),   # E col 5
    # Groupe 3 — Classification
    ('categorie_charge_id',        3, '7030A0'),   # F col 6
    ('type_flux_id',               3, '7030A0'),   # G col 7
    # Groupe 4 — Impact resultat
    ('code_impact',                4, 'C00000'),   # H col 8
    ('impact_resultat_reel',       4, 'C00000'),   # I col 9
    ('impact_resultat_comptable',  4, 'C00000'),   # J col 10
    ('prise_en_compta',            4, 'C00000'),   # K col 11
    # Groupe 5 — Qui paie
    ('associe_id',                 5, '974706'),   # L col 12
    ('mode_paiement_id',           5, '974706'),   # M col 13
    ('carte_id',                   5, '974706'),   # N col 14
    # Groupe 6 — Affectation
    ('affectation_type',           6, '1F3864'),   # O col 15
    ('logement_id',                6, '1F3864'),   # P col 16
    ('proprietaire_id',            6, '1F3864'),   # Q col 17
    # Groupe 7 — Conditions speciales
    ('reservation_id',             7, '833C00'),   # R col 18
    ('refacturable',               7, '833C00'),   # S col 19
    # Groupe 8 — Tracabilite
    ('source_flux',                8, '595959'),   # T col 20
    ('methode_traitement',         8, '595959'),   # U col 21
    ('paye_avec_montant_recupere', 8, '595959'),   # V col 22
    ('lien_virement_banque',       8, '595959'),   # W col 23
    # Groupe 9 — Controle
    ('statut_controle',            9, '7B0000'),   # X col 24
    ('niveau_anomalie',            9, '7B0000'),   # Y col 25
    ('code_anomalie',              9, '7B0000'),   # Z col 26
    ('statut_rapprochement',       9, '7B0000'),   # AA col 27
    # Groupe 10 — Documents
    ('justificatif',              10, '595959'),   # AB col 28
    ('commentaire',               10, '595959'),   # AC col 29
    # Groupe 11 — Systeme
    ('ROW_HASH',                  11, '262626'),   # AD col 30
    ('date_saisie',               11, '262626'),   # AE col 31
]

COL_WIDTHS = [
    28, 14, 12,  # A-C
    12, 18,      # D-E
    24, 30,      # F-G
    12, 22, 24,  # H-J
    14,          # K
    14, 18, 12,  # L-N
    22, 22, 22,  # O-Q
    20, 14,      # R-S
    24, 30,      # T-U
    24, 24,      # V-W
    18, 18,      # X-Y
    22, 24,      # Z-AA
    14, 28,      # AB-AC
    38, 14,      # AD-AE
]

# En-tetes
for col_i, (col_name, _grp, color) in enumerate(SAISIE_COLS, 1):
    c = ws_saisie.cell(row=1, column=col_i, value=col_name)
    c.font = Font(bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws_saisie.column_dimensions[get_column_letter(col_i)].width = COL_WIDTHS[col_i - 1]

ws_saisie.row_dimensions[1].height = 38
ws_saisie.freeze_panes = 'A2'

# Formules calculees sur 500 lignes
# C=mois (col3), I=impact_resultat_reel (col9), J=impact_resultat_comptable (col10), AD=ROW_HASH (col30)
N_FORMULA = 500
for row in range(2, N_FORMULA + 2):
    # mois
    ws_saisie.cell(row=row, column=3).value = f'=IF(B{row}="","",TEXT(B{row},"YYYY-MM"))'
    # impact_resultat_reel : IC→OUI, HC→OUI, HR→NON, vide→A_CONTROLER
    ws_saisie.cell(row=row, column=9).value = (
        f'=IF(H{row}="IC","OUI",'
        f'IF(H{row}="HC","OUI",'
        f'IF(H{row}="HR","NON","A_CONTROLER")))'
    )
    # impact_resultat_comptable : IC→OUI, HC→NON, HR→NON, vide→A_CONTROLER
    ws_saisie.cell(row=row, column=10).value = (
        f'=IF(H{row}="IC","OUI",'
        f'IF(H{row}="HC","NON",'
        f'IF(H{row}="HR","NON","A_CONTROLER")))'
    )
    # ROW_HASH : concatene date+montant+categorie+mode
    ws_saisie.cell(row=row, column=30).value = (
        f'=IF(A{row}="","",'
        f'TEXT(B{row},"YYYYMMDD")&"|"&TEXT(D{row},"0.00")&"|"&F{row}&"|"&M{row})'
    )
    # date_saisie : laissee vide — saisie manuelle (non volatile, voir README)

# Data Validations (Named Range → liste deroulante)
DV_MAP = [
    ('E',  'lst_Sens_Flux'),
    ('F',  'lst_Categories'),
    ('G',  'lst_TypesFlux_Lot3'),
    ('H',  'lst_Codes_Impact'),
    ('K',  'lst_Prise_Compta'),
    ('L',  'lst_Associes'),
    ('M',  'lst_ModesPaiement'),
    ('N',  'lst_Cartes'),
    ('O',  'lst_Affectation'),
    ('P',  'lst_Logements'),
    ('Q',  'lst_Proprietaires'),
    ('S',  'lst_Refacturable'),
    ('T',  'lst_Source_Flux'),
    ('U',  'lst_Methode'),
    ('X',  'lst_Statuts_Controle'),
    ('Y',  'lst_Niveau_Anomalie'),
    ('AA', 'lst_Statut_Rapprochement'),
    ('AB', 'lst_Justificatif'),
]

for col_l, lst_name in DV_MAP:
    dv = DataValidation(
        type="list",
        formula1=lst_name,
        allow_blank=True,
        showErrorMessage=True,
        error='Valeur hors liste. Utiliser la liste deroulante.',
        errorTitle='Valeur invalide',
    )
    ws_saisie.add_data_validation(dv)
    dv.sqref = f'{col_l}2:{col_l}1001'

# Mise en forme conditionnelle
red_fill    = PatternFill(start_color='FFAAAA', end_color='FFAAAA', fill_type='solid')
orange_fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
# Priorite : BLOQUANT (rouge) avant A_CONTROLER (orange)
ws_saisie.conditional_formatting.add(
    'A2:AE1001', FormulaRule(formula=['$Y2="BLOQUANT"'],    fill=red_fill)
)
ws_saisie.conditional_formatting.add(
    'A2:AE1001', FormulaRule(formula=['$Y2="A_CONTROLER"'], fill=orange_fill)
)
print("[OK] SAISIE — 31 colonnes + 18 DV + formules + MFC")

# ── 6. CONTROLES_SAISIE ───────────────────────────────────────────────────────
CTRL_HEADERS = ['Code', 'Libelle', 'Valeur', 'Seuil', 'Note']
for col_i, h in enumerate(CTRL_HEADERS, 1):
    c = ws_ctrl.cell(row=1, column=col_i, value=h)
    c.font = H_FONT
    c.fill = H_FILL

CONTROLS = [
    ('CTR-L3-01', 'Lignes saisies (charge_id renseigne)',
     '=MAX(0,COUNTA(SAISIE!A:A)-1)',
     'INFO', ''),
    ('CTR-L3-02', 'Lignes VALIDE (statut_controle)',
     '=COUNTIF(SAISIE!X:X,"VALIDE")',
     'INFO', ''),
    ('CTR-L3-03', 'Lignes A_CONTROLER (statut_controle)',
     '=COUNTIF(SAISIE!X:X,"A_CONTROLER")',
     'ATTENTION si > 0', ''),
    ('CTR-L3-04', 'Lignes EXCLU_RESULTAT',
     '=COUNTIF(SAISIE!X:X,"EXCLU_RESULTAT")',
     'INFO', ''),
    ('CTR-L3-05', 'Lignes A_VENTILER',
     '=COUNTIF(SAISIE!X:X,"A_VENTILER")',
     'ATTENTION si > 0', ''),
    ('CTR-L3-06', 'Lignes BLOQUANT (niveau_anomalie)',
     '=COUNTIF(SAISIE!Y:Y,"BLOQUANT")',
     'BLOQUANT si > 0', 'Doit etre 0 avant livraison vers MASTER'),
    ('CTR-L3-07', 'charge_id dupliques',
     '=SUMPRODUCT((COUNTIF(SAISIE!A2:A1001,SAISIE!A2:A1001)-1)*(SAISIE!A2:A1001<>""))',
     'BLOQUANT si > 0', 'PK unique obligatoire'),
    ('CTR-L3-08', 'Montant total IC',
     '=SUMIF(SAISIE!H:H,"IC",SAISIE!D:D)',
     'INFO', ''),
    ('CTR-L3-09', 'Montant total HC',
     '=SUMIF(SAISIE!H:H,"HC",SAISIE!D:D)',
     'INFO', ''),
    ('CTR-L3-10', 'date_charge vide (charge_id renseigne)',
     '=COUNTIFS(SAISIE!A2:A1001,"<>",SAISIE!B2:B1001,"")',
     'BLOQUANT si > 0', 'date_charge obligatoire si charge_id renseigne'),
    ('CTR-L3-11', 'CHG_021 sans reservation_id (D041)',
     '=COUNTIFS(SAISIE!F2:F1001,"CHG_021",SAISIE!R2:R1001,"")',
     'BLOQUANT si > 0', 'reservation_id obligatoire pour incidents voyageurs'),
    ('CTR-L3-12', 'code_impact vide (charge_id renseigne)',
     '=COUNTIFS(SAISIE!A2:A1001,"<>",SAISIE!H2:H1001,"")',
     'BLOQUANT si > 0', ''),
    ('CTR-L3-13', 'statut_controle vide (charge_id renseigne)',
     '=COUNTIFS(SAISIE!A2:A1001,"<>",SAISIE!X2:X1001,"")',
     'ATTENTION si > 0', 'Doit etre renseigne avant integration MASTER'),
]

for row_i, ctrl in enumerate(CONTROLS, 2):
    for col_i, val in enumerate(ctrl, 1):
        ws_ctrl.cell(row=row_i, column=col_i, value=val)

ws_ctrl.column_dimensions['A'].width = 14
ws_ctrl.column_dimensions['B'].width = 50
ws_ctrl.column_dimensions['C'].width = 65
ws_ctrl.column_dimensions['D'].width = 20
ws_ctrl.column_dimensions['E'].width = 50
ws_ctrl.freeze_panes = 'A2'
print("[OK] CONTROLES_SAISIE — 13 controles")

# ── 7. README ─────────────────────────────────────────────────────────────────
README_LINES = [
    ('SAISIE_Charges_Flux.xlsx — Guide de saisie (Lot 3)', True),
    ('', False),
    ('SOURCE : D026 — Source unique des achats, charges, consommables, linge, materiel.', False),
    ('EXCLUT : IK et virements associes (Lot 7). Acomptes proprietaires (Lot 5).', False),
    ('', False),
    ('REGLES OBLIGATOIRES', True),
    ('1. charge_id : PK unique. Nomenclature : CHG-AAAA-MM-IMPACT-ASSOC-NNN', False),
    ('   Exemple : CHG-2026-04-HC-WAFA-CB-001', False),
    ('2. montant : TOUJOURS positif (TTC). Le sens est porte par la colonne sens_flux.', False),
    ('3. code_impact determine les deux colonnes calculees (non surchargeables) :', False),
    ('   IC  →  impact_resultat_reel = OUI  /  impact_resultat_comptable = OUI', False),
    ('   HC  →  impact_resultat_reel = OUI  /  impact_resultat_comptable = NON', False),
    ('   HR  →  impact_resultat_reel = NON  /  impact_resultat_comptable = NON', False),
    ('4. statut_controle : passer a VALIDE avant integration dans MASTER_FACT_MAN_Charges.', False),
    ('5. CHG_021 (incident voyageur) : reservation_id OBLIGATOIRE (D041).', False),
    ('6. CHG_022 (AirCover refacture) : alimente charges_exceptionnelles_refacturees (D042).', False),
    ('7. Aucun montant fixe code en dur — forfaits dans REF_Charges_Recurrentes (D045).', False),
    ('8. date_saisie : saisir manuellement la date de creation de la ligne (formule TODAY()', False),
    ('   recalcule — utiliser Ctrl+; pour inserer la date du jour de facon non volatile).', False),
    ('', False),
    ('ONGLETS', True),
    ('- SAISIE            : 31 colonnes, listes deroulantes, MFC (rouge=BLOQUANT, orange=A_CONTROLER)', False),
    ('- REF_LOCALE        : 18 listes extraites de REF_Setup.xlsm (ne pas modifier manuellement)', False),
    ('- CONTROLES_SAISIE  : 13 controles automatiques — verifier avant chaque livraison', False),
    ('- README            : ce fichier', False),
    ('', False),
    ('VUE_ACHATS_MENAGE_VALIDES (filtree dans MASTER_FACT_MAN_Charges)', True),
    ('Condition : filtre_vue_menage = OUI ET statut_controle = VALIDE', False),
    ('Categories eligibles : CHG_003 (Blanchisserie), CHG_004 (Produits logement),', False),
    ('                       CHG_018 (Petit equipement), CHG_023 (Forfait local cave)', False),
    ('', False),
    ('ANTI-DOUBLE-COMPTAGE (D027 / D038)', True),
    ('Les achats saisies ici NE doivent PAS etre re-saisies dans M04.', False),
    ('M04 = main-d oeuvre uniquement. Achats, consommables, linge : SAISIE_Charges_Flux UNIQUEMENT.', False),
    ('', False),
    ('CONTROLES A FAIRE AVANT LIVRAISON (JOURNAL_CONTROLES D029)', True),
    ('CTR-L3-06 = 0  (aucun BLOQUANT)', False),
    ('CTR-L3-07 = 0  (aucun charge_id duplique)', False),
    ('CTR-L3-10 = 0  (aucune date_charge vide)', False),
    ('CTR-L3-11 = 0  (aucun CHG_021 sans reservation_id)', False),
    ('CTR-L3-12 = 0  (aucun code_impact vide)', False),
]

title_font  = Font(bold=True, size=12, color='1F497D')
normal_font = Font(size=10)
ws_readme.column_dimensions['A'].width = 90

for row_i, (line, bold) in enumerate(README_LINES, 1):
    c = ws_readme.cell(row=row_i, column=1, value=line)
    c.font = title_font if bold else normal_font

print("[OK] README")

# ── 8. Save ────────────────────────────────────────────────────────────────────
wb.save(OUT_PATH)
print(f"[SAVED] {OUT_PATH}")
print("\n=== SAISIE_Charges_Flux.xlsx cree avec succes ===")
