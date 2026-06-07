"""Lot 0 - Corrections REF_Setup.xlsm (B1-B6 + I1)"""
import openpyxl
from openpyxl import load_workbook
import unicodedata
from datetime import datetime

FILEPATH = r"C:/Users/Ewan/OneDrive/Documents/Conciergerie/Pilotage_Conciergerie/01_SOURCES_BRUTES/REF_Setup/REF_Setup.xlsm"

wb = load_workbook(FILEPATH, keep_vba=True, data_only=False)

# ── HELPERS ──────────────────────────────────────────────────

def fix_mojibake_str(s):
    """Fix UTF-8 bytes stored as latin-1 codepoints (e.g. Ã© -> é)."""
    if not isinstance(s, str) or '\xc3' not in s:
        return s
    result = []
    i = 0
    while i < len(s):
        cp = ord(s[i])
        if 0xC2 <= cp <= 0xDF and i + 1 < len(s):
            next_cp = ord(s[i+1])
            if 0x80 <= next_cp <= 0xBF:
                try:
                    fixed = bytes([cp, next_cp]).decode('utf-8')
                    result.append(fixed)
                    i += 2
                    continue
                except UnicodeDecodeError:
                    pass
        result.append(s[i])
        i += 1
    return ''.join(result)

def fix_sheet_mojibake(ws):
    count = 0
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and '\xc3' in cell.value:
                fixed = fix_mojibake_str(cell.value)
                if fixed != cell.value:
                    cell.value = fixed
                    count += 1
    return count

def normalize_name(name):
    if not name:
        return None
    n = unicodedata.normalize('NFD', name)
    n = ''.join(c for c in n if unicodedata.category(c) != 'Mn')
    return n.upper().strip()

def fix_type_intervenant(t):
    if t is None:
        return t
    mapping = {
        'interne': 'INTERNE', 'externe': 'EXTERNE',
        'autre': 'AUTRE', 'a_controler': 'A_CONTROLER'
    }
    return mapping.get(t.lower(), t.upper())

# ── B1 — FIX MOJIBAKE ────────────────────────────────────────
sheets_mojibake = ['REF_Associes', 'REF_Codes_Impact', 'REF_Types_Flux', 'REF_Types_Affectation']
for sname in sheets_mojibake:
    if sname in wb.sheetnames:
        n = fix_sheet_mojibake(wb[sname])
        print(f"[B1] {sname}: {n} cellules corrigees")

# ── B2 — REF_Statuts: add VALIDE, BLOQUANT, IGNORE_JUSTIFIE ──
ws_statuts = wb['REF_Statuts']
rows_s = list(ws_statuts.iter_rows(values_only=True))
existing_ids = [r[0] for r in rows_s[1:] if r[0]]
max_num = max(int(sid.replace('STAT_', '')) for sid in existing_ids if sid and sid.startswith('STAT_'))
new_statuts = [
    (f'STAT_{max_num+1:03d}', 'statut_controle', 'VALIDE',
     5, 'OUI', 'Ligne validee, incluse dans les calculs'),
    (f'STAT_{max_num+2:03d}', 'statut_controle', 'BLOQUANT',
     6, 'OUI', 'Anomalie bloquante - exclue des calculs et cloture impossible'),
    (f'STAT_{max_num+3:03d}', 'statut_controle', 'IGNORE_JUSTIFIE',
     7, 'OUI', 'Ignoree avec justification explicite - exclue des calculs mais tracable'),
]
for row_data in new_statuts:
    ws_statuts.append(row_data)
    print(f"[B2] REF_Statuts: ajout {row_data[2]}")

# ── B3 — CREATE REF_Statuts_Payout ───────────────────────────
if 'REF_Statuts_Payout' not in wb.sheetnames:
    ws_sp = wb.create_sheet('REF_Statuts_Payout')
    ws_sp.append(['statut_payout_id', 'statut_calcul_payout', 'description', 'actif', 'commentaire'])
    data_sp = [
        ('SPAY_001', 'NORMAL',             'Reservation avec payout calcule normalement',             'OUI', None),
        ('SPAY_002', 'ANNULE_SANS_PAYOUT', 'Annulation sans montant verse',                          'OUI', 'Code impact HR - aucun produit'),
        ('SPAY_003', 'ANNULE_AVEC_PAYOUT', 'Annulation avec payout partiel (D030)',                  'OUI', 'BaseCommission = CancellationPayout, sans deduction menage'),
        ('SPAY_004', 'PAYOUT_ABSENT',      'Payout attendu mais absent des donnees Hostaway',        'OUI', 'Controle BOOKING_PAYOUT_INCOMPLET'),
        ('SPAY_005', 'PAYOUT_INCOMPLET',   'Payout present mais champs incomplets ou incoherents',   'OUI', 'Controle BOOKING_PAYOUT_INCOMPLET'),
        ('SPAY_006', 'A_CONTROLER',        'Statut payout a verifier manuellement',                  'OUI', None),
    ]
    for row in data_sp:
        ws_sp.append(row)
    print(f"[B3] REF_Statuts_Payout: cree ({len(data_sp)} lignes)")

# ── B4 — CREATE REF_Cloture_Mensuelle (structure vide) ───────
if 'REF_Cloture_Mensuelle' not in wb.sheetnames:
    ws_cm = wb.create_sheet('REF_Cloture_Mensuelle')
    ws_cm.append([
        'mois', 'statut_mois', 'date_passage_controle', 'date_cloture',
        'nb_lignes_bancaires_non_classees', 'nb_controles_bloquants_ouverts', 'commentaire'
    ])
    print("[B4] REF_Cloture_Mensuelle: cree (structure vide)")

# ── B5 — REF_Parametres_Generaux: add 4 params ───────────────
ws_pg = wb['REF_Parametres_Generaux']
rows_pg = list(ws_pg.iter_rows(values_only=True))
existing_pg = [r[0] for r in rows_pg[1:] if r[0]]
max_pg = max(int(pid.replace('PARAM_', '')) for pid in existing_pg if pid and pid.startswith('PARAM_'))
date_debut = datetime(2026, 1, 1)
new_params = [
    (f'PARAM_{max_pg+1:03d}', 'TAUX_HORAIRE_MENAGE_INTERNE', '10',
     "Taux horaire menage interne (EUR/h) - M04, migration vers REF_Parametres_Generaux prevue",
     date_debut, None, 'OUI'),
    (f'PARAM_{max_pg+2:03d}', 'ARRONDI_DECIMALES', '2',
     'Nombre de decimales pour arrondi des montants (D035)',
     date_debut, None, 'OUI'),
    (f'PARAM_{max_pg+3:03d}', 'TOLERANCE_ARRONDI_LIGNE_EUR', '0.10',
     'Tolerance ecart arrondi par ligne (EUR) - D035',
     date_debut, None, 'OUI'),
    (f'PARAM_{max_pg+4:03d}', 'TOLERANCE_ARRONDI_CUMUL_EUR', '1.00',
     'Tolerance ecart arrondi cumule par facture/proprietaire/mois (EUR) - D035',
     date_debut, None, 'OUI'),
]
for row in new_params:
    ws_pg.append(row)
    print(f"[B5] REF_Parametres_Generaux: ajout {row[1]}")

# ── B6 — REF_Intervenants: add columns + fix type casse ──────
ws_int = wb['REF_Intervenants']
rows_int = list(ws_int.iter_rows(values_only=True))
headers_int = list(rows_int[0])
data_int = rows_int[1:]

new_headers = headers_int + ['nom_normalise', 'date_debut_validite', 'date_fin_validite']

# Clear sheet
for row in ws_int.iter_rows():
    for cell in row:
        cell.value = None

# Write new headers
for j, h in enumerate(new_headers, 1):
    ws_int.cell(row=1, column=j, value=h)

# Write corrected data rows
for i, row in enumerate(data_int, 2):
    d = dict(zip(headers_int, row))
    d['type_intervenant'] = fix_type_intervenant(d.get('type_intervenant'))
    d['nom_normalise'] = normalize_name(d.get('nom_intervenant'))
    d['date_debut_validite'] = datetime(2026, 1, 1)
    d['date_fin_validite'] = None
    for j, h in enumerate(new_headers, 1):
        ws_int.cell(row=i, column=j, value=d.get(h))
    print(f"[B6] {d['intervenant_id']}: type={d['type_intervenant']}, nom_normalise={d['nom_normalise']}")

# ── I1 — REF_Logements: add APPARTEMENT_DIVERS + LOGEMENT_DIVERS
ws_log = wb['REF_Logements']
rows_log = list(ws_log.iter_rows(values_only=True))
headers_log = list(rows_log[0])
existing_log_ids = [r[headers_log.index('logement_id')] for r in rows_log[1:] if r[0]]

new_logements = [
    {
        'logement_id': 'APPARTEMENT_DIVERS',
        'hostaway_listing_id': None,
        'nom_logement_officiel': 'Appartement divers (hors parc)',
        'nom_court': 'APPARTEMENT_DIVERS',
        'adresse': None, 'ville': None,
        'type_logement_id': None, 'proprietaire_id': None,
        'date_entree_gestion': None, 'date_sortie_gestion': None,
        'sur_hostaway': 'NON', 'dynamic_pricing': 'non',
        'actif': 'OUI',
        'commentaire': 'Code technique hors parc - jamais utilise pour masquer un mauvais mapping',
        'forfait_logiciel_consommables_mensuel': 0,
    },
    {
        'logement_id': 'LOGEMENT_DIVERS',
        'hostaway_listing_id': None,
        'nom_logement_officiel': 'Logement divers (hors parc)',
        'nom_court': 'LOGEMENT_DIVERS',
        'adresse': None, 'ville': None,
        'type_logement_id': None, 'proprietaire_id': None,
        'date_entree_gestion': None, 'date_sortie_gestion': None,
        'sur_hostaway': 'NON', 'dynamic_pricing': 'non',
        'actif': 'OUI',
        'commentaire': 'Code technique hors parc - jamais utilise pour masquer un mauvais mapping',
        'forfait_logiciel_consommables_mensuel': 0,
    },
]
for log in new_logements:
    if log['logement_id'] not in existing_log_ids:
        row_data = tuple(log.get(h) for h in headers_log)
        ws_log.append(row_data)
        print(f"[I1] REF_Logements: ajout {log['logement_id']}")

# ── SAVE ─────────────────────────────────────────────────────
wb.save(FILEPATH)
wb.close()
print("\nOK REF_Setup.xlsm sauvegarde - toutes corrections B1-B6 + I1 appliquees")
