"""
lot4ter_historiser_reservations_cloturees.py
Historisation des réservations des MOIS CLÔTURÉS — bloc réservations (après lot4bis).

Rôle :
  À la clôture d'un mois, archive TOUTES les réservations validées du mois
  (tous canaux : Airbnb, Booking, VRBO, Direct, hors Hostaway) dans une table
  historique unique et générique. Cette table devient la source de vérité du
  mois clôturé. Elle n'est PAS spécifique Hostaway ni VRBO.

Règles structurantes :
  - Mois clôturé = REF_Cloture_Mensuelle.statut_mois == "CLOTURE" UNIQUEMENT.
    Aucune règle automatique "mois < mois courant". REF vide => 0 mois clôturé.
  - Date de référence métier = check-in (date_arrivee). mois dérivé du check-in.
  - Upsert sur cle_historisation, JAMAIS de suppression : une réservation déjà
    historisée n'est pas réécrite et reste même si elle disparaît des sources live.
  - cle_historisation obligatoire sur toutes les lignes (HH incluses) :
    reservation_id (Hostaway) si disponible, sinon reservation_calc_id (clé HH stable).
  - ROW_HASH calculé sur les champs métier (pas les dates techniques de génération).
  - Backfill VRBO = simple origine de correction ponctuelle (origine_initiale=BACKFILL_VRBO),
    ensuite traité comme une réservation clôturée normale (canal=VRBO).
    Commission VRBO : assiette_commission = payout - coût ménage standard.

Sources (lecture seule) :
  - 02_TRAVAIL/Lot4bis_TableCommune/MASTER_CALC_Reservations.xlsx (onglet MASTER)
  - 02_TRAVAIL/Lot1_Hostaway/MASTER_CALC_HA_Payout.xlsx (financiers HA)
  - 01_SOURCES_BRUTES/REF_Setup/REF_Setup.xlsm
        (REF_Cloture_Mensuelle, REF_Logements, REF_Couts_Standards_Menage)
  - 01_SOURCES_BRUTES/VRBO/IMPORT_UNIQUE_Revenus_*.csv (backfill ponctuel)

Cible :
  - 02_DONNEES_NORMALISEES/historique_reservations/HIST_Reservations_Cloturees.xlsx
"""

import sys
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.stderr.reconfigure(encoding="utf-8", errors="replace")

import os
import csv
import glob
import hashlib
import datetime
import unicodedata
import collections

import openpyxl
from openpyxl.styles import Font, PatternFill

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PATH_RES  = os.path.join(ROOT, "02_TRAVAIL", "Lot4bis_TableCommune", "MASTER_CALC_Reservations.xlsx")
PATH_PAY  = os.path.join(ROOT, "02_TRAVAIL", "Lot1_Hostaway", "MASTER_CALC_HA_Payout.xlsx")
PATH_REF  = os.path.join(ROOT, "01_SOURCES_BRUTES", "REF_Setup", "REF_Setup.xlsm")
GLOB_VRBO = os.path.join(ROOT, "01_SOURCES_BRUTES", "VRBO", "IMPORT_UNIQUE_Revenus_*.csv")
OUT_DIR   = os.path.join(ROOT, "02_DONNEES_NORMALISEES", "historique_reservations")
OUT_FILE  = os.path.join(OUT_DIR, "HIST_Reservations_Cloturees.xlsx")
SHEET     = "HIST_Reservations_Cloturees"

FIGE_LE = datetime.datetime.now().strftime("%Y-%m-%dT%H:%M:%S")

# Valeurs référentielles génériques
SRC_HIST = "HIST_RESERVATIONS_CLOTUREES"
METHODE  = "HIST_PRIME_MOIS_CLOTURE"

COLS = [
    "cle_historisation", "reservation_calc_id",
    "reservation_id_hostaway", "reservation_hh_id",
    "canal", "logement_id", "proprietaire_id",
    "mois", "date_arrivee", "date_depart", "nuits",
    "montant_retenu", "payout_calcule", "menage_retenu", "assiette_commission",
    "code_impact", "impact_resultat_reel", "impact_resultat_comptable",
    "statut_controle", "niveau_anomalie", "code_anomalie",
    "origine_initiale", "source_ligne", "source_montant", "methode",
    "mois_cloture", "fige_le", "ROW_HASH",
]

# source (lot4bis) -> canal générique
CANAL_MAP = {
    "HOSTAWAY_AIRBNB": "AIRBNB",
    "HOSTAWAY_BOOKING": "BOOKING",
    "HOSTAWAY_VRBO_A_CONTROLER": "VRBO",
    "HOSTAWAY_VRBO_HH": "VRBO",
    "HOSTAWAY_DIRECT_HH": "DIRECT",
    "MANUEL_HORS_HOSTAWAY": "HH",
    "OWNERSTAY_EXCLU": "OWNERSTAY",
}


def norm(s):
    s = str(s or "").strip().lower()
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def load_sheet(path, sheet):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
    wb.close()
    if not rows:
        return []
    h = list(rows[0])
    return [dict(zip(h, r)) for r in rows[1:]]


def biz_hash(values):
    s = "|".join("" if v is None else str(v) for v in values)
    return hashlib.sha256(s.encode("utf-8")).hexdigest()[:16]


def parse_fr_date(s):
    months = {'janv': 1, 'fevr': 2, 'mars': 3, 'avr': 4, 'mai': 5, 'juin': 6,
              'juil': 7, 'aout': 8, 'sept': 9, 'oct': 10, 'nov': 11, 'dec': 12}
    s = norm(s).replace("\xa0", " ")
    for sep in ("-", " "):
        p = s.split(sep)
        if len(p) == 3:
            try:
                dd = int(p[0])
                mo = next((v for k, v in months.items() if p[1].startswith(k)), None)
                yy = int(p[2]); yy = 2000 + yy if yy < 100 else yy
                if mo:
                    return datetime.date(yy, mo, dd)
            except ValueError:
                pass
    return None


def closed_months():
    """Mois explicitement CLOTURE dans REF_Cloture_Mensuelle (aucune règle automatique)."""
    out = set()
    for d in load_sheet(PATH_REF, "REF_Cloture_Mensuelle"):
        if str(d.get("statut_mois") or "").strip().upper() == "CLOTURE" and d.get("mois"):
            out.add(str(d["mois"])[:7])
    return out


def cout_menage_standard():
    """(logement_id, date_ref) -> coût ménage standard, via type de logement + validité."""
    logs = {d["logement_id"]: d for d in load_sheet(PATH_REF, "REF_Logements")
            if d.get("logement_id") and str(d["logement_id"]) != "logement_id"}
    couts = [d for d in load_sheet(PATH_REF, "REF_Couts_Standards_Menage") if d.get("type_logement_id")]

    def to_d(v):
        if isinstance(v, (datetime.date, datetime.datetime)):
            return v.date() if isinstance(v, datetime.datetime) else v
        try:
            return datetime.date.fromisoformat(str(v)[:10])
        except (ValueError, TypeError):
            return None

    def lookup(logement_id, date_ref):
        lg = logs.get(logement_id)
        if not lg:
            return None
        type_id = lg.get("type_logement_id")
        dref = to_d(date_ref) or datetime.date.today()
        best = None
        for c in couts:
            if c.get("type_logement_id") != type_id or str(c.get("actif")) != "OUI":
                continue
            deb = to_d(c.get("date_debut_validite"))
            fin = to_d(c.get("date_fin_validite"))
            if deb and dref < deb:
                continue
            if fin and dref > fin:
                continue
            best = c.get("cout_standard_menage")
        return best
    return lookup


def build_vrbo_backfill():
    """reservation_id Hostaway -> net VRBO (csv), via logement + check-in + prénom.
    Réconcilie contre les lignes VRBO de lot4bis MASTER."""
    files = sorted(glob.glob(GLOB_VRBO))
    if not files:
        return {}
    cagg = collections.defaultdict(lambda: {"net": 0.0})
    for r in csv.reader(open(files[-1], encoding="utf-8-sig"), delimiter=";"):
        if not r or r[0] == "N° de propriété":
            continue
        code = r[3].strip()
        try:
            net = float(r[13].replace(",", "."))
        except (ValueError, IndexError):
            net = 0.0
        e = cagg[code]
        e["net"] += net
        e["log"] = r[1].strip()
        e["first"] = norm(r[4])
        e["arr"] = parse_fr_date(r[7])
    return cagg


def main():
    cmonths = closed_months()
    print(f"[lot4ter] mois CLOTURE (REF_Cloture_Mensuelle) : {sorted(cmonths) or 'AUCUN'}")

    res = load_sheet(PATH_RES, "MASTER")
    pay = {r["reservation_id"]: r for r in load_sheet(PATH_PAY, "data")}
    cout_men = cout_menage_standard()
    vrbo = build_vrbo_backfill()

    # Réconciliation backfill VRBO -> reservation_id : par (logement_id + check-in).
    # lot4bis MASTER ne porte pas le prénom voyageur ; (logement + check-in) est
    # suffisamment discriminant pour les lignes VRBO.
    vrbo_rows = [r for r in res
                 if CANAL_MAP.get(r.get("source")) == "VRBO" and r.get("reservation_id_hostaway")]
    csv_by_key = collections.defaultdict(list)
    for code, e in vrbo.items():
        if e.get("arr"):
            csv_by_key[(e["log"], str(e["arr"]))].append((code, round(e["net"], 2)))
    vrbo_net = {}
    for r in vrbo_rows:
        ci = str(r.get("date_arrivee"))[:10]
        cand = csv_by_key.get((r.get("logement_id"), ci))
        if cand:
            vrbo_net[r["reservation_id_hostaway"]] = cand[0][1]

    # HIST existant (upsert sans suppression)
    existing = {}
    if os.path.exists(OUT_FILE):
        for d in load_sheet(OUT_FILE, SHEET):
            existing[d["cle_historisation"]] = [d.get(c) for c in COLS]
    hist = dict(existing)

    added = vrbo_filled = skipped = already = 0
    for r in res:
        mois = str(r.get("mois") or str(r.get("date_arrivee"))[:7])[:7]
        if mois not in cmonths:                 # seul un mois CLOTURE est historisé
            skipped += 1
            continue
        rid_ha = r.get("reservation_id_hostaway")
        rid_hh = r.get("reservation_hh_id")
        calc_id = r.get("reservation_calc_id")
        # clé upsert STABLE (reservation_calc_id régénéré chaque run -> exclu sauf dernier recours)
        cle = str(rid_ha) if rid_ha else (str(rid_hh) if rid_hh else str(calc_id))
        if cle in hist:                          # déjà figé -> jamais réécrit
            already += 1
            continue

        canal = CANAL_MAP.get(r.get("source"), "INCONNU")
        montant = r.get("montant_retenu")
        p = pay.get(rid_ha, {}) if rid_ha else {}
        payout_c = p.get("payout_calcule")
        menage   = p.get("menage_retenu")
        assiette = p.get("assiette_commission")
        statut   = r.get("statut_controle")
        code_ano = r.get("code_anomalie")
        origine  = "API_HOSTAWAY" if rid_ha else "SAISIE_HH"

        # Correction backfill VRBO : payout + ménage standard + assiette
        if canal == "VRBO" and rid_ha in vrbo_net:
            payout_c = vrbo_net[rid_ha]
            menage = cout_men(r.get("logement_id"), r.get("date_arrivee")) or 0
            assiette = round(payout_c - menage, 2)
            montant = payout_c
            statut = "VALIDE"
            code_ano = None
            origine = "BACKFILL_VRBO"
            vrbo_filled += 1

        biz = biz_hash([cle, canal, r.get("logement_id"), mois, montant, payout_c, code_ano])
        hist[cle] = [
            cle, calc_id, rid_ha, rid_hh,
            canal, r.get("logement_id"), r.get("proprietaire_id"),
            mois, str(r.get("date_arrivee"))[:10], str(r.get("date_depart"))[:10], r.get("nuits"),
            montant, payout_c, menage, assiette,
            r.get("code_impact"), r.get("impact_resultat_reel"), r.get("impact_resultat_comptable"),
            statut, r.get("niveau_anomalie"), code_ano,
            origine, SRC_HIST, SRC_HIST, METHODE,
            mois, FIGE_LE, biz,
        ]
        added += 1

    rows_out = sorted(hist.values(), key=lambda x: (str(x[7]), str(x[8])))

    os.makedirs(OUT_DIR, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(COLS)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDDDDD")
    for r in rows_out:
        ws.append(r)
    wb.save(OUT_FILE)

    by_canal = collections.Counter(r[4] for r in rows_out)
    print(f"\n[lot4ter] HIST écrit : {OUT_FILE}")
    print(f"  lignes HIST totales        : {len(rows_out)}")
    print(f"  conservées (déjà figées)   : {len(existing)}")
    print(f"  nouvelles archivées        : {added}")
    print(f"  backfill VRBO appliqués     : {vrbo_filled}")
    print(f"  lignes ignorées (non clôt.) : {skipped}")
    print(f"  par canal                  : {dict(by_canal)}")
    if not cmonths:
        print("  AVERTISSEMENT : aucun mois CLOTURE dans REF_Cloture_Mensuelle -> HIST inchangé/vide.")


if __name__ == "__main__":
    main()
