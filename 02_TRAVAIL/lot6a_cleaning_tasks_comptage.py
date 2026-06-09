#!/usr/bin/env python3
"""
Lot 6a — Hostaway CleaningTasks : comptage ménages
=====================================================
Peuple MASTER_FACT_HA_CleaningTasks_Discovery.xlsx :
  data           : tâches brutes Hostaway (11 cols)
  MASTER_ENRICHI : enrichi — logement, mois, statut_menage, contrôles (21 cols)
  VUE_COMPTAGE   : agrégat mois×logement (11 cols)
  POWER_QUERY_CODE : code M de référence

Règles irrévocables :
  H6       — cost = NULL dans les onglets (API renvoie un cost sur 22/500 tâches — ignoré)
  D065     — Extraction segmentée par listingMapId (seul filtre efficace).
             Un seul appel global retourne 500 max (plafond API, offset ignoré).
             Par listing : chaque segment doit être < 500 sinon BLOQUANT.
             Déduplication par task_id après agrégation des segments.
  QM-L6a-02 — confirmed ≠ réalisé → statut_menage = 'prévu'
  QM-L6a-03 — pending → statut_menage = 'A_CONTROLER'
  QM-L6a-04 — type_ligne_menage default = TLM_001 MENAGE_STANDARD
"""

import os, sys, hashlib, logging
from collections import Counter, defaultdict
from copy import copy
from pathlib import Path
from datetime import datetime, timezone

import requests
import openpyxl as xl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)

# ══════════════════════════════════════════════════════════════
# PATHS & CONSTANTS
# ══════════════════════════════════════════════════════════════
ROOT     = Path(__file__).parent.parent
REF_PATH = ROOT / "01_SOURCES_BRUTES/REF_Setup/REF_Setup.xlsm"
OUT_PATH = ROOT / "02_TRAVAIL/Lot1_Hostaway/MASTER_FACT_HA_CleaningTasks_Discovery.xlsx"
BASE_URL = "https://api.hostaway.com"
H6_NOTE  = "cost=NULL_H6_comptage_uniquement"

# QM-L6a-02 / QM-L6a-03
STATUS_MAP = {
    "completed": "réalisé",
    "confirmed": "prévu",       # QM-L6a-02 : confirmed ≠ réalisé
    "pending":   "A_CONTROLER", # QM-L6a-03
    "cancelled": "annulé",
}

DEFAULT_TLM_ID  = "TLM_001"   # QM-L6a-04
DEFAULT_TLM_LIB = "MENAGE_STANDARD"
DEFAULT_CCM     = "OUI"

# ══════════════════════════════════════════════════════════════
# STYLES
# ══════════════════════════════════════════════════════════════
def _fill(c):
    return PatternFill("solid", fgColor=c)

_C_IDENT  = _fill("4472C4")  # bleu — identifiants
_C_RATT   = _fill("70AD47")  # vert — rattachement
_C_FLAG   = _fill("FFC000")  # or — flag H6
_C_STATUT = _fill("ED7D31")  # orange — contrôles
_C_META   = _fill("BFBFBF")  # gris — metadata
_C_CNT    = _fill("5B9BD5")  # bleu clair — comptages
_C_HDR    = _fill("1F3864")  # bleu foncé — headers
_WHITE_BOLD = Font(color="FFFFFF", bold=True, size=10)
_BOLD       = Font(bold=True, size=10)

def _style_headers(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.fill = copy(_C_HDR)
        cell.font = copy(_WHITE_BOLD)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _style_col(ws, col_idx, nrows, fill_style):
    for r in range(2, nrows + 2):
        ws.cell(r, col_idx).fill = copy(fill_style)

def _autowidth(ws, mn=10, mx=48):
    for col in ws.columns:
        w = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w + 2, mn), mx)

# ══════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════
def _now():
    return datetime.now(timezone.utc).isoformat(timespec="seconds")

def _hash(*vals):
    return hashlib.md5("|".join(str(v) for v in vals).encode()).hexdigest()[:16]

# ══════════════════════════════════════════════════════════════
# HOSTAWAY AUTH + API
# ══════════════════════════════════════════════════════════════
def _get_token():
    cid = os.getenv("HOSTAWAY_CLIENT_ID") or os.getenv("CLIENT_ID")
    sec = os.getenv("HOSTAWAY_CLIENT_SECRET") or os.getenv("CLIENT_SECRET")
    if not (cid and sec):
        raise RuntimeError("CLIENT_ID / CLIENT_SECRET absent du .env")
    r = requests.post(
        f"{BASE_URL}/v1/accessTokens",
        data={"grant_type": "client_credentials", "client_id": cid,
              "client_secret": sec, "scope": "general"},
        timeout=20,
    )
    r.raise_for_status()
    token = r.json().get("access_token", "")
    if not token:
        raise RuntimeError("Token vide — vérifier credentials")
    return token

def fetch_tasks_per_listing(token, account_id, listing_ids: list) -> tuple:
    """Extraction segmentée par listingMapId — D065 (méthode fiable).

    Retourne (tasks_deduped, extraction_report) où extraction_report est un dict
    contenant les stats par segment pour la validation anti-plafond.

    Règle : si un segment retourne >= 500 → BLOQUANT (données probablement tronquées).
    Déduplication : task_id (fallback : reservationId|listingMapId|canStartFrom|autoTaskId).
    """
    hdrs = {
        "Authorization": f"Bearer {token}",
        "account-id": str(account_id),
        "Content-Type": "application/json",
    }

    all_raw: list = []
    segment_counts: dict = {}
    plateaued: list = []

    for listing_id in listing_ids:
        r = requests.get(
            f"{BASE_URL}/v1/tasks",
            headers=hdrs,
            params={"listingMapId": str(listing_id), "limit": 500},
            timeout=30,
        )
        r.raise_for_status()
        batch = r.json().get("result", [])
        n = len(batch)
        segment_counts[listing_id] = n
        if n >= 500:
            plateaued.append(listing_id)
            log.error(f"  listing {listing_id}: {n} tâches — SEGMENT PLAFONNÉ (troncature probable)")
        else:
            log.info(f"  listing {listing_id}: {n} tâches")
        all_raw.extend(batch)

    # Déduplication par task_id
    seen: set = set()
    deduped: list = []
    for t in all_raw:
        tid = t.get("id")
        if tid is not None:
            key = tid
        else:
            key = (t.get("reservationId"), t.get("listingMapId"),
                   t.get("canStartFrom"), t.get("autoTaskId"))
        if key not in seen:
            seen.add(key)
            deduped.append(t)

    report = {
        "n_segments":        len(listing_ids),
        "n_requests":        len(listing_ids),
        "n_brut":            len(all_raw),
        "n_unique":          len(deduped),
        "n_doublons":        len(all_raw) - len(deduped),
        "segment_counts":    segment_counts,
        "plateaued":         plateaued,
    }
    log.info(f"API /v1/tasks — {len(listing_ids)} requêtes → brut={len(all_raw)}, "
             f"unique={len(deduped)}, doublons={len(all_raw)-len(deduped)}")
    return deduped, report

# ══════════════════════════════════════════════════════════════
# REF LOADING
# ══════════════════════════════════════════════════════════════
def load_ref_logements():
    """Retourne {hostaway_listing_id (int): (logement_id, proprietaire_id, actif_flag)}.
    Inclut logements inactifs (actif='NON') pour ne pas orpheliner les tâches historiques.
    """
    wb = xl.load_workbook(REF_PATH, read_only=True, data_only=True)
    ws = wb["REF_Logements"]
    rows = list(ws.iter_rows(values_only=True))
    h    = list(rows[0])
    i_lid, i_haid, i_pid, i_act = (
        h.index("logement_id"), h.index("hostaway_listing_id"),
        h.index("proprietaire_id"), h.index("actif"),
    )
    mapping = {
        int(r[i_haid]): (r[i_lid], r[i_pid], r[i_act])
        for r in rows[1:]
        if r[i_haid] is not None
    }
    wb.close()
    n_act = sum(1 for v in mapping.values() if v[2] == "OUI")
    n_ina = sum(1 for v in mapping.values() if v[2] != "OUI")
    log.info(f"REF_Logements : {n_act} actifs + {n_ina} inactifs mappés")
    return mapping

def load_ref_tlm():
    """Retourne {type_ligne_menage_id: compte_comme_menage}."""
    wb = xl.load_workbook(REF_PATH, read_only=True, data_only=True)
    ws = wb["REF_Types_Lignes_Menage"]
    rows = list(ws.iter_rows(values_only=True))
    h    = list(rows[0])
    i_id, i_ccm = h.index("type_ligne_menage_id"), h.index("compte_comme_menage")
    tlm = {r[i_id]: r[i_ccm] for r in rows[1:] if r[i_id]}
    wb.close()
    log.info(f"REF_Types_Lignes_Menage : {len(tlm)} types chargés")
    return tlm

# ══════════════════════════════════════════════════════════════
# SCHÉMAS COLONNES
# ══════════════════════════════════════════════════════════════
DATA_HEADERS = [
    "task_id", "reservation_id", "listingMapId", "title", "status",
    "canStartFrom", "assigneeUserId", "cost", "h6_note", "extrait_le", "ROW_HASH",
]

ENRICHI_HEADERS = [
    "task_id", "ROW_HASH",
    "mois", "logement_id", "proprietaire_id",
    "listingMapId", "reservation_id", "scheduled_date", "title",
    "status", "statut_menage",
    "type_ligne_menage_id", "type_ligne_menage_lib", "compte_comme_menage",
    "cost", "h6_note",
    "statut_controle", "niveau_anomalie", "code_anomalie",
    "extrait_le", "date_integration",
]

ENRICHI_COLORS = {
    "task_id": _C_IDENT, "ROW_HASH": _C_IDENT,
    "mois": _C_RATT, "logement_id": _C_RATT, "proprietaire_id": _C_RATT,
    "listingMapId": _C_RATT, "reservation_id": _C_RATT, "scheduled_date": _C_RATT,
    "title": _C_RATT, "status": _C_RATT, "statut_menage": _C_RATT,
    "type_ligne_menage_id": _C_RATT, "type_ligne_menage_lib": _C_RATT,
    "compte_comme_menage": _C_RATT,
    "cost": _C_FLAG, "h6_note": _C_FLAG,
    "statut_controle": _C_STATUT, "niveau_anomalie": _C_STATUT, "code_anomalie": _C_STATUT,
    "extrait_le": _C_META, "date_integration": _C_META,
}

COMPTAGE_HEADERS = [
    "mois", "logement_id", "proprietaire_id",
    "nb_menages_realises", "nb_menages_confirmes",
    "nb_menages_pending", "nb_menages_annules",
    "nb_taches_total",
    "statut_controle", "niveau_anomalie", "code_anomalie",
]

COMPTAGE_COLORS = {
    "mois": _C_RATT, "logement_id": _C_RATT, "proprietaire_id": _C_RATT,
    "nb_menages_realises": _C_CNT, "nb_menages_confirmes": _C_CNT,
    "nb_menages_pending": _C_CNT, "nb_menages_annules": _C_CNT,
    "nb_taches_total": _C_CNT,
    "statut_controle": _C_STATUT, "niveau_anomalie": _C_STATUT, "code_anomalie": _C_STATUT,
}

# ══════════════════════════════════════════════════════════════
# ROW BUILDERS
# ══════════════════════════════════════════════════════════════
def build_data_rows(tasks):
    extrait = _now()
    rows = []
    for t in tasks:
        rows.append([
            t.get("id"),
            t.get("reservationId"),
            t.get("listingMapId"),
            t.get("title", ""),
            t.get("status"),
            t.get("canStartFrom"),
            t.get("assigneeUserId"),
            None,        # cost — H6 irrévocable
            H6_NOTE,
            extrait,
            _hash(t.get("id")),
        ])
    return rows


def build_enrichi_rows(tasks, mapping, tlm):
    extrait = _now()
    rows = []
    for t in tasks:
        tid   = t.get("id")
        resid = t.get("reservationId")
        mapid = t.get("listingMapId")
        title = t.get("title", "")
        status = t.get("status", "")
        date_s = t.get("canStartFrom") or ""
        mois   = date_s[:7] if len(date_s) >= 7 else None

        statut_menage = STATUS_MAP.get(status, status)

        logement_id = None
        proprietaire_id = None
        logement_inactif = False
        if mapid is not None:
            entry = mapping.get(int(mapid))
            if entry:
                logement_id, proprietaire_id, actif_flag = entry
                logement_inactif = (actif_flag != "OUI")

        ccm = tlm.get(DEFAULT_TLM_ID, DEFAULT_CCM)

        # Controls
        codes = []
        if logement_id is None:
            codes.append("TASK_LOGEMENT_ABSENT")   # BLOQUANT : listingMapId inconnu
        elif logement_inactif:
            codes.append("TASK_LOGEMENT_INACTIF")  # A_CONTROLER : logement désactivé
        if resid is None:
            codes.append("TASK_SANS_RESERVATION")
        if status == "pending":
            codes.append("TASK_STATUT_PENDING")

        if "TASK_LOGEMENT_ABSENT" in codes:
            statut_c = "BLOQUANT"
            niveau   = "BLOQUANT"
        elif codes:
            statut_c = "A_CONTROLER"
            niveau   = "A_CONTROLER"
        else:
            statut_c = "OK"
            niveau   = ""

        rows.append([
            tid, _hash(tid),
            mois, logement_id, proprietaire_id,
            mapid, resid, date_s, title,
            status, statut_menage,
            DEFAULT_TLM_ID, DEFAULT_TLM_LIB, ccm,
            None,   # cost — H6
            H6_NOTE,
            statut_c, niveau, "|".join(codes) if codes else "",
            extrait, extrait,
        ])
    return rows


def build_comptage_rows(enrichi_rows):
    I = {h: i for i, h in enumerate(ENRICHI_HEADERS)}

    groups = defaultdict(list)
    for r in enrichi_rows:
        key = (r[I["mois"]], r[I["logement_id"]])
        groups[key].append(r)

    rows = []
    for (mois, logement_id), grp in sorted(
        groups.items(), key=lambda x: (x[0][0] or "", str(x[0][1] or ""))
    ):
        prop_id = next(
            (r[I["proprietaire_id"]] for r in grp if r[I["proprietaire_id"]]), None
        )

        def cnt(statut_val, ccm_filter=True):
            return sum(
                1 for r in grp
                if r[I["statut_menage"]] == statut_val
                and (not ccm_filter or r[I["compte_comme_menage"]] == "OUI")
            )

        nb_real = cnt("réalisé")
        nb_conf = cnt("prévu")
        nb_pend = cnt("A_CONTROLER")
        nb_ann  = cnt("annulé", ccm_filter=False)
        nb_tot  = len(grp)

        codes = []
        if logement_id is None:
            codes.append("COMPTAGE_LOGEMENT_ABSENT")

        if "COMPTAGE_LOGEMENT_ABSENT" in codes:
            statut_c, niveau = "BLOQUANT", "BLOQUANT"
        elif codes:
            statut_c, niveau = "A_CONTROLER", "A_CONTROLER"
        else:
            statut_c, niveau = "OK", ""

        rows.append([
            mois, logement_id, prop_id,
            nb_real, nb_conf, nb_pend, nb_ann, nb_tot,
            statut_c, niveau, "|".join(codes) if codes else "",
        ])

    return rows

# ══════════════════════════════════════════════════════════════
# POWER QUERY CODE DE RÉFÉRENCE
# ══════════════════════════════════════════════════════════════
PQ_CODE = """\
// ============================================================
// LOT 6A — POWER QUERY CODE DE RÉFÉRENCE
// Fichier : MASTER_FACT_HA_CleaningTasks_Discovery.xlsx
// Source  : Hostaway API /v1/tasks
// D065    : API retourne 500 max, offset ignoré → single call
// H6      : cost = NULL — comptage uniquement, jamais valorisation
// QM-L6a-02 : confirmed = 'prévu' (≠ réalisé)
// QM-L6a-03 : pending = 'A_CONTROLER'
// QM-L6a-04 : type_ligne_menage default = TLM_001 MENAGE_STANDARD
// ============================================================

// Q1 — REF_Logements
let
    Source  = Excel.Workbook(File.Contents("[CHEMIN_ABSOLU]/REF_Setup.xlsm"), null, true),
    Sheet   = Source{[Name="REF_Logements"]}[Data],
    Headers = Table.PromoteHeaders(Sheet, [PromoteAllScalars=true]),
    Actifs  = Table.SelectRows(Headers, each [actif] = "OUI"),
    Typed   = Table.TransformColumnTypes(Actifs, {
                  {"logement_id",          type text},
                  {"hostaway_listing_id",  Int64.Type},
                  {"proprietaire_id",      type text}
              })
in  Typed

// Q2 — data (tâches brutes)
let
    Source = Excel.CurrentWorkbook(){[Name="data"]}[Content],
    Typed  = Table.TransformColumnTypes(Source, {
                 {"task_id",        Int64.Type},
                 {"reservation_id", Int64.Type},
                 {"listingMapId",   Int64.Type},
                 {"title",          type text},
                 {"status",         type text},
                 {"canStartFrom",   type text},
                 {"cost",           type any}
             })
in  Typed

// Q3 — MASTER_ENRICHI
let
    Tasks     = Q2,
    Logements = Q1,
    Joined    = Table.NestedJoin(Tasks, {"listingMapId"},
                    Logements, {"hostaway_listing_id"}, "ref", JoinKind.Left),
    Expanded  = Table.ExpandTableColumn(Joined, "ref",
                    {"logement_id", "proprietaire_id"}),
    AddMois   = Table.AddColumn(Expanded, "mois",
                    each if [canStartFrom] <> null
                         then Text.Start([canStartFrom], 7) else null),
    AddStatut = Table.AddColumn(AddMois, "statut_menage",
                    each if [status] = "completed" then "réalisé"
                    else if [status] = "confirmed" then "prévu"
                    else if [status] = "pending"   then "A_CONTROLER"
                    else if [status] = "cancelled" then "annulé"
                    else [status]),
    AddTLM    = Table.AddColumn(AddStatut, "type_ligne_menage_id", each "TLM_001"),
    AddCCM    = Table.AddColumn(AddTLM,    "compte_comme_menage",  each "OUI"),
    AddCost   = Table.AddColumn(AddCCM,    "cost",    each null),
    AddH6     = Table.AddColumn(AddCost,   "h6_note", each "cost=NULL_H6_comptage_uniquement")
in  AddH6

// Q4 — VUE_COMPTAGE (agrégat mois × logement, compte_comme_menage=OUI)
let
    Enrichi = Q3,
    Comptes = Table.SelectRows(Enrichi, each [compte_comme_menage] = "OUI"),
    Grouped = Table.Group(Comptes, {"mois", "logement_id", "proprietaire_id"}, {
        {"nb_menages_realises",  each Table.RowCount(
            Table.SelectRows(_, each [statut_menage] = "réalisé")),   Int64.Type},
        {"nb_menages_confirmes", each Table.RowCount(
            Table.SelectRows(_, each [statut_menage] = "prévu")),     Int64.Type},
        {"nb_menages_pending",   each Table.RowCount(
            Table.SelectRows(_, each [statut_menage] = "A_CONTROLER")), Int64.Type},
        {"nb_menages_annules",   each Table.RowCount(
            Table.SelectRows(_, each [statut_menage] = "annulé")),    Int64.Type},
        {"nb_taches_total",      each Table.RowCount(_),              Int64.Type}
    }),
    Sorted  = Table.Sort(Grouped, {{"mois", Order.Ascending},
                                   {"logement_id", Order.Ascending}})
in  Sorted
"""

# ══════════════════════════════════════════════════════════════
# EXCEL WRITING
# ══════════════════════════════════════════════════════════════
def _write_sheet(wb, name, headers, rows, colors, position=None):
    if name in wb.sheetnames:
        del wb[name]
    if position is not None:
        ws = wb.create_sheet(name, position)
    else:
        ws = wb.create_sheet(name)

    ws.append(headers)
    for r in rows:
        ws.append(r)

    nrows = len(rows)
    for ci, h in enumerate(headers, 1):
        if h in colors:
            _style_col(ws, ci, nrows, colors[h])

    _style_headers(ws, len(headers))
    ws.freeze_panes = "B2"
    _autowidth(ws)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    log.info(f"  onglet {name}: {nrows} lignes, {len(headers)} cols")
    return ws


def _write_pq_sheet(wb):
    name = "POWER_QUERY_CODE"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.column_dimensions["A"].width = 110
    for i, line in enumerate(PQ_CODE.split("\n"), 1):
        c = ws.cell(i, 1, line)
        c.font = Font(name="Consolas", size=9)
    log.info("  onglet POWER_QUERY_CODE: écrit")
    return ws

# ══════════════════════════════════════════════════════════════
# VALIDATION
# ══════════════════════════════════════════════════════════════
def validate(data_rows, enrichi_rows, comptage_rows, tasks_raw, extraction_report):
    """Validation complète avec contrôle anti-plafond et stats exhaustives."""
    log.info("─" * 60)
    log.info("VALIDATION")
    fails = 0
    I   = {h: i for i, h in enumerate(ENRICHI_HEADERS)}
    I_c = {h: i for i, h in enumerate(COMPTAGE_HEADERS)}

    # ── Extraction anti-plafond ───────────────────────────────
    n_req  = extraction_report["n_requests"]
    n_brut = extraction_report["n_brut"]
    n_uniq = extraction_report["n_unique"]
    n_dup  = extraction_report["n_doublons"]
    plaf   = extraction_report["plateaued"]

    log.info(f"  Extraction : {n_req} requêtes API (1 par listing)")
    log.info(f"  Brut total : {n_brut} | Unique : {n_uniq} | Doublons supprimés : {n_dup}")
    if plaf:
        log.error(f"  FAIL anti-plafond : {len(plaf)} segment(s) plafonné(s) → {plaf}")
        fails += 1
    else:
        log.info(f"  OK anti-plafond : 0 segment plafonné (tous < 500)")

    # Comptes par segment (top 3 + min)
    sc = extraction_report["segment_counts"]
    non_zero = {k: v for k, v in sc.items() if v > 0}
    log.info(f"  Listings avec tâches : {len(non_zero)}/{len(sc)} "
             f"| max={max(sc.values())} | min_non_zero={min(non_zero.values()) if non_zero else 0}")

    # ── data onglet ───────────────────────────────────────────
    if not data_rows:
        log.error("  FAIL : data onglet vide")
        fails += 1
    else:
        log.info(f"  OK data : {len(data_rows)} lignes, {len(DATA_HEADERS)} cols")

    # H6 : cost=NULL dans les onglets
    cost_idx = DATA_HEADERS.index("cost")
    non_null = sum(1 for r in data_rows if r[cost_idx] is not None)
    if non_null:
        log.error(f"  FAIL H6 : {non_null} tâche(s) avec cost != NULL dans data onglet")
        fails += 1
    else:
        cost_api_nonnull = sum(1 for t in tasks_raw if t.get("cost") is not None)
        log.info(f"  OK H6 : cost=NULL dans onglets (API : {cost_api_nonnull}/500 avaient un coût — forcé NULL)")

    # ── Champs manquants (sur tâches brutes) ─────────────────
    no_res  = sum(1 for t in tasks_raw if not t.get("reservationId"))
    no_map  = sum(1 for t in tasks_raw if not t.get("listingMapId"))
    no_date = sum(1 for t in tasks_raw if not t.get("canStartFrom"))
    log.info(f"  Sans reservation_id : {no_res}")
    log.info(f"  Sans listingMapId   : {no_map}")
    log.info(f"  Sans scheduled_date : {no_date}")

    # ── Status + task_type ────────────────────────────────────
    raw_statuses = Counter(t.get("status") for t in tasks_raw)
    log.info(f"  Status API : {dict(raw_statuses)}")
    auto_ids = Counter(t.get("autoTaskId") for t in tasks_raw)
    log.info(f"  autoTaskId distincts : {len(auto_ids)}")

    # ── Période ───────────────────────────────────────────────
    dates = sorted(t.get("canStartFrom", "")[:7] for t in tasks_raw if t.get("canStartFrom"))
    if dates:
        log.info(f"  Période : {dates[0]} → {dates[-1]}")

    # ── MASTER_ENRICHI ────────────────────────────────────────
    n_blq = sum(1 for r in enrichi_rows if r[I["niveau_anomalie"]] == "BLOQUANT")
    n_act = sum(1 for r in enrichi_rows if r[I["niveau_anomalie"]] == "A_CONTROLER")
    n_ok  = len(enrichi_rows) - n_blq - n_act
    log.info(f"  MASTER_ENRICHI : {len(enrichi_rows)} lignes — BLOQUANT={n_blq} A_CONTROLER={n_act} OK={n_ok}")

    sc2 = Counter(r[I["statut_menage"]] for r in enrichi_rows)
    log.info(f"  Statuts : réalisé={sc2.get('réalisé',0)} prévu={sc2.get('prévu',0)} "
             f"A_CONTROLER={sc2.get('A_CONTROLER',0)} annulé={sc2.get('annulé',0)}")

    # ── VUE_COMPTAGE ──────────────────────────────────────────
    n_blq_c = sum(1 for r in comptage_rows if r[I_c["niveau_anomalie"]] == "BLOQUANT")
    log.info(f"  VUE_COMPTAGE : {len(comptage_rows)} lignes — BLOQUANT={n_blq_c}")

    mois_cnt = Counter(r[I_c["mois"]] for r in comptage_rows)
    log.info("  Distribution mois×logement :")
    for m in sorted(mois_cnt):
        log.info(f"    {m} : {mois_cnt[m]} logements")

    total_real = sum(r[I_c["nb_menages_realises"]] for r in comptage_rows)
    log.info(f"  Total ménages réalisés : {total_real}")

    log.info(f"  VALIDATION : {fails} FAIL(s)")
    log.info("─" * 60)
    return fails

# ══════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════
def main():
    log.info("=" * 60)
    log.info("LOT 6A — Hostaway CleaningTasks : comptage ménages")
    log.info("=" * 60)

    account_id = os.getenv("HOSTAWAY_ACCOUNT_ID") or os.getenv("ACCOUNT_ID")
    if not account_id:
        sys.exit("ERREUR : ACCOUNT_ID absent du .env")

    if not REF_PATH.exists():
        sys.exit(f"ERREUR : REF_Setup introuvable → {REF_PATH}")

    # 1. Référentiels
    log.info("Chargement REF_Setup...")
    mapping = load_ref_logements()
    tlm     = load_ref_tlm()

    # Liste de tous les listing IDs connus (actifs + inactifs)
    listing_ids = list(mapping.keys())
    log.info(f"Listings à interroger : {len(listing_ids)} (actifs + inactifs)")

    # 2. API Hostaway — extraction segmentée par listingMapId (D065)
    log.info("Authentification Hostaway...")
    token = _get_token()
    log.info("Authentification OK.")
    log.info("Extraction segmentée par listingMapId (D065 — anti-plafond)...")
    tasks, extraction_report = fetch_tasks_per_listing(token, account_id, listing_ids)
    if not tasks:
        sys.exit("ERREUR : aucune tâche retournée")

    # 3. Construction lignes
    log.info("Construction onglets...")
    data_rows     = build_data_rows(tasks)
    enrichi_rows  = build_enrichi_rows(tasks, mapping, tlm)
    comptage_rows = build_comptage_rows(enrichi_rows)

    # 4. Validation (avec rapport extraction)
    fails = validate(data_rows, enrichi_rows, comptage_rows, tasks, extraction_report)

    # 5. Écriture Excel
    log.info(f"Écriture → {OUT_PATH}")
    wb = xl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _write_sheet(wb, "data", DATA_HEADERS, data_rows,
                 {}, position=0)
    _write_sheet(wb, "MASTER_ENRICHI", ENRICHI_HEADERS, enrichi_rows,
                 ENRICHI_COLORS)
    _write_sheet(wb, "VUE_COMPTAGE", COMPTAGE_HEADERS, comptage_rows,
                 COMPTAGE_COLORS)
    _write_pq_sheet(wb)

    wb.save(OUT_PATH)
    log.info(f"Fichier sauvegardé : {OUT_PATH}")

    # Résumé
    log.info("=" * 60)
    log.info("RÉSUMÉ LOT 6A")
    log.info(f"  Requêtes API     : {extraction_report['n_requests']}")
    log.info(f"  Tâches brutes    : {extraction_report['n_brut']}")
    log.info(f"  Tâches uniques   : {extraction_report['n_unique']}")
    log.info(f"  Doublons supprim.: {extraction_report['n_doublons']}")
    log.info(f"  Segs. plafonnés  : {extraction_report['plateaued'] or 'aucun'}")
    log.info(f"  data             : {len(data_rows)} lignes")
    log.info(f"  MASTER_ENRICHI   : {len(enrichi_rows)} lignes — 21 cols")
    log.info(f"  VUE_COMPTAGE     : {len(comptage_rows)} lignes — 11 cols")
    log.info(f"  Onglets          : data | MASTER_ENRICHI | VUE_COMPTAGE | POWER_QUERY_CODE")
    if fails:
        log.warning(f"  ATTENTION : {fails} FAIL(s) — vérifier avant livraison")
        sys.exit(1)
    else:
        log.info("  STATUT : OK — 0 FAIL")
    log.info("=" * 60)


if __name__ == "__main__":
    main()
