#!/usr/bin/env python3
"""
lot5_master_acomptes_proprietaires.py
Lot 5 — Acomptes propriétaires
Crée :
  01_SOURCES_BRUTES/AcomptesProprietaires/SAISIE_AcomptesProprietaires.xlsx  (18 cols, 4 onglets)
  02_TRAVAIL/Lot5_AcomptesProprietaires/MASTER_FACT_MAN_AcomptesProprietaires.xlsx  (22 cols, 3 onglets)
"""

import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

BASE = r"C:\Users\Ewan\OneDrive\Documents\Conciergerie\Pilotage_Conciergerie"
REF_SETUP = os.path.join(BASE, "01_SOURCES_BRUTES", "REF_Setup", "REF_Setup.xlsm")
HH_MASTER = os.path.join(BASE, "02_TRAVAIL", "Lot4_ReservationsHH",
                          "MASTER_FACT_MAN_ReservationsHorsHostaway.xlsx")
SAISIE_OUT = os.path.join(BASE, "01_SOURCES_BRUTES", "AcomptesProprietaires",
                           "SAISIE_AcomptesProprietaires.xlsx")
MASTER_OUT = os.path.join(BASE, "02_TRAVAIL", "Lot5_AcomptesProprietaires",
                           "MASTER_FACT_MAN_AcomptesProprietaires.xlsx")

# ── Palette couleurs ──────────────────────────────────────────────────────────
C_IDENT  = "1F497D"
C_RATT   = "2E75B6"
C_FIN    = "7030A0"
C_MODE   = "4472C4"
C_IMPACT = "C00000"
C_STATUT = "7B0000"
C_COMM   = "595959"
C_PQ     = "404040"

# ── Colonnes SAISIE (18) ──────────────────────────────────────────────────────
# (nom, couleur_header, largeur_col)
SAISIE_COLS = [
    ("acompte_id",                C_IDENT,  16),
    ("ROW_HASH",                  C_IDENT,  34),
    ("mois",                      C_RATT,   10),
    ("proprietaire_id",           C_RATT,   16),
    ("logement_id",               C_RATT,   16),
    ("facture_ref",               C_RATT,   24),
    ("source_acompte",            C_RATT,   20),
    ("source_hh_id",              C_RATT,   24),
    ("montant_acompte",           C_FIN,    16),
    ("report_mois_precedent",     C_FIN,    22),
    ("mode_paiement_id",          C_MODE,   18),
    ("code_impact",               C_IMPACT, 12),
    ("impact_resultat_reel",      C_IMPACT, 18),
    ("impact_resultat_comptable", C_IMPACT, 22),
    ("statut_controle",           C_STATUT, 18),
    ("niveau_anomalie",           C_STATUT, 16),
    ("code_anomalie",             C_STATUT, 34),
    ("commentaire",               C_COMM,   42),
]

# ── Colonnes PQ supplémentaires MASTER (4) ────────────────────────────────────
PQ_COLS = [
    ("source_module",    C_PQ, 32),
    ("source_table",     C_PQ, 32),
    ("source_pk",        C_PQ, 24),
    ("date_integration", C_PQ, 22),
]

MASTER_COLS = SAISIE_COLS + PQ_COLS

# ── Contrôles (5 BLOQUANT + 5 A_CONTROLER) ───────────────────────────────────
CONTROLS = [
    ("ACOMPTE_NON_RATTACHE_FACTURE",     "BLOQUANT",
     "facture_ref vide pour ligne non EXCLU_RESULTAT",
     "facture_ref",
     "facture_ref null ou vide ET statut_controle <> EXCLU_RESULTAT"),
    ("ACOMPTE_HH_INCOHERENT",            "BLOQUANT",
     "montant_acompte ≠ acompte_facture dans MASTER_FACT_MAN HH",
     "montant_acompte",
     "source_acompte=HH_RESERVATION ET montant_acompte ≠ acompte_facture Lot 4"),
    ("ACOMPTE_CALC_ID_DUPLIQUE",         "BLOQUANT",
     "acompte_id dupliqué dans MASTER",
     "acompte_id",
     "acompte_id présent plusieurs fois dans la table"),
    ("ACOMPTE_PROPRIETAIRE_ABSENT",      "BLOQUANT",
     "proprietaire_id vide",
     "proprietaire_id",
     "proprietaire_id null ou vide"),
    ("ACOMPTE_MONTANT_INVALIDE",         "BLOQUANT",
     "montant_acompte null ou ≤ 0",
     "montant_acompte",
     "montant_acompte null ou <= 0"),
    ("ACOMPTE_LOGEMENT_ABSENT",          "A_CONTROLER",
     "logement_id vide alors que ventilation par logement attendue",
     "logement_id",
     "logement_id null ou vide"),
    ("ACOMPTE_SOURCE_HH_INTROUVABLE",    "A_CONTROLER",
     "source_hh_id renseigné mais absent de MASTER HH VALIDE + acompte>0",
     "source_hh_id",
     "source_hh_id non null ET absent de HH_Acomptes_Ref (VALIDE + acompte_facture > 0)"),
    ("ACOMPTE_REPORT_INCOHERENT",        "A_CONTROLER",
     "report_mois_precedent > 0 renseigné sans commentaire de justification",
     "report_mois_precedent",
     "report_mois_precedent > 0 ET commentaire null ou vide"),
    ("ACOMPTE_SOURCE_A_CONTROLER",       "A_CONTROLER",
     "source_acompte = AUTRE — vérification requise",
     "source_acompte",
     "source_acompte = AUTRE"),
    ("ACOMPTE_FACTURE_REF_FORMAT_INVALIDE", "A_CONTROLER",
     "facture_ref ne respecte pas le format FAC-AAAA-MM-PROP-NNN",
     "facture_ref",
     "facture_ref non null ET ne commence pas par FAC-"),
]


# ── Helpers styles ────────────────────────────────────────────────────────────
def hfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def hfont():
    return Font(bold=True, color="FFFFFF", size=10)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def write_headers(ws, cols):
    for i, (name, color, width) in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=name)
        c.fill = hfill(color)
        c.font = hfont()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


# ── Lecture REF_Setup ─────────────────────────────────────────────────────────
def read_ref_setup():
    wb = openpyxl.load_workbook(REF_SETUP, keep_vba=True, data_only=True)

    props = []
    for row in wb["REF_Proprietaires"].iter_rows(min_row=2, values_only=True):
        if row[0] and row[8] == "OUI":
            props.append((row[0], f"{row[1]} {row[2]}".strip()))

    logs = []
    for row in wb["REF_Logements"].iter_rows(min_row=2, values_only=True):
        if row[0] and row[12] == "OUI" and row[7]:  # actif OUI + proprietaire_id non null
            logs.append((row[0], row[3] or row[2] or row[0], row[7]))

    modes = []
    for row in wb["REF_Modes_Paiement"].iter_rows(min_row=2, values_only=True):
        if row[0] and row[5] == "OUI":
            modes.append((row[0], row[1]))

    return props, logs, modes


# ── Construction SAISIE ───────────────────────────────────────────────────────
def build_saisie(props, logs, modes):
    wb = Workbook()

    # ── Onglet SAISIE ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "SAISIE"
    write_headers(ws, SAISIE_COLS)

    # colonnes index (1-based)
    col = {name: i for i, (name, _, _) in enumerate(SAISIE_COLS, 1)}

    # DV : source_acompte (col G=7)
    dv_src = DataValidation(
        type="list",
        formula1='"HH_RESERVATION,VIREMENT_DIRECT,AUTRE"',
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Valeur invalide",
        error="Utilisez : HH_RESERVATION / VIREMENT_DIRECT / AUTRE"
    )
    dv_src.sqref = f"{get_column_letter(col['source_acompte'])}2:{get_column_letter(col['source_acompte'])}10000"
    ws.add_data_validation(dv_src)

    # DV : statut_controle (col O=15)
    dv_stat = DataValidation(
        type="list",
        formula1='"VALIDE,A_CONTROLER,EXCLU_RESULTAT"',
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Valeur invalide",
        error="Utilisez : VALIDE / A_CONTROLER / EXCLU_RESULTAT"
    )
    dv_stat.sqref = f"{get_column_letter(col['statut_controle'])}2:{get_column_letter(col['statut_controle'])}10000"
    ws.add_data_validation(dv_stat)

    # DV : proprietaire_id → REF_LOCALE col A
    prop_count = len(props)
    dv_prop = DataValidation(
        type="list",
        formula1=f"REF_LOCALE!$A$2:$A${prop_count + 1}",
        showDropDown=False
    )
    dv_prop.sqref = f"{get_column_letter(col['proprietaire_id'])}2:{get_column_letter(col['proprietaire_id'])}10000"
    ws.add_data_validation(dv_prop)

    # DV : logement_id → REF_LOCALE col C
    log_count = len(logs)
    dv_log = DataValidation(
        type="list",
        formula1=f"REF_LOCALE!$C$2:$C${log_count + 1}",
        showDropDown=False
    )
    dv_log.sqref = f"{get_column_letter(col['logement_id'])}2:{get_column_letter(col['logement_id'])}10000"
    ws.add_data_validation(dv_log)

    # DV : mode_paiement_id → REF_LOCALE col E
    mode_count = len(modes)
    dv_mode = DataValidation(
        type="list",
        formula1=f"REF_LOCALE!$E$2:$E${mode_count + 1}",
        showDropDown=False
    )
    dv_mode.sqref = f"{get_column_letter(col['mode_paiement_id'])}2:{get_column_letter(col['mode_paiement_id'])}10000"
    ws.add_data_validation(dv_mode)

    # ── Onglet REF_LOCALE ─────────────────────────────────────────────────────
    ws_ref = wb.create_sheet("REF_LOCALE")

    ref_headers = [
        ("proprietaire_id", "A", 16), ("nom_proprietaire", "B", 22),
        ("logement_id", "C", 16),     ("nom_court", "D", 28),
        ("mode_paiement_id", "E", 18), ("mode_paiement", "F", 24),
    ]
    for i, (name, _, width) in enumerate(ref_headers, 1):
        c = ws_ref.cell(row=1, column=i, value=name)
        c.font = Font(bold=True, size=10)
        c.fill = hfill("D9D9D9")
        ws_ref.column_dimensions[get_column_letter(i)].width = width

    for i, (pid, nom) in enumerate(props, 2):
        ws_ref.cell(row=i, column=1, value=pid)
        ws_ref.cell(row=i, column=2, value=nom)

    for i, (lid, nom_court, _) in enumerate(logs, 2):
        ws_ref.cell(row=i, column=3, value=lid)
        ws_ref.cell(row=i, column=4, value=nom_court)

    for i, (mid, mlabel) in enumerate(modes, 2):
        ws_ref.cell(row=i, column=5, value=mid)
        ws_ref.cell(row=i, column=6, value=mlabel)

    ws_ref.freeze_panes = "A2"

    # ── Onglet CONTROLES_SAISIE ───────────────────────────────────────────────
    ws_ctrl = wb.create_sheet("CONTROLES_SAISIE")

    ctrl_hdrs = ["code_controle", "niveau", "description", "colonne_cible", "regle"]
    ctrl_widths = [40, 14, 50, 20, 65]
    for i, (h, w) in enumerate(zip(ctrl_hdrs, ctrl_widths), 1):
        c = ws_ctrl.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, size=10)
        c.fill = hfill("D9D9D9")
        ws_ctrl.column_dimensions[get_column_letter(i)].width = w

    niveau_color = {"BLOQUANT": "C00000", "A_CONTROLER": "FF9900"}

    for i, (code, niveau, desc, col_cible, regle) in enumerate(CONTROLS, 2):
        ws_ctrl.cell(row=i, column=1, value=code).font = Font(bold=True, size=10)
        cn = ws_ctrl.cell(row=i, column=2, value=niveau)
        cn.fill = hfill(niveau_color[niveau])
        cn.font = Font(color="FFFFFF", bold=True, size=10)
        ws_ctrl.cell(row=i, column=3, value=desc)
        ws_ctrl.cell(row=i, column=4, value=col_cible)
        ws_ctrl.cell(row=i, column=5, value=regle)

    ws_ctrl.freeze_panes = "A2"

    # ── Onglet README ─────────────────────────────────────────────────────────
    ws_rm = wb.create_sheet("README")
    ws_rm.column_dimensions["A"].width = 95

    readme = [
        ("SAISIE_AcomptesProprietaires — Instructions Lot 5", True, 13),
        ("", False, 11),
        ("OBJECTIF", True, 11),
        ("Ce fichier centralise les acomptes versés aux propriétaires avant facturation définitive (Lot 12).", False, 11),
        ("Toutes les sources d'acomptes passent par ce fichier : HH, virements directs, autres.", False, 11),
        ("", False, 11),
        ("RÈGLES DE SAISIE", True, 11),
        ("1. acompte_id   : format ACC-AAAA-MM-NNN (ex: ACC-2026-05-001) — STABLE — ne jamais régénérer", False, 11),
        ("2. facture_ref  : format FAC-AAAA-MM-PROP-NNN — OBLIGATOIRE pour les lignes VALIDE", False, 11),
        ("3. source_acompte : HH_RESERVATION / VIREMENT_DIRECT / AUTRE (liste déroulante)", False, 11),
        ("4. source_hh_id : renseigner reservation_hh_id UNIQUEMENT si source_acompte = HH_RESERVATION", False, 11),
        ("5. montant_acompte : = acompte_facture du Lot 4 pour HH / montant saisi pour VIREMENT_DIRECT / AUTRE", False, 11),
        ("6. report_mois_precedent : informatif uniquement — commentaire OBLIGATOIRE si renseigné > 0", False, 11),
        ("7. code_impact / impact_resultat_reel / impact_resultat_comptable : HC / OUI / NON (fixes)", False, 11),
        ("", False, 11),
        ("TRAÇABILITÉ HH", True, 11),
        ("source_table = SAISIE_AcomptesProprietaires (toujours)", False, 11),
        ("source_pk    = acompte_id (toujours)", False, 11),
        ("source_hh_id = reservation_hh_id (uniquement si source_acompte = HH_RESERVATION)", False, 11),
        ("", False, 11),
        ("CONTRÔLES", True, 11),
        ("5 contrôles BLOQUANT — voir onglet CONTROLES_SAISIE", False, 11),
        ("5 contrôles A_CONTROLER — voir onglet CONTROLES_SAISIE", False, 11),
        ("", False, 11),
        ("FLUX POWER QUERY", True, 11),
        ("Ce fichier est lu par MASTER_FACT_MAN_AcomptesProprietaires.xlsx via 5 requêtes Power Query.", False, 11),
        ("Ne pas renommer les onglets ni les colonnes de la feuille SAISIE.", False, 11),
        ("", False, 11),
        ("INTERDICTIONS LOT 5", True, 11),
        ("- Ne pas calculer report_mois_suivant au Lot 5 (déféré Lot 10/12).", False, 11),
        ("- Ne pas inclure les acomptes dans revenu_net_exploitation_proprietaire (D031/D033).", False, 11),
        ("- Ne pas modifier les fichiers Lots 1, 3, 4 ou 4bis.", False, 11),
    ]

    for i, (text, bold, size) in enumerate(readme, 1):
        c = ws_rm.cell(row=i, column=1, value=text)
        c.font = Font(bold=bold, size=size)

    return wb


# ── Construction MASTER ───────────────────────────────────────────────────────
def build_master():
    wb = Workbook()

    # ── Onglet MASTER ─────────────────────────────────────────────────────────
    ws_m = wb.active
    ws_m.title = "MASTER"
    write_headers(ws_m, MASTER_COLS)

    # ── Onglet VUE_ACTIVE ─────────────────────────────────────────────────────
    ws_v = wb.create_sheet("VUE_ACTIVE")
    write_headers(ws_v, MASTER_COLS)
    note = ws_v.cell(row=2, column=1,
                     value="[Power Query] Filtre appliqué : statut_controle = VALIDE")
    note.font = Font(italic=True, color="595959", size=10)

    # ── Onglet POWER_QUERY_CODE ───────────────────────────────────────────────
    ws_pq = wb.create_sheet("POWER_QUERY_CODE")
    ws_pq.column_dimensions["A"].width = 130

    title = ws_pq.cell(row=1, column=1,
                       value="POWER_QUERY_CODE — Lot 5 Acomptes Propriétaires (5 requêtes)")
    title.font = Font(bold=True, size=13, color="1F497D")

    row = 3
    for q_name, q_code in _pq_queries():
        hdr = ws_pq.cell(row=row, column=1, value=f"{'='*3} {q_name} {'='*3}")
        hdr.font = Font(bold=True, size=11, color="1F497D")
        row += 1
        for line in q_code.strip().split("\n"):
            ws_pq.cell(row=row, column=1, value=line).font = Font(name="Consolas", size=9)
            row += 1
        row += 2

    return wb


# ── Requêtes M-code Power Query ───────────────────────────────────────────────
def _pq_queries():
    saisie_p = SAISIE_OUT.replace("\\", "\\\\")
    hh_p     = HH_MASTER.replace("\\", "\\\\")
    ref_p    = REF_SETUP.replace("\\", "\\\\")

    q1 = ("Q1_SAISIE_ACC_Source", f'''\
let
    Source   = Excel.Workbook(File.Contents("{SAISIE_OUT}"), null, true),
    Sheet    = Source{{[Item="SAISIE", Kind="Sheet"]}}[Data],
    Promoted = Table.PromoteHeaders(Sheet, [PromoteAllScalars=true]),
    Typed    = Table.TransformColumnTypes(Promoted, {{
        {{"acompte_id",                type text}},
        {{"ROW_HASH",                  type text}},
        {{"mois",                      type text}},
        {{"proprietaire_id",           type text}},
        {{"logement_id",               type text}},
        {{"facture_ref",               type text}},
        {{"source_acompte",            type text}},
        {{"source_hh_id",              type text}},
        {{"montant_acompte",           type number}},
        {{"report_mois_precedent",     type number}},
        {{"mode_paiement_id",          type text}},
        {{"code_impact",               type text}},
        {{"impact_resultat_reel",      type text}},
        {{"impact_resultat_comptable", type text}},
        {{"statut_controle",           type text}},
        {{"niveau_anomalie",           type text}},
        {{"code_anomalie",             type text}},
        {{"commentaire",               type text}}
    }}),
    // Exclure les lignes vides (sans acompte_id)
    Filtered = Table.SelectRows(Typed, each [acompte_id] <> null and [acompte_id] <> "")
in
    Filtered''')

    q2 = ("Q2_HH_Acomptes_Ref", f'''\
// Référence croisée Lot 4 — validation HH uniquement (pas chargé dans MASTER)
let
    Source   = Excel.Workbook(File.Contents("{HH_MASTER}"), null, true),
    Sheet    = Source{{[Item="MASTER", Kind="Sheet"]}}[Data],
    Promoted = Table.PromoteHeaders(Sheet, [PromoteAllScalars=true]),
    Typed    = Table.TransformColumnTypes(Promoted, {{
        {{"reservation_hh_id", type text}},
        {{"statut_controle",   type text}},
        {{"acompte_facture",   type number}},
        {{"proprietaire_id",   type text}},
        {{"logement_id",       type text}},
        {{"mois",              type text}}
    }}),
    // Filtre Correction 1 : VALIDE + acompte_facture > 0 uniquement
    Filtered = Table.SelectRows(Typed, each
        [statut_controle] = "VALIDE" and
        [acompte_facture] <> null and
        [acompte_facture] > 0
    ),
    Selected = Table.SelectColumns(Filtered, {{
        "reservation_hh_id", "acompte_facture",
        "proprietaire_id", "logement_id", "mois"
    }})
in
    Selected''')

    q3 = ("Q3_REF_Proprietaires_Source", f'''\
// Référence propriétaires actifs (validation croisée)
let
    Source   = Excel.Workbook(File.Contents("{REF_SETUP}"), null, true),
    Sheet    = Source{{[Item="REF_Proprietaires", Kind="Sheet"]}}[Data],
    Promoted = Table.PromoteHeaders(Sheet, [PromoteAllScalars=true]),
    Filtered = Table.SelectRows(Promoted, each [actif] = "OUI"),
    Selected = Table.SelectColumns(Filtered, {{"proprietaire_id", "nom_proprietaire"}})
in
    Selected''')

    q4 = ("Q4_MASTER_FACT_MAN_AcomptesProprietaires", '''\
let
    // ═══ 1. Sources ═══════════════════════════════════════════════════════════
    Saisie = Q1_SAISIE_ACC_Source,
    HH_Ref = Q2_HH_Acomptes_Ref,

    // ═══ 2. Doublons acompte_id ═══════════════════════════════════════════════
    ID_Grp  = Table.Group(Saisie, {"acompte_id"}, {{"_cnt", each Table.RowCount(_), Int64.Type}}),
    DupIDs  = Table.Column(Table.SelectRows(ID_Grp, each [_cnt] > 1), "acompte_id"),

    // ═══ 3. Contrôles par ligne ════════════════════════════════════════════════
    WithCtrl = Table.AddColumn(Saisie, "_ctrl", each
        let
            // --- BLOQUANTS ---
            is_dup       = List.Contains(DupIDs, [acompte_id]),
            prop_absent  = ([proprietaire_id] = null or [proprietaire_id] = ""),
            mt_invalide  = ([montant_acompte] = null or [montant_acompte] <= 0),
            fac_absent   = ([facture_ref] = null or [facture_ref] = "")
                            and [statut_controle] <> "EXCLU_RESULTAT",

            // Validation croisée HH
            hh_src       = ([source_acompte] = "HH_RESERVATION"),
            hh_id        = [source_hh_id],
            hh_rows      = if hh_src and hh_id <> null and hh_id <> ""
                           then Table.SelectRows(HH_Ref, each [reservation_hh_id] = hh_id)
                           else #table(
                               {"reservation_hh_id","acompte_facture","proprietaire_id","logement_id","mois"},
                               {}
                           ),
            hh_found     = Table.RowCount(hh_rows) > 0,
            hh_inconst   = hh_src
                            and hh_id <> null and hh_id <> ""
                            and hh_found
                            and let exp = hh_rows{0}[acompte_facture] in
                                [montant_acompte] <> null
                                and exp <> null
                                and Number.Abs([montant_acompte] - exp) > 0.01,

            // --- A_CONTROLER ---
            log_absent   = ([logement_id] = null or [logement_id] = ""),
            hh_introuvable = hh_src
                             and (hh_id = null or hh_id = "" or not hh_found),
            rpt_inconst  = ([report_mois_precedent] <> null and [report_mois_precedent] > 0)
                            and ([commentaire] = null or [commentaire] = ""),
            src_autre    = ([source_acompte] = "AUTRE"),
            fac_fmt_ko   = ([facture_ref] <> null and [facture_ref] <> "")
                            and not Text.StartsWith([facture_ref], "FAC-"),

            // --- Priorité BLOQUANT > A_CONTROLER ---
            bloquants = List.Select({
                if is_dup      then "ACOMPTE_CALC_ID_DUPLIQUE"     else null,
                if prop_absent then "ACOMPTE_PROPRIETAIRE_ABSENT"  else null,
                if mt_invalide then "ACOMPTE_MONTANT_INVALIDE"     else null,
                if fac_absent  then "ACOMPTE_NON_RATTACHE_FACTURE" else null,
                if hh_inconst  then "ACOMPTE_HH_INCOHERENT"        else null
            }, each _ <> null),
            a_ctrl = List.Select({
                if log_absent      then "ACOMPTE_LOGEMENT_ABSENT"             else null,
                if hh_introuvable  then "ACOMPTE_SOURCE_HH_INTROUVABLE"       else null,
                if rpt_inconst     then "ACOMPTE_REPORT_INCOHERENT"           else null,
                if src_autre       then "ACOMPTE_SOURCE_A_CONTROLER"          else null,
                if fac_fmt_ko      then "ACOMPTE_FACTURE_REF_FORMAT_INVALIDE" else null
            }, each _ <> null),

            niv  = if List.Count(bloquants) > 0 then "BLOQUANT"
                   else if List.Count(a_ctrl) > 0 then "A_CONTROLER"
                   else null,
            code = if List.Count(bloquants) > 0 then bloquants{0}
                   else if List.Count(a_ctrl) > 0 then a_ctrl{0}
                   else null
        in [niveau = niv, code = code]
    ),

    // ═══ 4. Application contrôles (PQ complète si SAISIE vide) ═══════════════
    WithNivPQ  = Table.AddColumn(WithCtrl, "_niv_pq", each [_ctrl][niveau], type text),
    WithCodPQ  = Table.AddColumn(WithNivPQ, "_cod_pq", each [_ctrl][code],  type text),

    ApplyNiv = Table.ReplaceValue(WithCodPQ,
        each [niveau_anomalie],
        each if ([niveau_anomalie] = null or [niveau_anomalie] = "")
             then [_niv_pq] else [niveau_anomalie],
        Replacer.ReplaceValue, {"niveau_anomalie"}),
    ApplyCod = Table.ReplaceValue(ApplyNiv,
        each [code_anomalie],
        each if ([code_anomalie] = null or [code_anomalie] = "")
             then [_cod_pq] else [code_anomalie],
        Replacer.ReplaceValue, {"code_anomalie"}),

    // ═══ 5. Colonnes système PQ ════════════════════════════════════════════════
    // source_table = SAISIE_AcomptesProprietaires (toujours)
    // source_pk    = acompte_id (toujours) — QM-L5-07 corrigé
    // source_hh_id = reservation_hh_id (dans colonne dédiée si HH_RESERVATION)
    WithModule = Table.AddColumn(ApplyCod,   "source_module",    each "LOT5_ACOMPTES_PROPRIETAIRES",    type text),
    WithTable  = Table.AddColumn(WithModule, "source_table",     each "SAISIE_AcomptesProprietaires",   type text),
    WithPK     = Table.AddColumn(WithTable,  "source_pk",        each [acompte_id],                     type text),
    WithDate   = Table.AddColumn(WithPK,     "date_integration", each DateTime.LocalNow(),               type datetime),

    // ═══ 6. Nettoyage + ordre final 22 colonnes ════════════════════════════════
    Cleaned  = Table.RemoveColumns(WithDate, {"_ctrl", "_niv_pq", "_cod_pq"}),
    Ordered  = Table.ReorderColumns(Cleaned, {
        "acompte_id", "ROW_HASH", "mois",
        "proprietaire_id", "logement_id", "facture_ref",
        "source_acompte", "source_hh_id",
        "montant_acompte", "report_mois_precedent",
        "mode_paiement_id",
        "code_impact", "impact_resultat_reel", "impact_resultat_comptable",
        "statut_controle", "niveau_anomalie", "code_anomalie", "commentaire",
        "source_module", "source_table", "source_pk", "date_integration"
    })
in
    Ordered''')

    q5 = ("Q5_VUE_ACTIVE_AcomptesProprietaires", '''\
let
    Source   = Q4_MASTER_FACT_MAN_AcomptesProprietaires,
    Filtered = Table.SelectRows(Source, each [statut_controle] = "VALIDE")
in
    Filtered''')

    return [q1, q2, q3, q4, q5]


# ── Contrôles structurels ─────────────────────────────────────────────────────
def check_structure(saisie_path, master_path):
    results = []

    wb_s = openpyxl.load_workbook(saisie_path)
    ws_s = wb_s["SAISIE"]
    saisie_cols = [c.value for c in ws_s[1] if c.value]
    expected_s  = [name for name, _, _ in SAISIE_COLS]

    results.append(("SAISIE_COLS_COUNT",
                    "OK" if len(saisie_cols) == 18 else "FAIL",
                    f"{len(saisie_cols)}/18"))
    results.append(("SAISIE_COLS_ORDRE",
                    "OK" if saisie_cols == expected_s else "FAIL",
                    str(saisie_cols) if saisie_cols != expected_s else "correct"))
    results.append(("SAISIE_ONGLETS",
                    "OK" if set(wb_s.sheetnames) >= {"SAISIE","REF_LOCALE","CONTROLES_SAISIE","README"} else "FAIL",
                    str(wb_s.sheetnames)))

    wb_m = openpyxl.load_workbook(master_path)
    ws_m = wb_m["MASTER"]
    master_cols = [c.value for c in ws_m[1] if c.value]
    expected_m  = [name for name, _, _ in MASTER_COLS]

    results.append(("MASTER_COLS_COUNT",
                    "OK" if len(master_cols) == 22 else "FAIL",
                    f"{len(master_cols)}/22"))
    results.append(("MASTER_COLS_ORDRE",
                    "OK" if master_cols == expected_m else "FAIL",
                    str(master_cols) if master_cols != expected_m else "correct"))
    results.append(("MASTER_ONGLETS",
                    "OK" if set(wb_m.sheetnames) >= {"MASTER","VUE_ACTIVE","POWER_QUERY_CODE"} else "FAIL",
                    str(wb_m.sheetnames)))

    ws_pq = wb_m["POWER_QUERY_CODE"]
    pq_content = " ".join(str(ws_pq.cell(r, 1).value) for r in range(1, ws_pq.max_row + 1))
    q_names = ["Q1_SAISIE_ACC_Source", "Q2_HH_Acomptes_Ref",
               "Q3_REF_Proprietaires_Source", "Q4_MASTER_FACT_MAN_AcomptesProprietaires",
               "Q5_VUE_ACTIVE_AcomptesProprietaires"]
    for qn in q_names:
        results.append((f"PQ_{qn}",
                        "OK" if qn in pq_content else "FAIL",
                        "présent" if qn in pq_content else "ABSENT"))

    # Contrôles dans CONTROLES_SAISIE
    ws_ctrl = wb_s["CONTROLES_SAISIE"]
    ctrl_codes = [ws_ctrl.cell(r, 1).value for r in range(2, ws_ctrl.max_row + 1)
                  if ws_ctrl.cell(r, 1).value]
    expected_ctrl = [c[0] for c in CONTROLS]
    results.append(("SAISIE_CONTROLES_COUNT",
                    "OK" if len(ctrl_codes) == 10 else "FAIL",
                    f"{len(ctrl_codes)}/10"))

    return results


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("LOT 5 — Acomptes propriétaires")
    print("=" * 60)

    # Répertoires
    os.makedirs(os.path.join(BASE, "01_SOURCES_BRUTES", "AcomptesProprietaires"), exist_ok=True)
    os.makedirs(os.path.join(BASE, "02_TRAVAIL", "Lot5_AcomptesProprietaires"), exist_ok=True)

    # Lecture REF_Setup
    print("\n[1/4] Lecture REF_Setup.xlsm...")
    props, logs, modes = read_ref_setup()
    print(f"  {len(props)} propriétaires actifs")
    print(f"  {len(logs)} logements actifs avec proprietaire_id")
    print(f"  {len(modes)} modes de paiement actifs")

    # SAISIE
    print("\n[2/4] Construction SAISIE_AcomptesProprietaires.xlsx...")
    saisie_wb = build_saisie(props, logs, modes)
    saisie_wb.save(SAISIE_OUT)
    print(f"  CRÉÉ: {SAISIE_OUT}")

    # MASTER
    print("\n[3/4] Construction MASTER_FACT_MAN_AcomptesProprietaires.xlsx...")
    master_wb = build_master()
    master_wb.save(MASTER_OUT)
    print(f"  CRÉÉ: {MASTER_OUT}")

    # Contrôles structurels
    print("\n[4/4] Contrôles structurels...")
    results = check_structure(SAISIE_OUT, MASTER_OUT)
    all_ok = True
    for name, status, detail in results:
        icon = "OK" if status == "OK" else "FAIL"
        print(f"  [{icon}] {name}: {detail}")
        if status != "OK":
            all_ok = False

    print("\n" + "=" * 60)
    print(f"RÉSULTAT: {'TOUT OK — 0 FAIL' if all_ok else 'ERREURS DÉTECTÉES'}")
    print("=" * 60)

    print(f"\nFichiers créés:")
    print(f"  {SAISIE_OUT}")
    print(f"  {MASTER_OUT}")


if __name__ == "__main__":
    main()
