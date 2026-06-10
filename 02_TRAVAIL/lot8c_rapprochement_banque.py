#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
lot8c_rapprochement_banque.py
Lot 8c - Rapprochement bancaire partiel (onglets d'attente)

Perimetre :
  - 39 lignes Airbnb : EN_ATTENTE_EXPORT_AIRBNB (sauf 4,97 EUR -> AJUSTEMENT)
  - 13 virements proprietaires : EN_ATTENTE_SAISIE_ACOMPTE (Lot 5 vide)
  - Aucun produit economique cree
  - Aucun rapprochement valide automatiquement
  - REF_Setup.xlsm non modifie
  - NORM_Banque non modifiee

Onglets ajoutes dans BANQUE_LOT8_IMPORT.xlsx :
  - RAPPROCH_AIRBNB_ATTENTE
  - RAPPROCH_PROPRIETAIRES_ATTENTE
  - CTRL_RAPPROCHEMENT_8C

Decisions : D-8c-01 a D-8c-07
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import shutil
import os
import re
from datetime import datetime

# ─────────────────────────────────────────────────────────────────────────────
# CHEMINS
# ─────────────────────────────────────────────────────────────────────────────

BASE = r"C:\Users\Ewan\OneDrive\Documents\Conciergerie\Pilotage_Conciergerie"
BANQUE_PATH = os.path.join(BASE, "02_TRAVAIL", "Lot8_Banque", "BANQUE_LOT8_IMPORT.xlsx")
ARCHIVE_DIR = os.path.join(BASE, "99_ARCHIVES", "LOT8_Banque")

IMPORT_ID = "IMP-BQ-CM-2026-03-001"

TS = datetime.now().strftime("%Y%m%d_%H%M%S")
NOW_STR = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# ─────────────────────────────────────────────────────────────────────────────
# STYLES
# ─────────────────────────────────────────────────────────────────────────────

HDR_FILL     = PatternFill("solid", fgColor="1F4E79")
HDR_FONT     = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
SUBHDR_FILL  = PatternFill("solid", fgColor="2E75B6")

FILL_ATTENTE = PatternFill("solid", fgColor="FFEB9C")  # jaune attente
FILL_AJUST   = PatternFill("solid", fgColor="FCE4D6")  # orange ajustement
FILL_INFO    = PatternFill("solid", fgColor="DEEAF1")  # bleu info
FILL_PROP    = PatternFill("solid", fgColor="E2EFDA")  # vert attente prop
FILL_BLOCK   = PatternFill("solid", fgColor="FFC7CE")  # rouge bloquant
FILL_WHITE   = PatternFill("solid", fgColor="FFFFFF")

DATA_FONT    = Font(name="Calibri", size=10)
BOLD_FONT    = Font(name="Calibri", bold=True, size=10)
ITALIC_FONT  = Font(name="Calibri", italic=True, size=10, color="888888")

ALN_CTR      = Alignment(horizontal="center", vertical="center")
ALN_LEFT     = Alignment(horizontal="left", vertical="center")
ALN_LEFT_WRP = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN_BORDER  = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTES METIER
# ─────────────────────────────────────────────────────────────────────────────

MONTANT_AJUSTEMENT_AIRBNB = 4.97  # ligne speciale -> AJUSTEMENT_AIRBNB_A_CONTROLER

PROP_IDS = {
    "PROP_0002", "PROP_0005", "PROP_0006", "PROP_0008",
    "PROP_0009", "PROP_0010", "FAMILLE_UZON_A_CONTROLER",
}

# Statuts rapprochement Lot 8c
STATUT_AIRBNB_ATTENTE  = "EN_ATTENTE_EXPORT_AIRBNB"
STATUT_AIRBNB_AJUST    = "AJUSTEMENT_AIRBNB_A_CONTROLER"
STATUT_PROP_ATTENTE    = "EN_ATTENTE_SAISIE_ACOMPTE"
METHODE_NON_RAPROCH    = "NON_RAPPROCHABLE_DETAIL_ABSENT"
METHODE_ATTENTE        = "EN_ATTENTE_PREREQUIS"
LOT_REF                = "LOT8C"

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def style_cell(cell, fill=None, font=None, alignment=None, border=None):
    if fill:      cell.fill = fill
    if font:      cell.font = font
    if alignment: cell.alignment = alignment
    if border:    cell.border = border


def write_header_row(ws, row, headers, col_widths=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row, i, h)
        style_cell(c, fill=HDR_FILL, font=HDR_FONT, alignment=ALN_CTR)
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def extract_gcode(libelle):
    """Extract G-XXXXXXXX reference from Airbnb bank libelle."""
    m = re.search(r'(G-[A-Z0-9]{10,})', libelle or "")
    return m.group(1) if m else None


def backup_file(src_path, label):
    os.makedirs(ARCHIVE_DIR, exist_ok=True)
    ext = os.path.splitext(src_path)[1]
    dst = os.path.join(ARCHIVE_DIR, "%s_PRE_LOT8C_%s%s" % (label, TS, ext))
    shutil.copy2(src_path, dst)
    print("  Backup: %s" % os.path.basename(dst))


def drop_sheet_if_exists(wb, name):
    if name in wb.sheetnames:
        del wb[name]


# ─────────────────────────────────────────────────────────────────────────────
# LECTURE NORM_BANQUE
# ─────────────────────────────────────────────────────────────────────────────

def read_norm_banque(wb):
    ws = wb["NORM_Banque"]
    hdr = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

    airbnb_rows = []
    prop_rows = []

    for r in range(2, ws.max_row + 1):
        tiers = ws.cell(r, hdr["tiers_detecte"]).value
        if not tiers:
            continue

        row = {
            "mouvement_id":  ws.cell(r, hdr["mouvement_id"]).value,
            "ligne_source":  ws.cell(r, hdr["ligne_source"]).value,
            "date_operation":ws.cell(r, hdr["date_operation"]).value,
            "date_valeur":   ws.cell(r, hdr["date_valeur"]).value,
            "libelle":       ws.cell(r, hdr["libelle"]).value or "",
            "montant":       float(ws.cell(r, hdr["montant"]).value or 0),
            "sens":          ws.cell(r, hdr["sens"]).value,
            "tiers_detecte": tiers,
            "categorie":     ws.cell(r, hdr["categorie"]).value,
            "regle_id":      ws.cell(r, hdr["regle_id_appliquee"]).value,
        }

        if tiers == "AIRBNB":
            airbnb_rows.append(row)
        elif tiers in PROP_IDS:
            prop_rows.append(row)

    return airbnb_rows, prop_rows


# ─────────────────────────────────────────────────────────────────────────────
# ONGLET RAPPROCH_AIRBNB_ATTENTE
# ─────────────────────────────────────────────────────────────────────────────

AIRBNB_HDR = [
    "mouvement_id", "date_operation", "libelle",
    "montant_banque", "sens",
    "reference_airbnb",
    "statut_rapprochement", "methode_rapprochement",
    "lot_rapprochement", "date_rapprochement",
    "commentaire",
]
AIRBNB_WIDTHS = [32, 14, 65, 14, 10, 40, 36, 34, 12, 20, 60]


def build_rapproch_airbnb(wb, airbnb_rows):
    drop_sheet_if_exists(wb, "RAPPROCH_AIRBNB_ATTENTE")
    ws = wb.create_sheet("RAPPROCH_AIRBNB_ATTENTE")
    write_header_row(ws, 1, AIRBNB_HDR, AIRBNB_WIDTHS)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    nb_attente = 0
    nb_ajust = 0
    total_attente = 0.0
    total_ajust = 0.0

    for row_idx, row in enumerate(sorted(airbnb_rows, key=lambda x: str(x["date_operation"])), 2):
        mt = row["montant"]
        gcode = extract_gcode(row["libelle"])

        if abs(mt - MONTANT_AJUSTEMENT_AIRBNB) < 0.01:
            statut  = STATUT_AIRBNB_AJUST
            methode = METHODE_NON_RAPROCH
            comment = ("Montant anormalement faible (%.2f EUR) — ajustement ou correction Airbnb "
                       "potentiel. Ne pas creer de produit. Verifier avec export Airbnb detaille." % mt)
            fill = FILL_AJUST
            nb_ajust += 1
            total_ajust += mt
        else:
            statut  = STATUT_AIRBNB_ATTENTE
            methode = METHODE_NON_RAPROCH
            comment = ("En attente export Airbnb detaille (reference %s). "
                       "Aucun rapprochement possible sans mapping G-code -> reservation_id. "
                       "Aucun produit economique cree (DC1/B8)." % (gcode or "?"))
            fill = FILL_ATTENTE
            nb_attente += 1
            total_attente += mt

        vals = [
            row["mouvement_id"],
            row["date_operation"],
            row["libelle"],
            mt,
            row["sens"],
            gcode,
            statut,
            methode,
            LOT_REF,
            NOW_STR,
            comment,
        ]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row_idx, c_idx, val)
            style_cell(cell, fill=fill, font=DATA_FONT,
                       alignment=ALN_LEFT_WRP if c_idx == 11 else ALN_LEFT,
                       border=THIN_BORDER)

    # Ligne info totaux (non valorisable — informatif uniquement)
    info_row = row_idx + 2 if airbnb_rows else 3
    ws.cell(info_row, 1, "TOTAL INFORMATIF").font = BOLD_FONT
    ws.cell(info_row, 1).fill = FILL_INFO
    ws.cell(info_row, 4, round(total_attente + total_ajust, 2)).fill = FILL_INFO
    ws.cell(info_row, 4).font = BOLD_FONT
    ws.cell(info_row, 7, "INFORMATIF_UNIQUEMENT").fill = FILL_INFO
    ws.cell(info_row, 7).font = ITALIC_FONT
    ws.cell(info_row, 11,
            "Total brut bancaire Airbnb (informatif). "
            "Non utilise pour rapprochement. Valide uniquement via export Airbnb detaille.").fill = FILL_INFO
    ws.cell(info_row, 11).font = ITALIC_FONT
    ws.row_dimensions[info_row].height = 30

    print("  RAPPROCH_AIRBNB_ATTENTE: %d en attente (%.2f EUR) + %d ajustement (%.2f EUR)" % (
        nb_attente, total_attente, nb_ajust, total_ajust))

    return nb_attente, nb_ajust, total_attente, total_ajust


# ─────────────────────────────────────────────────────────────────────────────
# ONGLET RAPPROCH_PROPRIETAIRES_ATTENTE
# ─────────────────────────────────────────────────────────────────────────────

PROP_HDR = [
    "mouvement_id", "date_operation", "libelle",
    "montant_banque", "sens",
    "proprietaire_id",
    "nature_presumee",
    "statut_rapprochement", "prerequis_rapprochement",
    "lot_rapprochement", "date_rapprochement",
    "commentaire",
]
PROP_WIDTHS = [32, 14, 60, 14, 10, 28, 36, 32, 38, 12, 20, 70]

PROP_LABELS = {
    "PROP_0002": "Cedrine Delrieu",
    "PROP_0005": "Florane Vassal",
    "PROP_0006": "Caroline Pons-Dinneweth",
    "PROP_0008": "David Toure",
    "PROP_0009": "Francois Maurer",
    "PROP_0010": "Noel Dureuil",
    "FAMILLE_UZON_A_CONTROLER": "Famille Uzon (Didier PROP_0001 ou Maryline PROP_0011)",
}


def build_rapproch_proprietaires(wb, prop_rows):
    drop_sheet_if_exists(wb, "RAPPROCH_PROPRIETAIRES_ATTENTE")
    ws = wb.create_sheet("RAPPROCH_PROPRIETAIRES_ATTENTE")
    write_header_row(ws, 1, PROP_HDR, PROP_WIDTHS)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    total_prop = 0.0
    for row_idx, row in enumerate(sorted(prop_rows, key=lambda x: str(x["date_operation"])), 2):
        mt = row["montant"]
        prop_id = row["tiers_detecte"]
        prop_label = PROP_LABELS.get(prop_id, prop_id)
        comment = ("Proprietaire identifie : %s (%s). "
                   "Nature comptable inconnue (acompte facture / remboursement charge / regularisation). "
                   "Aucune validation sans justificatif ou saisie Lot 5." % (prop_label, prop_id))

        vals = [
            row["mouvement_id"],
            row["date_operation"],
            row["libelle"],
            mt,
            row["sens"],
            prop_id,
            "ENCAISSEMENT_PROPRIETAIRE_A_VENTILER",
            STATUT_PROP_ATTENTE,
            "MASTER_FACT_MAN_AcomptesProprietaires vide — attendre saisie Lot 5",
            LOT_REF,
            NOW_STR,
            comment,
        ]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row_idx, c_idx, val)
            style_cell(cell, fill=FILL_PROP, font=DATA_FONT,
                       alignment=ALN_LEFT_WRP if c_idx == 12 else ALN_LEFT,
                       border=THIN_BORDER)
        total_prop += mt

    print("  RAPPROCH_PROPRIETAIRES_ATTENTE: %d virements (%.2f EUR)" % (len(prop_rows), total_prop))
    return total_prop


# ─────────────────────────────────────────────────────────────────────────────
# ONGLET CTRL_RAPPROCHEMENT_8C
# ─────────────────────────────────────────────────────────────────────────────

CTRL_8C_HDR = [
    "code_controle", "severite", "nb_lignes", "total_montant_eur",
    "description", "statut", "lot", "date_controle", "action_requise",
]
CTRL_8C_WIDTHS = [42, 16, 12, 20, 70, 28, 10, 20, 80]


def build_ctrl_rapprochement(wb, nb_attente, nb_ajust, total_attente, total_ajust, nb_prop, total_prop):
    drop_sheet_if_exists(wb, "CTRL_RAPPROCHEMENT_8C")
    ws = wb.create_sheet("CTRL_RAPPROCHEMENT_8C")
    write_header_row(ws, 1, CTRL_8C_HDR, CTRL_8C_WIDTHS)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    total_airbnb = round(total_attente + total_ajust, 2)

    ctrl_rows = [
        # (code, sev, nb, montant, desc, statut, action, fill)
        (
            "RAPPROCHEMENT_AIRBNB_EN_ATTENTE",
            "A_CONTROLER",
            nb_attente,
            round(total_attente, 2),
            ("%d lignes VIR AIRBNB PAYMENTS en attente export detaille Airbnb. "
             "G-codes extraits mais non matchables sans mapping G-code->reservation_id." % nb_attente),
            "EN_ATTENTE",
            "Fournir export Airbnb detaille (liste transactions par reference payout G-XXXX).",
            FILL_ATTENTE,
        ),
        (
            "AJUSTEMENT_AIRBNB_A_CONTROLER",
            "A_CONTROLER",
            nb_ajust,
            round(total_ajust, 2),
            ("Ligne Airbnb de %.2f EUR — montant anormalement faible, "
             "possible ajustement/correction Airbnb (D-8c-06)." % MONTANT_AJUSTEMENT_AIRBNB),
            "A_CONTROLER",
            "Verifier avec export Airbnb : correction/ajustement ou mini-reservation.",
            FILL_AJUST,
        ),
        (
            "TOTAL_AIRBNB_INFORMATIF",
            "INFO",
            nb_attente + nb_ajust,
            total_airbnb,
            ("Total brut bancaire Airbnb : %.2f EUR. "
             "INFORMATIF UNIQUEMENT — non utilise pour rapprochement. "
             "Ne pas crer de produit economique (REGLES DC1/B8)." % total_airbnb),
            "INFORMATIF",
            "Aucune action directe. Conserver pour controle apres reception export Airbnb.",
            FILL_INFO,
        ),
        (
            "RAPPROCHEMENT_PROPRIETAIRES_EN_ATTENTE",
            "A_CONTROLER",
            nb_prop,
            round(total_prop, 2),
            ("%d virements proprietaires entrants (%.2f EUR). "
             "Proprietaires identifies. Nature comptable inconnue. "
             "ENCAISSEMENT_PROPRIETAIRE_A_VENTILER." % (nb_prop, total_prop)),
            "EN_ATTENTE",
            "Saisir acomptes/factures dans Lot 5 (SAISIE_AcomptesProprietaires.xlsx) puis relancer.",
            FILL_PROP,
        ),
        (
            "LOT5_PREREQUIS_MANQUANT",
            "A_CONTROLER",
            nb_prop,
            round(total_prop, 2),
            ("Lot 5 / acomptes proprietaires non alimente ; "
             "rapprochement proprietaire complet impossible a ce stade. "
             "Les %d virements restent EN_ATTENTE_SAISIE_ACOMPTE." % nb_prop),
            "A_CONTROLER",
            "Alimenter SAISIE_AcomptesProprietaires.xlsx pour Lot 5, puis relancer Lot 8c.",
            FILL_ATTENTE,
        ),
    ]

    for row_idx, (code, sev, nb, mt, desc, statut, action, fill) in enumerate(ctrl_rows, 2):
        vals = [code, sev, nb, mt, desc, statut, LOT_REF, NOW_STR, action]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row_idx, c_idx, val)
            style_cell(cell, fill=fill, font=DATA_FONT,
                       alignment=ALN_LEFT_WRP if c_idx in (5, 9) else ALN_LEFT,
                       border=THIN_BORDER)
        ws.row_dimensions[row_idx].height = 40

    print("  CTRL_RAPPROCHEMENT_8C: %d entrees de controle" % len(ctrl_rows))


# ─────────────────────────────────────────────────────────────────────────────
# LOG_Traitement
# ─────────────────────────────────────────────────────────────────────────────

def update_log(wb, nb_airbnb, nb_prop, total_airbnb, total_prop):
    ws = wb["LOG_Traitement"]

    # Purge previous LOT8C entries (idempotence)
    for r in range(ws.max_row, 1, -1):
        if ws.cell(r, 3).value == "LOT8C_RAPPROCH":
            ws.delete_rows(r, 1)

    run_id = "RUN-LOT8C-%s" % TS
    log_row = [
        run_id, IMPORT_ID, "LOT8C_RAPPROCH", NOW_STR,
        nb_airbnb + nb_prop, None, None, None,
        0, nb_airbnb + nb_prop,
        0, None,
        ("lot8c_rapprochement_banque.py — %d Airbnb EN_ATTENTE (%.2f EUR), "
         "%d proprietaires EN_ATTENTE (%.2f EUR). "
         "Aucun produit economique cree. Aucun rapprochement valide." % (
             nb_airbnb, total_airbnb, nb_prop, total_prop)),
    ]
    log_row_idx = ws.max_row + 1
    for c_idx, val in enumerate(log_row, 1):
        cell = ws.cell(log_row_idx, c_idx, val)
        cell.font = Font(name="Calibri", size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER

    print("  LOG_Traitement: entree LOT8C_RAPPROCH ajoutee (row %d)" % log_row_idx)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("=" * 68)
    print("LOT 8C - Rapprochement bancaire partiel (onglets d'attente)")
    print("Timestamp: %s" % NOW_STR)
    print("=" * 68)

    # Backup
    backup_file(BANQUE_PATH, "BANQUE_LOT8_IMPORT")

    # Load workbook
    wb = openpyxl.load_workbook(BANQUE_PATH)

    print("\n--- Lecture NORM_Banque ---")
    airbnb_rows, prop_rows = read_norm_banque(wb)
    print("  Airbnb : %d lignes" % len(airbnb_rows))
    print("  Proprietaires : %d lignes" % len(prop_rows))

    print("\n--- Construction onglets ---")
    nb_attente, nb_ajust, total_attente, total_ajust = build_rapproch_airbnb(wb, airbnb_rows)
    total_prop = build_rapproch_proprietaires(wb, prop_rows)
    nb_prop = len(prop_rows)
    build_ctrl_rapprochement(wb, nb_attente, nb_ajust, total_attente, total_ajust, nb_prop, total_prop)

    print("\n--- LOG_Traitement ---")
    total_airbnb = round(total_attente + total_ajust, 2)
    update_log(wb, len(airbnb_rows), nb_prop, total_airbnb, total_prop)

    wb.save(BANQUE_PATH)
    print("  BANQUE_LOT8_IMPORT.xlsx sauvegarde")

    print("\n" + "=" * 68)
    print("BILAN LOT 8C")
    print("=" * 68)
    print("  1. Fichiers crees/modifies:")
    print("     02_TRAVAIL/Lot8_Banque/BANQUE_LOT8_IMPORT.xlsx (+3 onglets)")
    print("     02_TRAVAIL/lot8c_rapprochement_banque.py (ce script)")
    print()
    print("  2. Airbnb en attente export detaille : %d lignes" % nb_attente)
    print("  3. Total bancaire Airbnb (informatif) : %.2f EUR" % total_airbnb)
    print("     dont ajustement 4,97 EUR : %d ligne" % nb_ajust)
    print()
    print("  4. Virements proprietaires en attente  : %d lignes" % nb_prop)
    print("  5. Total virements proprietaires       : %.2f EUR" % total_prop)
    print()
    print("  6. Controles generes (CTRL_RAPPROCHEMENT_8C) :")
    print("     RAPPROCHEMENT_AIRBNB_EN_ATTENTE    -> A_CONTROLER (%d lignes)" % nb_attente)
    print("     AJUSTEMENT_AIRBNB_A_CONTROLER      -> A_CONTROLER (%d ligne)" % nb_ajust)
    print("     TOTAL_AIRBNB_INFORMATIF            -> INFORMATIF (%.2f EUR)" % total_airbnb)
    print("     RAPPROCHEMENT_PROPRIETAIRES_EN_ATTENTE -> A_CONTROLER (%d lignes)" % nb_prop)
    print("     LOT5_PREREQUIS_MANQUANT            -> A_CONTROLER (%d virements en attente)" % nb_prop)
    print()
    print("  7. Statuts de rapprochement utilises:")
    print("     EN_ATTENTE_EXPORT_AIRBNB           (%d lignes)" % nb_attente)
    print("     AJUSTEMENT_AIRBNB_A_CONTROLER      (%d ligne)" % nb_ajust)
    print("     EN_ATTENTE_SAISIE_ACOMPTE          (%d lignes)" % nb_prop)
    print()
    print("  8. Produit economique cree : AUCUN")
    print("     - VIR AIRBNB PAYMENTS : rapprochement uniquement (DC1/B8)")
    print("     - Virements proprietaires : ENCAISSEMENT_A_VENTILER, non valide")
    print()
    print("  Backup: BANQUE_LOT8_IMPORT_PRE_LOT8C_%s.xlsx" % TS)
    print()
    print("  REF_Setup.xlsm : non modifie")
    print("  NORM_Banque    : non modifiee")
    print("  Aucun git add/commit/push effectue.")
    print("=" * 68)


if __name__ == "__main__":
    main()
