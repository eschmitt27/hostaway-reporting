#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
lot7_ik_avantages.py
Construction Lot 7 — IK & Avantages associes
Cree MASTER_FACT_MAN_IK_Avantages.xlsx (6 onglets, structure vide)
et ajoute TYPE_FLUX_015 dans REF_Setup.xlsm
Decisions D-7-01 a D-7-07 (D089-D095) integrees.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import hashlib
import shutil
import os
from datetime import datetime

# ── Chemins ────────────────────────────────────────────────────────────────
BASE = os.path.dirname(os.path.abspath(__file__))          # 02_TRAVAIL/
ROOT = os.path.dirname(BASE)                                # Pilotage_Conciergerie/
REF_PATH   = os.path.join(ROOT, '01_SOURCES_BRUTES', 'REF_Setup', 'REF_Setup.xlsm')
OUT_DIR    = os.path.join(BASE, 'Lot7_IK_Avantages')
OUT_PATH   = os.path.join(OUT_DIR, 'MASTER_FACT_MAN_IK_Avantages.xlsx')
ARCHIVE_DIR= os.path.join(ROOT, '99_ARCHIVES', 'LOT7_IK_Avantages')
TS  = datetime.now().strftime('%Y%m%d_%H%M%S')
TAG = f'lot7_ik_avantages.py — {TS}'

os.makedirs(OUT_DIR,     exist_ok=True)
os.makedirs(ARCHIVE_DIR, exist_ok=True)

# ── Backup REF_Setup ───────────────────────────────────────────────────────
backup_path = os.path.join(ARCHIVE_DIR, f'REF_Setup_BACKUP_{TS}.xlsm')
shutil.copy2(REF_PATH, backup_path)
print(f'Backup : {backup_path}')

# ── TYPE_FLUX_015 ─────────────────────────────────────────────────────────
TYPE_FLUX_015_ROW = (
    'TYPE_FLUX_015',
    'INDEMNITE_KILOMETRIQUE',
    'Indemnite kilometrique associee',
    'IC', 'OUI', 'NON', 'OUI', 'OUI',
    'Source Lot 7 — indemnites kilomeItriques associees ; '
    'justificatif ou commentaire obligatoire.'
)

wb_ref = openpyxl.load_workbook(REF_PATH, keep_vba=True)
ws_tf = wb_ref['REF_Types_Flux']
existing_tf = [r[0] for r in ws_tf.iter_rows(min_row=2, values_only=True) if r[0]]
if 'TYPE_FLUX_015' not in existing_tf:
    ws_tf.append(TYPE_FLUX_015_ROW)
    print('TYPE_FLUX_015 ajoute dans REF_Types_Flux')
else:
    print('TYPE_FLUX_015 deja present')
wb_ref.save(REF_PATH)
print('REF_Setup.xlsm sauvegarde')

# ── Charger REF pour PARAMETRES ───────────────────────────────────────────
wb_ref2 = openpyxl.load_workbook(REF_PATH, keep_vba=True, data_only=True)

def sheet_to_rows(wb, name):
    return list(wb[name].iter_rows(values_only=True))

assoc_rows = sheet_to_rows(wb_ref2, 'REF_Associes')

RELEVANT_TF = {'TYPE_FLUX_001','TYPE_FLUX_002','TYPE_FLUX_003','TYPE_FLUX_004',
               'TYPE_FLUX_005','TYPE_FLUX_008','TYPE_FLUX_015'}
tf_all = sheet_to_rows(wb_ref2, 'REF_Types_Flux')
tf_rows = [tf_all[0]] + [r for r in tf_all[1:] if r[0] in RELEVANT_TF]

mp_rows = sheet_to_rows(wb_ref2, 'REF_Modes_Paiement')

# ── Styles ────────────────────────────────────────────────────────────────
FILL_HDR  = PatternFill('solid', fgColor='1F4E79')
FILL_SEC  = PatternFill('solid', fgColor='2E75B6')
FILL_INFO = PatternFill('solid', fgColor='DEEAF1')
FONT_HDR  = Font(bold=True, color='FFFFFF', size=10)
FONT_SEC  = Font(bold=True, color='FFFFFF', size=10)
FONT_INFO = Font(italic=True, size=9, color='555555')
FONT_BOLD = Font(bold=True, size=10)
FONT_NRM  = Font(size=10)

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.fill = FILL_HDR
        cell.font = FONT_HDR
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_section(ws, start, title, headers, data_rows):
    """Write titled section in PARAMETRES. Returns next free row."""
    c = ws.cell(start, 1, title)
    c.fill = FILL_SEC
    c.font = FONT_SEC
    start += 1
    for j, h in enumerate(headers, 1):
        cl = ws.cell(start, j, h)
        cl.fill = FILL_HDR
        cl.font = FONT_HDR
        cl.alignment = Alignment(horizontal='center')
    start += 1
    for row in data_rows:
        for j, v in enumerate(row, 1):
            ws.cell(start, j, v).font = FONT_NRM
        start += 1
    return start + 1

# ── Workbook ──────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ══════════════════════════════════════════════════════════════════════════
# ONGLET 1 — SOURCE_SAISIE
# ══════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet('SOURCE_SAISIE')

SRC_HDR = [
    'mois', 'associe_id', 'type_flux_id', 'type_remboursement',
    'nature', 'montant', 'mode_paiement_id', 'lien_origine', 'commentaire'
]
SRC_WID = [12, 15, 22, 30, 32, 12, 22, 40, 45]

ws1.append(SRC_HDR)
style_header(ws1, 1, len(SRC_HDR))
set_widths(ws1, SRC_WID)
ws1.freeze_panes = 'A2'

instr = [
    'AAAA-MM',
    'PERS_EWAN ou PERS_WAFA',
    'TYPE_FLUX_001 / 015 / 005 / (AJU: 001 nature=AVANCE)',
    'ASSOCIE_VERS_SOCIETE | SOCIETE_VERS_ASSOCIE | A_CONTROLER — si TYPE_FLUX_005 seulement',
    'Description courte (trajet, virement mensuel, avance, etc.)',
    '>0',
    'PAY_001..005',
    'Ref. bancaire ou justificatif',
    'Note libre — OBLIGATOIRE si IK (TYPE_FLUX_015)'
]
ws1.append(instr)
ws1.row_dimensions[2].height = 50
for c in range(1, len(SRC_HDR) + 1):
    cell = ws1.cell(2, c)
    cell.fill = FILL_INFO
    cell.font = FONT_INFO
    cell.alignment = Alignment(wrap_text=True)

# ══════════════════════════════════════════════════════════════════════════
# ONGLET 2 — PARAMETRES
# ══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('PARAMETRES')
r = 1

r = write_section(ws2, r, 'REF_Associes',
    list(assoc_rows[0]),
    [list(x) for x in assoc_rows[1:] if any(v is not None for v in x)])

r = write_section(ws2, r, 'REF_Types_Flux (pertinents Lot 7)',
    list(tf_rows[0]),
    [list(x) for x in tf_rows[1:] if any(v is not None for v in x)])

r = write_section(ws2, r, 'REF_Modes_Paiement',
    list(mp_rows[0]),
    [list(x) for x in mp_rows[1:] if any(v is not None for v in x)])

r = write_section(ws2, r, 'Valeurs statut_controle (Lot 7)',
    ['statut_controle', 'description'],
    [
        ('VALIDE',           'Ligne validee, aucune anomalie'),
        ('A_CONTROLER',      'Ligne a verifier — anomalie detectee'),
        ('EXCLU_RESULTAT',   'Ligne exclue du calcul resultat (justifiee)'),
        ('A_VENTILER',       'Ligne en attente de ventilation par associe'),
    ])

r = write_section(ws2, r, 'Valeurs niveau_anomalie',
    ['niveau_anomalie', 'description'],
    [
        ('INFO',        'Information, pas de blocage'),
        ('A_CONTROLER', 'Anomalie a corriger avant cloture'),
        ('BLOQUANT',    'Anomalie bloquante — statut_controle=A_CONTROLER obligatoire'),
    ])

r = write_section(ws2, r, 'Valeurs type_remboursement (si TYPE_FLUX_005)',
    ['type_remboursement', 'sens', 'effet_avantages_nets'],
    [
        ('ASSOCIE_VERS_SOCIETE', "L'associe rembourse la societe",  'Deduit (soustrait)'),
        ('SOCIETE_VERS_ASSOCIE', "La societe rembourse l'associe",  'Ajoute (additionne)'),
        ('A_CONTROLER',          'Sens non determine',               'A_CONTROLER'),
    ])

r = write_section(ws2, r, 'Format avantage_id (D-7-06 = D094)',
    ['prefixe', 'type_flux', 'exemple'],
    [
        ('VIR', 'TYPE_FLUX_001 VIREMENT_ASSOCIE',       'VIR-2026-05-PERS_EWAN-001'),
        ('IK',  'TYPE_FLUX_015 INDEMNITE_KILOMETRIQUE', 'IK-2026-05-PERS_WAFA-001'),
        ('REM', 'TYPE_FLUX_005 REMBOURSEMENT_ASSOCIE',  'REM-2026-05-PERS_EWAN-001'),
        ('AJU', 'Ajustements / avances / corrections',  'AJU-2026-05-PERS_WAFA-001'),
    ])

for col, w in zip('ABCDEFGHI', [30, 35, 28, 15, 12, 12, 12, 12, 80]):
    ws2.column_dimensions[col].width = w

# ══════════════════════════════════════════════════════════════════════════
# ONGLET 3 — MASTER_SAISIE (22 colonnes)
# ══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('MASTER_SAISIE')

MS_HDR = [
    'avantage_id',               #  1  PK — VIR/IK/REM/AJU-AAAA-MM-PERS_ID-NNN
    'ROW_HASH',                  #  2  SHA256[:16]
    'mois',                      #  3  AAAA-MM
    'associe_id',                #  4  FK REF_Associes — BLOQUANT si absent
    'type_flux_id',              #  5  FK REF_Types_Flux — BLOQUANT si absent/invalide
    'type_remboursement',        #  6  ASSOCIE_VERS_SOCIETE/SOCIETE_VERS_ASSOCIE/A_CONTROLER
                                 #     obligatoire si TYPE_FLUX_005, NULL sinon
    'nature',                    #  7  description courte ; obligatoire si TYPE_FLUX_015
    'montant',                   #  8  >0 — BLOQUANT sinon
    'code_impact',               #  9  IC/HC/HR — defaut REF_Types_Flux, override controle
    'impact_resultat_reel',      # 10  OUI/NON — defaut REF_Types_Flux
    'impact_resultat_comptable', # 11  OUI/NON — defaut REF_Types_Flux
    'mode_paiement_id',          # 12  FK REF_Modes_Paiement — optionnel
    'lien_origine',              # 13  ref bancaire/justificatif
                                 #     A_CONTROLER si TYPE_FLUX_005 sans lien
    'statut_controle',           # 14  VALIDE/A_CONTROLER/EXCLU_RESULTAT/A_VENTILER
    'niveau_anomalie',           # 15  INFO/A_CONTROLER/BLOQUANT
    'code_anomalie',             # 16  code erreur
    'commentaire',               # 17  note libre ; obligatoire si TYPE_FLUX_015
    'source_table',              # 18  constante = "SOURCE_SAISIE"
    'source_pk',                 # 19  = avantage_id (self-ref)
    'source_module',             # 20  constante = "LOT7_SAISIE_DIRECTE"
    'date_saisie',               # 21  date saisie utilisateur
    'date_integration',          # 22  date passage script
]
assert len(MS_HDR) == 22, f'MASTER_SAISIE: attendu 22 colonnes, obtenu {len(MS_HDR)}'

MS_WID = [32,18,10,14,22,30,32,10,12,20,24,18,38,20,16,38,40,20,28,24,14,16]
assert len(MS_WID) == 22

ws3.append(MS_HDR)
style_header(ws3, 1, 22)
set_widths(ws3, MS_WID)
ws3.freeze_panes = 'A2'
ws3.row_dimensions[1].height = 30

# ══════════════════════════════════════════════════════════════════════════
# ONGLET 4 — MASTER_CALC_AVANTAGES (15 colonnes)
# ══════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('MASTER_CALC_AVANTAGES')

MC_HDR = [
    'pk_id',                                 #  1  associe_id + '-' + mois
    'mois',                                  #  2  AAAA-MM
    'associe_id',                            #  3  FK REF_Associes
    'avantage_brut_virements',               #  4  MASTER_SAISIE TYPE_FLUX_001
    'avantage_brut_ik',                      #  5  MASTER_SAISIE TYPE_FLUX_015
    'avantage_brut_depenses_perso',          #  6  SAISIE_Charges_Flux TYPE_FLUX_002 — Lot 3
    'avantage_brut_montant_recupere_hh',     #  7  ReservationsHorsHostaway via reservation_hh_id — Lot 4
    'avantages_bruts_total',                 #  8  = col4+col5+col6+col7
    'charges_payees_pour_societe',           #  9  SAISIE_Charges_Flux TYPE_FLUX_004+008 — Lot 3
    'remboursements_associe_vers_societe',   # 10  MASTER_SAISIE TYPE_FLUX_005 ASSOC→SOC (deduit)
    'remboursements_societe_vers_associe',   # 11  MASTER_SAISIE TYPE_FLUX_005 SOC→ASSOC (ajoute)
    'avantages_nets',                        # 12  = col8 - col9 - col10 + col11
    'detail_sources',                        # 13  liste refs
    'statut_controle',                       # 14  VALIDE/A_CONTROLER/EXCLU_RESULTAT/A_VENTILER
    'code_anomalie',                         # 15  code erreur
]
assert len(MC_HDR) == 15, f'MASTER_CALC_AVANTAGES: attendu 15 colonnes, obtenu {len(MC_HDR)}'

MC_WID = [30,10,14,26,20,30,36,24,28,38,38,18,50,20,38]
assert len(MC_WID) == 15

ws4.append(MC_HDR)
style_header(ws4, 1, 15)
set_widths(ws4, MC_WID)
ws4.freeze_panes = 'A2'
ws4.row_dimensions[1].height = 30

# Ligne de formule de reference (info)
formula_ref = [
    '# formule',
    '',
    '',
    'MASTER_SAISIE TYPE_FLUX_001',
    'MASTER_SAISIE TYPE_FLUX_015',
    'Lot3 SAISIE_Charges_Flux TYPE_FLUX_002',
    'Lot4 ReservationsHH via reservation_hh_id',
    '= col4+col5+col6+col7',
    'Lot3 SAISIE_Charges_Flux TYPE_FLUX_004+008',
    'MASTER_SAISIE TYPE_FLUX_005 type_rem=ASSOC_VERS_SOC',
    'MASTER_SAISIE TYPE_FLUX_005 type_rem=SOC_VERS_ASSOC',
    '= col8 - col9 - col10 + col11',
    '',
    '',
    '',
]
ws4.append(formula_ref)
ws4.row_dimensions[2].height = 45
for c in range(1, 16):
    cell = ws4.cell(2, c)
    cell.fill = FILL_INFO
    cell.font = FONT_INFO
    cell.alignment = Alignment(wrap_text=True)

# ══════════════════════════════════════════════════════════════════════════
# ONGLET 5 — POWER_QUERY_CODE
# ══════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet('POWER_QUERY_CODE')
ws5.column_dimensions['A'].width = 38
ws5.column_dimensions['B'].width = 120

ws5.append(['requete', 'code_M'])
style_header(ws5, 1, 2)

PQ_QUERIES = [
    ('Q1 — Charger SOURCE_SAISIE',
"""let
    Source = Excel.CurrentWorkbook(){[Name="SOURCE_SAISIE"]}[Content],
    Typed = Table.TransformColumnTypes(Source, {
        {"mois", type text}, {"associe_id", type text},
        {"type_flux_id", type text}, {"type_remboursement", type text},
        {"nature", type text}, {"montant", type number},
        {"mode_paiement_id", type text}, {"lien_origine", type text},
        {"commentaire", type text}
    }),
    // Exclure ligne instruction (ligne 2 vide ou format mois invalide)
    Filtered = Table.SelectRows(Typed, each
        Text.Length([mois] ?? "") = 7
        and Text.Contains([mois] ?? "", "-")
        and [associe_id] <> null
    )
in
    Filtered"""),

    ('Q2 — Enrichir depuis REF_Types_Flux',
"""let
    Source = Q1,
    RefTF = Excel.CurrentWorkbook(){[Name="REF_TYPES_FLUX_LOT7"]}[Content],
    Joined = Table.NestedJoin(Source, "type_flux_id",
        RefTF, "type_flux_id", "ref", JoinKind.LeftOuter),
    Exp = Table.ExpandTableColumn(Joined, "ref",
        {"code_impact_defaut","avantage_brut_defaut","deduit_avantage_defaut"},
        {"code_impact_defaut","avantage_brut_defaut","deduit_avantage_defaut"}),
    WithCI = Table.AddColumn(Exp, "code_impact", each
        [code_impact_defaut] ?? "A_CONTROLER"),
    WithIR = Table.AddColumn(WithCI, "impact_resultat_reel", each
        if ([code_impact_defaut] ?? "") = "HR" then "NON" else "OUI"),
    WithIC = Table.AddColumn(WithIR, "impact_resultat_comptable", each
        if ([code_impact_defaut] ?? "") = "IC" then "OUI" else "NON")
in
    WithIC"""),

    ('Q3 — Produire MASTER_SAISIE (22 colonnes)',
"""let
    Source = Q2,
    // Controles metier
    WithCtrl = Table.AddColumn(Source, "_ctrl_code", each
        if [associe_id] = null then "AVANTAGE_ASSOCIE_SANS_ASSOCIE_ID|BLOQUANT"
        else if [type_flux_id] = null then "AVANTAGE_SANS_TYPE_FLUX|BLOQUANT"
        else if [montant] = null or [montant] <= 0 then "AVANTAGE_ASSOCIE_MONTANT_INVALIDE|BLOQUANT"
        else if [type_flux_id] = "TYPE_FLUX_005"
             and ([type_remboursement] = null or [type_remboursement] = "")
             then "REMBOURSEMENT_SENS_ABSENT|A_CONTROLER"
        else if [type_flux_id] = "TYPE_FLUX_005"
             and ([lien_origine] = null or [lien_origine] = "")
             then "REMBOURSEMENT_SANS_LIEN_ORIGINE|A_CONTROLER"
        else if [type_flux_id] = "TYPE_FLUX_001"
             and ([lien_origine] = null or [lien_origine] = "")
             then "VIREMENT_ASSOCIE_SANS_LIEN_BANQUE|A_CONTROLER"
        else if [type_flux_id] = "TYPE_FLUX_015"
             and ([commentaire] = null or [commentaire] = "")
             and ([lien_origine] = null or [lien_origine] = "")
             then "IK_SANS_JUSTIFICATIF|A_CONTROLER"
        else "|"
    ),
    Split = Table.SplitColumn(WithCtrl, "_ctrl_code",
        Splitter.SplitTextByDelimiter("|"),
        {"_code_anomalie", "_niveau_anomalie"}),
    WithSC = Table.AddColumn(Split, "statut_controle", each
        if [_niveau_anomalie] <> "" then "A_CONTROLER" else "VALIDE"),
    WithNA = Table.AddColumn(WithSC, "niveau_anomalie", each
        if [_niveau_anomalie] = "BLOQUANT" then "BLOQUANT"
        else if [_niveau_anomalie] = "A_CONTROLER" then "A_CONTROLER"
        else ""),
    WithCA = Table.AddColumn(WithNA, "code_anomalie", each
        if [_code_anomalie] = "" then "" else [_code_anomalie]),
    // Colonnes source
    WithST = Table.AddColumn(WithCA, "source_table",  each "SOURCE_SAISIE"),
    WithSM = Table.AddColumn(WithST, "source_module", each "LOT7_SAISIE_DIRECTE"),
    WithDI = Table.AddColumn(WithSM, "date_integration", each DateTime.LocalNow()),
    // Selection colonnes finales dans l ordre
    Final = Table.SelectColumns(WithDI, {
        "mois","associe_id","type_flux_id","type_remboursement",
        "nature","montant","code_impact","impact_resultat_reel","impact_resultat_comptable",
        "mode_paiement_id","lien_origine","statut_controle","niveau_anomalie","code_anomalie",
        "commentaire","source_table","source_module","date_integration"
    })
    // Note : avantage_id, ROW_HASH, source_pk et date_saisie sont renseignes
    // manuellement dans SOURCE_SAISIE ou ajoutes ici selon configuration.
in
    Final"""),

    ('Q4 — Sources derivees Lot 3 + Lot 4',
"""// Q4 — Sources derivees
// Lot 3 : MASTER_FACT_MAN_Charges.xlsx
// Lot 4 : MASTER_FACT_MAN_ReservationsHorsHostaway.xlsx
// Adapter chemins selon emplacement. Exemples ci-dessous.

// Q4a — Depenses perso + charges payees (Lot 3)
let
    Path3 = "..\\Lot3_Charges\\MASTER_FACT_MAN_Charges.xlsx",
    WB3 = Excel.Workbook(File.Contents(Path3), null, true),
    MASTER3 = WB3{[Name="MASTER"]}[Data],
    DepPerso = Table.SelectRows(MASTER3, each [type_flux_id] = "TYPE_FLUX_002"),
    AggDP = Table.Group(DepPerso, {"mois","associe_id"}, {
        {"avantage_brut_depenses_perso", each List.Sum([montant]), type number}
    }),
    Charges = Table.SelectRows(MASTER3, each
        [type_flux_id] = "TYPE_FLUX_004" or [type_flux_id] = "TYPE_FLUX_008"),
    AggCH = Table.Group(Charges, {"mois","associe_id"}, {
        {"charges_payees_pour_societe", each List.Sum([montant]), type number}
    })
in
    {AggDP, AggCH}

// Q4b — Montants recuperes HH (Lot 4)
let
    Path4 = "..\\Lot4_ReservationsHH\\MASTER_FACT_MAN_ReservationsHorsHostaway.xlsx",
    WB4 = Excel.Workbook(File.Contents(Path4), null, true),
    MASTER4 = WB4{[Name="MASTER"]}[Data],
    HH = Table.SelectRows(MASTER4, each
        [montant_recupere] <> null and [montant_recupere] > 0),
    AggHH = Table.Group(HH, {"mois","associe_id_recuperateur"}, {
        {"avantage_brut_montant_recupere_hh", each List.Sum([montant_recupere]), type number}
    }),
    // Renommer pour join sur associe_id
    Renamed = Table.RenameColumns(AggHH, {{"associe_id_recuperateur","associe_id"}})
in
    Renamed"""),

    ('Q5 — Produire MASTER_CALC_AVANTAGES (15 colonnes)',
"""// Formule :
// avantages_bruts_total = virements + IK + depenses_perso + montant_recupere_hh
// avantages_nets = bruts_total - charges_payees - rem_ASSOC_VERS_SOC + rem_SOC_VERS_ASSOC

let
    SAISIE = Q3,
    // Agregations MASTER_SAISIE
    VIR = Table.Group(
        Table.SelectRows(SAISIE, each [type_flux_id] = "TYPE_FLUX_001"),
        {"mois","associe_id"}, {{"v", each List.Sum([montant]), type number}}),
    IK = Table.Group(
        Table.SelectRows(SAISIE, each [type_flux_id] = "TYPE_FLUX_015"),
        {"mois","associe_id"}, {{"k", each List.Sum([montant]), type number}}),
    REM_AVS = Table.Group(
        Table.SelectRows(SAISIE, each
            [type_flux_id] = "TYPE_FLUX_005"
            and [type_remboursement] = "ASSOCIE_VERS_SOCIETE"),
        {"mois","associe_id"}, {{"ravs", each List.Sum([montant]), type number}}),
    REM_SVA = Table.Group(
        Table.SelectRows(SAISIE, each
            [type_flux_id] = "TYPE_FLUX_005"
            and [type_remboursement] = "SOCIETE_VERS_ASSOCIE"),
        {"mois","associe_id"}, {{"rsva", each List.Sum([montant]), type number}}),
    // Grille mois x associe
    Pairs = Table.Distinct(Table.SelectColumns(SAISIE, {"mois","associe_id"})),
    // Jointures
    J1 = Table.NestedJoin(Pairs,{"mois","associe_id"},VIR,{"mois","associe_id"},"_v",JoinKind.LeftOuter),
    J1e= Table.ExpandTableColumn(J1,"_v",{"v"},{"avantage_brut_virements"}),
    J2 = Table.NestedJoin(J1e,{"mois","associe_id"},IK,{"mois","associe_id"},"_k",JoinKind.LeftOuter),
    J2e= Table.ExpandTableColumn(J2,"_k",{"k"},{"avantage_brut_ik"}),
    // Sources derivees (Lot 3 et Lot 4 — a 0 jusqu au peuplement)
    // Ajouter ici: J3 AggDP, J4 AggCH, J5 AggHH quand disponibles
    // Initialisation a 0 dans l attente
    WZ = Table.TransformColumns(J2e, {
        {"avantage_brut_virements", each _ ?? 0, type number},
        {"avantage_brut_ik",        each _ ?? 0, type number}
    }),
    WDP  = Table.AddColumn(WZ,  "avantage_brut_depenses_perso",         each 0.0),
    WHH  = Table.AddColumn(WDP, "avantage_brut_montant_recupere_hh",    each 0.0),
    WCH  = Table.AddColumn(WHH, "charges_payees_pour_societe",          each 0.0),
    // Remboursements
    JR1  = Table.NestedJoin(WCH,{"mois","associe_id"},REM_AVS,{"mois","associe_id"},"_r1",JoinKind.LeftOuter),
    JR1e = Table.ExpandTableColumn(JR1,"_r1",{"ravs"},{"_ravs"}),
    JR2  = Table.NestedJoin(JR1e,{"mois","associe_id"},REM_SVA,{"mois","associe_id"},"_r2",JoinKind.LeftOuter),
    JR2e = Table.ExpandTableColumn(JR2,"_r2",{"rsva"},{"_rsva"}),
    WRZ  = Table.TransformColumns(JR2e, {
        {"_ravs", each _ ?? 0, type number},
        {"_rsva", each _ ?? 0, type number}
    }),
    // Calculs finaux
    WBruts = Table.AddColumn(WRZ, "avantages_bruts_total", each
        [avantage_brut_virements] + [avantage_brut_ik]
        + [avantage_brut_depenses_perso] + [avantage_brut_montant_recupere_hh]),
    WNets = Table.AddColumn(WBruts, "avantages_nets", each
        [avantages_bruts_total]
        - [charges_payees_pour_societe]
        - [_ravs]
        + [_rsva]),
    WPK = Table.AddColumn(WNets, "pk_id", each [associe_id] & "-" & [mois]),
    WSC = Table.AddColumn(WPK, "statut_controle", each
        if [avantages_nets] < 0 then "A_CONTROLER" else "VALIDE"),
    WAN = Table.AddColumn(WSC, "code_anomalie", each
        if [avantages_nets] < 0 then "AVANTAGE_NET_NEGATIF" else ""),
    WDS = Table.AddColumn(WAN, "detail_sources", each
        "SAISIE_LOT7=" & Text.From([avantage_brut_virements] + [avantage_brut_ik])
        & " | LOT3=" & Text.From([avantage_brut_depenses_perso] + [charges_payees_pour_societe])
        & " | LOT4=" & Text.From([avantage_brut_montant_recupere_hh])),
    // Selection et renommage colonnes finales
    Sel = Table.SelectColumns(WDS, {
        "pk_id","mois","associe_id",
        "avantage_brut_virements","avantage_brut_ik",
        "avantage_brut_depenses_perso","avantage_brut_montant_recupere_hh",
        "avantages_bruts_total","charges_payees_pour_societe",
        "_ravs","_rsva",
        "avantages_nets","detail_sources","statut_controle","code_anomalie"
    }),
    Ren = Table.RenameColumns(Sel, {
        {"_ravs", "remboursements_associe_vers_societe"},
        {"_rsva", "remboursements_societe_vers_associe"}
    })
in
    Ren"""),
]

for q_name, q_code in PQ_QUERIES:
    row_i = ws5.max_row + 1
    ws5.cell(row_i, 1, q_name).font = Font(bold=True, size=10)
    ws5.cell(row_i, 2, q_code).font = Font(name='Courier New', size=9)
    ws5.cell(row_i, 2).alignment = Alignment(wrap_text=True, vertical='top')
    ws5.row_dimensions[row_i].height = 220

# ══════════════════════════════════════════════════════════════════════════
# ONGLET 6 — README
# ══════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet('README')
ws6.column_dimensions['A'].width = 28
ws6.column_dimensions['B'].width = 105

README_LINES = [
    ('Lot',           'Lot 7 — IK & Avantages associes'),
    ('Fichier',       'MASTER_FACT_MAN_IK_Avantages.xlsx'),
    ('Produit par',   TAG),
    ('Pre-requis',    'Lots 3 (SAISIE_Charges_Flux) et 4 (ReservationsHorsHostaway)'),
    ('',              ''),
    ('ONGLETS',       ''),
    ('SOURCE_SAISIE', 'Saisie directe brute — flux non disponibles ailleurs (virements, IK, remboursements, ajustements)'),
    ('PARAMETRES',    'Referentiels locaux — associes, types flux, modes paiement, valeurs fermees'),
    ('MASTER_SAISIE', 'Lignes normalisees — 22 colonnes — PK stable — controles'),
    ('MASTER_CALC_AVANTAGES', 'Consolidation calculee par associe_id x mois — 15 colonnes — formule avantages_nets'),
    ('POWER_QUERY_CODE', 'Code M requetes Q1-Q5'),
    ('README',        'Ce fichier'),
    ('',              ''),
    ('SAISIE RESIDUELLE', 'MASTER_SAISIE = saisie strictement residuelle'),
    ('NE PAS RESSAISIR', 'DEPENSE_PERSO_COMPTE_PRO → SAISIE_Charges_Flux Lot 3 (TYPE_FLUX_002)'),
    ('NE PAS RESSAISIR', 'CHARGE_PAYEE_PERSO_OU_LIQUIDE → SAISIE_Charges_Flux Lot 3 (TYPE_FLUX_004)'),
    ('NE PAS RESSAISIR', 'PAIEMENT_PRESTATAIRE_LIQUIDE → SAISIE_Charges_Flux Lot 3 (TYPE_FLUX_008)'),
    ('NE PAS RESSAISIR', 'montant_recupere HH → ReservationsHorsHostaway Lot 4 via reservation_hh_id'),
    ('',              ''),
    ('FORMULE',       'avantages_bruts_total = virements + IK + depenses_perso + montant_recupere_hh'),
    ('FORMULE',       'avantages_nets = bruts_total - charges_payees - rem_ASSOC_VERS_SOC + rem_SOC_VERS_ASSOC'),
    ('',              ''),
    ('CONTROLES BLOQUANTS',   'niveau_anomalie=BLOQUANT, statut_controle=A_CONTROLER'),
    ('',  'AVANTAGE_ASSOCIE_SANS_ASSOCIE_ID      — associe_id IS NULL'),
    ('',  'AVANTAGE_ASSOCIE_MONTANT_INVALIDE      — montant <= 0'),
    ('',  'AVANTAGE_SANS_TYPE_FLUX                — type_flux_id absent ou hors REF_Types_Flux'),
    ('',  'MONTANT_RECUPERE_HH_NON_REPRIS_AVANTAGES — montant_recupere HH > 0 sans MASTER_CALC'),
    ('',  'DOUBLE_COMPTAGE_SAISIE_ET_DERIVE        — flux present ici ET dans Lot 3/4'),
    ('',  'SAISIE_LOT7_SOURCE_DEJA_EXISTANTE       — source canonique existe deja en Lot 3, 4 ou Lot 8'),
    ('',              ''),
    ('CONTROLES A_CONTROLER', 'niveau_anomalie=A_CONTROLER, statut_controle=A_CONTROLER'),
    ('',  'REMBOURSEMENT_SANS_LIEN_ORIGINE    — TYPE_FLUX_005 sans lien_origine'),
    ('',  'REMBOURSEMENT_SENS_ABSENT          — TYPE_FLUX_005 sans type_remboursement valide'),
    ('',  'VIREMENT_ASSOCIE_SANS_LIEN_BANQUE  — TYPE_FLUX_001 sans lien_origine'),
    ('',  'IK_SANS_JUSTIFICATIF              — TYPE_FLUX_015 sans commentaire ET sans lien_origine'),
    ('',  'AVANTAGE_NET_NEGATIF              — avantages_nets < 0'),
    ('',  'TYPE_FLUX_IMPACT_INCOHERENT       — code_impact diverge du defaut REF_Types_Flux sans commentaire'),
    ('',              ''),
    ('TYPE_FLUX_015',   'INDEMNITE_KILOMETRIQUE — IC — avantage_brut=OUI — ajoute REF_Setup.xlsm Lot 7'),
    ('D089-D095',       'Decisions D-7-01 a D-7-07 (voir DECISIONS_METIER.md)'),
    ('D096',            'Correction LOG_0016 T2→T3 (Cyprien/Clarisse) — faite avant Lot 7'),
    ('',                ''),
    ('Peuplement',      'Structure vide — aucune donnee fictive — a saisir via SOURCE_SAISIE'),
    ('Lot 8 (banque)',  'Virements associes du compte pro rapproches sans ressaisie au Lot 8'),
    ('Lot 3 (charges)', 'DEPENSE_PERSO et CHARGE_PAYEE_PERSO derives sans ressaisie quand peuple'),
    ('Lot 4 (HH)',      'montant_recupere derive via reservation_hh_id sans ressaisie'),
]

for i, (k, v) in enumerate(README_LINES, 1):
    cell_k = ws6.cell(i, 1, k)
    cell_v = ws6.cell(i, 2, v)
    cell_v.alignment = Alignment(wrap_text=True)
    if k in ('ONGLETS','SAISIE RESIDUELLE','FORMULE','CONTROLES BLOQUANTS','CONTROLES A_CONTROLER'):
        cell_k.fill = FILL_SEC
        cell_k.font = FONT_SEC
    elif k:
        cell_k.font = FONT_BOLD
    else:
        cell_k.font = FONT_NRM
    cell_v.font = FONT_NRM

# ── Sauvegarde Excel ───────────────────────────────────────────────────────
wb.save(OUT_PATH)

# ── Rapport ────────────────────────────────────────────────────────────────
print(f'\nCree     : {OUT_PATH}')
print(f'Onglets  : {[s.title for s in wb.worksheets]}')
print(f'MASTER_SAISIE          : {len(MS_HDR)} colonnes')
print(f'MASTER_CALC_AVANTAGES  : {len(MC_HDR)} colonnes')
print(f'SOURCE_SAISIE          : {len(SRC_HDR)} colonnes saisie')
print(f'Power Query            : {len(PQ_QUERIES)} requetes (Q1-Q5)')
print(f'TYPE_FLUX_015          : INDEMNITE_KILOMETRIQUE — IC — ajoutee')
print(f'Backup REF_Setup       : {backup_path}')
print(f'Tag script             : {TAG}')
