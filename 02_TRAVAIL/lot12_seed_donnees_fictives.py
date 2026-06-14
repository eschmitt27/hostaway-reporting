#!/usr/bin/env python3
"""
lot12_seed_donnees_fictives.py
Insertion de DONNÉES FICTIVES de test pour valider le pipeline jusqu'au Lot 12.

RÈGLE ABSOLUE : aucune donnée fictive ne doit être confondue avec une vraie donnée.
  - Tous les IDs fictifs commencent par : ZZ_TEST_
  - Tag obligatoire dans une colonne commentaire / code_anomalie :
        __TEST_FICTIF_LOT12_A_SUPPRIMER__
  - Montants reconnaissables : 111.11 / 222.22 / 333.33

Sécurité :
  - Sauvegarde de chaque fichier modifié dans 99_ARCHIVES/LOT12_TEST_DATA/ avant écriture.
  - Idempotent : ne réinsère pas une ligne dont l'ID ZZ_TEST_ est déjà présent.
  - N'écrit QUE des lignes nouvelles. Ne modifie aucune ligne réelle existante.

Périmètre (décisions humaines 2026-06-14) :
  - Lot 9 d'abord seul : Charges + M04 + HH + Ménage externe (sources lues par lot9 étendu / lot4bis).
  - IK exclu (décision « IK hors Flux »).
  - Acomptes différés (passe Lot 10 ultérieure).
  - Lignes fictives en statut_controle = VALIDE (pour traverser le pipeline).

Dépendances : pip install openpyxl
"""

import os
import shutil
import datetime
import openpyxl

TAG = "__TEST_FICTIF_LOT12_A_SUPPRIMER__"
COMMENT = f"Donnée fictive de test Lot 12 — {TAG}"
NOW = datetime.datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
TODAY = datetime.date.today().isoformat()

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BACKUP_DIR = os.path.join(ROOT, "99_ARCHIVES", "LOT12_TEST_DATA")

# Commun à toutes les lignes fictives
SRC_COMMON = {
    "statut_controle": "VALIDE",
    "niveau_anomalie": "INFO",
    "source_module":   "LOT12_TEST_FICTIF",
    "source_table":    "SEED_FICTIF",
    "date_integration": NOW,
}

# ── Définition des lignes fictives par fichier ─────────────────────────────────

CHARGES = [
    {  # 1 charge logement IC
        "charge_id": "ZZ_TEST_CHARGE_001", "date_charge": "2026-03-15", "mois": "2026-03",
        "montant": 111.11, "sens_flux": "CHARGE", "sens": "CHARGE",
        "categorie_charge_id": "CHG_001", "type_flux_id": "TYPE_FLUX_011", "code_impact": "IC",
        "impact_resultat_reel": "OUI", "impact_resultat_comptable": "OUI", "prise_en_compta": "OUI",
        "affectation_type": "LOGEMENT", "logement_id": "LOG_0001", "proprietaire_id": "PROP_0001",
        "commentaire": COMMENT, "source_pk": "ZZ_TEST_CHARGE_001", "date_saisie": NOW,
    },
    {  # 1 charge ménage HC
        "charge_id": "ZZ_TEST_CHARGE_002", "date_charge": "2026-03-15", "mois": "2026-03",
        "montant": 111.11, "sens_flux": "CHARGE", "sens": "CHARGE",
        "categorie_charge_id": "CHG_001", "type_flux_id": "TYPE_FLUX_009", "code_impact": "HC",
        "impact_resultat_reel": "OUI", "impact_resultat_comptable": "NON", "prise_en_compta": "NON",
        "affectation_type": "LOGEMENT", "logement_id": "LOG_0001", "proprietaire_id": "PROP_0001",
        "commentaire": COMMENT, "source_pk": "ZZ_TEST_CHARGE_002", "date_saisie": NOW,
    },
    {  # 1 dépense perso compte pro IC
        "charge_id": "ZZ_TEST_CHARGE_003", "date_charge": "2026-03-16", "mois": "2026-03",
        "montant": 222.22, "sens_flux": "CHARGE", "sens": "CHARGE",
        "categorie_charge_id": "CHG_015", "type_flux_id": "TYPE_FLUX_011", "code_impact": "IC",
        "impact_resultat_reel": "OUI", "impact_resultat_comptable": "OUI", "prise_en_compta": "OUI",
        "affectation_type": "GLOBAL", "associe_id": "PERS_EWAN", "mode_paiement_id": "PAY_001",
        "commentaire": COMMENT, "source_pk": "ZZ_TEST_CHARGE_003", "date_saisie": NOW,
    },
    {  # 1 charge payée perso/liquide HC
        "charge_id": "ZZ_TEST_CHARGE_004", "date_charge": "2026-03-16", "mois": "2026-03",
        "montant": 222.22, "sens_flux": "CHARGE", "sens": "CHARGE",
        "categorie_charge_id": "CHG_015", "type_flux_id": "TYPE_FLUX_010", "code_impact": "HC",
        "impact_resultat_reel": "OUI", "impact_resultat_comptable": "NON", "prise_en_compta": "NON",
        "affectation_type": "GLOBAL", "associe_id": "PERS_WAFA", "mode_paiement_id": "PAY_002",
        "commentaire": COMMENT, "source_pk": "ZZ_TEST_CHARGE_004", "date_saisie": NOW,
    },
    {  # 1 charge non affectable IC
        "charge_id": "ZZ_TEST_CHARGE_005", "date_charge": "2026-03-17", "mois": "2026-03",
        "montant": 333.33, "sens_flux": "CHARGE", "sens": "CHARGE",
        "categorie_charge_id": "CHG_001", "type_flux_id": "TYPE_FLUX_011", "code_impact": "IC",
        "impact_resultat_reel": "OUI", "impact_resultat_comptable": "OUI", "prise_en_compta": "OUI",
        "affectation_type": "NON_AFFECTABLE",
        "commentaire": COMMENT, "source_pk": "ZZ_TEST_CHARGE_005", "date_saisie": NOW,
    },
]

M04 = [
    {  # 1 ménage interne fictif (tag dans code_anomalie : pas de colonne commentaire en M04)
        "menage_calc_id": "ZZ_TEST_MENAGE_INT_001", "mois": "2026-03", "annee": 2026, "mois_num": 3,
        "logement_id": "LOG_0001", "proprietaire_id": "PROP_0001",
        "intervenant_id": "INT_0001", "nom_intervenant": "Imene TEST", "type_intervenant": "INTERNE",
        "type_menage": "STANDARD", "nb_menages": 1, "nb_heures": 11.11,
        "taux_horaire_intervenant": 10, "cout_execution_total": 111.11,
        "cout_execution_unitaire": 111.11, "total_execution": 111.11,
        "type_flux_id": "TYPE_FLUX_013", "sens": "CHARGE", "code_impact": "HC",
        "impact_resultat_reel": "OUI", "impact_resultat_comptable": "NON",
        "code_anomalie": TAG, "source_pk": "ZZ_TEST_MENAGE_INT_001",
    },
]

HH = [
    {  # réservation directe
        "reservation_hh_id": "ZZ_TEST_HH_001", "mois": "2026-03", "canal_id": "DIRECT",
        "source_financiere": "DIRECT", "proprietaire_id": "PROP_0001", "logement_id": "LOG_0001",
        "date_arrivee": "2026-03-10", "date_depart": "2026-03-13", "nuits": 3,
        "total_percu": 333.33, "menage": 111.11, "taux_commission": 0.18, "commission": 40.00,
        "montant_recupere": 111.11, "associe_id_recuperateur": "PERS_EWAN",
        "montant_reverse_proprietaire": 200.00, "mode_paiement_id": "PAY_002", "acompte_facture": 0,
        "code_impact": "IC", "impact_resultat_reel": "OUI", "impact_resultat_comptable": "OUI",
        "commentaire": COMMENT, "source_pk": "ZZ_TEST_HH_001",
    },
    {  # réservation type VRBO
        "reservation_hh_id": "ZZ_TEST_HH_002", "mois": "2026-03", "canal_id": "VRBO",
        "source_financiere": "VRBO", "proprietaire_id": "PROP_0002", "logement_id": "LOG_0002",
        "date_arrivee": "2026-03-20", "date_depart": "2026-03-22", "nuits": 2,
        "total_percu": 222.22, "menage": 111.11, "taux_commission": 0.15, "commission": 16.67,
        "montant_recupere": 0, "montant_reverse_proprietaire": 100.00, "mode_paiement_id": "PAY_001",
        "acompte_facture": 0, "code_impact": "IC",
        "impact_resultat_reel": "OUI", "impact_resultat_comptable": "OUI",
        "commentaire": COMMENT, "source_pk": "ZZ_TEST_HH_002",
    },
]

MENEXT = [
    {  # ménage externe avec date de prestation
        "menage_externe_id": "ZZ_TEST_MENEXT_001", "facture_id": "ZZ_TEST_FACTURE_001",
        "date_facture": "2026-03-20", "date_menage": "2026-03-18", "precision_date_menage": "EXACTE",
        "mois": "2026-03", "annee": 2026, "prestataire_id": "INT_0003", "nom_prestataire": "Mounir TEST",
        "type_intervenant": "EXTERNE", "regime_tva_prestataire": "FRANCHISE_TVA",
        "logement_id": "LOG_0001", "proprietaire_id": "PROP_0001",
        "type_ligne_menage_id": "TLM_001", "nombre_menages": 1,
        "montant_ligne_ht": 111.11, "taux_tva": 0, "montant_tva": 0, "montant_ligne_ttc": 111.11,
        "montant_facture_total_ht": 111.11, "montant_facture_total_ttc": 111.11,
        "somme_lignes_facture_ttc": 111.11, "ecart_reconciliation_ttc": 0,
        "code_impact": "IC", "prise_en_compta": "OUI",
        "impact_resultat_reel": "OUI", "impact_resultat_comptable": "OUI",
        "type_flux_id": "TYPE_FLUX_014", "sens": "CHARGE",
        "commentaire": COMMENT, "source_pk": "ZZ_TEST_MENEXT_001",
    },
]

ACOMPTES = [
    {  # acompte rattaché à une facture (VALIDE -> injecté en REGLEMENT par lot10)
        "acompte_id": "ZZ_TEST_ACOMPTE_001", "mois": "2026-03",
        "proprietaire_id": "PROP_0001", "logement_id": "LOG_0001",
        "facture_ref": "ZZ_TEST_FACTURE_001", "source_acompte": "VIREMENT_DIRECT",
        "montant_acompte": 222.22, "mode_paiement_id": "PAY_001", "code_impact": "HC",
        "impact_resultat_reel": "OUI", "impact_resultat_comptable": "NON",
        "commentaire": COMMENT, "source_pk": "ZZ_TEST_ACOMPTE_001",
    },
    {  # acompte NON rattaché (A_CONTROLER -> exclu de l'injection lot10, teste contrôle)
        "acompte_id": "ZZ_TEST_ACOMPTE_002", "mois": "2026-03",
        "proprietaire_id": "PROP_0002", "logement_id": "LOG_0002",
        "facture_ref": None, "source_acompte": "AUTRE",
        "montant_acompte": 111.11, "mode_paiement_id": "PAY_001", "code_impact": "HC",
        "impact_resultat_reel": "OUI", "impact_resultat_comptable": "NON",
        "statut_controle": "A_CONTROLER",
        "commentaire": COMMENT, "source_pk": "ZZ_TEST_ACOMPTE_002",
    },
]

# (fichier, onglet, colonne_id, lignes)
# PASSE LOT 10 : Charges + M04 + MenExt + HH + Acomptes.
# HH géré par lot10 via montant saisi (pas payout Hostaway).
# Acomptes injectés uniquement dans REGLEMENT (jamais exploitation).
TARGETS = [
    ("Charges",  "02_TRAVAIL/Lot3_Charges/MASTER_FACT_MAN_Charges.xlsx",                         "MASTER", "charge_id",         CHARGES),
    ("M04",      "02_DONNEES_NORMALISEES/menages/M04_MENAGES_PowerQuery.xlsx",                    "MASTER", "menage_calc_id",    M04),
    ("MenExt",   "02_TRAVAIL/Lot6c_MenagesExternes/MASTER_FACT_MEN_MenagesExternes.xlsx",         "MASTER", "menage_externe_id", MENEXT),
    ("HH",       "02_TRAVAIL/Lot4_ReservationsHH/MASTER_FACT_MAN_ReservationsHorsHostaway.xlsx",  "MASTER", "reservation_hh_id", HH),
    ("Acomptes", "02_TRAVAIL/Lot5_AcomptesProprietaires/MASTER_FACT_MAN_AcomptesProprietaires.xlsx", "MASTER", "acompte_id",    ACOMPTES),
]

# Alias historique (suppression couvre les mêmes cibles)
TARGETS_ALL = TARGETS


def backup(path):
    os.makedirs(BACKUP_DIR, exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    base = os.path.basename(path)
    dst = os.path.join(BACKUP_DIR, f"{base}.BACKUP_{ts}")
    shutil.copy2(path, dst)
    return dst


def seed_file(name, relpath, sheet_name, id_col, rows):
    path = os.path.join(ROOT, relpath)
    if not os.path.exists(path):
        print(f"  [{name}] ABSENT : {path} — ignoré")
        return 0
    wb = openpyxl.load_workbook(path)
    if sheet_name not in wb.sheetnames:
        print(f"  [{name}] onglet {sheet_name} absent — ignoré")
        wb.close()
        return 0
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    if id_col not in headers:
        print(f"  [{name}] colonne {id_col} absente — ignoré")
        wb.close()
        return 0
    id_idx = headers.index(id_col)

    # IDs déjà présents (idempotence)
    existing = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        v = r[id_idx]
        if v is not None:
            existing.add(str(v))

    added = 0
    skipped = 0
    for row in rows:
        rid = row[id_col]
        if rid in existing:
            skipped += 1
            continue
        full = dict(SRC_COMMON)
        full.update(row)
        # source_table commun sauf si déjà défini
        full.setdefault("source_table", "SEED_FICTIF")
        ordered = [full.get(h) for h in headers]
        ws.append(ordered)
        added += 1

    if added > 0:
        bkp = backup(path)
        wb.save(path)
        print(f"  [{name}] +{added} lignes fictives (skip {skipped}) — backup: {os.path.basename(bkp)}")
    else:
        print(f"  [{name}] 0 ajout (déjà présent : {skipped})")
    wb.close()
    return added


def verify_tag(name, relpath, sheet_name, id_col):
    """Vérifie que chaque ligne ZZ_TEST_ porte le tag quelque part."""
    path = os.path.join(ROOT, relpath)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    headers = [c for c in next(ws.iter_rows(values_only=True))]
    id_idx = headers.index(id_col)
    bad = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        v = r[id_idx]
        if v is not None and str(v).startswith("ZZ_TEST_"):
            has_tag = any(TAG in str(c) for c in r if c is not None)
            if not has_tag:
                bad.append(str(v))
    wb.close()
    if bad:
        print(f"  [{name}] !! lignes ZZ_TEST_ SANS tag : {bad}")
        return False
    return True


def main():
    print("=" * 60)
    print("SEED DONNÉES FICTIVES LOT 12")
    print("=" * 60)
    print(f"Tag      : {TAG}")
    print(f"Backups  : {BACKUP_DIR}")
    print()

    total = 0
    for name, relpath, sheet, id_col, rows in TARGETS:
        total += seed_file(name, relpath, sheet, id_col, rows)

    print()
    print("Vérification tag sur toutes les lignes ZZ_TEST_ :")
    all_ok = True
    for name, relpath, sheet, id_col, rows in TARGETS:
        if os.path.exists(os.path.join(ROOT, relpath)):
            ok = verify_tag(name, relpath, sheet, id_col)
            if ok:
                print(f"  [{name}] OK — toutes lignes fictives taguées")
            all_ok = all_ok and ok

    print()
    print(f"Total lignes fictives ajoutées : {total}")
    print(f"Vérification tag : {'OK' if all_ok else 'ÉCHEC'}")
    print("Rappel : ne JAMAIS committer avec des données fictives présentes (règle #10).")
    print("Nettoyage : py 02_TRAVAIL/lot12_remove_donnees_fictives.py")


if __name__ == "__main__":
    main()
