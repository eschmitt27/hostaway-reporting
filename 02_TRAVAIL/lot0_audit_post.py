"""Lot 0 - Audit post-correction REF_Setup.xlsm"""
import openpyxl
from collections import Counter

FILEPATH = r"C:/Users/Ewan/OneDrive/Documents/Conciergerie/Pilotage_Conciergerie/01_SOURCES_BRUTES/REF_Setup/REF_Setup.xlsm"

wb = openpyxl.load_workbook(FILEPATH, read_only=True, keep_vba=True, data_only=True)

def get_dicts(wb, sname):
    ws = wb[sname]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = rows[0]
    return headers, [dict(zip(headers, r)) for r in rows[1:]]

print("=" * 60)
print("AUDIT POST-CORRECTION — REF_Setup.xlsm")
print("=" * 60)

# 1. Onglets
print(f"\n[ONGLETS] Total: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    ws = wb[s]
    rows = list(ws.iter_rows(values_only=True))
    n = len(rows) - 1
    print(f"  {s}: {n} lignes")

# 2. Mojibake
print("\n[B1 MOJIBAKE]")
found = False
for sname in wb.sheetnames:
    ws = wb[sname]
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, str) and '\xc3' in cell:
                print(f"  ENCORE PRESENT [{sname}]: {repr(cell[:60])}")
                found = True
if not found:
    print("  OK: Aucun mojibake detecte")

# 3. REF_Statuts valeurs statut_controle
print("\n[B2 REF_Statuts]")
_, statuts = get_dicts(wb, 'REF_Statuts')
sc_vals = [r['statut'] for r in statuts if r.get('famille_statut') == 'statut_controle']
print(f"  Famille statut_controle: {sc_vals}")
for v in ['VALIDE', 'BLOQUANT', 'IGNORE_JUSTIFIE', 'A_CONTROLER']:
    found_v = any(r['statut'] == v for r in statuts)
    print(f"  {v}: {'OK' if found_v else 'MANQUANT'}")

# 4. REF_Statuts_Payout
print("\n[B3 REF_Statuts_Payout]")
if 'REF_Statuts_Payout' in wb.sheetnames:
    _, spay = get_dicts(wb, 'REF_Statuts_Payout')
    vals = [r.get('statut_calcul_payout') for r in spay]
    print(f"  OK: {len(spay)} lignes: {vals}")
else:
    print("  MANQUANT")

# 5. REF_Cloture_Mensuelle
print("\n[B4 REF_Cloture_Mensuelle]")
if 'REF_Cloture_Mensuelle' in wb.sheetnames:
    ws_cm = wb['REF_Cloture_Mensuelle']
    headers_cm = list(ws_cm.iter_rows(values_only=True))[0]
    print(f"  OK: colonnes = {list(headers_cm)}")
else:
    print("  MANQUANT")

# 6. REF_Parametres_Generaux
print("\n[B5 REF_Parametres_Generaux]")
_, params = get_dicts(wb, 'REF_Parametres_Generaux')
param_names = [r['nom_parametre'] for r in params]
print(f"  Total: {len(params)} params")
for p in ['TAUX_HORAIRE_MENAGE_INTERNE', 'ARRONDI_DECIMALES',
          'TOLERANCE_ARRONDI_LIGNE_EUR', 'TOLERANCE_ARRONDI_CUMUL_EUR']:
    found_p = p in param_names
    if found_p:
        val = next(r['valeur'] for r in params if r['nom_parametre'] == p)
        print(f"  OK: {p} = {val}")
    else:
        print(f"  MANQUANT: {p}")

# 7. REF_Intervenants
print("\n[B6 REF_Intervenants]")
ws_int = wb['REF_Intervenants']
rows_int = list(ws_int.iter_rows(values_only=True))
headers_int = list(rows_int[0])
print(f"  Colonnes: {headers_int}")
for col in ['nom_normalise', 'date_debut_validite', 'date_fin_validite']:
    print(f"  {col}: {'OK' if col in headers_int else 'MANQUANT'}")
for row in rows_int[1:]:
    d = dict(zip(headers_int, row))
    tv = d.get('type_intervenant', '')
    nn = d.get('nom_normalise', '')
    ok_type = tv in ('INTERNE', 'EXTERNE', 'AUTRE', 'A_CONTROLER')
    print(f"  {d['intervenant_id']}: type={tv} ({'OK' if ok_type else 'ERREUR'}), nom_normalise={nn}")

# 8. REF_Logements codes hors-parc
print("\n[I1 REF_Logements codes hors-parc]")
_, logs = get_dicts(wb, 'REF_Logements')
ids = [r['logement_id'] for r in logs]
print(f"  Total logements: {len(logs)}")
for code in ['APPARTEMENT_DIVERS', 'LOGEMENT_DIVERS']:
    print(f"  {code}: {'OK' if code in ids else 'MANQUANT'}")

# 9. Clés uniques globales
print("\n[CLES UNIQUES]")
pk_map = {
    'REF_Logements': 'logement_id',
    'REF_Proprietaires': 'proprietaire_id',
    'REF_Statuts': 'statut_id',
    'REF_Statuts_Payout': 'statut_payout_id',
    'REF_Parametres_Generaux': 'parametre_id',
    'REF_Intervenants': 'intervenant_id',
    'REF_Codes_Impact': 'code_impact',
    'REF_Types_Flux': 'type_flux_id',
}
all_ok = True
for sname, pk_col in pk_map.items():
    if sname not in wb.sheetnames:
        print(f"  [{sname}] ABSENT")
        continue
    ws = wb[sname]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        continue
    headers = list(rows[0])
    if pk_col not in headers:
        print(f"  [{sname}] CLE ABSENTE: {pk_col}")
        all_ok = False
        continue
    idx = headers.index(pk_col)
    keys = [r[idx] for r in rows[1:] if r[idx] is not None]
    dupes = [k for k, cnt in Counter(keys).items() if cnt > 1]
    if dupes:
        print(f"  [{sname}] DOUBLONS: {dupes}")
        all_ok = False
    else:
        print(f"  [{sname}] OK ({len(keys)} cles uniques)")
if all_ok:
    print("  --> Aucun doublon detecte sur tous les onglets verifies")

# 10. Dates serie brutes
print("\n[DATES SERIE]")
found_dates = False
for sname in wb.sheetnames:
    ws = wb[sname]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        for j, cell in enumerate(row):
            if isinstance(cell, int) and 40000 < cell < 60000:
                print(f"  [{sname}] row{i+1} col{j+1}: {cell}")
                found_dates = True
if not found_dates:
    print("  OK: Aucune date serie brute")

print("\n" + "=" * 60)
print("FIN AUDIT POST-CORRECTION")
print("=" * 60)

wb.close()
