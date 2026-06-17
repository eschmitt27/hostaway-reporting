"""
lot6d_rapprochement_menages.py — PARTIE 1 : rapprochement du NOMBRE de ménages
================================================================================
Compare, SANS AUCUN COÛT, par mois × logement × intervenant :
  ménages Hostaway Tasks réalisés (completed)
  vs ménages externes facturés (lot6c, lignes TLM compte_comme_menage=OUI)
  vs ménages internes M04 déclarés.

Décisions : D099, D100, D104 (révisée — mapping via REF_Intervenants.hostaway_assigneeUserId).

DRY-RUN : sortie de test 02_TRAVAIL/Lot6_DryRun/DRYRUN_Rapprochement_Menages_Complet.xlsx
  (non définitive). N'écrit RIEN dans Flux / résultats. Ne touche pas M04 / banque / VRBO.

Périmètre courant : MOIS = 2026-05.
"""

import sys, os, glob, hashlib, datetime, collections, unicodedata, warnings
warnings.filterwarnings("ignore")
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import openpyxl
from openpyxl.styles import Font, PatternFill

MONTH = "2026-05"
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REF  = os.path.join(ROOT, "01_SOURCES_BRUTES", "REF_Setup", "REF_Setup.xlsm")
OUTD = os.path.join(ROOT, "02_TRAVAIL", "Lot6_DryRun")
OUT  = os.path.join(OUTD, "DRYRUN_Rapprochement_Menages_Complet.xlsx")

def sh(p, s):
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True); ws = wb[s]
    rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]; wb.close()
    h = list(rows[0]); return h, [dict(zip(h, r)) for r in rows[1:]]

def norm(s):
    s = str(s or "").strip().lower()
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")

def rowhash(*v):
    return hashlib.sha256("|".join("" if x is None else str(x) for x in v).encode()).hexdigest()[:16]

# ── Référentiels ────────────────────────────────────────────────────────────
_, ref_int = sh(REF, "REF_Intervenants")
assignee2int = {}    # assigneeUserId -> (intervenant_id, nom, type)
int_by_id = {}
for d in ref_int:
    int_by_id[d["intervenant_id"]] = d
    a = d.get("hostaway_assigneeUserId")
    if a is not None and str(d.get("hostaway_mapping_actif")).upper() == "OUI":
        assignee2int[a] = (d["intervenant_id"], d.get("nom_intervenant"), d.get("type_intervenant"))
KNOWN = {norm(d["nom_normalise"]): d["intervenant_id"] for d in ref_int if d.get("nom_normalise")}

_, ref_clo = sh(REF, "REF_Cloture_Mensuelle")
cloture = {str(d["mois"])[:7] for d in ref_clo if str(d.get("statut_mois")).upper() == "CLOTURE" and d.get("mois")}
mois_historique = MONTH in cloture          # mois clôturé => non-assignés = INFO historique

_, ref_log = sh(REF, "REF_Logements")
log_info = {d["logement_id"]: d for d in ref_log if d.get("logement_id") and str(d["logement_id"]) != "logement_id"}

# ── A. Hostaway Tasks réalisés (completed) du mois ───────────────────────────
tf = glob.glob(os.path.join(ROOT, "02_TRAVAIL", "**", "MASTER_FACT_HA_CleaningTasks_Discovery.xlsx"), recursive=True)[0]
_, t_data = sh(tf, "data")
_, t_enr  = sh(tf, "MASTER_ENRICHI")
assignee_by_task = {d["task_id"]: d.get("assigneeUserId") for d in t_data}
title_by_task    = {d["task_id"]: d.get("title") for d in t_data}

def title_names(t):
    n = norm(str(t).split(" - ")[0])
    import re
    return {k for k in KNOWN if re.search(r"\b" + re.escape(k) + r"\b", n)}

tasks = collections.Counter()       # (logement_id, intervenant_id) -> nb réalisés
controls = []
conflits = nonmap_hist = nonmap_futur = 0
for d in t_enr:
    if str(d.get("mois"))[:7] != MONTH:
        continue
    if d.get("statut_menage") != "réalisé":   # completed uniquement
        continue
    tid = d["task_id"]; a = assignee_by_task.get(tid)
    lg = d.get("logement_id")
    if a in assignee2int:
        iid, nom, typ = assignee2int[a]
        # contrôle secondaire title
        tn = title_names(title_by_task.get(tid))
        exp = norm(nom)
        others = {x for x in tn if x != exp}
        if others and exp not in tn:
            conflits += 1
            controls.append({"type": "CONFLIT_TITLE_ASSIGNEE", "niveau": "A_CONTROLER",
                "detail": f"task {tid} assignee->{iid}({nom}) title={sorted(tn)}"})
    elif a in (None, 0):
        if mois_historique:
            iid = "NON_ATTRIBUE"; nonmap_hist += 1
            controls.append({"type": "TASK_NON_ASSIGNEE_HISTORIQUE_IGNOREE", "niveau": "INFO",
                "detail": f"task {tid} logement {lg} — historique, ignorée"})
        else:
            iid = "NON_ATTRIBUE"; nonmap_futur += 1
            controls.append({"type": "TASK_FUTURE_SANS_INTERVENANT_ASSIGNE", "niveau": "A_CONTROLER",
                "detail": f"task {tid} logement {lg} — mois ouvert/futur, à assigner"})
    else:
        iid = "ASSIGNEE_NON_MAPPE"
        controls.append({"type": "MENAGE_ASSIGNEE_NON_MAPPE", "niveau": "A_CONTROLER",
            "detail": f"task {tid} assigneeUserId={a} inconnu du référentiel"})
    tasks[(lg, iid)] += 1

# ── C. Ménages externes déclarés (facturés) du mois ──────────────────────────
fc = glob.glob(os.path.join(ROOT, "02_TRAVAIL", "**", "MASTER_FACT_MEN_MenagesExternes.xlsx"), recursive=True)[0]
_, ext = sh(fc, "MASTER")
ext_cnt = collections.Counter()     # (logement_id, intervenant_id) -> nb ménages
for d in ext:
    if str(d.get("mois"))[:7] != MONTH:
        continue
    if str(d.get("type_ligne_menage_id")) not in ("TLM_001", "TLM_002"):   # compte_comme_menage=OUI
        continue
    q = d.get("nombre_menages") or 0
    m = d.get("montant_ligne_ttc") or 0
    if (q or 0) == 0 and (m or 0) == 0:   # ligne 0€/q0 exclue (INFO)
        controls.append({"type": "EXCLU_VOLUME", "niveau": "INFO",
            "detail": f"facture {d.get('facture_id')} ligne 0€/q0 exclue du comptage"})
        continue
    ext_cnt[(d.get("logement_id"), d.get("prestataire_id"))] += q

# ── D. Ménages internes déclarés du mois ─────────────────────────────────────
# Priorité à la source normalisée DRY-RUN (Google Sheet) si présente, sinon M04 réel.
# (M04 réel jamais modifié ici — lecture seule.)
int_cnt = collections.Counter()     # (logement_id, intervenant_id) -> nb ménages
m04_alimente = False
src_interne = "AUCUNE"
DRY_M04 = os.path.join(OUTD, "DRYRUN_M04_Declarations_Internes_Normalisees.xlsx")
if os.path.exists(DRY_M04):
    src_interne = "DRYRUN_SHEET"
    _, decl = sh(DRY_M04, "MASTER_NORMALISE")
    for d in decl:
        if str(d.get("mois"))[:7] != MONTH:
            continue
        m04_alimente = True
        int_cnt[(d.get("logement_id"), d.get("intervenant_id"))] += (d.get("nb_menages") or 0)
else:
    src_interne = "M04_REEL"
    m04f = glob.glob(os.path.join(ROOT, "02_DONNEES_NORMALISEES", "menages", "M04_MENAGES_PowerQuery.xlsx"))[0]
    _, m04 = sh(m04f, "MASTER")
    for d in m04:
        if str(d.get("mois"))[:7] != MONTH:
            continue
        m04_alimente = True
        int_cnt[(d.get("logement_id"), d.get("intervenant_id"))] += (d.get("nb_menages") or 0)

# ── Tableau comparaison mois × logement × intervenant ────────────────────────
keys = set(tasks) | set(ext_cnt) | set(int_cnt)
def nom_app(lg): return (log_info.get(lg) or {}).get("nom_logement_officiel")
def prop(lg):    return (log_info.get(lg) or {}).get("proprietaire_id")
def i_nom(iid):  return (int_by_id.get(iid) or {}).get("nom_intervenant") or iid
def i_typ(iid):  return (int_by_id.get(iid) or {}).get("type_intervenant")

comp = []
cnt_statut = collections.Counter()
for (lg, iid) in sorted(keys, key=lambda x: (str(x[0]), str(x[1]))):
    nb_t = tasks.get((lg, iid), 0)
    nb_e = ext_cnt.get((lg, iid), 0)
    nb_i = int_cnt.get((lg, iid), 0)
    tot_dec = nb_e + nb_i
    ecart = tot_dec - nb_t
    typ = i_typ(iid)
    statut, code, comm = "VALIDE", "", ""
    if iid == "NON_ATTRIBUE":
        if mois_historique:
            statut, code, comm = "INFO", "TASK_NON_ASSIGNEE_HISTORIQUE_IGNOREE", "Tasks non assignées (historique), ignorées pour blocage"
        else:
            statut, code, comm = "A_CONTROLER", "TASK_FUTURE_SANS_INTERVENANT_ASSIGNE", "Tasks non assignées sur mois ouvert/futur"
    elif iid == "ASSIGNEE_NON_MAPPE":
        statut, code, comm = "A_CONTROLER", "MENAGE_ASSIGNEE_NON_MAPPE", "assigneeUserId inconnu du référentiel"
    elif typ == "INTERNE" and nb_t > 0 and nb_i == 0:
        statut, code, comm = "A_CONTROLER", "MENAGE_M04_NON_ALIMENTE", "Tasks Hostaway internes mais M04 vide"
    elif nb_t > 0 and tot_dec < nb_t:
        statut, code, comm = "A_CONTROLER", "MENAGE_TOTAL_ECART_HOSTAWAY", "Tasks Hostaway > déclaré total"
    elif typ == "EXTERNE" and nb_e > nb_t:
        statut, code, comm = "A_CONTROLER", "MENAGE_PRESTATAIRE_ECART_HOSTAWAY", "Facturé externe > tasks Hostaway"
    elif ecart != 0:
        statut, code, comm = "A_CONTROLER", "MENAGE_ECART_NOMBRE", "Écart à expliquer"
    cnt_statut[statut] += 1
    comp.append({"mois": MONTH, "nom_appartement": nom_app(lg), "logement_id": lg, "proprietaire_id": prop(lg),
        "intervenant_id": iid, "nom_intervenant": i_nom(iid), "type_intervenant": typ,
        "source_mapping_hostaway": "REF_Intervenants.hostaway_assigneeUserId",
        "nb_menages_tasks_hostaway_completed": nb_t,
        "nb_menages_declares_externe": nb_e, "nb_menages_declares_interne_m04": nb_i,
        "total_menages_declares": tot_dec, "ecart": ecart,
        "statut_controle": statut, "code_controle": code, "commentaire": comm,
        "ROW_HASH": rowhash(MONTH, lg, iid, nb_t, tot_dec)})

# ── Résumés ──────────────────────────────────────────────────────────────────
def resume(keyf):
    agg = collections.defaultdict(lambda: [0, 0, 0])
    for r in comp:
        k = keyf(r); agg[k][0] += r["nb_menages_tasks_hostaway_completed"]
        agg[k][1] += r["nb_menages_declares_externe"]; agg[k][2] += r["nb_menages_declares_interne_m04"]
    return agg
res_app = resume(lambda r: (r["logement_id"], r["nom_appartement"]))
res_int = resume(lambda r: (r["intervenant_id"], r["nom_intervenant"], r["type_intervenant"]))

# ── Écriture ──────────────────────────────────────────────────────────────────
os.makedirs(OUTD, exist_ok=True)
wb = openpyxl.Workbook()
def write(ws, cols, rows):
    ws.append(cols)
    for c in ws[1]: c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="DDDDDD")
    for r in rows: ws.append([r.get(c) if isinstance(r, dict) else r[i] for i, c in enumerate(cols)])

ws1 = wb.active; ws1.title = "TABLEAU_COMPARAISON"
COLS1 = ["mois","nom_appartement","logement_id","proprietaire_id","intervenant_id","nom_intervenant",
    "type_intervenant","source_mapping_hostaway","nb_menages_tasks_hostaway_completed",
    "nb_menages_declares_externe","nb_menages_declares_interne_m04","total_menages_declares",
    "ecart","statut_controle","code_controle","commentaire"]
write(ws1, COLS1, comp)

ws2 = wb.create_sheet("RESUME_APPARTEMENT")
r2 = [{"mois":MONTH,"nom_appartement":k[1],"logement_id":k[0],"tasks_hostaway_total":v[0],
    "declares_externes_total":v[1],"declares_internes_m04_total":v[2],"total_declares":v[1]+v[2],
    "ecart_total":v[1]+v[2]-v[0],"statut_controle":"VALIDE" if v[1]+v[2]-v[0]==0 else "A_CONTROLER",
    "commentaire":""} for k,v in sorted(res_app.items())]
write(ws2, ["mois","nom_appartement","logement_id","tasks_hostaway_total","declares_externes_total",
    "declares_internes_m04_total","total_declares","ecart_total","statut_controle","commentaire"], r2)

ws3 = wb.create_sheet("RESUME_INTERVENANT")
r3 = [{"mois":MONTH,"intervenant_id":k[0],"nom_intervenant":k[1],"type_intervenant":k[2],
    "tasks_hostaway_total":v[0],"declares_externes_total":v[1],"declares_internes_m04_total":v[2],
    "total_declares":v[1]+v[2],"ecart_total":v[1]+v[2]-v[0],
    "statut_controle":"VALIDE" if v[1]+v[2]-v[0]==0 else "A_CONTROLER","commentaire":""} for k,v in sorted(res_int.items())]
write(ws3, ["mois","intervenant_id","nom_intervenant","type_intervenant","tasks_hostaway_total",
    "declares_externes_total","declares_internes_m04_total","total_declares","ecart_total","statut_controle","commentaire"], r3)

ws4 = wb.create_sheet("CONTROLES")
cc = collections.Counter(c["type"] for c in controls)
write(ws4, ["code_controle","niveau","nb","exemple"],
    [{"code_controle":k,"niveau":next(c["niveau"] for c in controls if c["type"]==k),"nb":n,
      "exemple":next(c["detail"] for c in controls if c["type"]==k)} for k,n in cc.most_common()])
try:
    wb.save(OUT)
except PermissionError:
    OUT = OUT.replace(".xlsx", "_MAJ.xlsx")
    wb.save(OUT)
    print(f"[lot6d] AVERTISSEMENT : sortie originale verrouillée (Excel ouvert) -> écrit dans {os.path.basename(OUT)}")

# ── Rapport console ──────────────────────────────────────────────────────────
print(f"[lot6d] DRY-RUN mois={MONTH} (clôturé={mois_historique}) -> {OUT}")
print(f"  TABLEAU_COMPARAISON : {len(comp)} lignes")
print(f"  statuts : {dict(cnt_statut)}")
print(f"  interne mai : alimenté={m04_alimente} source={src_interne}")
print(f"  conflits title : {conflits} | non assignés histo : {nonmap_hist} | futur : {nonmap_futur}")
print("\n  RESUME_APPARTEMENT:")
for r in r2: print(f"    {r['logement_id']:8} {str(r['nom_appartement'])[:26]:26} tasks={r['tasks_hostaway_total']:3} ext={r['declares_externes_total']:3} m04={r['declares_internes_m04_total']:3} ecart={r['ecart_total']:4} {r['statut_controle']}")
print("\n  RESUME_INTERVENANT:")
for r in r3: print(f"    {str(r['intervenant_id']):16} {str(r['nom_intervenant'])[:10]:10} {str(r['type_intervenant']):8} tasks={r['tasks_hostaway_total']:3} ext={r['declares_externes_total']:3} m04={r['declares_internes_m04_total']:3} ecart={r['ecart_total']:4} {r['statut_controle']}")
print("\n  CONTROLES:")
for k,n in cc.most_common(): print(f"    {k}: {n}")
"""END"""
