#!/usr/bin/env python3
"""
lot12_remove_donnees_fictives.py
Suppression de TOUTES les données fictives de test Lot 12.

Une ligne est fictive si :
  - son ID (1re colonne clé) commence par  ZZ_TEST_
  - OU une de ses cellules contient le tag  __TEST_FICTIF_LOT12_A_SUPPRIMER__

Sécurité :
  - Sauvegarde de chaque fichier modifié dans 99_ARCHIVES/LOT12_TEST_DATA/ avant écriture.
  - Affiche le nombre de lignes supprimées par fichier.
  - Vérifie qu'il ne reste AUCUNE donnée fictive après nettoyage.

Dépendances : pip install openpyxl
"""

import os
import shutil
import datetime
import openpyxl

TAG = "__TEST_FICTIF_LOT12_A_SUPPRIMER__"
ID_PREFIX = "ZZ_TEST_"

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BACKUP_DIR = os.path.join(ROOT, "99_ARCHIVES", "LOT12_TEST_DATA")

# (fichier, onglet, colonne_id)
TARGETS = [
    ("Charges",  "02_TRAVAIL/Lot3_Charges/MASTER_FACT_MAN_Charges.xlsx",                         "MASTER", "charge_id"),
    ("M04",      "02_DONNEES_NORMALISEES/menages/M04_MENAGES_PowerQuery.xlsx",                    "MASTER", "menage_calc_id"),
    ("HH",       "02_TRAVAIL/Lot4_ReservationsHH/MASTER_FACT_MAN_ReservationsHorsHostaway.xlsx",  "MASTER", "reservation_hh_id"),
    ("MenExt",   "02_TRAVAIL/Lot6c_MenagesExternes/MASTER_FACT_MEN_MenagesExternes.xlsx",         "MASTER", "menage_externe_id"),
    ("Acomptes", "02_TRAVAIL/Lot5_AcomptesProprietaires/MASTER_FACT_MAN_AcomptesProprietaires.xlsx", "MASTER", "acompte_id"),
]


def backup(path):
    os.makedirs(BACKUP_DIR, exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    base = os.path.basename(path)
    dst = os.path.join(BACKUP_DIR, f"{base}.PRE_REMOVE_{ts}")
    shutil.copy2(path, dst)
    return dst


def is_fictif(row_values, id_idx):
    v = row_values[id_idx]
    if v is not None and str(v).startswith(ID_PREFIX):
        return True
    if any(c is not None and TAG in str(c) for c in row_values):
        return True
    return False


def remove_file(name, relpath, sheet_name, id_col):
    path = os.path.join(ROOT, relpath)
    if not os.path.exists(path):
        print(f"  [{name}] ABSENT — ignoré")
        return 0
    wb = openpyxl.load_workbook(path)
    if sheet_name not in wb.sheetnames:
        print(f"  [{name}] onglet {sheet_name} absent — ignoré")
        wb.close()
        return 0
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    if id_col not in headers:
        print(f"  [{name}] colonne {id_col} absente — ignoré")
        wb.close()
        return 0
    id_idx = headers.index(id_col)

    # Repérer lignes fictives (indices Excel, en partant du bas pour delete sûr)
    to_delete = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if is_fictif(row, id_idx):
            to_delete.append(r_idx)

    if not to_delete:
        print(f"  [{name}] 0 ligne fictive")
        wb.close()
        return 0

    bkp = backup(path)
    for r_idx in sorted(to_delete, reverse=True):
        ws.delete_rows(r_idx, 1)
    wb.save(path)
    print(f"  [{name}] -{len(to_delete)} lignes fictives supprimées — backup: {os.path.basename(bkp)}")
    wb.close()
    return len(to_delete)


def verify_clean(name, relpath, sheet_name, id_col):
    path = os.path.join(ROOT, relpath)
    if not os.path.exists(path):
        return True
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    headers = [c for c in next(ws.iter_rows(values_only=True))]
    id_idx = headers.index(id_col)
    remaining = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if is_fictif(row, id_idx):
            remaining += 1
    wb.close()
    return remaining == 0


def main():
    print("=" * 60)
    print("SUPPRESSION DONNÉES FICTIVES LOT 12")
    print("=" * 60)
    print(f"Critères : ID commence par {ID_PREFIX}  OU  cellule contient {TAG}")
    print()

    total = 0
    for name, relpath, sheet, id_col in TARGETS:
        total += remove_file(name, relpath, sheet, id_col)

    print()
    print("Vérification post-nettoyage (0 fictif attendu) :")
    all_clean = True
    for name, relpath, sheet, id_col in TARGETS:
        clean = verify_clean(name, relpath, sheet, id_col)
        print(f"  [{name}] {'PROPRE' if clean else '!! FICTIF RESTANT'}")
        all_clean = all_clean and clean

    print()
    print(f"Total lignes fictives supprimées : {total}")
    print(f"État final : {'PROPRE — aucune donnée fictive' if all_clean else 'ÉCHEC — fictif restant'}")


if __name__ == "__main__":
    main()
