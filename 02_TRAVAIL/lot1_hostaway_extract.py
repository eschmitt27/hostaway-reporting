#!/usr/bin/env python3
"""
lot1_hostaway_extract.py — Module Hostaway — Extraction Lot 1

Usage:
  python lot1_hostaway_extract.py --check-auth   # Test auth uniquement (aucun secret affiché)
  python lot1_hostaway_extract.py --dry-run      # Auth + comptage, aucune écriture
  python lot1_hostaway_extract.py                # Extraction complète depuis 2026-01-01

Sorties dans 02_TRAVAIL/Lot1_Hostaway/ :
  MASTER_REF_HA_Listings.xlsx
  MASTER_FACT_HA_Reservations.xlsx
  MASTER_FACT_HA_ReservationDetails.xlsx
  MASTER_FACT_HA_ReservationFinanceFields.xlsx
  MASTER_FACT_HA_ReservationFees.xlsx
  MASTER_CALC_HA_Payout.xlsx
  MASTER_CTRL_HA_Anomalies.xlsx
  MASTER_FACT_HA_CleaningTasks_Discovery.xlsx
  MASTER_RUN_Log.xlsx

Sécurité : CLIENT_SECRET, access_token, Authorization header jamais loggés ni affichés.
"""

import os
import sys
import re
import json
import time
import hashlib
import logging
import argparse
from datetime import datetime, timezone
from pathlib import Path

# ── Vérification dépendances ─────────────────────────────────
_missing = []
try:
    import requests
except ImportError:
    _missing.append("requests")
try:
    import pandas as pd
except ImportError:
    _missing.append("pandas")
try:
    import openpyxl
except ImportError:
    _missing.append("openpyxl")
try:
    from dotenv import load_dotenv
except ImportError:
    _missing.append("python-dotenv")

if _missing:
    print(f"[ERREUR] Dépendances manquantes : {', '.join(_missing)}")
    print(f"         Installer : pip install {' '.join(_missing)}")
    sys.exit(1)

# ── Constantes ────────────────────────────────────────────────
DATE_FROM       = "2026-01-01"
PAGE_SIZE       = 100
MAX_RETRIES     = 3
RETRY_BASE_WAIT = 2   # secondes (exponentiel)

DETAILS_MINIMAL = "minimal"   # détail uniquement si payout impossible depuis liste
DETAILS_FULL    = "full"      # détail systématique (comportement original)

# Statuts Hostaway (§6.4)
STATUS_ACTIVE   = {"new", "modified"}
STATUS_OWNER    = {"ownerStay"}
STATUS_CANCEL   = {"cancelled"}
STATUS_SKIP     = {
    "inquiry", "declined", "expired",
    "inquiryPreapproved", "inquiryNotPossible",
    "closedByHost", "unavailable",
}

# Canaux (pour logique payout H1/H2/H3)
CH_AIRBNB  = {"airbnbOfficial", "airbnb2", "airbnb"}
CH_BOOKING = {"bookingcom", "booking.com"}
CH_VRBO    = {"vrboical", "vrbo", "homeaway"}
CH_DIRECT  = {"direct", "manuel", "owner"}

# ── Chemins ───────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent.parent
ENV_FILE = BASE_DIR / ".env"
OUT_DIR  = BASE_DIR / "02_TRAVAIL" / "Lot1_Hostaway"
LOG_DIR  = BASE_DIR / "04_LOGS"
REF_FILE = BASE_DIR / "01_SOURCES_BRUTES" / "REF_Setup" / "REF_Setup.xlsm"


# ═══════════════════════════════════════════════════════════════
# LOGGING SÉCURISÉ
# ═══════════════════════════════════════════════════════════════
class _SecretFilter(logging.Filter):
    """Bloque les lignes contenant des valeurs de secrets (pas les noms de variables)."""
    _PAT = re.compile(
        r"(access_token\s*[:=]\s*\S|bearer\s+\S{8,}|authorization\s*:\s*bearer)",
        re.IGNORECASE,
    )
    def filter(self, record):
        return not self._PAT.search(record.getMessage())


def setup_logging(run_id: str, silent_file: bool = False) -> logging.Logger:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    fmt = "%(asctime)s [%(levelname)s] %(message)s"
    handlers: list[logging.Handler] = [logging.StreamHandler(sys.stdout)]
    if not silent_file:
        fh = logging.FileHandler(LOG_DIR / f"lot1_ha_{run_id}.log", encoding="utf-8")
        fh.setFormatter(logging.Formatter(fmt))
        handlers.append(fh)
    logging.basicConfig(level=logging.INFO, format=fmt, handlers=handlers, force=True)
    log = logging.getLogger("lot1_ha")
    for h in log.handlers:
        h.addFilter(_SecretFilter())
    log.addFilter(_SecretFilter())
    return log


# ═══════════════════════════════════════════════════════════════
# CHECK-LISTINGS — test léger endpoint /v1/listings
# ═══════════════════════════════════════════════════════════════
def check_listings_mode(env_file: Path) -> None:
    """
    Mode --check-listings : auth + appel /v1/listings uniquement.
    Affiche : HTTP OK, nombre de listings, champs disponibles, 3 premiers IDs.
    Aucun token ni header Authorization affiché. sys.exit(0|1|2).
    """
    print("[LISTINGS] Démarrage vérification endpoint listings")

    if not env_file.exists():
        print(f"[LISTINGS] ÉCHEC — .env introuvable : {env_file}")
        sys.exit(1)

    load_dotenv(env_file)
    client_id     = os.getenv("HOSTAWAY_CLIENT_ID", "")
    client_secret = os.getenv("HOSTAWAY_CLIENT_SECRET", "")
    account_id    = os.getenv("HOSTAWAY_ACCOUNT_ID", "")
    base_url      = os.getenv("HOSTAWAY_BASE_URL", "https://api.hostaway.com").rstrip("/")

    missing = [k for k, v in [
        ("HOSTAWAY_CLIENT_ID", client_id),
        ("HOSTAWAY_CLIENT_SECRET", client_secret),
        ("HOSTAWAY_ACCOUNT_ID", account_id),
    ] if not v]
    if missing:
        print(f"[LISTINGS] ÉCHEC — variables manquantes : {', '.join(missing)}")
        sys.exit(1)

    # Auth
    print("[LISTINGS] Authentification OAuth2...")
    try:
        resp_auth = requests.post(
            f"{base_url}/v1/accessTokens",
            data={
                "grant_type": "client_credentials",
                "client_id": client_id,
                "client_secret": client_secret,
                "scope": "general",
            },
            timeout=15,
        )
        resp_auth.raise_for_status()
        token = resp_auth.json().get("access_token", "")
        if not token:
            print("[LISTINGS] ÉCHEC — token absent de la réponse auth")
            sys.exit(1)
    except requests.exceptions.ConnectionError:
        print(f"[LISTINGS] ÉCHEC — erreur réseau vers {base_url}")
        sys.exit(1)
    except requests.exceptions.Timeout:
        print("[LISTINGS] ÉCHEC — timeout auth (15s)")
        sys.exit(1)
    except Exception as e:
        print(f"[LISTINGS] ÉCHEC — auth : {type(e).__name__}: {e}")
        sys.exit(1)

    print("[LISTINGS] Auth OK (token non affiché)")

    # Appel /v1/listings
    listings_url = f"{base_url}/v1/listings"
    print(f"[LISTINGS] GET {listings_url} ...")
    try:
        resp = requests.get(
            listings_url,
            headers={
                "Authorization": f"Bearer {token}",
                "account-id": account_id,
                "Content-Type": "application/json",
            },
            params={"limit": 200, "includeResources": 0},
            timeout=30,
        )
    except requests.exceptions.ConnectionError:
        print(f"[LISTINGS] ÉCHEC — erreur réseau sur {listings_url}")
        sys.exit(1)
    except requests.exceptions.Timeout:
        print("[LISTINGS] ÉCHEC — timeout (30s) sur /v1/listings")
        sys.exit(1)
    except Exception as e:
        print(f"[LISTINGS] ÉCHEC — {type(e).__name__}: {e}")
        sys.exit(2)

    print(f"[LISTINGS] HTTP {resp.status_code}")
    if resp.status_code == 404:
        print(f"[LISTINGS] ÉCHEC — endpoint introuvable (HTTP 404) : {listings_url}")
        sys.exit(1)
    if resp.status_code == 401:
        print("[LISTINGS] ÉCHEC — token rejeté (HTTP 401)")
        sys.exit(1)
    if resp.status_code >= 400:
        print(f"[LISTINGS] ÉCHEC — erreur HTTP {resp.status_code}")
        sys.exit(1)

    try:
        body = resp.json()
    except Exception:
        print("[LISTINGS] ÉCHEC — réponse JSON invalide")
        sys.exit(1)

    results = body.get("result", [])
    count   = body.get("count", len(results))
    print(f"[LISTINGS] HTTP OK")
    print(f"[LISTINGS] Nombre de listings : {count} (reçus dans cette page : {len(results)})")

    if not results:
        print("[LISTINGS] Aucun listing retourné — compte vide ou paramètres incorrects")
        sys.exit(1)

    # Champs disponibles (1er objet)
    fields = list(results[0].keys())
    print(f"[LISTINGS] Champs disponibles ({len(fields)}) : {fields}")

    # IDs non sensibles — ce sont des entiers internes Hostaway
    ids_sample = [r.get("id") for r in results[:3]]
    names_sample = [r.get("name", r.get("internalListingName", "?"))[:40] for r in results[:3]]
    print(f"[LISTINGS] Premiers IDs (listing.id = listingMapId réservations) : {ids_sample}")
    print(f"[LISTINGS] Noms correspondants : {names_sample}")

    # Vérifier si listingMapId est un champ distinct
    has_map_field = any("listingMapId" in r for r in results[:5])
    print(f"[LISTINGS] Champ listingMapId présent dans listings : {'OUI' if has_map_field else 'NON — utiliser id'}")

    print("[LISTINGS] OK — endpoint /v1/listings opérationnel")
    sys.exit(0)


# ═══════════════════════════════════════════════════════════════
# CHECK-RESERVATIONS-SAMPLE — diagnostic léger sur 5 réservations
# ═══════════════════════════════════════════════════════════════
def check_reservations_sample_mode(env_file: Path) -> None:
    """
    Mode --check-reservations-sample : auth + 5 premières réservations.
    Affiche : clés disponibles, champs canal présents, nb avec canal vide.
    Aucune donnée personnelle ni secret affiché. sys.exit(0|1|2).
    """
    print("[SAMPLE] Diagnostic réservations — 5 premières depuis 2026-01-01")

    if not env_file.exists():
        print(f"[SAMPLE] ÉCHEC — .env introuvable : {env_file}")
        sys.exit(1)

    load_dotenv(env_file)
    client_id     = os.getenv("HOSTAWAY_CLIENT_ID", "")
    client_secret = os.getenv("HOSTAWAY_CLIENT_SECRET", "")
    account_id    = os.getenv("HOSTAWAY_ACCOUNT_ID", "")
    base_url      = os.getenv("HOSTAWAY_BASE_URL", "https://api.hostaway.com").rstrip("/")

    missing = [k for k, v in [
        ("HOSTAWAY_CLIENT_ID", client_id),
        ("HOSTAWAY_CLIENT_SECRET", client_secret),
        ("HOSTAWAY_ACCOUNT_ID", account_id),
    ] if not v]
    if missing:
        print(f"[SAMPLE] ÉCHEC — variables manquantes : {', '.join(missing)}")
        sys.exit(1)

    # Auth
    try:
        resp_auth = requests.post(
            f"{base_url}/v1/accessTokens",
            data={"grant_type": "client_credentials", "client_id": client_id,
                  "client_secret": client_secret, "scope": "general"},
            timeout=15,
        )
        resp_auth.raise_for_status()
        token = resp_auth.json().get("access_token", "")
        if not token:
            print("[SAMPLE] ÉCHEC — token absent")
            sys.exit(1)
    except Exception as e:
        print(f"[SAMPLE] ÉCHEC — auth : {type(e).__name__}")
        sys.exit(1)

    print("[SAMPLE] Auth OK")

    # Fetch 5 réservations
    headers = {
        "Authorization": f"Bearer {token}",
        "account-id": account_id,
        "Content-Type": "application/json",
    }
    try:
        resp = requests.get(
            f"{base_url}/v1/reservations",
            headers=headers,
            params={"dateFrom": DATE_FROM, "limit": 5, "offset": 0},
            timeout=30,
        )
    except Exception as e:
        print(f"[SAMPLE] ÉCHEC — requête réservations : {type(e).__name__}")
        sys.exit(1)

    print(f"[SAMPLE] HTTP {resp.status_code}")
    if resp.status_code != 200:
        print(f"[SAMPLE] ÉCHEC — HTTP {resp.status_code} sur /v1/reservations")
        sys.exit(1)

    try:
        body = resp.json()
    except Exception:
        print("[SAMPLE] ÉCHEC — JSON invalide")
        sys.exit(1)

    results = body.get("result", [])
    total   = body.get("count", "?")
    print(f"[SAMPLE] Total réservations (API count) : {total}")
    print(f"[SAMPLE] Réservations dans ce lot      : {len(results)}")

    if not results:
        print("[SAMPLE] Aucune réservation retournée")
        sys.exit(1)

    # Clés disponibles
    all_keys = set()
    for r in results:
        all_keys.update(r.keys())
    print(f"[SAMPLE] Clés disponibles ({len(all_keys)}) : {sorted(all_keys)}")

    # Champs canal potentiels présents
    canal_fields_present = [f for f in _CHANNEL_FIELDS if f in all_keys]
    canal_fields_absent  = [f for f in _CHANNEL_FIELDS if f not in all_keys]
    print(f"[SAMPLE] Champs canal présents  : {canal_fields_present}")
    print(f"[SAMPLE] Champs canal absents   : {canal_fields_absent}")

    # Canal résolu par réservation (sans afficher données perso)
    nb_unknown = 0
    for i, r in enumerate(results):
        channel, source_brute = resolve_channel(r)
        res_id = r.get("id", "?")
        status = r.get("status", "?")
        if channel == "UNKNOWN":
            nb_unknown += 1
            print(f"[SAMPLE]   resa #{i+1} id={res_id} status={status} -> canal=UNKNOWN (champs vides)")
        else:
            print(f"[SAMPLE]   resa #{i+1} id={res_id} status={status} -> canal={channel} (source='{source_brute}')")

    print(f"[SAMPLE] Réservations avec canal UNKNOWN dans ce lot : {nb_unknown}/{len(results)}")

    if nb_unknown == len(results):
        print("[SAMPLE] ATTENTION : tous les canaux sont UNKNOWN — vérifier les champs Hostaway")
    elif nb_unknown > 0:
        print("[SAMPLE] Certains canaux absents — anomalie CHANNEL_ABSENT sera créée")
    else:
        print("[SAMPLE] Tous les canaux résolus — OK")

    print("[SAMPLE] OK — diagnostic réservations terminé")
    sys.exit(0)


# ═══════════════════════════════════════════════════════════════
# CHECK-AUTH — diagnostic complet, print() uniquement, jamais log
# ═══════════════════════════════════════════════════════════════
def check_auth_mode(env_file: Path) -> None:
    """
    Mode --check-auth : test OAuth2 avec diagnostic non-sensible.
    Toujours sys.exit(0|1|2). Ne retourne jamais.
    Utilise print() uniquement — bypasse le logging et son filtre.
    """
    print("[AUTH] Démarrage vérification authentification Hostaway")
    print(f"[AUTH] Fichier .env : {env_file}")

    # 1. Existence .env
    if not env_file.exists():
        print(f"[AUTH] ÉCHEC — .env introuvable : {env_file}")
        print("[AUTH] Créer le fichier .env à la racine du projet avec :")
        print("         HOSTAWAY_CLIENT_ID=...")
        print("         HOSTAWAY_CLIENT_SECRET=...")
        print("         HOSTAWAY_ACCOUNT_ID=...")
        print("         HOSTAWAY_BASE_URL=https://api.hostaway.com")
        sys.exit(1)

    load_dotenv(env_file)
    client_id     = os.getenv("HOSTAWAY_CLIENT_ID", "")
    client_secret = os.getenv("HOSTAWAY_CLIENT_SECRET", "")
    account_id    = os.getenv("HOSTAWAY_ACCOUNT_ID", "")
    base_url      = os.getenv("HOSTAWAY_BASE_URL", "https://api.hostaway.com")

    # 2. Variables manquantes
    missing = [k for k, v in [
        ("HOSTAWAY_CLIENT_ID",     client_id),
        ("HOSTAWAY_CLIENT_SECRET", client_secret),
        ("HOSTAWAY_ACCOUNT_ID",    account_id),
    ] if not v]
    if missing:
        print(f"[AUTH] ÉCHEC — variable(s) manquante(s) dans .env : {', '.join(missing)}")
        sys.exit(1)

    # Affichage masqué total — jamais de valeur réelle
    print("[AUTH] Variables .env : CLIENT_ID=*** CLIENT_SECRET=*** ACCOUNT_ID=*** — toutes présentes")
    print(f"[AUTH] BASE_URL      : {base_url}")

    # 3. Tentative OAuth2
    token_url = f"{base_url.rstrip('/')}/v1/accessTokens"
    print(f"[AUTH] Token URL     : {token_url}")
    print("[AUTH] Envoi requête OAuth2 (timeout=15s)...")

    try:
        resp = requests.post(
            token_url,
            data={
                "grant_type":    "client_credentials",
                "client_id":     client_id,
                "client_secret": client_secret,
                "scope":         "general",
            },
            timeout=15,
        )
    except requests.exceptions.ConnectionError:
        print(f"[AUTH] ÉCHEC — erreur réseau : impossible de joindre {base_url}")
        print("[AUTH] Vérifier : connexion internet, BASE_URL dans .env, pare-feu")
        sys.exit(1)
    except requests.exceptions.Timeout:
        print(f"[AUTH] ÉCHEC — timeout réseau (15s) — {token_url} ne répond pas")
        sys.exit(1)
    except requests.exceptions.RequestException as e:
        print(f"[AUTH] ÉCHEC — erreur réseau inattendue : {type(e).__name__}")
        sys.exit(1)
    except Exception as e:
        print(f"[AUTH] ÉCHEC — erreur technique inattendue : {type(e).__name__}: {e}")
        sys.exit(2)

    print(f"[AUTH] HTTP {resp.status_code}")

    # 4. Diagnostic HTTP
    if resp.status_code == 401:
        print("[AUTH] ÉCHEC — credentials invalides (HTTP 401)")
        print("[AUTH] Vérifier HOSTAWAY_CLIENT_ID et HOSTAWAY_CLIENT_SECRET dans .env")
        sys.exit(1)
    if resp.status_code == 403:
        print("[AUTH] ÉCHEC — accès refusé (HTTP 403) — compte ou scope incorrect")
        sys.exit(1)
    if resp.status_code == 404:
        print(f"[AUTH] ÉCHEC — endpoint incorrect (HTTP 404) : {token_url}")
        print("[AUTH] Vérifier HOSTAWAY_BASE_URL dans .env")
        sys.exit(1)
    if resp.status_code >= 500:
        print(f"[AUTH] ÉCHEC — erreur serveur Hostaway (HTTP {resp.status_code})")
        print("[AUTH] Réessayer dans quelques minutes")
        sys.exit(1)
    if resp.status_code >= 400:
        body_safe = resp.text[:300] if resp.text else "(corps vide)"
        print(f"[AUTH] ÉCHEC — erreur HTTP {resp.status_code}")
        print(f"[AUTH] Réponse (tronquée, sans secrets) : {body_safe[:200]}")
        sys.exit(1)

    # 5. Parsing JSON
    try:
        body = resp.json()
    except Exception:
        print("[AUTH] ÉCHEC — réponse JSON invalide malgré HTTP 200")
        sys.exit(1)

    if "access_token" not in body:
        print("[AUTH] ÉCHEC — champ 'access_token' absent de la réponse JSON")
        print(f"[AUTH] Clés reçues : {list(body.keys())}")
        sys.exit(1)

    expires_in = body.get("expires_in", "?")
    print(f"[AUTH] Token reçu (non affiché) — expire dans {expires_in}s")
    print("[AUTH] OK — authentification réussie, credentials valides")
    sys.exit(0)


# ═══════════════════════════════════════════════════════════════
# AUTHENTIFICATION OAuth2 — secrets jamais loggés
# ═══════════════════════════════════════════════════════════════
class HostawayAuth:
    def __init__(self, base_url: str, client_id: str, client_secret: str):
        self._base   = base_url.rstrip("/")
        self._cid    = client_id
        self._csec   = client_secret   # JAMAIS loggé
        self._token  = None
        self._exp    = 0.0

    def get_token(self) -> str:
        if self._token and time.time() < self._exp - 60:
            return self._token
        self._refresh()
        return self._token

    def _refresh(self):
        resp = requests.post(
            f"{self._base}/v1/accessTokens",
            data={
                "grant_type":    "client_credentials",
                "client_id":     self._cid,
                "client_secret": self._csec,
                "scope":         "general",
            },
            timeout=30,
        )
        resp.raise_for_status()
        body         = resp.json()
        self._token  = body["access_token"]       # jamais affiché
        self._exp    = time.time() + int(body.get("expires_in", 3600))

    def test(self) -> bool:
        try:
            self._refresh()
            return bool(self._token)
        except Exception:
            return False


# ═══════════════════════════════════════════════════════════════
# CLIENT API
# ═══════════════════════════════════════════════════════════════
class HostawayClient:
    def __init__(self, auth: HostawayAuth, account_id: str, log):
        self._auth  = auth
        self._aid   = str(account_id)
        self._base  = auth._base
        self._log   = log

    def _headers(self) -> dict:
        # Authorization header jamais loggé
        return {
            "Authorization": f"Bearer {self._auth.get_token()}",
            "account-id":    self._aid,
            "Content-Type":  "application/json",
        }

    def _get(self, path: str, params: dict = None) -> dict:
        url = f"{self._base}{path}"
        for attempt in range(1, MAX_RETRIES + 1):
            try:
                r = requests.get(url, headers=self._headers(), params=params, timeout=60)
                if r.status_code == 429:
                    wait = RETRY_BASE_WAIT * (2 ** attempt)
                    self._log.warning(f"Rate limit 429 — attente {wait}s (tentative {attempt})")
                    time.sleep(wait)
                    continue
                # Erreurs HTTP : message propre, sans traceback brut
                if r.status_code == 404:
                    raise RuntimeError(
                        f"Endpoint introuvable (HTTP 404) : {path}\n"
                        f"  URL complète : {url}\n"
                        f"  Vérifier que cet endpoint existe dans l'API Hostaway."
                    )
                if r.status_code == 401:
                    raise RuntimeError(f"Token expiré ou invalide (HTTP 401) : {path}")
                if r.status_code >= 400:
                    raise RuntimeError(f"Erreur HTTP {r.status_code} sur {path}")
                r.raise_for_status()
                return r.json()
            except RuntimeError:
                raise  # propager sans retry
            except requests.exceptions.ConnectionError:
                if attempt == MAX_RETRIES:
                    raise RuntimeError(f"Erreur réseau : impossible de joindre {self._base}")
                wait = RETRY_BASE_WAIT * attempt
                self._log.warning(f"Erreur réseau (t.{attempt}/{MAX_RETRIES}) — retry {wait}s")
                time.sleep(wait)
            except requests.exceptions.Timeout:
                if attempt == MAX_RETRIES:
                    raise RuntimeError(f"Timeout (60s) sur {path}")
                wait = RETRY_BASE_WAIT * attempt
                self._log.warning(f"Timeout (t.{attempt}/{MAX_RETRIES}) — retry {wait}s")
                time.sleep(wait)
            except requests.RequestException as exc:
                if attempt == MAX_RETRIES:
                    raise RuntimeError(f"Erreur requête sur {path} : {type(exc).__name__}")
                wait = RETRY_BASE_WAIT * attempt
                self._log.warning(f"Erreur requête (t.{attempt}/{MAX_RETRIES}) : {exc} — retry {wait}s")
                time.sleep(wait)
        return {}

    # ── Endpoints ────────────────────────────────────────────

    def get_listings(self) -> list:
        """
        Retourne la liste des listings via GET /v1/listings.
        Chaque listing : id = identifiant interne Hostaway = ce que les réservations
        appellent listingMapId. Pas de /v1/listingMaps (endpoint inexistant).
        """
        data = self._get("/v1/listings", {"limit": 200, "includeResources": 0})
        return data.get("result", [])

    def count_reservations(self, date_from: str) -> int:
        data = self._get("/v1/reservations", {
            "dateFrom": date_from, "limit": 1, "offset": 0,
        })
        return data.get("count", 0)

    def get_reservations_page(self, date_from: str, offset: int) -> list:
        data = self._get("/v1/reservations", {
            "dateFrom":       date_from,
            "limit":          PAGE_SIZE,
            "offset":         offset,
        })
        return data.get("result", [])

    def get_reservation_detail(self, res_id: int) -> dict:
        data = self._get(f"/v1/reservations/{res_id}")
        return data.get("result", {})

    def get_tasks(self, date_from: str) -> list:
        results, offset = [], 0
        while True:
            data  = self._get("/v1/tasks", {
                "dateFrom": date_from, "limit": PAGE_SIZE, "offset": offset,
            })
            batch = data.get("result", [])
            results.extend(batch)
            if len(batch) < PAGE_SIZE:
                break
            offset += PAGE_SIZE
            time.sleep(0.2)
        return results


# ═══════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════
def row_hash(*values) -> str:
    raw = "|".join(str(v) for v in values).encode()
    return hashlib.md5(raw).hexdigest()[:16]

def safe_float(v, default: float = 0.0) -> float:
    try:
        return float(v) if v is not None else default
    except (TypeError, ValueError):
        return default

def now_utc() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")

def _first(*values):
    """Retourne la première valeur non-None non-zéro."""
    for v in values:
        if v is not None and v != "" and v != 0 and v != 0.0:
            return v
    return None

def parse_finance_fields(money: dict) -> list:
    """Extraire la liste [{name, value}] depuis money.financeFields ou money.currencies."""
    if not money:
        return []
    ff = money.get("financeFields") or money.get("finance_fields", [])
    if ff:
        return ff
    # Fallback : money.currencies[0].fields
    currencies = money.get("currencies", [])
    if currencies and isinstance(currencies[0], dict):
        return currencies[0].get("fields", [])
    return []

def ff_map(ff_list: list) -> dict:
    """Dict {name: float_value} depuis la liste finance fields."""
    return {
        f.get("name", ""): safe_float(f.get("value"))
        for f in ff_list
        if f.get("name")
    }

# Champs potentiels de canal dans la réponse Hostaway (ordre de priorité)
# source + channelName confirmés présents ; originalChannel présent aussi
_CHANNEL_FIELDS = (
    "source", "channelName", "originalChannel",
    "channel", "origin", "reservationChannel", "channelSource",
)

def detect_channel(source) -> str:
    """Normalise une valeur de canal. Robuste si None, vide ou non-string."""
    s = str(source or "").lower().strip()
    if not s:
        return "UNKNOWN"
    if s in {c.lower() for c in CH_AIRBNB}:
        return "AIRBNB"
    if s in {c.lower() for c in CH_BOOKING}:
        return "BOOKING"
    if s in {c.lower() for c in CH_VRBO}:
        return "VRBO"
    return "DIRECT"

def resolve_channel(res: dict) -> tuple:
    """
    Retourne (channel_normalise, source_brute) en essayant plusieurs champs.
    Si aucun champ ne contient de valeur → ("UNKNOWN", "").
    """
    for field in _CHANNEL_FIELDS:
        val = res.get(field)
        if val and str(val).strip():
            return detect_channel(val), str(val).strip()
    return "UNKNOWN", ""

def source_financiere(channel: str) -> str:
    return {
        "AIRBNB":  "HOSTAWAY_AIRBNB",
        "BOOKING": "HOSTAWAY_BOOKING",
        "VRBO":    "A_CONTROLER",
        "DIRECT":  "MANUEL_HORS_HOSTAWAY",
    }.get(channel, "INCONNU")


def _ff_from_res(res: dict) -> list:
    """
    Extrait la liste [{name, value}] depuis le champ financeField de la réponse liste.
    financeField peut être un dict {name: value} ou une liste [{name, value}].
    Retourne toujours une liste normalisée.
    """
    raw = res.get("financeField")
    if not raw:
        return []
    if isinstance(raw, list):
        return raw  # déjà au bon format
    if isinstance(raw, dict):
        return [{"name": k, "value": v} for k, v in raw.items()]
    return []

def _needs_detail(res: dict, channel: str, ff_list: list) -> bool:
    """
    Retourne True si l'appel détail est nécessaire pour calculer le payout.
    En mode minimal : évite le détail quand les données liste sont suffisantes.
    - AIRBNB : suffisant si airbnbExpectedPayoutAmount ou airbnbPayoutSum disponible
    - BOOKING : suffisant si totalPriceFromChannel dans ff_list
    - VRBO Unknown : détail inutile (payout impossible de toute façon)
    - Cancelled : suffisant si cancellationAmount déjà dans la liste
    - ownerStay : pas de payout, détail inutile
    """
    status = str(res.get("status") or "")
    if status in STATUS_OWNER:
        return False
    if status in STATUS_CANCEL:
        cancel = safe_float(_first(res.get("cancellationAmount"), res.get("cancellationPayout")))
        return cancel <= 0  # si montant présent dans liste → pas de détail
    if channel == "AIRBNB":
        return not (safe_float(res.get("airbnbExpectedPayoutAmount"))
                    or ff_map(ff_list).get("airbnbPayoutSum"))
    if channel == "BOOKING":
        ffd = ff_map(ff_list)
        return not ffd.get("totalPriceFromChannel")
    if channel == "VRBO":
        pay_status = str(res.get("paymentStatus") or "").lower()
        return pay_status not in ("", "unknown")  # unknown → détail inutile
    return False  # DIRECT → pas de valorisation Hostaway


# ═══════════════════════════════════════════════════════════════
# CALCUL PAYOUT (H1 / H2 / H3)
# ═══════════════════════════════════════════════════════════════
_META_NON_APPLICABLE = {
    "menage_retenu_source": "NON_APPLICABLE",
    "cout_standard_id": None, "cout_standard_menage_snapshot": None,
    "cout_standard_date_debut_validite": None, "cout_standard_date_fin_validite": None,
    "logement_id_snapshot": None, "type_logement_id_snapshot": None,
    "date_reference_cout_menage": None,
}


class PayoutCalculator:
    """
    Calcule payout et ménage retenu canal par canal.
    Returns 5-tuple: (payout, source_payout, statut_calcul_payout, menage_retenu, meta_dict)
    meta_dict contient les 8 colonnes de traçabilité du coût standard.
    H3: Direct/VRBO-Unknown → None, jamais valorisé depuis Hostaway.
    menage_retenu = cout_standard REF_Couts_Standards_Menage sélectionné par date d'arrivée.
    """

    def __init__(self, cost_ref_df: "pd.DataFrame" = None):
        self._cost_ref_df = cost_ref_df if cost_ref_df is not None else pd.DataFrame()

    def _menage_standard(self, res: dict) -> dict:
        """Lookup date-aware depuis REF_Couts_Standards_Menage. Retourne dict traçabilité."""
        map_id    = res.get("listingMapId")
        check_in  = res.get("arrivalDate") or res.get("checkInDate")
        return _lookup_menage_by_date(self._cost_ref_df, map_id, check_in)

    def has_cost_standard(self, map_id) -> bool:
        """Vrai si listingMapId a au moins un cout_standard dans le référentiel."""
        if self._cost_ref_df is None or self._cost_ref_df.empty:
            return False
        try:
            key = int(float(map_id))
        except (TypeError, ValueError):
            return False
        return key in self._cost_ref_df["listingMapId_num"].values

    def calc(self, res: dict, ff_list: list, fees: list, channel: str = None) -> tuple:
        if channel is None:
            channel, _ = resolve_channel(res)
        status = res.get("status", "")
        ffd    = ff_map(ff_list)

        cancel = safe_float(
            _first(res.get("cancellationAmount"), res.get("cancellationPayout"))
        )
        if status in STATUS_CANCEL:
            if cancel > 0:
                return cancel, "cancellationAmount", "ANNULE_AVEC_PAYOUT", 0.0, dict(_META_NON_APPLICABLE)
            return 0.0, "ANNULE", "ANNULE_SANS_PAYOUT", 0.0, dict(_META_NON_APPLICABLE)

        if channel == "AIRBNB":
            return self._airbnb(res, ffd)
        if channel == "BOOKING":
            return self._booking(res, ffd, fees)
        if channel == "VRBO":
            return self._vrbo(res, ffd)
        return None, "DIRECT_HORS_HOSTAWAY", "A_CONTROLER", 0.0, dict(_META_NON_APPLICABLE)

    def _airbnb(self, res: dict, ffd: dict) -> tuple:
        # H1 : airbnbExpectedPayoutAmount > fallback airbnbPayoutSum
        payout = safe_float(res.get("airbnbExpectedPayoutAmount")) or None
        src    = "airbnbExpectedPayoutAmount"
        if not payout:
            payout = ffd.get("airbnbPayoutSum") or None
            src    = "airbnbPayoutSum_fallback"
        if not payout:
            return None, "ABSENT", "PAYOUT_ABSENT", 0.0, dict(_META_NON_APPLICABLE)
        meta = self._menage_standard(res)
        return payout, src, "NORMAL", meta["menage_retenu"], meta

    def _booking(self, res: dict, ffd: dict, fees: list) -> tuple:
        # H2 : formule finance fields
        total = ffd.get("totalPriceFromChannel")
        if total:
            city   = ffd.get("cityTax", 0.0)
            ota    = ffd.get("otaPaymentProcessingFee", 0.0)
            ch_fee = ffd.get("hostChannelFee", 0.0)
            payout = total - city - ota - ch_fee
            meta   = self._menage_standard(res)
            return payout, "totalPriceFromChannel_formula", "NORMAL", meta["menage_retenu"], meta
        total_p = safe_float(res.get("totalPrice"))
        if total_p > 0:
            ch_comm = safe_float(res.get("channelCommissionAmount", 0))
            tax_fee = sum(
                safe_float(f.get("amount"))
                for f in fees
                if f.get("type", "").lower() in ("city_tax", "citytax", "tax", "tourist_tax")
            )
            payout = total_p - ch_comm - tax_fee
            meta   = self._menage_standard(res)
            return payout, "totalPrice_fallback", "PAYOUT_INCOMPLET", meta["menage_retenu"], meta
        return None, "ABSENT", "PAYOUT_ABSENT", 0.0, dict(_META_NON_APPLICABLE)

    def _vrbo(self, res: dict, ffd: dict) -> tuple:
        pay_status = (res.get("paymentStatus") or "").lower()
        if pay_status == "unknown" or not pay_status:
            return None, "VRBO_UNKNOWN", "A_CONTROLER", 0.0, dict(_META_NON_APPLICABLE)
        payout = ffd.get("totalPriceFromChannel") or safe_float(res.get("totalPrice")) or None
        if payout:
            menage = ffd.get("cleaningFee", 0.0)
            return payout, "vrbo_totalPrice", "PAYOUT_INCOMPLET", menage, dict(_META_NON_APPLICABLE)
        return None, "ABSENT", "A_CONTROLER", 0.0, dict(_META_NON_APPLICABLE)


# ═══════════════════════════════════════════════════════════════
# ANOMALY DETECTOR
# ═══════════════════════════════════════════════════════════════
class AnomalyDetector:
    def __init__(self, known_ids: set):
        self._known = known_ids
        self._rows  = []

    def check_channel_absent(self, res_id):
        self._add(res_id, "CHANNEL_ABSENT", "A_CONTROLER",
                  "Aucun champ canal trouvé dans la réservation — classé UNKNOWN")

    def check_reservation(self, res_id, channel, payout_status, map_id):
        if channel == "UNKNOWN":
            self.check_channel_absent(res_id)
        if channel == "BOOKING" and payout_status == "PAYOUT_ABSENT":
            self._add(res_id, "BOOKING_PAYOUT_INCOMPLET", "BLOQUANT",
                      "Réservation Booking active sans payout calculable (H2 impossible)")
        if channel == "VRBO" and payout_status == "A_CONTROLER":
            self._add(res_id, "VRBO_MONTANT_NON_RENSEIGNE", "A_CONTROLER",
                      "VRBO paymentStatus=Unknown — saisie manuelle requise au Lot 4")
        if payout_status == "PAYOUT_INCOMPLET":
            self._add(res_id, "PAYOUT_INCOMPLET", "A_CONTROLER",
                      "Payout calculé via fallback — champs financiers partiels")
        if map_id and map_id not in self._known:
            self._add(res_id, "LISTING_ORPHELIN_A_CONTROLER", "A_CONTROLER",
                      f"listingMapId {map_id} absent de REF_Logements — Lot 2 requis")

    def check_listing(self, map_id):
        if map_id and map_id not in self._known:
            self._add(None, "LISTING_ORPHELIN_A_CONTROLER", "A_CONTROLER",
                      f"Listing {map_id} dans API Hostaway absent de REF_Logements")

    def _add(self, res_id, code, sev, desc):
        self._rows.append({
            "reservation_id":  res_id,
            "code_anomalie":   code,
            "severite":        sev,
            "description":     desc,
            "statut":          "OUVERT",
            "date_detection":  now_utc(),
            "ROW_HASH":        row_hash(code, str(res_id)),
        })

    def to_df(self) -> "pd.DataFrame":
        if self._rows:
            return pd.DataFrame(self._rows).drop_duplicates(subset=["code_anomalie", "reservation_id"])
        return pd.DataFrame(columns=[
            "reservation_id", "code_anomalie", "severite",
            "description", "statut", "date_detection", "ROW_HASH",
        ])

    def bloquants(self) -> int:
        df = self.to_df()
        return len(df[df["severite"] == "BLOQUANT"]) if not df.empty else 0

    def a_controler(self) -> int:
        df = self.to_df()
        return len(df[df["severite"] == "A_CONTROLER"]) if not df.empty else 0


# ═══════════════════════════════════════════════════════════════
# WRITER
# ═══════════════════════════════════════════════════════════════
def write_excel(df: "pd.DataFrame", path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="data", index=False)


# ═══════════════════════════════════════════════════════════════
# CHARGEMENT RÉFÉRENTIELS REF_SETUP
# ═══════════════════════════════════════════════════════════════
def load_known_listing_ids(ref_path: Path, log) -> set:
    """Charge les hostaway_listing_id de REF_Logements pour flag orphelin."""
    try:
        import openpyxl as xl
        wb = xl.load_workbook(ref_path, read_only=True, data_only=True)
        ws = wb["REF_Logements"]
        rows = list(ws.iter_rows(values_only=True))
        headers = list(rows[0])
        idx = headers.index("hostaway_listing_id")
        ids = {r[idx] for r in rows[1:] if r[idx] is not None}
        wb.close()
        log.info(f"REF_Logements : {len(ids)} IDs Hostaway connus (flag orphelin actif)")
        return ids
    except Exception as e:
        log.warning(f"REF_Logements illisible ({e}) — flag orphelin désactivé")
        return set()


def _build_cost_ref_df(ref_path: Path) -> "pd.DataFrame":
    """
    Construit DataFrame {listingMapId_num, logement_id, nom_court, type_logement_id,
    cout_standard_id, cout_std, date_debut_validite, date_fin_validite}
    depuis REF_Setup.xlsm via jointure :
      REF_Mapping_Logements (Hostaway, listingMapId, actif=OUI)
      → REF_Logements (type_logement_id)
      → REF_Couts_Standards_Menage (actif=OUI, avec dates de validité).
    Résultat : 1 ligne par (listingMapId, période de validité).
    """
    import openpyxl as xl
    wb = xl.load_workbook(ref_path, read_only=True, data_only=True)

    def _ws_df(name):
        ws  = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        return pd.DataFrame(rows[1:], columns=rows[0])

    df_map  = _ws_df("REF_Mapping_Logements")
    df_log  = _ws_df("REF_Logements")
    df_cout = _ws_df("REF_Couts_Standards_Menage")
    wb.close()

    df_ha = df_map[
        (df_map["source"] == "Hostaway") &
        (df_map["champ_source"] == "listingMapId") &
        (df_map["actif"] == "OUI")
    ][["valeur_source", "logement_id"]].copy()
    df_ha.columns = ["listingMapId_num", "logement_id"]
    df_ha["listingMapId_num"] = pd.to_numeric(df_ha["listingMapId_num"], errors="coerce")

    df_log = df_log[["logement_id", "type_logement_id", "nom_court"]].copy()

    df_cout_act = df_cout[df_cout["actif"] == "OUI"][
        ["cout_standard_id", "type_logement_id", "cout_standard_menage",
         "date_debut_validite", "date_fin_validite"]
    ].copy()
    df_cout_act = df_cout_act.rename(columns={"cout_standard_menage": "cout_std"})
    df_cout_act["cout_std"] = pd.to_numeric(df_cout_act["cout_std"], errors="coerce")
    df_cout_act["date_debut_validite"] = pd.to_datetime(df_cout_act["date_debut_validite"], errors="coerce")
    df_cout_act["date_fin_validite"]   = pd.to_datetime(df_cout_act["date_fin_validite"],   errors="coerce")

    result = (
        df_ha
        .merge(df_log, on="logement_id", how="left")
        .merge(df_cout_act, on="type_logement_id", how="left")
    )[["listingMapId_num", "logement_id", "nom_court", "type_logement_id",
       "cout_standard_id", "cout_std", "date_debut_validite", "date_fin_validite"]]

    return result.dropna(subset=["listingMapId_num"])


def load_menage_cost_ref(ref_path: Path, log) -> "pd.DataFrame":
    """
    Retourne DataFrame [listingMapId_num, logement_id, nom_court, type_logement_id,
    cout_standard_id, cout_std, date_debut_validite, date_fin_validite].
    Une ligne par (listingMapId, période de validité cout_standard).
    Source : REF_Couts_Standards_Menage via REF_Mapping_Logements → REF_Logements.
    """
    _empty_cols = ["listingMapId_num", "logement_id", "nom_court", "type_logement_id",
                   "cout_standard_id", "cout_std", "date_debut_validite", "date_fin_validite"]
    try:
        df = _build_cost_ref_df(ref_path)
        n_map = int(df["listingMapId_num"].nunique())
        n_row = len(df)
        log.info(f"REF_Couts_Standards_Menage : {n_map} listingMapId, {n_row} ligne(s) de validité chargées")
        return df
    except Exception as e:
        log.warning(f"REF_Couts_Standards_Menage illisible ({e}) — menage_retenu=0 par défaut")
        return pd.DataFrame(columns=_empty_cols)


def _lookup_menage_by_date(cost_ref_df: "pd.DataFrame", map_id, check_in) -> dict:
    """
    Sélectionne le coût standard ménage valide à la date d'arrivée de la réservation.
    Règle : date_debut_validite <= check_in
            AND (date_fin_validite IS NULL OR check_in <= date_fin_validite)
    Date de référence = checkInDate de la réservation (jamais date du jour / date recalcul).
    Retourne un dict avec tous les champs de traçabilité.
    Anomalie BLOQUANTE si doublon de validité pour la même date.
    """
    _empty = {
        "menage_retenu": 0.0, "found": False, "doublon": False,
        "menage_retenu_source": "MENAGE_SANS_COUT_STANDARD",
        "cout_standard_id": None, "cout_standard_menage_snapshot": None,
        "cout_standard_date_debut_validite": None, "cout_standard_date_fin_validite": None,
        "logement_id_snapshot": None, "type_logement_id_snapshot": None,
        "date_reference_cout_menage": None,
    }

    if cost_ref_df is None or cost_ref_df.empty:
        return _empty.copy()

    try:
        key = int(float(map_id))
    except (TypeError, ValueError):
        return _empty.copy()

    try:
        ts_in = pd.Timestamp(check_in)
        if pd.isna(ts_in):
            return _empty.copy()
    except Exception:
        return _empty.copy()

    df_lid = cost_ref_df[cost_ref_df["listingMapId_num"] == key]
    if df_lid.empty:
        r = _empty.copy()
        r["date_reference_cout_menage"] = str(ts_in.date())
        return r

    mask_debut = df_lid["date_debut_validite"].notna() & (df_lid["date_debut_validite"] <= ts_in)
    mask_fin   = df_lid["date_fin_validite"].isna() | (df_lid["date_fin_validite"] >= ts_in)
    df_valid   = df_lid[mask_debut & mask_fin]

    date_ref_str = str(ts_in.date())

    if df_valid.empty:
        r = _empty.copy()
        r["date_reference_cout_menage"] = date_ref_str
        return r

    if len(df_valid) > 1:
        ids = list(df_valid["cout_standard_id"].values)
        return {
            "menage_retenu": 0.0, "found": False, "doublon": True,
            "menage_retenu_source": "COUT_STANDARD_MENAGE_DOUBLON_VALIDITE",
            "cout_standard_id": ",".join(str(i) for i in ids),
            "cout_standard_menage_snapshot": None,
            "cout_standard_date_debut_validite": None,
            "cout_standard_date_fin_validite": None,
            "logement_id_snapshot": str(df_valid.iloc[0]["logement_id"]) if pd.notna(df_valid.iloc[0]["logement_id"]) else None,
            "type_logement_id_snapshot": str(df_valid.iloc[0]["type_logement_id"]) if pd.notna(df_valid.iloc[0]["type_logement_id"]) else None,
            "date_reference_cout_menage": date_ref_str,
        }

    row   = df_valid.iloc[0]
    fin_v = row["date_fin_validite"]
    return {
        "menage_retenu": float(row["cout_std"]) if pd.notna(row["cout_std"]) else 0.0,
        "found": pd.notna(row["cout_std"]),
        "doublon": False,
        "menage_retenu_source": "REF_COUT_STANDARD_MENAGE",
        "cout_standard_id": str(row["cout_standard_id"]) if pd.notna(row["cout_standard_id"]) else None,
        "cout_standard_menage_snapshot": float(row["cout_std"]) if pd.notna(row["cout_std"]) else None,
        "cout_standard_date_debut_validite": str(row["date_debut_validite"].date()) if pd.notna(row["date_debut_validite"]) else None,
        "cout_standard_date_fin_validite": str(fin_v.date()) if pd.notna(fin_v) else None,
        "logement_id_snapshot": str(row["logement_id"]) if pd.notna(row["logement_id"]) else None,
        "type_logement_id_snapshot": str(row["type_logement_id"]) if pd.notna(row["type_logement_id"]) else None,
        "date_reference_cout_menage": date_ref_str,
    }


# ═══════════════════════════════════════════════════════════════
# CLEANING TASKS — fonction isolée, non-bloquante
# ═══════════════════════════════════════════════════════════════
def _extract_cleaning_tasks(client, date_from: str, detector, log) -> tuple:
    """
    Extrait les tâches ménage. Retourne (rows_tasks, statut).
    statut: "OK" | "INCOMPLETE" | "FAILED"
    Non-bloquant : toujours retourne quelque chose, même vide.
    H6 : cost = NULL irrévocablement.
    """
    log.info("Extraction taches menage (cleaning tasks, H6)...")
    rows_tasks = []
    statut     = "FAILED"
    try:
        tasks_raw = client.get_tasks(date_from)
        for t in tasks_raw:
            rows_tasks.append({
                "task_id":        t.get("id"),
                "reservation_id": t.get("reservationId"),
                "listingMapId":   t.get("listingMapId"),
                "task_type":      t.get("taskType") or t.get("type"),
                "status":         t.get("status"),
                "scheduled_date": t.get("scheduledDate") or t.get("date"),
                "assignee":       t.get("assigneeName") or t.get("assignee"),
                "cost":           None,  # H6 : jamais valorise
                "h6_note":        "cost=NULL_H6_comptage_uniquement",
                "extrait_le":     now_utc(),
                "ROW_HASH":       row_hash(t.get("id")),
            })
        statut = "OK"
        log.info(f"  {len(rows_tasks)} taches menage extraites.")
    except Exception as e:
        log.warning(f"Extraction taches menage echouee : {type(e).__name__}: {e}")
        log.warning("  Les tables principales restent valides. Relancer avec --only-cleaning-tasks.")
        detector._add(None, "CLEANING_TASKS_EXTRACTION_INCOMPLETE", "A_CONTROLER",
                       f"Extraction taches menage incomplete : {type(e).__name__}")
        statut = "INCOMPLETE" if rows_tasks else "FAILED"
    return rows_tasks, statut


# ═══════════════════════════════════════════════════════════════
# RECALC PAYOUT ONLY — sans API, depuis fichiers Excel existants
# ═══════════════════════════════════════════════════════════════
def recalc_payout_only(out_dir: Path, log, payout_source: Path = None) -> None:
    """
    --recalc-payout-only : recalcule menage_retenu depuis REF_Couts_Standards_Menage.
    Source : payout_source (backup recommandé) ou MASTER_CALC_HA_Payout.xlsx si None.
    Lookup date-aware : cout_standard sélectionné selon checkInDate de chaque réservation.
    Ajoute 8 colonnes de traçabilité. Applique pour AIRBNB NORMAL et BOOKING NORMAL.
    Annulations avec payout : menage_retenu = 0 conservé (D030).
    N'appelle pas l'API. Imprime les 14+ contrôles obligatoires.
    """
    log.info("=" * 60)
    log.info("RECALC PAYOUT ONLY — correctif #3 : REF_Couts_Standards_Menage + date-aware")
    log.info("Source menage_retenu : cout_standard REF_Setup selectionne par checkInDate")
    log.info("API Hostaway : NON appelee")
    log.info("=" * 60)

    prod_path   = out_dir / "MASTER_CALC_HA_Payout.xlsx"
    source_path = payout_source if payout_source else prod_path

    if not source_path.exists():
        log.error(f"Source payout introuvable : {source_path}")
        sys.exit(1)

    log.info(f"Source payout    : {source_path.name}")
    log.info(f"Destination      : {prod_path.name}")

    # ── 1. Charger source payout ─────────────────────────────────
    log.info("Chargement source payout...")
    df_pay = pd.read_excel(source_path, sheet_name="data")
    n_total = len(df_pay)
    log.info(f"  {n_total} lignes chargees")

    # ── 2. Charger checkInDate depuis MASTER_FACT_HA_Reservations ─
    res_path = out_dir / "MASTER_FACT_HA_Reservations.xlsx"
    if not res_path.exists():
        log.error(f"MASTER_FACT_HA_Reservations introuvable : {res_path}")
        log.error("Impossible de recuperer checkInDate — recalcul annule.")
        sys.exit(1)
    log.info("Chargement checkInDate depuis MASTER_FACT_HA_Reservations...")
    df_res     = pd.read_excel(res_path, sheet_name="data")
    df_res_ci  = df_res[["reservation_id", "checkInDate", "cleaningFee_res"]].copy()
    df_res_ci["checkInDate"] = pd.to_datetime(df_res_ci["checkInDate"], errors="coerce")
    log.info(f"  {len(df_res_ci)} reservations, checkInDate pret")

    # ── 3. Coûts standards depuis REF_Setup (avec dates) ─────────
    log.info("Chargement REF_Couts_Standards_Menage depuis REF_Setup (avec dates validite)...")
    cost_ref_df = _build_cost_ref_df(REF_FILE)
    n_map = int(cost_ref_df["listingMapId_num"].nunique())
    log.info(f"  {n_map} listingMapId, {len(cost_ref_df)} ligne(s) de validite")

    # ── 4. Merge checkInDate dans df_pay ─────────────────────────
    df_pay = df_pay.merge(df_res_ci[["reservation_id", "checkInDate"]], on="reservation_id", how="left")

    # ── 5. Masques ────────────────────────────────────────────────
    mask_ab_norm = (df_pay["channel_type"] == "AIRBNB")  & (df_pay["statut_calcul_payout"] == "NORMAL")
    mask_bk_norm = (df_pay["channel_type"] == "BOOKING") & (df_pay["statut_calcul_payout"] == "NORMAL")
    mask_cancel  = df_pay["statut_calcul_payout"] == "ANNULE_AVEC_PAYOUT"
    mask_normal  = mask_ab_norm | mask_bk_norm

    # ── 6. Stats AVANT ────────────────────────────────────────────
    men_ab_avant = float(df_pay.loc[mask_ab_norm, "menage_retenu"].fillna(0).sum())
    men_bk_avant = float(df_pay.loc[mask_bk_norm, "menage_retenu"].fillna(0).sum())
    ass_ab_avant = float(df_pay.loc[mask_ab_norm, "assiette_commission"].fillna(0).sum())
    ass_bk_avant = float(df_pay.loc[mask_bk_norm, "assiette_commission"].fillna(0).sum())
    cancel_avant = float(df_pay.loc[mask_cancel,  "menage_retenu"].fillna(0).sum())

    # ── 7. Lookup date-aware + remplissage 8 colonnes ─────────────
    log.info("Lookup date-aware cout_standard par reservation...")
    _8cols = [
        "menage_retenu_source", "cout_standard_id", "cout_standard_menage_snapshot",
        "cout_standard_date_debut_validite", "cout_standard_date_fin_validite",
        "logement_id_snapshot", "type_logement_id_snapshot", "date_reference_cout_menage",
    ]
    for col in _8cols:
        if col not in df_pay.columns:
            df_pay[col] = None

    # Traçabilité pour colonnes non-NORMAL : NON_APPLICABLE
    for col in _8cols:
        df_pay.loc[~mask_normal, col] = "NON_APPLICABLE" if col == "menage_retenu_source" else None

    n_doublon   = 0
    n_missing   = 0
    n_ab_corr   = 0
    n_bk_corr   = 0
    rows_idx_normal = df_pay.index[mask_normal]

    for idx in rows_idx_normal:
        row       = df_pay.loc[idx]
        map_id    = row["listingMapId"]
        check_in  = row.get("checkInDate")
        meta      = _lookup_menage_by_date(cost_ref_df, map_id, check_in)

        if meta["doublon"]:
            n_doublon += 1
            log.warning(
                f"  DOUBLON VALIDITE BLOQUANT — reservation_id={row['reservation_id']} "
                f"listingMapId={map_id} checkIn={check_in} "
                f"cout_standard_id={meta['cout_standard_id']}"
            )
        elif not meta["found"]:
            n_missing += 1
            log.warning(
                f"  SANS COUT STANDARD — reservation_id={row['reservation_id']} "
                f"listingMapId={map_id} checkIn={check_in}"
            )

        # Mise à jour menage_retenu + assiette
        menage_new = meta["menage_retenu"]
        df_pay.at[idx, "menage_retenu"]       = menage_new
        df_pay.at[idx, "assiette_commission"] = (
            float(row["payout_calcule"]) - menage_new
            if pd.notna(row["payout_calcule"]) else None
        )

        # 8 colonnes de traçabilité
        for col in _8cols:
            df_pay.at[idx, col] = meta.get(col.replace("menage_retenu_source", "menage_retenu_source"))

        if row["channel_type"] == "AIRBNB":
            n_ab_corr += 1
        else:
            n_bk_corr += 1

    # ── 8. Stats APRÈS ────────────────────────────────────────────
    men_ab_apres = float(df_pay.loc[mask_ab_norm, "menage_retenu"].fillna(0).sum())
    men_bk_apres = float(df_pay.loc[mask_bk_norm, "menage_retenu"].fillna(0).sum())
    ass_ab_apres = float(df_pay.loc[mask_ab_norm, "assiette_commission"].fillna(0).sum())
    ass_bk_apres = float(df_pay.loc[mask_bk_norm, "assiette_commission"].fillna(0).sum())
    cancel_apres = float(df_pay.loc[mask_cancel,  "menage_retenu"].fillna(0).sum())
    cancel_ok    = abs(cancel_avant - cancel_apres) < 0.01

    # ── 9. Contrôle traçabilité colonnes NORMAL ───────────────────
    n_ab_with_id = int(df_pay.loc[mask_ab_norm, "cout_standard_id"].notna().sum())
    n_bk_with_id = int(df_pay.loc[mask_bk_norm, "cout_standard_id"].notna().sum())
    n_snap_ok    = int(
        (df_pay.loc[mask_normal, "cout_standard_menage_snapshot"] ==
         df_pay.loc[mask_normal, "menage_retenu"]).sum()
    )
    n_date_ok    = int(df_pay.loc[mask_normal, "date_reference_cout_menage"].notna().sum())

    # ── 10. Écart cleaningFee_res vs cout_std (informel) ──────────
    ecart_cln_vs_std = None
    if res_path.exists():
        df_check      = df_pay[mask_ab_norm].merge(
            df_res_ci[["reservation_id", "cleaningFee_res"]], on="reservation_id", how="left"
        )
        cln_total         = float(df_check["cleaningFee_res"].fillna(0).sum())
        std_total         = float(df_check["cout_standard_menage_snapshot"].fillna(0).sum())
        ecart_cln_vs_std  = cln_total - std_total

    delta_ab          = ass_ab_avant - ass_ab_apres
    delta_bk          = ass_bk_avant - ass_bk_apres
    commission_impact = (delta_ab + delta_bk) * 0.15

    # ── 11. Supprimer colonne helper ──────────────────────────────
    df_pay.drop(columns=["checkInDate"], inplace=True, errors="ignore")

    # ── 12. 14+ CONTRÔLES OBLIGATOIRES ───────────────────────────
    log.info("\n--- 14+ CONTROLES OBLIGATOIRES CTR-2026-06-018 (correctif #3 date-aware) ---")
    log.info(f"  CTR-1   Lignes Airbnb NORMAL traitees         : {n_ab_corr}")
    log.info(f"  CTR-2   Lignes Booking NORMAL traitees        : {n_bk_corr}")
    log.info(f"  CTR-3   menage_retenu Airbnb AVANT (source)   : {men_ab_avant:.2f} E")
    log.info(f"  CTR-4   menage_retenu Airbnb APRES REF_Setup  : {men_ab_apres:.2f} E")
    log.info(f"  CTR-5   menage_retenu Booking AVANT           : {men_bk_avant:.2f} E")
    log.info(f"  CTR-6   menage_retenu Booking APRES REF_Setup : {men_bk_apres:.2f} E")
    log.info(f"  CTR-7   assiette Airbnb AVANT                 : {ass_ab_avant:.2f} E")
    log.info(f"  CTR-8   assiette Airbnb APRES                 : {ass_ab_apres:.2f} E")
    log.info(f"  CTR-9   assiette Booking AVANT                : {ass_bk_avant:.2f} E")
    log.info(f"  CTR-10  assiette Booking APRES                : {ass_bk_apres:.2f} E")
    if ecart_cln_vs_std is not None:
        log.info(f"  CTR-11  Ecart cleaningFee_res vs cout_std AB  : {ecart_cln_vs_std:+.2f} E")
    else:
        log.info(f"  CTR-11  Ecart cleaningFee_res vs cout_std     : (MASTER_FACT absent)")
    log.info(f"  CTR-12  Lignes NORMAL sans cout_standard      : {n_missing}")
    log.info(f"  CTR-13  Annulations avec payout intactes      : {'OK' if cancel_ok else 'ECHEC'}")
    log.info(f"           menage_retenu annule AVANT           : {cancel_avant:.2f} E")
    log.info(f"           menage_retenu annule APRES           : {cancel_apres:.2f} E")
    log.info(f"  CTR-14  Impact estime commissions (~15%)      : {commission_impact:+.2f} E")
    log.info(f"           API non relancee                     : OUI")
    log.info(f"           Source utilisee                      : {source_path.name}")
    log.info(f"  CTR-15  DOUBLONS VALIDITE (BLOQUANT)         : {n_doublon}")
    log.info(f"  CTR-16  Airbnb NORMAL avec cout_standard_id  : {n_ab_with_id} / {n_ab_corr}")
    log.info(f"  CTR-17  Booking NORMAL avec cout_standard_id : {n_bk_with_id} / {n_bk_corr}")
    log.info(f"  CTR-18  Snapshot == menage_retenu (NORMAL)   : {n_snap_ok} / {int(mask_normal.sum())}")
    log.info(f"  CTR-19  date_reference non vide (NORMAL)     : {n_date_ok} / {int(mask_normal.sum())}")

    if not cancel_ok:
        log.error("ECHEC CTR-13 : annulations avec payout alterees — ecriture annulee")
        sys.exit(1)

    if n_doublon > 0:
        log.error(f"ECHEC CTR-15 : {n_doublon} doublon(s) de validite BLOQUANT — corriger REF_Setup avant validation")
        sys.exit(1)

    # ── 13. Écriture ─────────────────────────────────────────────
    write_excel(df_pay, prod_path)
    log.info(f"\nMaster payout ecrit : {prod_path}")
    log.info(f"  {len(df_pay)} lignes, 8 colonnes tracabilite ajoutees")
    log.info("RECALC #3 OK — EN_ATTENTE_VALIDATION_HUMAINE")
    log.info("Lot 10 reste bloque jusqu'a validation humaine du correctif final.")
    log.info("=" * 60)


# ═══════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(
        description="Lot 1 — Extraction Hostaway API",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--check-auth", action="store_true",
        help="Test OAuth2 uniquement. Aucun secret affiché.",
    )
    parser.add_argument(
        "--check-listings", action="store_true",
        help="Auth + test /v1/listings : compte, champs, 3 IDs. Aucune écriture.",
    )
    parser.add_argument(
        "--check-reservations-sample", action="store_true",
        help="Auth + 5 réservations : clés, champs canal, nb UNKNOWN. Aucune écriture.",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Auth + comptage réservations. Aucune écriture disque.",
    )
    parser.add_argument(
        "--details-mode", choices=[DETAILS_MINIMAL, DETAILS_FULL],
        default=DETAILS_MINIMAL,
        help="minimal (défaut) : détail uniquement si payout impossible. full : toujours.",
    )
    parser.add_argument(
        "--skip-cleaning-tasks", action="store_true",
        help="Ignorer l'extraction des tâches ménage (plus rapide).",
    )
    parser.add_argument(
        "--only-cleaning-tasks", action="store_true",
        help="Extraire uniquement les tâches ménage (nécessite tables réservations existantes).",
    )
    parser.add_argument(
        "--recalc-payout-only", action="store_true",
        help="Recalcule MASTER_CALC_HA_Payout depuis fichiers Excel existants. N'appelle pas l'API.",
    )
    parser.add_argument(
        "--payout-source", type=str, default=None,
        help="Fichier payout source pour --recalc-payout-only (backup recommandé). "
             "Défaut : fichier production MASTER_CALC_HA_Payout.xlsx.",
    )
    parser.add_argument(
        "--limit", type=int, default=0,
        help="Limiter à N réservations (0 = pas de limite). Pour tests.",
    )
    args = parser.parse_args()

    # ── Modes diagnostic : sys.exit() direct ─────────────────────
    if args.check_auth:
        check_auth_mode(ENV_FILE)
        # jamais atteint
    if args.check_listings:
        check_listings_mode(ENV_FILE)
    if args.check_reservations_sample:
        check_reservations_sample_mode(ENV_FILE)
        # jamais atteint

    run_id    = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_start = datetime.now(timezone.utc)
    log       = setup_logging(run_id, silent_file=args.dry_run)

    # ── Recalcul payout sans API ─────────────────────────────────
    if getattr(args, "recalc_payout_only", False):
        src = Path(args.payout_source) if getattr(args, "payout_source", None) else None
        recalc_payout_only(OUT_DIR, log, payout_source=src)
        return

    details_mode = getattr(args, "details_mode", DETAILS_MINIMAL)
    skip_tasks   = getattr(args, "skip_cleaning_tasks", False)
    only_tasks   = getattr(args, "only_cleaning_tasks", False)
    limit        = getattr(args, "limit", 0)

    log.info("=" * 60)
    log.info(f"LOT 1 — EXTRACTION HOSTAWAY — run_id={run_id}")
    if args.dry_run:
        mode_str = "--dry-run"
    elif only_tasks:
        mode_str = "--only-cleaning-tasks"
    else:
        mode_str = f"EXTRACTION (details={details_mode}{', limit='+str(limit) if limit else ''}{', skip-tasks' if skip_tasks else ''})"
    log.info(f"Mode     : {mode_str}")
    log.info(f"Date from: {DATE_FROM}")
    log.info("=" * 60)

    # ── Charger .env ─────────────────────────────────────────
    if not ENV_FILE.exists():
        log.error(f".env introuvable : {ENV_FILE}")
        sys.exit(1)

    load_dotenv(ENV_FILE)
    client_id     = os.getenv("HOSTAWAY_CLIENT_ID", "")
    client_secret = os.getenv("HOSTAWAY_CLIENT_SECRET", "")
    account_id    = os.getenv("HOSTAWAY_ACCOUNT_ID", "")
    base_url      = os.getenv("HOSTAWAY_BASE_URL", "https://api.hostaway.com")

    missing_vars = [k for k, v in [
        ("HOSTAWAY_CLIENT_ID", client_id),
        ("HOSTAWAY_CLIENT_SECRET", client_secret),
        ("HOSTAWAY_ACCOUNT_ID", account_id),
    ] if not v]

    if missing_vars:
        log.error(f"Variables manquantes dans .env : {', '.join(missing_vars)}")
        sys.exit(1)

    # Jamais de valeur affichée — CLIENT_ID, ACCOUNT_ID, CLIENT_SECRET masqués
    log.info("Variables .env : CLIENT_ID=*** ACCOUNT_ID=*** — présentes")
    log.info(f"BASE_URL     : {base_url}")

    # ── Auth ─────────────────────────────────────────────────
    log.info("Authentification OAuth2...")
    auth = HostawayAuth(base_url, client_id, client_secret)
    ok   = auth.test()
    if not ok:
        log.error("Authentification echouee. Lancer --check-auth pour diagnostic détaillé.")
        sys.exit(1)
    log.info("Authentification OK. Token valide (non affiché).")

    # ── Client ───────────────────────────────────────────────
    client = HostawayClient(auth, account_id, log)

    # ── Comptage (dry-run s'arrête ici) ──────────────────────
    log.info(f"Comptage réservations depuis {DATE_FROM}...")
    total = client.count_reservations(DATE_FROM)
    log.info(f"API Reservations count = {total}")

    if args.dry_run:
        log.info("--dry-run termine. Aucune ecriture. Lancer sans --dry-run pour extraire.")
        return

    # ── REF_Logements (flag orphelin) ────────────────────────
    known_ids    = load_known_listing_ids(REF_FILE, log)
    cost_ref_df  = load_menage_cost_ref(REF_FILE, log)
    detector     = AnomalyDetector(known_ids)
    payout_c     = PayoutCalculator(cost_ref_df)

    # Variables de suivi run (initialisées tôt pour le try/finally)
    rows_listings = []
    rows_res      = []
    rows_details  = []
    rows_ff       = []
    rows_fees     = []
    rows_payout   = []
    rows_tasks    = []
    processed     = 0
    skipped       = 0
    nb_details    = 0
    tasks_statut  = "SKIPPED" if (skip_tasks or only_tasks) else "PENDING"
    statut_run    = "FAILED"

    try:
        # ── ONLY-CLEANING-TASKS : branche dédiée ─────────────
        if only_tasks:
            log.info("Mode --only-cleaning-tasks : skip réservations/listings.")
            rows_tasks, tasks_statut = _extract_cleaning_tasks(client, DATE_FROM, detector, log)
            log.info("Ecriture MASTER_FACT_HA_CleaningTasks_Discovery...")
            OUT_DIR.mkdir(parents=True, exist_ok=True)
            write_excel(pd.DataFrame(rows_tasks),
                        OUT_DIR / "MASTER_FACT_HA_CleaningTasks_Discovery.xlsx")
            log.info(f"  {len(rows_tasks)} taches menage.")
            statut_run = "TERMINE_OK" if tasks_statut == "OK" else "PARTIAL_OK"
            total = 0
        else:
            # ── LISTINGS ─────────────────────────────────────
            log.info("Extraction listings via /v1/listings...")
            try:
                listings_raw = client.get_listings()
            except RuntimeError as e:
                log.error(f"Echec recuperation listings : {e}")
                sys.exit(1)

            for lst in listings_raw:
                lid    = lst.get("id")
                map_id = lst.get("listingMapId") or lid
                detector.check_listing(map_id)
                special = lst.get("specialStatus") or ""
                actif   = "NON" if special.strip() else "OUI"
                rows_listings.append({
                    "listingMapId":      map_id,
                    "listing_id_ha":     lid,
                    "nom_listing":       lst.get("name"),
                    "internalName":      lst.get("internalListingName") or lst.get("name"),
                    "ville":             lst.get("city"),
                    "actif":             actif,
                    "special_status":    special or None,
                    "airbnb_status":     lst.get("airbnbExportStatus"),
                    "bookingcom_status": lst.get("bookingcomExportStatus"),
                    "sur_hostaway":      "OUI",
                    "extrait_le":        now_utc(),
                    "ROW_HASH":          row_hash(map_id, lid),
                })
            log.info(f"  {len(rows_listings)} listings.")

            # ── RESERVATIONS (pagination) ─────────────────────
            log.info(f"Extraction reservations depuis {DATE_FROM} ({total} attendues, "
                     f"details={details_mode}{', limit='+str(limit) if limit else ''})...")
            offset = 0
            stopped_by_limit = False

            while True:
                batch = client.get_reservations_page(DATE_FROM, offset)
                if not batch:
                    break

                for res in batch:
                    # Limite de test
                    if limit and processed >= limit:
                        stopped_by_limit = True
                        break

                    res_id = res.get("id")
                    status = str(res.get("status") or "")

                    if status in STATUS_SKIP:
                        skipped += 1
                        continue

                    channel, source_brute = resolve_channel(res)
                    if channel == "UNKNOWN":
                        log.warning(f"  Reservation {res_id} — canal absent")

                    is_active    = status in STATUS_ACTIVE
                    is_cancelled = status in STATUS_CANCEL
                    is_owner     = status in STATUS_OWNER
                    cancel_amt   = safe_float(
                        _first(res.get("cancellationAmount"), res.get("cancellationPayout"))
                    )
                    include_financial = (is_active or (is_cancelled and cancel_amt > 0)) and not is_owner
                    include_any       = is_active or is_owner or (is_cancelled and cancel_amt > 0)

                    if not include_any:
                        skipped += 1
                        continue

                    try:
                        map_id = res.get("listingMapId")

                        rows_res.append({
                            "reservation_id":       res_id,
                            "listingMapId":         map_id,
                            "source":               source_brute,
                            "channel_type":         channel,
                            "source_financiere":    source_financiere(channel),
                            "status":               status,
                            "paymentStatus":        res.get("paymentStatus"),
                            "checkInDate":          res.get("arrivalDate") or res.get("checkInDate"),
                            "checkOutDate":         res.get("departureDate") or res.get("checkOutDate"),
                            "nights":               res.get("nights"),
                            "guestCount":           res.get("guestCount"),
                            "guestName":            res.get("guestName"),
                            "totalPrice":           safe_float(res.get("totalPrice")),
                            "cleaningFee_res":      safe_float(res.get("cleaningFee")),
                            "channelCommission":    safe_float(res.get("channelCommissionAmount")),
                            "airbnbExpectedPayout": safe_float(res.get("airbnbExpectedPayoutAmount")),
                            "is_ownerStay":         "OUI" if is_owner else "NON",
                            "inclure_resultat":     "OUI" if include_financial else "NON",
                            "updatedOn":            res.get("updatedOn"),
                            "createdOn":            res.get("createdOn"),
                            "extrait_le":           now_utc(),
                            "ROW_HASH":             row_hash(res_id, res.get("updatedOn")),
                        })

                        # Finance fields depuis la liste (sans appel détail)
                        ff_list_base = _ff_from_res(res)
                        fees_list    = res.get("reservationFees") or []

                        # Décision appel détail
                        fetch_detail = (
                            details_mode == DETAILS_FULL
                            or (not is_owner and _needs_detail(res, channel, ff_list_base))
                        )

                        if fetch_detail:
                            try:
                                detail = client.get_reservation_detail(res_id)
                                nb_details += 1
                                time.sleep(0.1)
                            except Exception as e:
                                log.warning(f"  Detail resa {res_id} echoue : {e}")
                                detail = {}

                            ff_list  = _ff_from_res(detail) or ff_list_base
                            fees_list = detail.get("reservationFees") or fees_list

                            rows_details.append({
                                "reservation_id": res_id,
                                "json_snapshot":  json.dumps(detail, ensure_ascii=False)[:4000],
                                "extrait_le":     now_utc(),
                                "ROW_HASH":       row_hash(res_id, "detail"),
                            })
                        else:
                            ff_list = ff_list_base

                        # Finance fields → table
                        for ff_item in ff_list:
                            name = ff_item.get("name", "")
                            if not name:
                                continue
                            rows_ff.append({
                                "reservation_id":     res_id,
                                "financeField_name":  name,
                                "financeField_value": safe_float(ff_item.get("value")),
                                "currency":           ff_item.get("currency", "EUR"),
                                "ROW_HASH":           row_hash(res_id, name),
                            })

                        # Fees → table
                        for fee in fees_list:
                            fee_id = fee.get("id") or row_hash(res_id, fee.get("name"), str(fee.get("amount")))
                            rows_fees.append({
                                "reservation_id": res_id,
                                "fee_id":         fee_id,
                                "fee_name":       fee.get("name"),
                                "fee_type":       fee.get("type"),
                                "amount":         safe_float(fee.get("amount")),
                                "currency":       fee.get("currency", "EUR"),
                                "ROW_HASH":       row_hash(res_id, str(fee_id)),
                            })

                        # Payout H1/H2/H3
                        if not is_owner:
                            payout, payout_src, payout_status, menage, meta = payout_c.calc(
                                res, ff_list, fees_list, channel=channel
                            )
                            detector.check_reservation(res_id, channel, payout_status, map_id)
                            if channel in ("AIRBNB", "BOOKING") and not payout_c.has_cost_standard(map_id):
                                detector._add(res_id, "COUT_STANDARD_MENAGE_ABSENT", "A_CONTROLER",
                                              f"listingMapId {map_id} sans cout_standard REF_Couts_Standards_Menage")
                            if meta.get("doublon"):
                                detector._add(res_id, "COUT_STANDARD_MENAGE_DOUBLON_VALIDITE", "BLOQUANT",
                                              f"Doublons validité cout_standard pour listingMapId {map_id} "
                                              f"à checkInDate {meta.get('date_reference_cout_menage')} : "
                                              f"cout_standard_id={meta.get('cout_standard_id')}")
                            assiette = (payout - menage) if payout is not None else None
                            rows_payout.append({
                                "reservation_id":                    res_id,
                                "listingMapId":                      map_id,
                                "source":                            source_brute,
                                "channel_type":                      channel,
                                "statut_calcul_payout":              payout_status,
                                "payout_calcule":                    payout,
                                "source_payout":                     payout_src,
                                "menage_retenu":                     menage,
                                "assiette_commission":               assiette,
                                "menage_retenu_source":              meta.get("menage_retenu_source"),
                                "cout_standard_id":                  meta.get("cout_standard_id"),
                                "cout_standard_menage_snapshot":     meta.get("cout_standard_menage_snapshot"),
                                "cout_standard_date_debut_validite": meta.get("cout_standard_date_debut_validite"),
                                "cout_standard_date_fin_validite":   meta.get("cout_standard_date_fin_validite"),
                                "logement_id_snapshot":              meta.get("logement_id_snapshot"),
                                "type_logement_id_snapshot":         meta.get("type_logement_id_snapshot"),
                                "date_reference_cout_menage":        meta.get("date_reference_cout_menage"),
                                "inclure_resultat_auto":             "OUI" if payout_status == "NORMAL" else "NON",
                                "extrait_le":                        now_utc(),
                                "ROW_HASH":                          row_hash(res_id, payout_status),
                            })

                        processed += 1

                    except Exception as exc:
                        log.warning(f"  Resa {res_id} ignoree — {type(exc).__name__}: {exc}")
                        detector._add(res_id, "RESERVATION_SCHEMA_INATTENDU", "A_CONTROLER",
                                       f"Schema inattendu : {type(exc).__name__}")
                        skipped += 1

                if stopped_by_limit:
                    log.info(f"  --limit {limit} atteint — arret pagination.")
                    break

                log.info(f"  offset={offset:4d} : lot {len(batch)} — traite={processed} "
                         f"saute={skipped} details={nb_details}")
                if len(batch) < PAGE_SIZE:
                    break
                offset += PAGE_SIZE
                time.sleep(0.1)

            log.info(f"Reservations : {processed} traitees, {skipped} sautees, "
                     f"{nb_details} appels detail sur {processed} ({100*nb_details//max(processed,1)}%)")

            # ── ÉCRITURE TABLES PRINCIPALES (avant cleaning tasks) ──
            log.info("Ecriture tables principales Excel...")
            OUT_DIR.mkdir(parents=True, exist_ok=True)
            main_tables = {
                "MASTER_REF_HA_Listings":                 pd.DataFrame(rows_listings),
                "MASTER_FACT_HA_Reservations":             pd.DataFrame(rows_res),
                "MASTER_FACT_HA_ReservationDetails":       pd.DataFrame(rows_details),
                "MASTER_FACT_HA_ReservationFinanceFields": pd.DataFrame(rows_ff),
                "MASTER_FACT_HA_ReservationFees":          pd.DataFrame(rows_fees),
                "MASTER_CALC_HA_Payout":                   pd.DataFrame(rows_payout),
                "MASTER_CTRL_HA_Anomalies":                detector.to_df(),
            }
            for name, df in main_tables.items():
                write_excel(df, OUT_DIR / f"{name}.xlsx")
                log.info(f"  {name} : {len(df)} lignes")

            # ── CLEANING TASKS (non-bloquant) ─────────────────
            if skip_tasks:
                log.info("--skip-cleaning-tasks : taches menage ignorees.")
                tasks_statut = "SKIPPED"
            else:
                rows_tasks, tasks_statut = _extract_cleaning_tasks(client, DATE_FROM, detector, log)
                if tasks_statut != "OK":
                    log.warning("Taches menage incompletes — voir anomalie CLEANING_TASKS_EXTRACTION_INCOMPLETE.")

            write_excel(pd.DataFrame(rows_tasks),
                        OUT_DIR / "MASTER_FACT_HA_CleaningTasks_Discovery.xlsx")
            log.info(f"  MASTER_FACT_HA_CleaningTasks_Discovery : {len(rows_tasks)} lignes [{tasks_statut}]")

            # Réécrire anomalies avec éventuelles anomalies cleaning
            write_excel(detector.to_df(), OUT_DIR / "MASTER_CTRL_HA_Anomalies.xlsx")

            nb_bloq = detector.bloquants()
            if nb_bloq > 0 or tasks_statut not in ("OK", "SKIPPED"):
                statut_run = "PARTIAL_OK" if tasks_statut not in ("OK", "SKIPPED") else "TERMINE_AVEC_BLOQUANTS"
            else:
                statut_run = "TERMINE_OK"

    except Exception as fatal:
        log.error(f"Erreur fatale run : {type(fatal).__name__}: {fatal}")
        statut_run = "FAILED"

    finally:
        # ── MASTER_RUN_LOG — toujours écrit ──────────────────
        run_end  = datetime.now(timezone.utc)
        duree    = round((run_end - run_start).total_seconds(), 1)
        anom_df  = detector.to_df()
        nb_bloq  = detector.bloquants()
        nb_actrl = detector.a_controler()
        OUT_DIR.mkdir(parents=True, exist_ok=True)

        df_log = pd.DataFrame([{
            "run_id":                   run_id,
            "date_extraction":          run_start.isoformat(),
            "date_fin":                 run_end.isoformat(),
            "duree_secondes":           duree,
            "date_from":                DATE_FROM,
            "details_mode":             details_mode,
            "skip_cleaning_tasks":      str(skip_tasks),
            "limit":                    limit or "none",
            "nb_listings":              len(rows_listings),
            "nb_reservations_total_api":total if not only_tasks else 0,
            "nb_reservations_traitees": processed,
            "nb_reservations_sautees":  skipped,
            "nb_details_calls":         nb_details,
            "pct_details":              f"{100*nb_details//max(processed,1)}%",
            "nb_finance_fields":        len(rows_ff),
            "nb_fees":                  len(rows_fees),
            "nb_payout_calcules":       len(rows_payout),
            "nb_payout_NORMAL":         len([r for r in rows_payout if r.get("statut_calcul_payout") == "NORMAL"]),
            "nb_payout_INCOMPLET":      len([r for r in rows_payout if r.get("statut_calcul_payout") == "PAYOUT_INCOMPLET"]),
            "nb_payout_ABSENT":         len([r for r in rows_payout if r.get("statut_calcul_payout") == "PAYOUT_ABSENT"]),
            "nb_cleaning_tasks":        len(rows_tasks),
            "tasks_statut":             tasks_statut,
            "nb_anomalies_total":       len(anom_df),
            "nb_anomalies_BLOQUANT":    nb_bloq,
            "nb_anomalies_A_CONTROLER": nb_actrl,
            "statut_run":               statut_run,
            "lot1_validable":           "NON" if nb_bloq > 0 or statut_run == "FAILED"
                                        else "OUI_sous_reserve_controle",
            "commentaire":              f"Lot1 {details_mode} depuis {DATE_FROM}",
        }])
        try:
            write_excel(df_log, OUT_DIR / "MASTER_RUN_Log.xlsx")
        except Exception as e:
            log.error(f"Impossible d'ecrire MASTER_RUN_Log : {e}")

    # ── RÉSUMÉ CONSOLE (dans finally les vars sont disponibles) ──
    run_end_summary = datetime.now(timezone.utc)
    duree_s = round((run_end_summary - run_start).total_seconds(), 1)
    nb_bloq_s  = detector.bloquants()
    nb_actrl_s = detector.a_controler()
    log.info("=" * 60)
    log.info(f"RUN {statut_run} — {run_id}  ({duree_s}s)")
    log.info(f"  Listings           : {len(rows_listings)}")
    log.info(f"  Reservations API   : {total if not only_tasks else 'n/a'}")
    log.info(f"  Traitees           : {processed}")
    log.info(f"  Sautees (skip)     : {skipped}")
    log.info(f"  Appels detail      : {nb_details} ({100*nb_details//max(processed,1)}%)")
    log.info(f"  Finance fields     : {len(rows_ff)}")
    log.info(f"  Fees               : {len(rows_fees)}")
    log.info(f"  Payouts calcules   : {len(rows_payout)}")
    log.info(f"    NORMAL           : {len([r for r in rows_payout if r.get('statut_calcul_payout')=='NORMAL'])}")
    log.info(f"    INCOMPLET        : {len([r for r in rows_payout if r.get('statut_calcul_payout')=='PAYOUT_INCOMPLET'])}")
    log.info(f"    ABSENT           : {len([r for r in rows_payout if r.get('statut_calcul_payout')=='PAYOUT_ABSENT'])}")
    log.info(f"    A_CONTROLER      : {len([r for r in rows_payout if r.get('statut_calcul_payout')=='A_CONTROLER'])}")
    log.info(f"  Taches menage      : {len(rows_tasks)} [{tasks_statut}]")
    log.info(f"  Anomalies BLOQUANT : {nb_bloq_s}")
    log.info(f"  Anomalies A_CTRL   : {nb_actrl_s}")
    log.info("=" * 60)
    if nb_bloq_s > 0:
        log.warning(f"  ATTENTION : {nb_bloq_s} anomalie(s) BLOQUANTE(s) — Lot 1 NON validable.")
        log.warning("  Voir MASTER_CTRL_HA_Anomalies.xlsx")
    elif statut_run in ("TERMINE_OK", "PARTIAL_OK"):
        log.info("  Aucun bloquant. Verification humaine requise avant marquage FAIT.")
    log.info(f"  Tables dans : {OUT_DIR}")


if __name__ == "__main__":
    main()
