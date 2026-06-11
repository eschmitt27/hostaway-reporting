"""
Lot 4bis — Correctif : peuplement de MASTER_CALC_Reservations.xlsx
Décisions : D052, D053, D054, D055, D056, D057

Sources :
  - MASTER_FACT_HA_Reservations.xlsx   (02_TRAVAIL/Lot1_Hostaway/)
  - MASTER_CALC_HA_Payout.xlsx         (02_TRAVAIL/Lot1_Hostaway/)
  - MASTER_FACT_MAN_ReservationsHorsHostaway.xlsx (02_TRAVAIL/Lot4_ReservationsHH/)
  - REF_Setup.xlsm                     (01_SOURCES_BRUTES/REF_Setup/)

Cible :
  - 02_TRAVAIL/Lot4bis_TableCommune/MASTER_CALC_Reservations.xlsx
    onglets MASTER + VUE_FLUX (POWER_QUERY_CODE conservé)

Règles métier :
  S1  AIRBNB HA pur              → HOSTAWAY_AIRBNB,           HOSTAWAY_PAYOUT, IC, VALIDE
  S2  BOOKING HA pur             → HOSTAWAY_BOOKING,          HOSTAWAY_PAYOUT, IC, VALIDE
  S3  DIRECT HA + HH liée        → HOSTAWAY_DIRECT_HH,        MANUEL_HH,       du HH, du HH (ligne HA exclue)
  S4  VRBO HA + HH renseignée    → HOSTAWAY_VRBO_HH,          MANUEL_VRBO,     HC, du HH (ligne HA exclue)
  S5  VRBO HA sans HH            → HOSTAWAY_VRBO_A_CONTROLER, A_CONTROLER,     HC, A_CONTROLER
  S6  HH pure                    → MANUEL_HORS_HOSTAWAY,      MANUEL_HH,       du HH
  S7  ownerStay                  → OWNERSTAY_EXCLU,           NON_CONCERNE,    HR, EXCLU_RESULTAT

Cas non couvert par D054 (proposition validée) :
  DIRECT HA sans HH              → HOSTAWAY_DIRECT_HH, A_CONTROLER, HC, A_CONTROLER,
                                   code_anomalie=DIRECT_SANS_SAISIE_HH

Bloquants : RESERVATION_DOUBLON_HOSTAWAY_HH, RESERVATION_CALC_ID_DUPLIQUE,
            LOGEMENT_NON_MAPPE, LOGEMENT_MAPPING_MULTIPLE
"""

import hashlib
import datetime
import os
import sys
from collections import defaultdict, Counter

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.stderr.reconfigure(encoding="utf-8", errors="replace")

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# ---------------------------------------------------------------------------
# Chemins
# ---------------------------------------------------------------------------
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

PATH_HA_RES  = os.path.join(ROOT, "02_TRAVAIL", "Lot1_Hostaway", "MASTER_FACT_HA_Reservations.xlsx")
PATH_HA_PAY  = os.path.join(ROOT, "02_TRAVAIL", "Lot1_Hostaway", "MASTER_CALC_HA_Payout.xlsx")
PATH_HH_FACT = os.path.join(ROOT, "02_TRAVAIL", "Lot4_ReservationsHH", "MASTER_FACT_MAN_ReservationsHorsHostaway.xlsx")
PATH_REF     = os.path.join(ROOT, "01_SOURCES_BRUTES", "REF_Setup", "REF_Setup.xlsm")
PATH_TARGET  = os.path.join(ROOT, "02_TRAVAIL", "Lot4bis_TableCommune", "MASTER_CALC_Reservations.xlsx")

DATE_INTEGRATION = datetime.datetime.now().isoformat(timespec="seconds")
SOURCE_MODULE    = "lot4bis"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def row_hash(values):
    s = "|".join("" if v is None else str(v) for v in values)
    return hashlib.sha256(s.encode("utf-8")).hexdigest()[:16]


def date_to_str(val):
    if val is None:
        return None
    if isinstance(val, (datetime.date, datetime.datetime)):
        return str(val)[:10]
    return str(val)[:10]


def load_sheet(path, sheet_name):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    non_empty = [r for r in rows if any(c is not None for c in r)]
    if not non_empty:
        return [], []
    headers = list(non_empty[0])
    data = non_empty[1:]
    return headers, data


def rows_to_dicts(headers, data):
    return [dict(zip(headers, r)) for r in data]


def abort(msg):
    print(f"\n[BLOQUANT] {msg}")
    print("Script arrêté — MASTER_CALC_Reservations.xlsx NON modifié.")
    sys.exit(1)


# ---------------------------------------------------------------------------
# 1. Lecture des sources
# ---------------------------------------------------------------------------
print("=== Lot 4bis — Chargement réservations ===")
print(f"Date : {DATE_INTEGRATION}")
print()

print("[1/5] Lecture MASTER_FACT_HA_Reservations...")
h_res, d_res = load_sheet(PATH_HA_RES, "data")
res_dicts = rows_to_dicts(h_res, d_res)
print(f"      {len(res_dicts)} lignes")

print("[2/5] Lecture MASTER_CALC_HA_Payout...")
h_pay, d_pay = load_sheet(PATH_HA_PAY, "data")
pay_dicts = rows_to_dicts(h_pay, d_pay)
print(f"      {len(pay_dicts)} lignes")

print("[3/5] Lecture MASTER_FACT_MAN_ReservationsHorsHostaway...")
h_hh, d_hh = load_sheet(PATH_HH_FACT, "MASTER")
hh_dicts_raw = rows_to_dicts(h_hh, d_hh)
# Filtre : uniquement les lignes VALIDE (D053 — RESERVATION_HH_NON_VALIDE)
hh_valides = [r for r in hh_dicts_raw
              if r.get("statut_controle") == "VALIDE"
              and r.get("reservation_hh_id") is not None
              and not str(r.get("reservation_hh_id", "")).startswith("[")]
hh_non_valides = len(hh_dicts_raw) - len(hh_valides)
print(f"      {len(hh_dicts_raw)} lignes brutes, {len(hh_valides)} VALIDE retenues")

print("[4/5] Lecture REF_Setup (Mapping + Logements)...")
h_map, d_map = load_sheet(PATH_REF, "REF_Mapping_Logements")
h_log, d_log = load_sheet(PATH_REF, "REF_Logements")
map_dicts = rows_to_dicts(h_map, d_map)
log_dicts  = rows_to_dicts(h_log, d_log)
print(f"      {len(map_dicts)} mappings, {len(log_dicts)} logements")

# Index mapping : listingMapId → logement_id (Hostaway, actif=OUI)
ha_map_index = defaultdict(list)
for r in map_dicts:
    if r.get("source") == "Hostaway" and r.get("champ_source") == "listingMapId" and r.get("actif") == "OUI":
        ha_map_index[r["valeur_source"]].append(r["logement_id"])

# Index logements : logement_id → {proprietaire_id, actif}
log_index = {r["logement_id"]: r for r in log_dicts}

print("[5/5] Index payout par reservation_id...")
pay_index = {r["reservation_id"]: r for r in pay_dicts}

# ---------------------------------------------------------------------------
# 2. Index HH : reservation_id_hostaway → hh ligne (pour anti-doublon)
# ---------------------------------------------------------------------------
hh_by_ha_id = {}   # reservation_id_hostaway → ligne HH (S3/S4)
hh_pure = []       # lignes HH sans reservation_id_hostaway (S6)

for hh in hh_valides:
    ha_id = hh.get("reservation_id_hostaway")
    if ha_id:
        if ha_id in hh_by_ha_id:
            abort(
                f"RESERVATION_DOUBLON_HOSTAWAY_HH — reservation_id_hostaway={ha_id} "
                f"rattaché à 2+ lignes HH actives."
            )
        hh_by_ha_id[ha_id] = hh
    else:
        hh_pure.append(hh)

print(f"\nIndex HH : {len(hh_by_ha_id)} liens HA→HH, {len(hh_pure)} HH pures")

# ---------------------------------------------------------------------------
# 3. Construction du MASTER
# ---------------------------------------------------------------------------
print("\n--- Construction MASTER ---")

master_rows = []
counters_by_month_ha  = defaultdict(int)   # (mois, "HA") → compteur
counters_by_month_hh  = defaultdict(int)   # (mois, "HH") → compteur

ha_ids_linked_to_hh = set(hh_by_ha_id.keys())

stats = Counter()

def resolve_logement(listing_map_id, date_arrivee_str=None, date_depart_str=None):
    """Retourne (logement_id, proprietaire_id, anomalie_code, anomalie_msg) ou abort si BLOQUANT.

    Règle date_sortie_gestion (validée 2026-06-11) :
      date_arrivee < date_sortie AND date_depart <= date_sortie → VALIDE (réservation historique)
      date_arrivee < date_sortie AND date_depart >  date_sortie → A_CONTROLER SEJOUR_CHEVAUCHE_SORTIE_GESTION
      date_arrivee >= date_sortie                               → A_CONTROLER LOGEMENT_INACTIF
      date_sortie absente ET actif=NON                          → A_CONTROLER LOGEMENT_INACTIF
    """
    matches = ha_map_index.get(listing_map_id, [])
    if not matches:
        abort(f"LOGEMENT_NON_MAPPE — listingMapId={listing_map_id} absent de REF_Mapping_Logements.")
    if len(matches) > 1:
        abort(f"LOGEMENT_MAPPING_MULTIPLE — listingMapId={listing_map_id} → {matches}.")
    logement_id = matches[0]
    log_row = log_index.get(logement_id, {})
    proprietaire_id = log_row.get("proprietaire_id")
    actif = log_row.get("actif")
    date_sortie_raw = log_row.get("date_sortie_gestion")
    ano_code = None
    ano_msg = None

    if actif == "NON":
        if date_sortie_raw is not None:
            try:
                d_sortie = (date_sortie_raw.date()
                            if isinstance(date_sortie_raw, datetime.datetime)
                            else date_sortie_raw)
                d_arr = (datetime.date.fromisoformat(str(date_arrivee_str)[:10])
                         if date_arrivee_str else None)
                d_dep = (datetime.date.fromisoformat(str(date_depart_str)[:10])
                         if date_depart_str else None)

                if d_arr is None:
                    ano_code = "LOGEMENT_INACTIF"
                    ano_msg  = f"{logement_id} actif=NON, date arrivée absente"
                elif d_arr >= d_sortie:
                    ano_code = "LOGEMENT_INACTIF"
                    ano_msg  = f"{logement_id} arrivée {d_arr} >= sortie {d_sortie}"
                elif d_dep is not None and d_dep > d_sortie:
                    ano_code = "SEJOUR_CHEVAUCHE_SORTIE_GESTION"
                    ano_msg  = f"{logement_id} départ {d_dep} > sortie {d_sortie}"
                # sinon : arrivée ET départ avant sortie → pas d'anomalie logement
            except (ValueError, TypeError):
                ano_code = "LOGEMENT_INACTIF"
                ano_msg  = f"{logement_id} actif=NON (dates non parsables)"
        else:
            # date_sortie_gestion absente → logement inactif sans date → flaguer
            ano_code = "LOGEMENT_INACTIF"
            ano_msg  = f"{logement_id} actif=NON, date_sortie_gestion absente"

    if not proprietaire_id:
        ano_code = "PROPRIETAIRE_ABSENT"
        ano_msg  = f"Logement {logement_id} sans proprietaire_id"

    return logement_id, proprietaire_id, ano_code, ano_msg


def make_row_ha(res, payout, source_val, source_montant, montant_retenu,
                code_impact, statut_controle, niveau_anomalie, code_anomalie, commentaire,
                logement_id, proprietaire_id):
    check_in = res.get("checkInDate")
    mois = str(check_in)[:7] if check_in else "0000-00"
    branch = "HA"
    counters_by_month_ha[(mois, branch)] += 1
    n = counters_by_month_ha[(mois, branch)]
    reservation_calc_id = f"RES-{mois}-HA-{n:03d}"

    impact_reel  = "A_CONTROLER" if code_impact not in ("IC", "HC", "HR") else ("NON" if code_impact == "HR" else "OUI")
    impact_compta = "A_CONTROLER" if code_impact not in ("IC", "HC", "HR") else ("OUI" if code_impact == "IC" else "NON")
    if statut_controle == "EXCLU_RESULTAT":
        impact_reel   = "NON"
        impact_compta = "NON"

    hash_keys = [reservation_calc_id, source_val, res.get("reservation_id"), mois,
                 logement_id, montant_retenu, code_impact]

    return {
        "reservation_calc_id":     reservation_calc_id,
        "ROW_HASH":                row_hash(hash_keys),
        "source":                  source_val,
        "reservation_id_hostaway": res.get("reservation_id"),
        "reservation_hh_id":       None,
        "mois":                    mois,
        "logement_id":             logement_id,
        "proprietaire_id":         proprietaire_id,
        "date_arrivee":            date_to_str(res.get("checkInDate")),
        "date_depart":             date_to_str(res.get("checkOutDate")),
        "nuits":                   res.get("nights"),
        "montant_retenu":          montant_retenu,
        "source_montant":          source_montant,
        "code_impact":             code_impact,
        "impact_resultat_reel":    impact_reel,
        "impact_resultat_comptable": impact_compta,
        "statut_controle":         statut_controle,
        "niveau_anomalie":         niveau_anomalie,
        "code_anomalie":           code_anomalie,
        "commentaire":             commentaire,
        "source_module":           SOURCE_MODULE,
        "source_table":            "MASTER_FACT_HA_Reservations",
        "source_pk":               str(res.get("reservation_id")),
        "date_integration":        DATE_INTEGRATION,
    }


def make_row_hh(hh, src_override=None):
    mois = str(hh.get("mois") or "0000-00")[:7]
    branch = "HH"
    counters_by_month_hh[(mois, branch)] += 1
    n = counters_by_month_hh[(mois, branch)]
    reservation_calc_id = f"RES-{mois}-HH-{n:03d}"
    source_val = src_override or "MANUEL_HORS_HOSTAWAY"
    code_impact = hh.get("code_impact") or "HC"
    statut = hh.get("statut_controle") or "A_CONTROLER"
    niveau = hh.get("niveau_anomalie") or "INFO"
    montant = hh.get("total_percu") or 0

    impact_reel  = "NON" if code_impact == "HR" else "OUI"
    impact_compta = "OUI" if code_impact == "IC" else "NON"

    hash_keys = [reservation_calc_id, source_val, hh.get("reservation_hh_id"), mois,
                 hh.get("logement_id"), montant, code_impact]

    source_montant = "MANUEL_VRBO" if source_val == "HOSTAWAY_VRBO_HH" else "MANUEL_HH"

    return {
        "reservation_calc_id":     reservation_calc_id,
        "ROW_HASH":                row_hash(hash_keys),
        "source":                  source_val,
        "reservation_id_hostaway": hh.get("reservation_id_hostaway"),
        "reservation_hh_id":       hh.get("reservation_hh_id"),
        "mois":                    mois,
        "logement_id":             hh.get("logement_id"),
        "proprietaire_id":         hh.get("proprietaire_id"),
        "date_arrivee":            date_to_str(hh.get("date_arrivee")),
        "date_depart":             date_to_str(hh.get("date_depart")),
        "nuits":                   hh.get("nuits"),
        "montant_retenu":          montant,
        "source_montant":          source_montant,
        "code_impact":             code_impact,
        "impact_resultat_reel":    impact_reel,
        "impact_resultat_comptable": impact_compta,
        "statut_controle":         statut,
        "niveau_anomalie":         niveau,
        "code_anomalie":           hh.get("code_anomalie"),
        "commentaire":             hh.get("commentaire"),
        "source_module":           SOURCE_MODULE,
        "source_table":            "MASTER_FACT_MAN_ReservationsHorsHostaway",
        "source_pk":               str(hh.get("reservation_hh_id")),
        "date_integration":        DATE_INTEGRATION,
    }


# --- Branche HA ---
for res in res_dicts:
    rid = res.get("reservation_id")
    channel = res.get("channel_type") or ""
    sf      = res.get("source_financiere") or ""
    incl    = res.get("inclure_resultat")
    status  = res.get("status") or ""
    total   = res.get("totalPrice") or 0
    payout  = pay_index.get(rid)

    # S7 — ownerStay
    if incl == "NON" or status == "ownerStay":
        logement_id, proprietaire_id, ano_code, ano_msg = resolve_logement(
            res["listingMapId"],
            date_arrivee_str=date_to_str(res.get("checkInDate")),
            date_depart_str=date_to_str(res.get("checkOutDate")),
        )
        stats["OWNERSTAY_EXCLU"] += 1
        commentaire = "Hostaway ownerStay — exclu résultat"
        if status == "modified":
            commentaire += " | Hostaway status=modified"
        row = make_row_ha(res, payout, "OWNERSTAY_EXCLU", "NON_CONCERNE", 0,
                          "HR", "EXCLU_RESULTAT", "INFO", None, commentaire,
                          logement_id, proprietaire_id)
        master_rows.append(row)
        continue

    # Ligne HA déjà couverte par HH (S3/S4) → exclure
    if rid in ha_ids_linked_to_hh:
        stats["HA_EXCLU_PAR_HH"] += 1
        continue

    logement_id, proprietaire_id, ano_code, ano_msg = resolve_logement(
        res["listingMapId"],
        date_arrivee_str=date_to_str(res.get("checkInDate")),
        date_depart_str=date_to_str(res.get("checkOutDate")),
    )

    statut_logement  = None
    niveau_logement  = None
    if ano_code in ("LOGEMENT_INACTIF", "SEJOUR_CHEVAUCHE_SORTIE_GESTION", "PROPRIETAIRE_ABSENT"):
        statut_logement = "A_CONTROLER"
        niveau_logement = "A_CONTROLER"

    # Payout info
    payout_calcule  = payout.get("payout_calcule") if payout else None
    statut_payout   = payout.get("statut_calcul_payout") if payout else None

    # Commentaire modified
    extra_comment = ""
    if status == "modified":
        extra_comment = f" | Hostaway status=modified, payout {statut_payout or 'absent'}"

    # S1 — AIRBNB
    if channel == "AIRBNB":
        if statut_payout == "NORMAL" and statut_logement is None:
            commentaire = f"Airbnb HA normal{extra_comment}"
            row = make_row_ha(res, payout, "HOSTAWAY_AIRBNB", "HOSTAWAY_PAYOUT",
                              payout_calcule, "IC", "VALIDE", "INFO", None,
                              commentaire, logement_id, proprietaire_id)
            stats["S1_AIRBNB_VALIDE"] += 1
        else:
            commentaire = f"Airbnb HA — payout={statut_payout}{extra_comment}"
            code_ano = ano_code or f"PAYOUT_{statut_payout}"
            row = make_row_ha(res, payout, "HOSTAWAY_AIRBNB", "HOSTAWAY_PAYOUT",
                              payout_calcule or 0, "IC",
                              statut_logement or "A_CONTROLER",
                              niveau_logement or "A_CONTROLER",
                              code_ano, commentaire, logement_id, proprietaire_id)
            stats["S1_AIRBNB_A_CONTROLER"] += 1
        master_rows.append(row)

    # S2 — BOOKING
    elif channel == "BOOKING":
        if statut_payout == "NORMAL" and statut_logement is None:
            commentaire = f"Booking HA normal{extra_comment}"
            row = make_row_ha(res, payout, "HOSTAWAY_BOOKING", "HOSTAWAY_PAYOUT",
                              payout_calcule, "IC", "VALIDE", "INFO", None,
                              commentaire, logement_id, proprietaire_id)
            stats["S2_BOOKING_VALIDE"] += 1
        else:
            commentaire = f"Booking HA — payout={statut_payout}{extra_comment}"
            code_ano = ano_code or f"PAYOUT_{statut_payout}"
            row = make_row_ha(res, payout, "HOSTAWAY_BOOKING", "HOSTAWAY_PAYOUT",
                              payout_calcule or 0, "IC",
                              statut_logement or "A_CONTROLER",
                              niveau_logement or "A_CONTROLER",
                              code_ano, commentaire, logement_id, proprietaire_id)
            stats["S2_BOOKING_A_CONTROLER"] += 1
        master_rows.append(row)

    # S5 — VRBO sans HH
    elif channel == "VRBO":
        commentaire = f"VRBO Hostaway sans saisie HH — montant en attente{extra_comment}"
        row = make_row_ha(res, payout, "HOSTAWAY_VRBO_A_CONTROLER", "A_CONTROLER",
                          0, "HC", "A_CONTROLER", "A_CONTROLER",
                          "VRBO_MONTANT_NON_RENSEIGNE",
                          commentaire, logement_id, proprietaire_id)
        master_rows.append(row)
        stats["S5_VRBO_A_CONTROLER"] += 1

    # DIRECT sans HH (cas non couvert D054 — proposition validée)
    elif channel == "DIRECT":
        commentaire = f"Réservation directe Hostaway sans saisie hors Hostaway en face{extra_comment}"
        row = make_row_ha(res, payout, "HOSTAWAY_DIRECT_HH", "A_CONTROLER",
                          0, "HC", "A_CONTROLER", "A_CONTROLER",
                          "DIRECT_SANS_SAISIE_HH",
                          commentaire, logement_id, proprietaire_id)
        master_rows.append(row)
        stats["DIRECT_SANS_SAISIE_HH"] += 1

    else:
        commentaire = f"Canal inconnu: {channel}"
        row = make_row_ha(res, payout, f"INCONNU_{channel}", "A_CONTROLER",
                          0, "HC", "A_CONTROLER", "A_CONTROLER",
                          "CANAL_INCONNU", commentaire, logement_id, proprietaire_id)
        master_rows.append(row)
        stats["CANAL_INCONNU"] += 1


# --- Branche HH — lignes liées à HA (S3/S4) ---
for ha_id, hh in hh_by_ha_id.items():
    canal = (hh.get("canal_id") or "").upper()
    src = "HOSTAWAY_VRBO_HH" if "VRBO" in canal else "HOSTAWAY_DIRECT_HH"
    row = make_row_hh(hh, src_override=src)
    master_rows.append(row)
    stats["S3_S4_DIRECT_VRBO_HH"] += 1

# --- Branche HH — lignes pures (S6) ---
for hh in hh_pure:
    row = make_row_hh(hh)
    master_rows.append(row)
    stats["S6_HH_PURE"] += 1

# HH non-valides signalées
if hh_non_valides > 0:
    print(f"  AVERTISSEMENT : {hh_non_valides} lignes HH non-VALIDE exclues (RESERVATION_HH_NON_VALIDE)")

print(f"\nTotal lignes construites : {len(master_rows)}")

# ---------------------------------------------------------------------------
# 4. Contrôle BLOQUANT : RESERVATION_CALC_ID_DUPLIQUE
# ---------------------------------------------------------------------------
calc_ids = [r["reservation_calc_id"] for r in master_rows]
dup = [cid for cid, cnt in Counter(calc_ids).items() if cnt > 1]
if dup:
    abort(f"RESERVATION_CALC_ID_DUPLIQUE — {len(dup)} IDs dupliqués : {dup[:5]}")

# ---------------------------------------------------------------------------
# 5. Contrôle : aucune donnée personnelle voyageur
# ---------------------------------------------------------------------------
PERSONAL_FIELDS = {"guestname", "guest_name", "guest_email", "email", "telephone",
                   "phone", "adresse_voyageur", "guest_comment", "commentaire_voyageur",
                   "iban", "nom_complet", "prenom"}
all_cols = {k.lower() for r in master_rows for k in r.keys()}
found_personal = PERSONAL_FIELDS & all_cols
if found_personal:
    abort(f"DONNEE_PERSONNELLE_DETECTEE — colonnes: {found_personal}. Arrêt avant écriture.")

print("[OK] Aucune donnée personnelle voyageur dans les colonnes produites.")

# ---------------------------------------------------------------------------
# 6. Définition VUE_FLUX (VALIDE + impact_reel=OUI + montant≠0)
# ---------------------------------------------------------------------------
vue_rows = [r for r in master_rows
            if r["statut_controle"] == "VALIDE"
            and r["impact_resultat_reel"] == "OUI"
            and (r["montant_retenu"] or 0) != 0]

print(f"VUE_FLUX : {len(vue_rows)} lignes (VALIDE + impact_reel=OUI + montant≠0)")

# ---------------------------------------------------------------------------
# 7. Écriture dans MASTER_CALC_Reservations.xlsx
# ---------------------------------------------------------------------------
HEADERS = [
    "reservation_calc_id", "ROW_HASH", "source", "reservation_id_hostaway",
    "reservation_hh_id", "mois", "logement_id", "proprietaire_id",
    "date_arrivee", "date_depart", "nuits", "montant_retenu", "source_montant",
    "code_impact", "impact_resultat_reel", "impact_resultat_comptable",
    "statut_controle", "niveau_anomalie", "code_anomalie", "commentaire",
    "source_module", "source_table", "source_pk", "date_integration",
]
assert len(HEADERS) == 24, f"Attendu 24 colonnes, trouvé {len(HEADERS)}"

print(f"\n[6/6] Écriture dans {PATH_TARGET}...")

# Load workbook preserving POWER_QUERY_CODE
wb = load_workbook(PATH_TARGET)

# Rebuild MASTER sheet
if "MASTER" in wb.sheetnames:
    del wb["MASTER"]
ws_master = wb.create_sheet("MASTER", 0)

ws_master.append(HEADERS)
for r in master_rows:
    ws_master.append([r.get(h) for h in HEADERS])

# Rebuild VUE_FLUX sheet
if "VUE_FLUX" in wb.sheetnames:
    del wb["VUE_FLUX"]
ws_vue = wb.create_sheet("VUE_FLUX", 1)

ws_vue.append(HEADERS)
for r in vue_rows:
    ws_vue.append([r.get(h) for h in HEADERS])

wb.save(PATH_TARGET)
wb.close()
print("Fichier sauvegardé.")

# ---------------------------------------------------------------------------
# 8. Rapport final
# ---------------------------------------------------------------------------
print("\n" + "=" * 60)
print("RAPPORT LOT 4BIS — MASTER_CALC_Reservations")
print("=" * 60)

print(f"\nLignes MASTER      : {len(master_rows)}")
print(f"Lignes VUE_FLUX    : {len(vue_rows)}")

print("\nRépartition par source :")
src_count = Counter(r["source"] for r in master_rows)
for s, c in sorted(src_count.items()):
    print(f"  {s:<35} {c}")

print("\nRépartition par statut_controle :")
statut_count = Counter(r["statut_controle"] for r in master_rows)
for s, c in sorted(statut_count.items()):
    print(f"  {s:<20} {c}")

print("\nRépartition par code_impact :")
ci_count = Counter(r["code_impact"] for r in master_rows)
for s, c in sorted(ci_count.items()):
    print(f"  {s:<10} {c}")

print("\nAnomalies A_CONTROLER :")
ano_count = Counter(r["code_anomalie"] for r in master_rows if r["code_anomalie"])
for s, c in sorted(ano_count.items()):
    print(f"  {s:<40} {c}")

print(f"\nDIRECT_SANS_SAISIE_HH              : {stats['DIRECT_SANS_SAISIE_HH']}")
print(f"VRBO_MONTANT_NON_RENSEIGNE         : {stats['S5_VRBO_A_CONTROLER']}")
print(f"ownerStay EXCLU_RESULTAT           : {stats['OWNERSTAY_EXCLU']}")

print("\n[SECURITE] Vérification données personnelles voyageur :")
print("  Colonnes présentes dans MASTER :", HEADERS)
print("  Intersection avec champs personnels :", found_personal or "AUCUNE ✓")

print("\nBloquants détectés : 0")
print("Script terminé avec succès.")
print("RAPPEL : aucun git add/commit sans validation humaine.")
