"""
lot9_construire_flux.py
Table de flux unifiée MASTER_CALC_Flux — Lot 9

Sources :
  - MASTER_CALC_Reservations VUE_FLUX   → TYPE_FLUX_017 PRODUIT IC
  - MASTER_FACT_MEN_MenagesExternes     → TYPE_FLUX_014 CHARGE IC  (VALIDE seulement)
  - BANQUE_LOT8_IMPORT NORM_Banque      → TYPE_FLUX_016 CHARGE IC  (TYPE_FLUX_016 + VALIDE seulement)

Sécurité bancaire : aucune donnée brute (libellé, compte, IBAN) dans la sortie.

Règles anti-doublon (composite key) : source_module + source_table + source_pk + type_flux_id + sens
Convention montant : toujours positif. Sens explicite : PRODUIT / CHARGE / NEUTRALISATION.
"""

import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.stderr.reconfigure(encoding='utf-8', errors='replace')

import os
import datetime
import hashlib
import openpyxl
from openpyxl.styles import PatternFill, Font

# ── CHEMINS ───────────────────────────────────────────────────────────────────
ROOT    = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC_RES = os.path.join(ROOT, '02_TRAVAIL', 'Lot4quater_SourceResolue', 'MASTER_CALC_Reservations_Resolues.xlsx')  # source résolue open/closed (lot4quater)
SRC_MEN = os.path.join(ROOT, '02_TRAVAIL', 'Lot6c_MenagesExternes', 'MASTER_FACT_MEN_MenagesExternes.xlsx')
SRC_BNQ = os.path.join(ROOT, '02_TRAVAIL', 'Lot8_Banque',           'BANQUE_LOT8_IMPORT.xlsx')
# Lot 9 correctif (2026-06-14) : ingestion charges (Lot 3) + ménages internes M04 (Lot 6b)
SRC_CHG = os.path.join(ROOT, '02_TRAVAIL', 'Lot3_Charges',          'MASTER_FACT_MAN_Charges.xlsx')
SRC_M04 = os.path.join(ROOT, '02_DONNEES_NORMALISEES', 'menages',   'M04_MENAGES_PowerQuery.xlsx')
OUT_DIR = os.path.join(ROOT, '02_TRAVAIL', 'Lot9_FluxUnifie')
OUT_FILE = os.path.join(OUT_DIR, 'MASTER_CALC_Flux.xlsx')

DATE_INTEGRATION = datetime.datetime.now().strftime('%Y-%m-%dT%H:%M:%S')

# 22 colonnes officielles MASTER_CALC_Flux
COLS = [
    'flux_id', 'ROW_HASH',
    'source_module', 'source_table', 'source_pk',
    'date_flux', 'mois',
    'logement_id', 'proprietaire_id', 'associe_id',
    'type_flux_id', 'sens', 'montant', 'code_impact',
    'inclure_resultat_reel', 'inclure_resultat_comptable', 'inclure_resultat_hors_compta',
    'statut_controle', 'niveau_anomalie', 'code_anomalie', 'commentaire',
    'date_integration',
]

# Champs bancaires bruts interdits dans la sortie (sécurité)
BANK_FORBIDDEN = {'libelle', 'libelle_brut', 'compte_id', 'iban'}

# ── HELPERS ───────────────────────────────────────────────────────────────────
def impact_flags(code_impact):
    return {
        'IC': ('OUI', 'OUI', 'NON'),
        'HC': ('OUI', 'NON', 'OUI'),
        'HR': ('NON', 'NON', 'NON'),
    }.get(code_impact, ('NON', 'NON', 'NON'))


def make_row_hash(source_module, source_table, source_pk, type_flux_id, sens):
    s = f'{source_module}|{source_table}|{source_pk}|{type_flux_id}|{sens}'
    return hashlib.sha256(s.encode('utf-8')).hexdigest()[:16]


def to_date_str(v):
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    s = str(v)
    return s[:10] if len(s) >= 10 else s


def to_mois(v):
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime('%Y-%m')
    return str(v)[:7]


def mois_to_first_day(mois_str):
    if mois_str and len(mois_str) == 7:
        return f'{mois_str}-01'
    return mois_str


def load_sheet(path, sheet_name):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
    wb.close()
    if not rows:
        return []
    h = list(rows[0])
    return [dict(zip(h, r)) for r in rows[1:]]


# ── CTR-9-001 : fichiers sources existent ────────────────────────────────────
missing = []
for p, n in [(SRC_RES, 'MASTER_CALC_Reservations'), (SRC_MEN, 'MASTER_FACT_MEN_MenagesExternes'), (SRC_BNQ, 'BANQUE_LOT8_IMPORT')]:
    if not os.path.exists(p):
        missing.append(f'{n} ({p})')
if missing:
    for m in missing:
        print(f'BLOQUANT [CTR-9-001] Source manquante : {m}')
    sys.exit(1)

# ── CHARGEMENT ────────────────────────────────────────────────────────────────
print('Chargement VUE_FLUX...')
vue_flux = load_sheet(SRC_RES, 'VUE_FLUX')
print(f'  {len(vue_flux)} lignes')

print('Chargement MenagesExternes MASTER...')
men_all    = load_sheet(SRC_MEN, 'MASTER')
men_valide = [r for r in men_all if r.get('statut_controle') == 'VALIDE']
print(f'  {len(men_valide)} VALIDE / {len(men_all)} total')

print('Chargement NORM_Banque...')
bnq_all    = load_sheet(SRC_BNQ, 'NORM_Banque')
bnq_valide = [r for r in bnq_all if r.get('type_flux_id') == 'TYPE_FLUX_016' and r.get('statut_controle') == 'VALIDE']
print(f'  {len(bnq_valide)} TYPE_FLUX_016 VALIDE / {len(bnq_all)} total')


def _is_placeholder_id(v):
    """Ligne d'instruction Power Query / formule, pas une vraie clé métier."""
    if v is None:
        return True
    s = str(v).strip()
    return s == '' or s[0] in '#[<←-*'


# Charges (Lot 3) — VALIDE seulement, hors lignes placeholder Power Query
print('Chargement MASTER_FACT_MAN_Charges...')
chg_all = load_sheet(SRC_CHG, 'MASTER') if os.path.exists(SRC_CHG) else []
chg_valide = [r for r in chg_all
              if r.get('statut_controle') == 'VALIDE'
              and not _is_placeholder_id(r.get('charge_id'))]
print(f'  {len(chg_valide)} VALIDE / {len(chg_all)} total')

# Ménages internes M04 (Lot 6b) — VALIDE seulement, HC obligatoire, hors placeholder
print('Chargement M04_MENAGES_PowerQuery MASTER...')
m04_all = load_sheet(SRC_M04, 'MASTER') if os.path.exists(SRC_M04) else []
m04_valide = [r for r in m04_all
              if r.get('statut_controle') == 'VALIDE'
              and not _is_placeholder_id(r.get('menage_calc_id'))]
print(f'  {len(m04_valide)} VALIDE / {len(m04_all)} total')

# ── CTR-9-002 : VUE_FLUX non vide ────────────────────────────────────────────
if not vue_flux:
    print('BLOQUANT [CTR-9-002] VUE_FLUX vide')
    sys.exit(1)

# ── CTR-9-003 : Volume VUE_FLUX ≥ 1000 ───────────────────────────────────────
if len(vue_flux) < 1000:
    print(f'BLOQUANT [CTR-9-003] VUE_FLUX volume suspect : {len(vue_flux)} lignes (attendu >= 1000)')
    sys.exit(1)

# ── CTR-9-010 (anticipé) : sécurité colonnes sortie ─────────────────────────
collision = BANK_FORBIDDEN & set(COLS)
if collision:
    print(f'BLOQUANT [CTR-9-010] Colonnes sensibles dans COLS : {collision}')
    sys.exit(1)

# ── CONSTRUCTION DES FLUX ─────────────────────────────────────────────────────
flux_rows  = []
seen_keys  = set()
doublons   = []
counters   = {}  # module_code -> counter

def add_flux(module_code, source_table, source_pk, date_flux, mois,
             logement_id, proprietaire_id, associe_id,
             type_flux_id, sens, montant, code_impact,
             statut_controle, commentaire,
             niveau_anomalie='INFO', code_anomalie=None):

    SOURCE_MODULE = 'lot9'

    key = (SOURCE_MODULE, source_table, str(source_pk) if source_pk is not None else '', type_flux_id, sens)
    if key in seen_keys:
        doublons.append(key)
        return
    seen_keys.add(key)

    counters[module_code] = counters.get(module_code, 0) + 1
    n = counters[module_code]

    reel, compta, hors = impact_flags(code_impact)

    d = {
        'flux_id':                      f'FLUX-{mois}-{code_impact}-{module_code}-{n:04d}',
        'ROW_HASH':                     make_row_hash(SOURCE_MODULE, source_table, str(source_pk) if source_pk is not None else '', type_flux_id, sens),
        'source_module':                SOURCE_MODULE,
        'source_table':                 source_table,
        'source_pk':                    str(source_pk) if source_pk is not None else None,
        'date_flux':                    date_flux,
        'mois':                         mois,
        'logement_id':                  logement_id,
        'proprietaire_id':              proprietaire_id,
        'associe_id':                   associe_id,
        'type_flux_id':                 type_flux_id,
        'sens':                         sens,
        'montant':                      abs(float(montant)) if montant is not None else None,
        'code_impact':                  code_impact,
        'inclure_resultat_reel':        reel,
        'inclure_resultat_comptable':   compta,
        'inclure_resultat_hors_compta': hors,
        'statut_controle':              statut_controle,
        'niveau_anomalie':              niveau_anomalie,
        'code_anomalie':                code_anomalie,
        'commentaire':                  commentaire,
        'date_integration':             DATE_INTEGRATION,
    }
    flux_rows.append(d)


# 1. Réservations — VUE_FLUX triée par reservation_calc_id pour déterminisme
print('\nModule RES — Réservations Hostaway...')
vue_flux_sorted = sorted(vue_flux, key=lambda r: r.get('reservation_calc_id') or '')
for r in vue_flux_sorted:
    date_flux = to_date_str(r.get('date_arrivee'))
    mois      = r.get('mois')
    if date_flux is None:
        date_flux = mois_to_first_day(mois)

    add_flux(
        module_code    = 'RES',
        source_table   = 'MASTER_CALC_Reservations_VUE_FLUX',
        source_pk      = r.get('reservation_calc_id'),
        date_flux      = date_flux,
        mois           = mois,
        logement_id    = r.get('logement_id'),
        proprietaire_id= r.get('proprietaire_id'),
        associe_id     = None,
        type_flux_id   = 'TYPE_FLUX_017',
        sens           = 'PRODUIT',
        montant        = r.get('montant_retenu'),
        code_impact    = r.get('code_impact', 'IC'),
        statut_controle= r.get('statut_controle', 'A_CONTROLER'),
        commentaire    = r.get('commentaire'),
    )
res_count = counters.get('RES', 0)
print(f'  {res_count} flux RES')


# 2. Ménages externes — VALIDE seulement
print('\nModule MEN — Ménages externes...')
men_sorted = sorted(men_valide, key=lambda r: r.get('source_pk') or r.get('menage_externe_id') or '')
for r in men_sorted:
    date_flux = to_date_str(r.get('date_facture') or r.get('date_menage'))
    mois      = r.get('mois')
    if date_flux is None:
        date_flux = mois_to_first_day(mois)

    add_flux(
        module_code    = 'MEN',
        source_table   = 'MASTER_FACT_MEN_MenagesExternes',
        source_pk      = r.get('source_pk') or r.get('menage_externe_id'),
        date_flux      = date_flux,
        mois           = mois,
        logement_id    = r.get('logement_id'),
        proprietaire_id= r.get('proprietaire_id'),
        associe_id     = None,
        type_flux_id   = r.get('type_flux_id', 'TYPE_FLUX_014'),
        sens           = r.get('sens', 'CHARGE'),
        montant        = r.get('montant_ligne_ttc'),
        code_impact    = r.get('code_impact', 'IC'),
        statut_controle= 'VALIDE',
        commentaire    = 'Ménage externe validé Lot 6c',
    )
men_count = counters.get('MEN', 0)
print(f'  {men_count} flux MEN')


# 3. Banque — TYPE_FLUX_016 VALIDE uniquement, commentaire générique
print('\nModule BNQ — Frais bancaires...')
bnq_sorted = sorted(bnq_valide, key=lambda r: r.get('mouvement_id') or '')
for r in bnq_sorted:
    date_op   = r.get('date_operation')
    date_flux = to_date_str(date_op)
    mois      = to_mois(date_op)

    add_flux(
        module_code    = 'BNQ',
        source_table   = 'BANQUE_LOT8_IMPORT_NORM_Banque',
        source_pk      = r.get('mouvement_id'),
        date_flux      = date_flux,
        mois           = mois,
        logement_id    = None,
        proprietaire_id= None,
        associe_id     = None,
        type_flux_id   = 'TYPE_FLUX_016',
        sens           = 'CHARGE',
        montant        = r.get('montant'),
        code_impact    = r.get('code_impact', 'IC'),
        statut_controle= 'VALIDE',
        commentaire    = 'Frais bancaires validés Lot 8',
    )
bnq_count = counters.get('BNQ', 0)
print(f'  {bnq_count} flux BNQ')


# 4. Charges (Lot 3) — VALIDE seulement. sens/code_impact/type portés par la ligne.
print('\nModule CHG — Charges...')
chg_sorted = sorted(chg_valide, key=lambda r: str(r.get('charge_id') or ''))
for r in chg_sorted:
    date_flux = to_date_str(r.get('date_charge'))
    mois      = r.get('mois') or to_mois(r.get('date_charge'))
    if date_flux is None:
        date_flux = mois_to_first_day(mois)

    add_flux(
        module_code    = 'CHG',
        source_table   = 'MASTER_FACT_MAN_Charges',
        source_pk      = r.get('charge_id'),
        date_flux      = date_flux,
        mois           = mois,
        logement_id    = r.get('logement_id'),
        proprietaire_id= r.get('proprietaire_id'),
        associe_id     = r.get('associe_id'),
        type_flux_id   = r.get('type_flux_id'),
        sens           = r.get('sens') or r.get('sens_flux') or 'CHARGE',
        montant        = r.get('montant'),
        code_impact    = r.get('code_impact', 'IC'),
        statut_controle= 'VALIDE',
        commentaire    = r.get('commentaire'),
    )
chg_count = counters.get('CHG', 0)
print(f'  {chg_count} flux CHG')


# 5. Ménages internes M04 (TYPE_FLUX_013) — ANALYTIQUE UNIQUEMENT (D105 révisée).
# NE PAS injecter le coût MO interne réel comme charge résultat : il ne sert qu'au
# calcul analytique du coût complet (lot6f). Aucun impact résultat/compta direct.
print('\nModule MEN_INT — M04 NON injecté (TYPE_FLUX_013 = analytique seul, D105 révisée)')
men_int_count = 0

# 6. Écart analytique gain/perte ménage (Lot 6f) — TYPE_FLUX_018, HC, impacte HORS_COMPTA.
#    Seul l'ÉCART (standard − coût complet) est injecté, jamais le coût complet entier.
#    gain (écart>0) -> PRODUIT HC (augmente HC) ; perte (écart<0) -> CHARGE HC (diminue HC).
print('\nModule GPM — Écart analytique gain/perte ménage (TYPE_FLUX_018 HC)...')
SRC_GPM = os.path.join(ROOT, '02_TRAVAIL', 'Lot6f_CoutComplet_Menages', 'MASTER_CALC_CoutComplet_Menages.xlsx')
gpm_rows = load_sheet(SRC_GPM, 'DETAIL_COUT_COMPLET') if os.path.exists(SRC_GPM) else []
for r in sorted(gpm_rows, key=lambda x: (str(x.get('mois')), str(x.get('logement_id')), str(x.get('intervenant_id')))):
    ec = r.get('ecart_vs_standard_total')
    if ec is None or float(ec) == 0:
        continue
    ec = float(ec)
    mois = r.get('mois')
    add_flux(
        module_code    = 'GPM',
        source_table   = 'MASTER_CALC_CoutComplet_Menages',
        source_pk      = f"{mois}|{r.get('logement_id')}|{r.get('intervenant_id')}",
        date_flux      = mois_to_first_day(mois),
        mois           = mois,
        logement_id    = r.get('logement_id'),
        proprietaire_id= r.get('proprietaire_id'),
        associe_id     = None,
        type_flux_id   = 'TYPE_FLUX_018',
        sens           = 'PRODUIT' if ec > 0 else 'CHARGE',
        montant        = abs(ec),
        code_impact    = 'HC',
        statut_controle= 'VALIDE',
        commentaire    = f"Écart analytique ménage {r.get('statut_ecart')} (standard − coût complet)",
    )
gpm_count = counters.get('GPM', 0)
print(f'  {gpm_count} flux GPM (écart analytique)')


# ── CONTRÔLES POST-CONSTRUCTION ───────────────────────────────────────────────
print('\nContrôles post-construction...')

# CTR-9-004 : montants positifs
neg = [r for r in flux_rows if r.get('montant') is not None and r['montant'] < 0]
if neg:
    print(f'BLOQUANT [CTR-9-004] {len(neg)} montants négatifs')
    for x in neg[:5]:
        print(f'  {x["flux_id"]} montant={x["montant"]}')
    sys.exit(1)
print(f'  CTR-9-004 OK — aucun montant négatif')

# CTR-9-005 : sens valides
bad_sens = [r for r in flux_rows if r.get('sens') not in ('PRODUIT', 'CHARGE', 'NEUTRALISATION')]
if bad_sens:
    print(f'BLOQUANT [CTR-9-005] {len(bad_sens)} valeurs de sens invalides')
    for x in bad_sens[:5]:
        print(f'  {x["flux_id"]} sens={x["sens"]}')
    sys.exit(1)
print(f'  CTR-9-005 OK — tous sens valides')

# CTR-9-006 : code_impact valides
bad_impact = [r for r in flux_rows if r.get('code_impact') not in ('IC', 'HC', 'HR')]
if bad_impact:
    print(f'BLOQUANT [CTR-9-006] {len(bad_impact)} code_impact invalides')
    for x in bad_impact[:5]:
        print(f'  {x["flux_id"]} code_impact={x["code_impact"]}')
    sys.exit(1)
print(f'  CTR-9-006 OK — tous code_impact valides')

# CTR-9-007 : flux_id unique
ids = [r['flux_id'] for r in flux_rows]
if len(ids) != len(set(ids)):
    dupes = [i for i in set(ids) if ids.count(i) > 1]
    print(f'BLOQUANT [CTR-9-007] flux_id non unique — {len(dupes)} ID en doublon : {dupes[:5]}')
    sys.exit(1)
print(f'  CTR-9-007 OK — {len(ids)} flux_id uniques')

# CTR-9-008 : doublons source technique
if doublons:
    print(f'BLOQUANT [CTR-9-008] {len(doublons)} doublons source technique (composite key)')
    for d in doublons[:5]:
        print(f'  {d}')
    sys.exit(1)
print(f'  CTR-9-008 OK — aucun doublon source technique')

# CTR-9-009 : volume total = somme des sources (M04 013 NON injecté ; GPM 018 ajouté)
expected = len(vue_flux) + len(men_valide) + len(bnq_valide) + len(chg_valide) + gpm_count
if len(flux_rows) != expected:
    print(f'BLOQUANT [CTR-9-009] Volume inattendu : {len(flux_rows)} flux vs {expected} attendu '
          f'({len(vue_flux)} RES + {len(men_valide)} MEN + {len(bnq_valide)} BNQ '
          f'+ {len(chg_valide)} CHG + {gpm_count} GPM)')
    sys.exit(1)
print(f'  CTR-9-009 OK — {len(flux_rows)} flux = {len(vue_flux)} RES + {len(men_valide)} MEN '
      f'+ {len(bnq_valide)} BNQ + {len(chg_valide)} CHG + {gpm_count} GPM (M04 013 analytique, non injecté)')

# CTR-9-011 (garde-fou régression) : portion RES (TYPE_FLUX_017) inchangée
res_flux_count = sum(1 for r in flux_rows if r.get('type_flux_id') == 'TYPE_FLUX_017')
if res_flux_count != len(vue_flux):
    print(f'BLOQUANT [CTR-9-011] Portion RES modifiée : {res_flux_count} TYPE_FLUX_017 vs {len(vue_flux)} attendu')
    sys.exit(1)
print(f'  CTR-9-011 OK — portion RES inchangée ({res_flux_count} TYPE_FLUX_017)')

print(f'\nTous contrôles BLOQUANTS OK.')

# ── ÉCRITURE EXCEL ────────────────────────────────────────────────────────────
print(f'\nÉcriture {OUT_FILE}...')
os.makedirs(OUT_DIR, exist_ok=True)

wb_out = openpyxl.Workbook()
ws_m   = wb_out.active
ws_m.title = 'MASTER'

HDR_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HDR_FONT = Font(bold=True, color='FFFFFF')

ws_m.append(COLS)
for cell in ws_m[1]:
    cell.fill = HDR_FILL
    cell.font = HDR_FONT

for r in flux_rows:
    ws_m.append([r.get(c) for c in COLS])

ws_m.freeze_panes = 'A2'
ws_m.auto_filter.ref = ws_m.dimensions

COL_WIDTHS = {
    'flux_id': 36, 'ROW_HASH': 20,
    'source_module': 14, 'source_table': 44, 'source_pk': 36,
    'date_flux': 14, 'mois': 10,
    'logement_id': 14, 'proprietaire_id': 16, 'associe_id': 14,
    'type_flux_id': 20, 'sens': 14, 'montant': 12, 'code_impact': 13,
    'inclure_resultat_reel': 21, 'inclure_resultat_comptable': 24, 'inclure_resultat_hors_compta': 26,
    'statut_controle': 17, 'niveau_anomalie': 17, 'code_anomalie': 26, 'commentaire': 44,
    'date_integration': 22,
}
for i, col in enumerate(COLS, 1):
    ws_m.column_dimensions[openpyxl.utils.get_column_letter(i)].width = COL_WIDTHS.get(col, 18)

wb_out.save(OUT_FILE)

# ── BILAN ─────────────────────────────────────────────────────────────────────
print(f'\n{"="*60}')
print(f'BILAN LOT 9 — MASTER_CALC_Flux')
print(f'{"="*60}')
print(f'Fichier    : {OUT_FILE}')
print(f'Colonnes   : {len(COLS)}')
print(f'Total flux : {len(flux_rows)}')
print(f'  RES      : {res_count}  (TYPE_FLUX_017 PRODUIT IC)')
print(f'  MEN      : {men_count}  (TYPE_FLUX_014 CHARGE  IC)')
print(f'  BNQ      : {bnq_count}  (TYPE_FLUX_016 CHARGE  IC)')
print(f'  CHG      : {chg_count}  (Charges Lot 3 — sens/impact par ligne)')
print(f'  MEN_INT  : {men_int_count}  (TYPE_FLUX_013 — analytique seul, NON injecté)')
print(f'  GPM      : {gpm_count}  (TYPE_FLUX_018 écart analytique HC — gain/perte ménage)')

from collections import Counter
by_sens     = Counter(r['sens']            for r in flux_rows)
by_impact   = Counter(r['code_impact']     for r in flux_rows)
by_statut   = Counter(r['statut_controle'] for r in flux_rows)
by_module   = Counter(r['source_module']   for r in flux_rows)
by_mois     = Counter(r['mois']            for r in flux_rows)

print(f'\nSens       : {dict(by_sens)}')
print(f'Code impact: {dict(by_impact)}')
print(f'Statut     : {dict(by_statut)}')
print(f'\nPar mois:')
for mois in sorted(by_mois):
    print(f'  {mois} : {by_mois[mois]}')

print(f'\nSécurité   : aucune colonne bancaire sensible dans COLS [OK]')
print(f'Doublons   : {len(doublons)} [OK]')
print(f'{"="*60}')
