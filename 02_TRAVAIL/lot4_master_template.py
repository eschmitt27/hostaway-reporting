"""
lot4_master_template.py
Lot 4 — Creation MASTER_FACT_MAN_ReservationsHorsHostaway.xlsx (squelette + M-code Power Query)
- Onglet MASTER     : 34 colonnes (30 SAISIE + 4 colonnes PQ)
- Onglet VUE_ACTIVE : memes 34 colonnes — filtre statut_controle=VALIDE
- Onglet POWER_QUERY_CODE : M-code a copier dans Excel (3 requetes)

Colonnes PQ ajoutees (4) :
  source_module    — constante LOT4_HH
  source_table     — constante SAISIE_ReservationsHorsHostaway
  source_pk        — = reservation_hh_id
  date_integration — DateTime.LocalNow() a chaque refresh PQ
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE     = "C:/Users/Ewan/OneDrive/Documents/Conciergerie/Pilotage_Conciergerie"
OUT_DIR  = f"{BASE}/02_TRAVAIL/Lot4_ReservationsHH"
OUT_PATH = f"{OUT_DIR}/MASTER_FACT_MAN_ReservationsHorsHostaway.xlsx"

os.makedirs(OUT_DIR, exist_ok=True)

# ── 34 colonnes MASTER ────────────────────────────────────────────────────────
# (nom_colonne, couleur_hex) — meme groupement que SAISIE + groupe PQ
MASTER_COLS = [
    # Groupe 1 — Identification
    ('reservation_hh_id',            '1F497D'),
    ('ROW_HASH',                     '1F497D'),
    # Groupe 2 — Rattachement
    ('mois',                         '2E75B6'),
    ('canal_id',                     '2E75B6'),
    ('source_financiere',            '2E75B6'),
    ('proprietaire_id',              '2E75B6'),
    ('logement_id',                  '2E75B6'),
    ('reservation_id_hostaway',      '2E75B6'),
    # Groupe 3 — Sejour
    ('date_arrivee',                 '375623'),
    ('date_depart',                  '375623'),
    ('nuits',                        '375623'),
    # Groupe 4 — Financier encaissement
    ('total_percu',                  '7030A0'),
    ('menage',                       '7030A0'),
    ('taux_commission',              '7030A0'),
    ('taux_commission_source',       '7030A0'),
    ('commentaire_taux_commission',  '7030A0'),
    ('commission',                   '7030A0'),
    # Groupe 5 — Associe recuperateur
    ('montant_recupere',             '974706'),
    ('associe_id_recuperateur',      '974706'),
    # Groupe 6 — Reversement proprietaire
    ('montant_reverse_proprietaire', 'C55A11'),
    ('mode_paiement_id',             'C55A11'),
    # Groupe 7 — Acompte
    ('acompte_facture',              '833C00'),
    # Groupe 8 — Statut / impact
    ('code_impact',                  'C00000'),
    ('comptabilisation',             'C00000'),
    ('impact_resultat_reel',         'C00000'),
    ('impact_resultat_comptable',    'C00000'),
    ('statut_controle',              '7B0000'),
    ('niveau_anomalie',              '7B0000'),
    ('code_anomalie',                '7B0000'),
    # Groupe 9 — Note
    ('commentaire',                  '595959'),
    # Groupe 10 — Systeme integration PQ [PQ]
    ('source_module',                '404040'),
    ('source_table',                 '404040'),
    ('source_pk',                    '404040'),
    ('date_integration',             '404040'),
]

assert len(MASTER_COLS) == 34, f"Attendu 34 colonnes, got {len(MASTER_COLS)}"

# ── M-code Power Query ────────────────────────────────────────────────────────

MCODE_SAISIE = r"""
// Requete : SAISIE_HH_Source
// Lit SAISIE_ReservationsHorsHostaway.xlsx, typage fort, filtre lignes vides
// Mode : Connexion uniquement (ne pas charger dans une feuille)

let
    BASE_PATH = "C:\Users\Ewan\OneDrive\Documents\Conciergerie\Pilotage_Conciergerie\",
    Source = Excel.Workbook(
        File.Contents(BASE_PATH & "01_SOURCES_BRUTES\ReservationsHH\SAISIE_ReservationsHorsHostaway.xlsx"),
        null, true
    ),
    SAISIE_Sheet = Source{[Item="SAISIE",Kind="Sheet"]}[Data],
    Promoted = Table.PromoteHeaders(SAISIE_Sheet, [PromoteAllScalars=true]),
    Typed = Table.TransformColumnTypes(Promoted, {
        {"reservation_hh_id",            type text},
        {"ROW_HASH",                     type text},
        {"mois",                         type text},
        {"canal_id",                     type text},
        {"source_financiere",            type text},
        {"proprietaire_id",              type text},
        {"logement_id",                  type text},
        {"reservation_id_hostaway",      type text},
        {"date_arrivee",                 type date},
        {"date_depart",                  type date},
        {"nuits",                        type number},
        {"total_percu",                  type number},
        {"menage",                       type number},
        {"taux_commission",              type number},
        {"taux_commission_source",       type text},
        {"commentaire_taux_commission",  type text},
        {"commission",                   type number},
        {"montant_recupere",             type number},
        {"associe_id_recuperateur",      type text},
        {"montant_reverse_proprietaire", type number},
        {"mode_paiement_id",             type text},
        {"acompte_facture",              type number},
        {"code_impact",                  type text},
        {"comptabilisation",             type text},
        {"impact_resultat_reel",         type text},
        {"impact_resultat_comptable",    type text},
        {"statut_controle",              type text},
        {"niveau_anomalie",              type text},
        {"code_anomalie",                type text},
        {"commentaire",                  type text}
    }),
    Filtered = Table.SelectRows(Typed, each
        [reservation_hh_id] <> null and [reservation_hh_id] <> ""
    )
in
    Filtered
""".strip()

MCODE_MASTER = r"""
// Requete : MASTER_FACT_MAN_ReservationsHorsHostaway
// 34 colonnes = 30 SAISIE + 4 colonnes PQ
// Charger dans : onglet MASTER (Clic droit > Charger dans > Table > feuille MASTER)

let
    Source = SAISIE_HH_Source,

    // Colonnes systeme d integration PQ
    AddSourceModule = Table.AddColumn(Source,          "source_module",    each "LOT4_HH",                          type text),
    AddSourceTable  = Table.AddColumn(AddSourceModule, "source_table",     each "SAISIE_ReservationsHorsHostaway",  type text),
    AddSourcePK     = Table.AddColumn(AddSourceTable,  "source_pk",        each [reservation_hh_id],                type text),
    AddDateInteg    = Table.AddColumn(AddSourcePK,     "date_integration", each DateTime.LocalNow(),                type datetime),

    // Reordonnement final — 34 colonnes
    Reordered = Table.ReorderColumns(AddDateInteg, {
        "reservation_hh_id", "ROW_HASH",
        "mois", "canal_id", "source_financiere",
        "proprietaire_id", "logement_id", "reservation_id_hostaway",
        "date_arrivee", "date_depart", "nuits",
        "total_percu", "menage", "taux_commission", "taux_commission_source",
        "commentaire_taux_commission", "commission",
        "montant_recupere", "associe_id_recuperateur",
        "montant_reverse_proprietaire", "mode_paiement_id",
        "acompte_facture",
        "code_impact", "comptabilisation",
        "impact_resultat_reel", "impact_resultat_comptable",
        "statut_controle", "niveau_anomalie", "code_anomalie",
        "commentaire",
        "source_module", "source_table", "source_pk", "date_integration"
    })
in
    Reordered
""".strip()

MCODE_VUE = r"""
// Requete : VUE_ACTIVE_ReservationsHH
// Filtre MASTER : statut_controle=VALIDE — 34 colonnes identiques
// Charger dans : onglet VUE_ACTIVE (Clic droit > Charger dans > Table > feuille VUE_ACTIVE)
// Alimente : Lot 4bis (MASTER_CALC_Reservations) et Lot 5 (acomptes proprietaires)

let
    Source = MASTER_FACT_MAN_ReservationsHorsHostaway,
    Filtre = Table.SelectRows(Source, each
        [statut_controle] = "VALIDE"
    )
in
    Filtre
""".strip()

INSTRUCTIONS = """
INSTRUCTIONS POWER QUERY — MASTER_FACT_MAN_ReservationsHorsHostaway.xlsx
===========================================================================
34 colonnes = 30 colonnes SAISIE_ReservationsHorsHostaway + 4 colonnes PQ :
  source_module, source_table, source_pk, date_integration

1. Ouvrir MASTER_FACT_MAN_ReservationsHorsHostaway.xlsx dans Excel.
2. Onglet Donnees > Obtenir des donnees > Editeur Power Query.

CREER 3 REQUETES dans cet ordre exact :

Req 1 : SAISIE_HH_Source
  Donnees > Requete vide > coller M-code REQUETE 1
  Mode : Connexion uniquement (ne PAS charger dans une feuille)

Req 2 : MASTER_FACT_MAN_ReservationsHorsHostaway
  Donnees > Requete vide > coller M-code REQUETE 2
  Puis : Clic droit > Charger dans > Table > Feuille MASTER

Req 3 : VUE_ACTIVE_ReservationsHH
  Donnees > Requete vide > coller M-code REQUETE 3
  Puis : Clic droit > Charger dans > Table > Feuille VUE_ACTIVE

NOTES :
- Adapter BASE_PATH si le projet change de lecteur ou d utilisateur.
- VUE_ACTIVE alimente Lot 4bis (table commune) et Lot 5 (acomptes proprietaires).
- Ne pas inclure les lignes statut_controle=A_CONTROLER dans le flux — elles restent en attente.
""".strip()

# ── Workbook ──────────────────────────────────────────────────────────────────
wb        = Workbook()
ws_master = wb.active
ws_master.title = 'MASTER'
ws_vue    = wb.create_sheet('VUE_ACTIVE')
ws_pq     = wb.create_sheet('POWER_QUERY_CODE')

H_FONT = Font(bold=True, color='FFFFFF', size=10)

def write_headers(ws, cols):
    for col_i, (col_name, color) in enumerate(cols, 1):
        c = ws.cell(row=1, column=col_i, value=col_name)
        c.font = H_FONT
        c.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(col_i)].width = max(len(col_name) + 2, 14)
    ws.row_dimensions[1].height = 42
    ws.freeze_panes = 'A2'

# MASTER — 34 colonnes
write_headers(ws_master, MASTER_COLS)
ws_master.cell(row=2, column=1,
    value='[Charge par Power Query — requete MASTER_FACT_MAN_ReservationsHorsHostaway]'
).font = Font(italic=True, color='808080')
print(f"[OK] MASTER — {len(MASTER_COLS)} colonnes")

# VUE_ACTIVE — memes 34 colonnes
write_headers(ws_vue, MASTER_COLS)
ws_vue.cell(row=2, column=1,
    value='[Vue par Power Query — requete VUE_ACTIVE_ReservationsHH — filtre statut_controle=VALIDE]'
).font = Font(italic=True, color='808080')
print(f"[OK] VUE_ACTIVE — {len(MASTER_COLS)} colonnes")

# POWER_QUERY_CODE
ws_pq.column_dimensions['A'].width = 36
ws_pq.column_dimensions['C'].width = 112

title_font = Font(bold=True, size=12, color='1F497D')
code_font  = Font(name='Courier New', size=9, color='262626')
note_font  = Font(italic=True, size=10, color='595959')

ws_pq.cell(row=1, column=1,
    value='POWER QUERY M-CODE — MASTER_FACT_MAN_ReservationsHorsHostaway').font = title_font
ws_pq.cell(row=2, column=1,
    value='Ne pas modifier — reference uniquement').font = note_font

row_off = 4
for line in INSTRUCTIONS.split('\n'):
    ws_pq.cell(row=row_off, column=1, value=line).font = Font(size=9)
    row_off += 1
row_off += 2

for label, mcode in [
    ('REQUETE 1 : SAISIE_HH_Source',                         MCODE_SAISIE),
    ('REQUETE 2 : MASTER_FACT_MAN_ReservationsHorsHostaway', MCODE_MASTER),
    ('REQUETE 3 : VUE_ACTIVE_ReservationsHH',                MCODE_VUE),
]:
    ws_pq.cell(row=row_off, column=1, value=label).font = Font(bold=True, size=10, color='7030A0')
    row_off += 1
    for line in mcode.split('\n'):
        ws_pq.cell(row=row_off, column=3, value=line).font = code_font
        ws_pq.row_dimensions[row_off].height = 14
        row_off += 1
    row_off += 2

print("[OK] POWER_QUERY_CODE — 3 requetes M-code")

wb.save(OUT_PATH)
print(f"[SAVED] {OUT_PATH}")
print(f"\n=== MASTER_FACT_MAN_ReservationsHorsHostaway.xlsx cree avec succes ({len(MASTER_COLS)} colonnes) ===")
