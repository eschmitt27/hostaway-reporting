#!/usr/bin/env python3
"""
Lot 10 — Calcul resultats, commissions et net proprietaire
Pilotage_Conciergerie

Sources (lecture seule):
  MASTER_CALC_Flux.xlsx           (Lot 9)
  MASTER_CALC_Reservations.xlsx   (Lot 4bis)
  MASTER_CALC_HA_Payout.xlsx      (Lot 1)
  REF_Setup.xlsm                  (REF_Logements, REF_Proprietaires)

Sorties:
  02_TRAVAIL/Lot10_Resultats/MASTER_CALC_Commissions.xlsx
  02_TRAVAIL/Lot10_Resultats/MASTER_CALC_Resultats.xlsx
  02_TRAVAIL/Lot10_Resultats/MASTER_CALC_NetProprietaire.xlsx

Dependances:
  pip install pandas openpyxl

Jointure:
  MASTER_CALC_Flux.source_pk
    -> MASTER_CALC_Reservations.reservation_calc_id
    -> MASTER_CALC_Reservations.reservation_id_hostaway
    -> MASTER_CALC_HA_Payout.reservation_id
"""
import sys
import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# Paths
# ─────────────────────────────────────────────────────────────────────────────
BASE        = Path(__file__).resolve().parent.parent
FLUX_FILE   = BASE / "02_TRAVAIL/Lot9_FluxUnifie/MASTER_CALC_Flux.xlsx"
RES_FILE    = BASE / "02_TRAVAIL/Lot4bis_TableCommune/MASTER_CALC_Reservations.xlsx"
PAYOUT_FILE = BASE / "02_TRAVAIL/Lot1_Hostaway/MASTER_CALC_HA_Payout.xlsx"
HH_FILE     = BASE / "02_TRAVAIL/Lot4_ReservationsHH/MASTER_FACT_MAN_ReservationsHorsHostaway.xlsx"
ACC_FILE    = BASE / "02_TRAVAIL/Lot5_AcomptesProprietaires/MASTER_FACT_MAN_AcomptesProprietaires.xlsx"
REF_FILE    = BASE / "01_SOURCES_BRUTES/REF_Setup/REF_Setup.xlsm"
OUT_DIR     = BASE / "02_TRAVAIL/Lot10_Resultats"

# Constante : tolérance d'arrondi par ligne (D035)
TOL_LIGNE   = 0.10
# Sentinelle charges sans logement/proprietaire (D-LOT10C-03)
SENTINEL_GLOBAL = "GLOBAL_NON_AFFECTE"
COMM_FILE   = OUT_DIR / "MASTER_CALC_Commissions.xlsx"
RESULT_FILE = OUT_DIR / "MASTER_CALC_Resultats.xlsx"
NET_FILE    = OUT_DIR / "MASTER_CALC_NetProprietaire.xlsx"

# ─────────────────────────────────────────────────────────────────────────────
# Logging
# ─────────────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)-8s %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("lot10")

# ─────────────────────────────────────────────────────────────────────────────
# Excel styles
# ─────────────────────────────────────────────────────────────────────────────
FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FILL_WARN   = PatternFill("solid", fgColor="FFEB9C")
FONT_HEADER = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
FONT_DATA   = Font(name="Calibri", size=10)

# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────
def _read_sheet(path: Path, sheet=None, keep_vba: bool = False) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=keep_vba)
    ws = wb[sheet] if sheet else wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = list(next(rows_iter))
    rows = [dict(zip(headers, r)) for r in rows_iter]
    wb.close()
    return pd.DataFrame(rows)


def _write_sheet(wb: Workbook, name: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(name)
    cols = list(df.columns)
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(1, ci, col)
        cell.font = FILL_HEADER and FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for ri, row_vals in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row_vals, 1):
            ws.cell(ri, ci, val).font = FONT_DATA
    ws.freeze_panes = "A2"
    for ci, col in enumerate(cols, 1):
        max_len = len(str(col))
        if len(df) > 0:
            col_len = df[col].astype(str).str.len().max()
            max_len = max(max_len, int(col_len) if pd.notna(col_len) else max_len)
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 45)


def _mois_range(first_mois: str, last_mois: str) -> list:
    """Return ['YYYY-MM', ...] inclusive between first and last."""
    try:
        periods = pd.period_range(start=first_mois, end=last_mois, freq="M")
        return [str(p) for p in periods]
    except Exception:
        return [first_mois]


def _n(val) -> float:
    """Coerce to float, 0.0 on failure."""
    try:
        return float(val) if val is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


def _is_placeholder_id(v) -> bool:
    """Ligne d'instruction Power Query / formule, pas une vraie cle metier."""
    if v is None:
        return True
    s = str(v).strip()
    return s == "" or s[0] in "#[<←-*"


# ─────────────────────────────────────────────────────────────────────────────
# 1. Load sources
# ─────────────────────────────────────────────────────────────────────────────
def load_sources():
    log.info("=== Chargement sources (lecture seule) ===")

    df_flux   = _read_sheet(FLUX_FILE)
    df_res    = _read_sheet(RES_FILE, sheet="MASTER")
    df_payout = _read_sheet(PAYOUT_FILE, sheet="data")
    df_log    = _read_sheet(REF_FILE, sheet="REF_Logements",     keep_vba=True)
    df_prop   = _read_sheet(REF_FILE, sheet="REF_Proprietaires", keep_vba=True)

    # Sources HH + Acomptes (lecture seule, hors placeholder Power Query)
    df_hh  = _read_sheet(HH_FILE,  sheet="MASTER") if HH_FILE.exists()  else pd.DataFrame()
    df_acc = _read_sheet(ACC_FILE, sheet="MASTER") if ACC_FILE.exists() else pd.DataFrame()
    if len(df_hh) > 0 and "reservation_hh_id" in df_hh.columns:
        df_hh = df_hh[~df_hh["reservation_hh_id"].map(_is_placeholder_id)].reset_index(drop=True)
    if len(df_acc) > 0 and "acompte_id" in df_acc.columns:
        df_acc = df_acc[~df_acc["acompte_id"].map(_is_placeholder_id)].reset_index(drop=True)

    # Filter duplicate-header rows in xlsm sheets
    df_log  = df_log[df_log["logement_id"].astype(str) != "logement_id"].reset_index(drop=True)
    df_prop = df_prop[df_prop["proprietaire_id"].astype(str) != "proprietaire_id"].reset_index(drop=True)

    # Numeric conversions
    df_prop["taux_commission"] = pd.to_numeric(df_prop["taux_commission"], errors="coerce")
    df_log["forfait_logiciel_consommables_mensuel"] = pd.to_numeric(
        df_log["forfait_logiciel_consommables_mensuel"], errors="coerce"
    ).fillna(0.0)

    log.info(f"  Flux          : {len(df_flux)} lignes")
    log.info(f"  Reservations  : {len(df_res)} lignes")
    log.info(f"  Payout        : {len(df_payout)} lignes")
    log.info(f"  HH saisie     : {len(df_hh)} lignes (hors placeholder)")
    log.info(f"  Acomptes      : {len(df_acc)} lignes (hors placeholder)")
    log.info(f"  REF_Logements : {len(df_log)} logements")
    log.info(f"  REF_Prop      : {len(df_prop)} proprietaires")
    return df_flux, df_res, df_payout, df_hh, df_acc, df_log, df_prop


# ─────────────────────────────────────────────────────────────────────────────
# 2. Build commissions
# ─────────────────────────────────────────────────────────────────────────────
def build_commissions(df_flux, df_res, df_payout, df_hh, df_log, df_prop):
    log.info("=== Construction commissions (routage HA / HH) ===")
    hh_controls = []

    # ── 2a. Filter TYPE_FLUX_017 ──
    df_017 = df_flux[df_flux["type_flux_id"] == "TYPE_FLUX_017"].copy()
    log.info(f"  TYPE_FLUX_017 : {len(df_017)} lignes dans Flux")

    dup = df_017[df_017.duplicated("source_pk", keep=False)]
    if len(dup) > 0:
        log.error(f"BLOQUANT DOUBLON_RESERVATION_FLUX — {len(dup)} lignes dupliquees source_pk")
        sys.exit(1)

    df_017_sel = df_017[[
        "source_pk", "flux_id", "mois", "logement_id", "proprietaire_id",
    ]].rename(columns={
        "mois":           "mois_flux",
        "logement_id":    "logement_id_flux",
        "proprietaire_id":"proprietaire_id_flux",
    })

    # ── 2b. Join Flux -> Reservations (+ champs de routage) ──
    df_res_num = df_res.copy()
    df_res_num["reservation_id_hostaway"] = pd.to_numeric(
        df_res_num["reservation_id_hostaway"], errors="coerce"
    )
    df_res_sel = df_res_num[[
        "reservation_calc_id", "reservation_id_hostaway", "reservation_hh_id",
        "source", "source_montant", "montant_retenu",
        "logement_id", "proprietaire_id", "date_arrivee", "date_depart", "nuits",
    ]]
    df_j = df_017_sel.merge(
        df_res_sel, left_on="source_pk", right_on="reservation_calc_id", how="left",
    )
    missing_res = df_j["reservation_calc_id"].isna()
    if missing_res.any():
        log.error(f"BLOQUANT JOINTURE_RESERVATIONS_MANQUANTE — {missing_res.sum()} lignes")
        sys.exit(1)
    log.info("  Jointure Flux <-> Reservations OK")

    # logement/proprietaire effectifs (Reservations prioritaire, fallback Flux)
    df_j["logement_id_eff"]     = df_j["logement_id"].combine_first(df_j["logement_id_flux"])
    df_j["proprietaire_id_eff"] = df_j["proprietaire_id"].combine_first(df_j["proprietaire_id_flux"])

    # ── 2c. Routage HA vs HH (D-LOT10C-05) ──
    HA_SOURCES = {"HOSTAWAY_AIRBNB", "HOSTAWAY_BOOKING"}
    is_ha = df_j["reservation_id_hostaway"].notna() & df_j["source"].isin(HA_SOURCES)
    df_ha = df_j[is_ha].copy()
    df_hhb = df_j[~is_ha].copy()
    log.info(f"  Routage : {len(df_ha)} Hostaway / {len(df_hhb)} HH")

    df_prop_sel = df_prop[["proprietaire_id", "taux_commission"]].copy()

    # ===================== BRANCHE HOSTAWAY =====================
    df_pay_sel = df_payout[[
        "reservation_id", "channel_type", "statut_calcul_payout",
        "payout_calcule", "source_payout", "menage_retenu", "assiette_commission",
        "inclure_resultat_auto", "menage_retenu_source", "cout_standard_id",
        "cout_standard_menage_snapshot", "cout_standard_date_debut_validite",
        "cout_standard_date_fin_validite", "logement_id_snapshot",
        "type_logement_id_snapshot", "date_reference_cout_menage",
    ]].copy()
    df_pay_sel["reservation_id"] = pd.to_numeric(df_pay_sel["reservation_id"], errors="coerce")

    df_ha = df_ha.merge(
        df_pay_sel, left_on="reservation_id_hostaway", right_on="reservation_id", how="left",
    )
    # BLOQUANT JOINTURE_PAYOUT_MANQUANTE : branche Hostaway uniquement
    missing_pay = df_ha["reservation_id"].isna()
    if missing_pay.any():
        log.error(f"BLOQUANT JOINTURE_PAYOUT_MANQUANTE — {missing_pay.sum()} lignes Hostaway")
        sys.exit(1)
    log.info("  Jointure HA <-> Payout OK")

    df_ha = df_ha.merge(
        df_prop_sel, left_on="proprietaire_id_eff", right_on="proprietaire_id",
        how="left", suffixes=("", "_ref"),
    )
    normal_ha = df_ha["statut_calcul_payout"] == "NORMAL"
    if (normal_ha & df_ha["proprietaire_id_eff"].isna()).any():
        log.error("BLOQUANT COMMISSION_LOGEMENT_SANS_PROPRIETAIRE — HA")
        sys.exit(1)
    if (normal_ha & df_ha["taux_commission"].isna()).any():
        log.error("BLOQUANT COMMISSION_SANS_TAUX — HA")
        sys.exit(1)
    for col in ["payout_calcule", "menage_retenu", "assiette_commission", "taux_commission"]:
        df_ha[col] = pd.to_numeric(df_ha[col], errors="coerce")
    if (normal_ha & (df_ha["assiette_commission"].fillna(0) < 0)).any():
        log.error("BLOQUANT ASSIETTE_NEGATIVE — HA")
        sys.exit(1)
    df_ha["commission_conciergerie"] = None
    df_ha["net_proprietaire"]        = None
    df_ha.loc[normal_ha, "commission_conciergerie"] = (
        df_ha.loc[normal_ha, "assiette_commission"] * df_ha.loc[normal_ha, "taux_commission"]
    ).round(2)
    df_ha.loc[normal_ha, "net_proprietaire"] = (
        df_ha.loc[normal_ha, "payout_calcule"] - df_ha.loc[normal_ha, "menage_retenu"]
        - df_ha.loc[normal_ha, "commission_conciergerie"]
    ).round(2)
    df_ha["source_type"] = "HOSTAWAY"
    df_ha_norm = df_ha[normal_ha].copy()

    # ===================== BRANCHE HH =====================
    # D-LOT10C-01 : commission HH = (total_percu - menage) x taux REF_Prop
    df_hh_norm = pd.DataFrame()
    n_hh_exclus = 0
    if len(df_hhb) > 0:
        if len(df_hh) > 0:
            df_hh_sel = df_hh[[
                "reservation_hh_id", "total_percu", "menage", "commission", "taux_commission",
            ]].rename(columns={
                "commission": "commission_saisie", "taux_commission": "taux_hh_saisie",
            })
            df_hhb = df_hhb.merge(df_hh_sel, on="reservation_hh_id", how="left")
        else:
            for c in ["total_percu", "menage", "commission_saisie", "taux_hh_saisie"]:
                df_hhb[c] = None
        df_hhb = df_hhb.merge(
            df_prop_sel, left_on="proprietaire_id_eff", right_on="proprietaire_id",
            how="left", suffixes=("", "_ref"),
        )
        df_hhb["total_percu"]     = pd.to_numeric(df_hhb["total_percu"], errors="coerce")
        df_hhb["menage"]          = pd.to_numeric(df_hhb["menage"], errors="coerce").fillna(0.0)
        df_hhb["taux_commission"] = pd.to_numeric(df_hhb["taux_commission"], errors="coerce")

        # D-LOT10C-05 : integrer seulement si total_percu renseigne (>0) et taux dispo
        ok = df_hhb["total_percu"].notna() & (df_hhb["total_percu"] > 0) & df_hhb["taux_commission"].notna()
        df_hh_ok  = df_hhb[ok].copy()
        n_hh_exclus = int((~ok).sum())

        if len(df_hh_ok) > 0:
            df_hh_ok["payout_calcule"]      = df_hh_ok["total_percu"].round(2)
            df_hh_ok["menage_retenu"]       = df_hh_ok["menage"].round(2)
            df_hh_ok["assiette_commission"] = (df_hh_ok["total_percu"] - df_hh_ok["menage"]).round(2)
            if (df_hh_ok["assiette_commission"] < 0).any():
                log.error("BLOQUANT ASSIETTE_NEGATIVE — HH")
                sys.exit(1)
            df_hh_ok["commission_conciergerie"] = (
                df_hh_ok["assiette_commission"] * df_hh_ok["taux_commission"]
            ).round(2)
            df_hh_ok["net_proprietaire"] = (
                df_hh_ok["total_percu"] - df_hh_ok["menage"] - df_hh_ok["commission_conciergerie"]
            ).round(2)
            df_hh_ok["statut_calcul_payout"] = "NORMAL"
            df_hh_ok["channel_type"]         = df_hh_ok["source"]
            df_hh_ok["source_type"]          = "HH"
            df_hh_ok["menage_retenu_source"] = "SAISIE_HH"
            # Ecart commission saisie vs recalcul (D-LOT10C-01)
            cs = pd.to_numeric(df_hh_ok["commission_saisie"], errors="coerce")
            ecart = (cs - df_hh_ok["commission_conciergerie"]).abs()
            for _, r in df_hh_ok[ecart > TOL_LIGNE].iterrows():
                hh_controls.append({
                    "reservation": r["reservation_calc_id"],
                    "code_anomalie": "COMMISSION_HH_SAISIE_DIFFERE_RECALCUL",
                    "niveau": "A_CONTROLER",
                    "message": (
                        f"{r['reservation_calc_id']}: commission saisie={cs.loc[r.name]} "
                        f"!= recalcul={r['commission_conciergerie']} (>0.10 EUR)"
                    ),
                })
            df_hh_norm = df_hh_ok

        if n_hh_exclus > 0:
            hh_controls.append({
                "reservation": None,
                "code_anomalie": "HH_SANS_MONTANT_SAISI",
                "niveau": "A_CONTROLER",
                "message": f"{n_hh_exclus} reservations HH dans Flux sans total_percu/taux exploitable",
            })

    # ── 2h. Sortie COMMISSIONS (NORMAL HA + HH) ──
    COMM_OUT_COLS = [
        "source_pk", "reservation_calc_id", "reservation_id_hostaway",
        "logement_id_eff", "proprietaire_id_eff", "mois_flux",
        "date_arrivee", "date_depart", "nuits", "channel_type", "source_type",
        "statut_calcul_payout", "payout_calcule", "menage_retenu", "menage_retenu_source",
        "cout_standard_id", "cout_standard_menage_snapshot", "date_reference_cout_menage",
        "assiette_commission", "taux_commission", "commission_conciergerie",
        "net_proprietaire", "inclure_resultat_auto",
        "logement_id_snapshot", "type_logement_id_snapshot",
    ]

    def _shape(df):
        if len(df) == 0:
            return pd.DataFrame(columns=COMM_OUT_COLS)
        for c in COMM_OUT_COLS:
            if c not in df.columns:
                df[c] = None
        return df[COMM_OUT_COLS].copy()

    df_comm = pd.concat([_shape(df_ha_norm), _shape(df_hh_norm)], ignore_index=True)
    df_comm = df_comm.rename(columns={
        "logement_id_eff":     "logement_id",
        "proprietaire_id_eff": "proprietaire_id",
        "mois_flux":           "mois",
        "source_pk":           "flux_source_pk",
    }).reset_index(drop=True)

    # ── 2i. A_CONTROLER from Payout (Hostaway hors Flux) ──
    df_ac = df_payout[df_payout["statut_calcul_payout"] == "A_CONTROLER"].copy()
    df_ac["code_anomalie_lot10"] = "RESERVATION_EXCLUE_A_CONTROLER"
    df_ac = df_ac.reset_index(drop=True)

    log.info(f"  NORMAL integres  : {len(df_comm)} (HA {len(df_ha_norm)} + HH {len(df_hh_norm)})")
    log.info(f"  HH exclus        : {n_hh_exclus} (sans montant saisi)")
    log.info(f"  A_CONTROLER (pay): {len(df_ac)}")
    return df_comm, df_ac, hh_controls


# ─────────────────────────────────────────────────────────────────────────────
# 3. Build charge fixe grid (Option A)
# ─────────────────────────────────────────────────────────────────────────────
def build_charge_fixe(df_flux, df_log):
    log.info("=== Construction grille charge fixe (Option A) ===")
    controls = []

    df_017 = df_flux[df_flux["type_flux_id"] == "TYPE_FLUX_017"].copy()
    log_first = df_017.groupby("logement_id")["mois"].min().to_dict()
    log_last  = df_017.groupby("logement_id")["mois"].max().to_dict()

    rows = []
    for _, lr in df_log.iterrows():
        log_id  = lr["logement_id"]
        forfait = _n(lr["forfait_logiciel_consommables_mensuel"])
        actif   = lr.get("actif", "OUI")
        d_sortie = lr.get("date_sortie_gestion")
        d_entree = lr.get("date_entree_gestion")
        prop_id  = lr.get("proprietaire_id")

        # Skip APPARTEMENT_DIVERS / LOGEMENT_DIVERS (no proprietaire)
        if not prop_id:
            continue

        # Skip forfait = 0
        if forfait == 0.0:
            continue

        # Skip if no TYPE_FLUX_017 for this logement
        if log_id not in log_first:
            controls.append({
                "logement_id":   log_id,
                "code_anomalie": "LOG_SANS_FLUX_017",
                "niveau":        "A_CONTROLER",
                "message":       (
                    f"{log_id}: forfait={forfait} EUR "
                    f"mais aucune reservation TYPE_FLUX_017 dans MASTER_CALC_Flux"
                ),
            })
            continue

        first_mois = log_first[log_id]
        last_mois  = log_last[log_id]

        # Override last_mois with date_sortie_gestion if logement is inactive
        if d_sortie and actif == "NON":
            if hasattr(d_sortie, "strftime"):
                last_mois = d_sortie.strftime("%Y-%m")
            else:
                last_mois = str(d_sortie)[:7]

        # Control: first_mois < date_entree_gestion
        if d_entree:
            d_entree_str = (
                d_entree.strftime("%Y-%m")
                if hasattr(d_entree, "strftime")
                else str(d_entree)[:7]
            )
            if first_mois < d_entree_str:
                controls.append({
                    "logement_id":   log_id,
                    "code_anomalie": "CHARGE_FIXE_DATE_ENTREE_GESTION_INCOHERENTE",
                    "niveau":        "A_CONTROLER",
                    "message":       (
                        f"{log_id}: premier mois Flux={first_mois} "
                        f"< date_entree_gestion REF={d_entree_str}"
                    ),
                })

        for mois in _mois_range(first_mois, last_mois):
            rows.append({
                "mois":                 mois,
                "logement_id":          log_id,
                "proprietaire_id":      prop_id,
                "charge_fixe_mensuelle":forfait,
                "charge_fixe_source":   "REF_Logements.forfait_logiciel_consommables_mensuel",
            })

    df_cfix = pd.DataFrame(rows) if rows else pd.DataFrame(columns=[
        "mois", "logement_id", "proprietaire_id",
        "charge_fixe_mensuelle", "charge_fixe_source",
    ])

    n_log = df_cfix["logement_id"].nunique() if len(df_cfix) > 0 else 0
    n_inco = sum(1 for c in controls if c["code_anomalie"] == "CHARGE_FIXE_DATE_ENTREE_GESTION_INCOHERENTE")
    n_sans = sum(1 for c in controls if c["code_anomalie"] == "LOG_SANS_FLUX_017")
    log.info(f"  Charge fixe  : {len(df_cfix)} lignes / {n_log} logements")
    log.info(f"  INCOHERENT   : {n_inco} logements (premier mois Flux < date_entree_gestion REF)")
    log.info(f"  SANS_FLUX_017: {n_sans} logements avec forfait>0 mais aucune reservation")
    for c in controls:
        log.info(f"    [{c['niveau']}] {c['code_anomalie']}: {c['message']}")

    return df_cfix, controls


# ─────────────────────────────────────────────────────────────────────────────
# 4. Build resultats from MASTER_CALC_Flux
# ─────────────────────────────────────────────────────────────────────────────
def build_resultats(df_flux):
    log.info("=== Construction resultats (depuis MASTER_CALC_Flux) ===")

    df = df_flux.copy()
    df["montant"] = pd.to_numeric(df["montant"], errors="coerce").fillna(0.0)

    def _agg_vision(mask, vision_label):
        df_v = df[mask].copy()
        if len(df_v) == 0:
            return pd.DataFrame(columns=[
                "mois", "logement_id", "proprietaire_id",
                "total_produits", "total_charges", "resultat",
                "nb_flux", "vision", "commentaire",
            ])
        # D-LOT10C-03 : charges sans logement/proprietaire -> ligne dediee GLOBAL_NON_AFFECTE
        # (sinon perdues par groupby dropna=True)
        df_v["logement_id"]     = df_v["logement_id"].fillna(SENTINEL_GLOBAL)
        df_v["proprietaire_id"] = df_v["proprietaire_id"].fillna(SENTINEL_GLOBAL)
        df_v.loc[df_v["logement_id"].astype(str).str.strip() == "", "logement_id"] = SENTINEL_GLOBAL
        df_v.loc[df_v["proprietaire_id"].astype(str).str.strip() == "", "proprietaire_id"] = SENTINEL_GLOBAL
        grp = df_v.groupby(["mois", "logement_id", "proprietaire_id", "sens"]).agg(
            montant=("montant", "sum"),
            nb=("flux_id", "count"),
        ).reset_index()
        prod = grp[grp["sens"] == "PRODUIT"].groupby(
            ["mois", "logement_id", "proprietaire_id"]
        ).agg(total_produits=("montant", "sum"), nb_prod=("nb", "sum")).reset_index()
        chg  = grp[grp["sens"] == "CHARGE"].groupby(
            ["mois", "logement_id", "proprietaire_id"]
        ).agg(total_charges=("montant", "sum"), nb_chg=("nb", "sum")).reset_index()
        res = prod.merge(chg, on=["mois", "logement_id", "proprietaire_id"], how="outer").fillna(0)
        res["resultat"]   = (res["total_produits"] - res["total_charges"]).round(2)
        res["nb_flux"]    = res["nb_prod"] + res["nb_chg"]
        res["vision"]     = vision_label
        res["commentaire"]= ""
        return res[["mois", "logement_id", "proprietaire_id",
                    "total_produits", "total_charges", "resultat",
                    "nb_flux", "vision", "commentaire"]]

    df_reel  = _agg_vision(df["inclure_resultat_reel"]        == "OUI", "REEL")
    df_compt = _agg_vision(df["inclure_resultat_comptable"]   == "OUI", "COMPTABLE")
    df_hc    = _agg_vision(df["inclure_resultat_hors_compta"] == "OUI", "HORS_COMPTA")

    if len(df_hc) == 0:
        # Pas de flux HC -> placeholder explicite (sources HC vides)
        df_hc = pd.DataFrame([{
            "mois": "N/A", "logement_id": "N/A", "proprietaire_id": "N/A",
            "total_produits": 0.0, "total_charges": 0.0, "resultat": 0.0,
            "nb_flux": 0, "vision": "HORS_COMPTA",
            "commentaire": "HC_ZERO_SOURCES_VIDES — aucun flux HC",
        }])
    else:
        df_hc["commentaire"] = "Flux HC presents"

    hc_tot = df_hc["resultat"].sum() if len(df_hc) > 0 else 0.0
    log.info(f"  REEL        : {len(df_reel)} lignes  total={df_reel['resultat'].sum():,.2f} EUR")
    log.info(f"  COMPTABLE   : {len(df_compt)} lignes  total={df_compt['resultat'].sum():,.2f} EUR")
    log.info(f"  HORS_COMPTA : {len(df_hc)} lignes  total={hc_tot:,.2f} EUR")

    return df_reel, df_compt, df_hc


# ─────────────────────────────────────────────────────────────────────────────
# 5. Build net proprietaire
# ─────────────────────────────────────────────────────────────────────────────
def build_net_proprietaire(df_comm, df_cfix, df_acc):
    log.info("=== Construction net proprietaire ===")

    # ── 5a. EXPLOITATION (per reservation) — acomptes JAMAIS ici (D031/D033) ──
    df_exploit = df_comm.copy()
    df_exploit["charge_fixe_mensuelle"] = 0.0
    df_exploit["commentaire_charge_fixe"] = "Charge fixe 1x/mois dans REGLEMENT"
    df_exploit["revenu_net_exploitation"] = pd.to_numeric(
        df_exploit["net_proprietaire"], errors="coerce"
    ).round(2)

    # ── 5a-bis. Index acomptes (D-LOT10C-02) — bloc REGLEMENT uniquement ──
    # acc_by_log : (mois, logement_id) -> (montant, proprietaire_id)   [acomptes avec logement]
    # acc_by_prop: (mois, proprietaire_id) -> montant                  [acomptes sans logement]
    acc_by_log, acc_by_prop = {}, {}
    if len(df_acc) > 0:
        dfa = df_acc.copy()
        dfa = dfa[dfa.get("statut_controle", "VALIDE").astype(str) == "VALIDE"]
        dfa["montant_acompte"] = pd.to_numeric(dfa["montant_acompte"], errors="coerce").fillna(0.0)
        for _, a in dfa.iterrows():
            mois = a.get("mois")
            prop = a.get("proprietaire_id")
            logid = a.get("logement_id")
            mt = _n(a.get("montant_acompte"))
            if logid and str(logid).strip() and str(logid) != "None":
                k = (mois, logid)
                cur = acc_by_log.get(k, (0.0, prop))
                acc_by_log[k] = (round(cur[0] + mt, 2), prop)
            else:
                k = (mois, prop)
                acc_by_prop[k] = round(acc_by_prop.get(k, 0.0) + mt, 2)
        log.info(f"  Acomptes indexes : {len(acc_by_log)} (mois x logement) / {len(acc_by_prop)} (mois x prop seul)")

    # ── 5b. Aggregate reservations per mois x logement ──
    df_num = df_comm.copy()
    for col in ["payout_calcule", "menage_retenu", "commission_conciergerie", "net_proprietaire"]:
        df_num[col] = pd.to_numeric(df_num[col], errors="coerce").fillna(0.0)

    df_agg_res = pd.DataFrame()
    if len(df_num) > 0:
        df_agg_res = df_num.groupby(["mois", "logement_id", "proprietaire_id"]).agg(
            total_payout_mois                =("payout_calcule",        "sum"),
            total_menage_mois                =("menage_retenu",         "sum"),
            total_commission_mois            =("commission_conciergerie","sum"),
            net_proprietaire_avant_charge_mois=("net_proprietaire",     "sum"),
            nb_reservations                  =("reservation_calc_id",   "count"),
        ).reset_index().round(2)

    # ── 5c. Build REGLEMENT: outer join cfix x agg_res ──
    cfix_keys = (
        set(zip(df_cfix["mois"], df_cfix["logement_id"])) if len(df_cfix) > 0 else set()
    )
    res_keys  = (
        set(zip(df_agg_res["mois"], df_agg_res["logement_id"])) if len(df_agg_res) > 0 else set()
    )
    # Inclure les acomptes (mois x logement) pour qu'un acompte sans resa/cfix reste visible
    all_keys = cfix_keys | res_keys | set(acc_by_log.keys())

    # Index for fast lookup
    cfix_idx = (
        df_cfix.set_index(["mois", "logement_id"]) if len(df_cfix) > 0 else pd.DataFrame()
    )
    res_idx  = (
        df_agg_res.set_index(["mois", "logement_id"]) if len(df_agg_res) > 0 else pd.DataFrame()
    )

    reg_rows = []
    for (mois, log_id) in sorted(all_keys):
        row = {"mois": mois, "logement_id": log_id}

        # Charge fixe
        if (mois, log_id) in cfix_keys:
            cr = cfix_idx.loc[(mois, log_id)]
            if isinstance(cr, pd.DataFrame):
                cr = cr.iloc[0]
            row["proprietaire_id"]      = cr["proprietaire_id"]
            row["charge_fixe_mensuelle"] = _n(cr["charge_fixe_mensuelle"])
            row["charge_fixe_source"]   = cr["charge_fixe_source"]
        else:
            row["charge_fixe_mensuelle"] = 0.0
            row["charge_fixe_source"]   = "NON_APPLICABLE_FORFAIT_ZERO"

        # Reservation aggregates
        if (mois, log_id) in res_keys:
            rr = res_idx.loc[(mois, log_id)]
            if isinstance(rr, pd.DataFrame):
                rr = rr.iloc[0]
            if "proprietaire_id" not in row or not row.get("proprietaire_id"):
                row["proprietaire_id"] = rr["proprietaire_id"]
            row["total_payout_mois"]                 = _n(rr["total_payout_mois"])
            row["total_menage_mois"]                 = _n(rr["total_menage_mois"])
            row["total_commission_mois"]             = _n(rr["total_commission_mois"])
            row["net_proprietaire_avant_charge_mois"]= _n(rr["net_proprietaire_avant_charge_mois"])
            row["nb_reservations"]                   = int(_n(rr["nb_reservations"]))
        else:
            row.setdefault("proprietaire_id", None)
            row["total_payout_mois"]                 = 0.0
            row["total_menage_mois"]                 = 0.0
            row["total_commission_mois"]             = 0.0
            row["net_proprietaire_avant_charge_mois"]= 0.0
            row["nb_reservations"]                   = 0

        # Acompte propriétaire (mois x logement) — bloc REGLEMENT uniquement (D-LOT10C-02)
        acc_amt = 0.0
        if (mois, log_id) in acc_by_log:
            acc_amt = acc_by_log[(mois, log_id)][0]
            if not row.get("proprietaire_id"):
                row["proprietaire_id"] = acc_by_log[(mois, log_id)][1]

        # Bloc reglement (n'impacte JAMAIS revenu_net_exploitation — D031/D033)
        row["montant_du_conciergerie"] = round(
            row["total_commission_mois"]
            + row["total_menage_mois"]
            + row["charge_fixe_mensuelle"],
            2,
        )
        row["acompte_conciergerie_recu_via_airbnb"] = 0.0
        row["autres_acomptes_recus"]                = round(acc_amt, 2)
        row["paiement_deja_recu"]                   = 0.0
        row["reste_a_payer_conciergerie"] = round(
            row["montant_du_conciergerie"]
            - row["acompte_conciergerie_recu_via_airbnb"]
            - row["autres_acomptes_recus"]
            - row["paiement_deja_recu"],
            2,
        )
        row["net_proprietaire_apres_charge_mois"] = round(
            row["net_proprietaire_avant_charge_mois"] - row["charge_fixe_mensuelle"], 2
        )
        # A_CONTROLER: mois with charge fixe but no reservation
        if row["nb_reservations"] == 0 and row["charge_fixe_mensuelle"] > 0:
            row["statut_reglement"] = "MOIS_ACTIF_SANS_RESERVATION"
        else:
            row["statut_reglement"] = "A_CONTROLER"
        reg_rows.append(row)

    # Acomptes sans logement (fallback proprietaire) -> ligne dediee GLOBAL_NON_AFFECTE
    for (mois, prop), amt in acc_by_prop.items():
        reg_rows.append({
            "mois": mois, "logement_id": SENTINEL_GLOBAL, "proprietaire_id": prop,
            "charge_fixe_mensuelle": 0.0, "charge_fixe_source": "NON_APPLICABLE",
            "total_payout_mois": 0.0, "total_menage_mois": 0.0, "total_commission_mois": 0.0,
            "net_proprietaire_avant_charge_mois": 0.0, "nb_reservations": 0,
            "montant_du_conciergerie": 0.0,
            "acompte_conciergerie_recu_via_airbnb": 0.0,
            "autres_acomptes_recus": round(amt, 2),
            "paiement_deja_recu": 0.0,
            "reste_a_payer_conciergerie": round(-amt, 2),
            "net_proprietaire_apres_charge_mois": 0.0,
            "statut_reglement": "ACOMPTE_SANS_LOGEMENT",
        })

    df_reg = pd.DataFrame(reg_rows)

    # ── 5d. VUE_MOIS (per mois x proprietaire) ──
    if len(df_reg) > 0:
        sum_cols = {
            "total_payout_mois":                 "sum",
            "total_menage_mois":                 "sum",
            "total_commission_mois":             "sum",
            "charge_fixe_mensuelle":             "sum",
            "montant_du_conciergerie":           "sum",
            "reste_a_payer_conciergerie":        "sum",
            "net_proprietaire_avant_charge_mois":"sum",
            "net_proprietaire_apres_charge_mois":"sum",
            "nb_reservations":                   "sum",
        }
        df_vue = df_reg.groupby(["mois", "proprietaire_id"]).agg(
            **{k: (k, v) for k, v in sum_cols.items() if k in df_reg.columns}
        ).reset_index().round(2)
        df_vue = df_vue.sort_values(["mois", "proprietaire_id"]).reset_index(drop=True)
    else:
        df_vue = pd.DataFrame(columns=[
            "mois", "proprietaire_id",
            "total_payout_mois", "total_commission_mois", "charge_fixe_mensuelle",
        ])

    log.info(f"  EXPLOITATION     : {len(df_exploit)} reservations")
    log.info(f"  REGLEMENT        : {len(df_reg)} lignes mois x logement")
    log.info(f"  VUE_MOIS         : {len(df_vue)} lignes mois x proprietaire")

    return df_exploit, df_reg, df_vue


# ─────────────────────────────────────────────────────────────────────────────
# 6. Write Excel outputs
# ─────────────────────────────────────────────────────────────────────────────
def write_all(df_comm, df_ac, df_reel, df_compt, df_hc, df_exploit, df_reg, df_vue):
    log.info("=== Ecriture fichiers Excel ===")
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # ── MASTER_CALC_Commissions ──
    wb1 = Workbook()
    wb1.remove(wb1.active)
    _write_sheet(wb1, "COMMISSIONS", df_comm)
    _write_sheet(wb1, "A_CONTROLER", df_ac)
    wb1.save(COMM_FILE)
    log.info(f"  {COMM_FILE.name} : OK")

    # ── MASTER_CALC_Resultats ──
    # PAR_MOIS_LOGEMENT: all three visions stacked
    df_par_ml = pd.concat([df_reel, df_compt, df_hc], ignore_index=True)

    # PAR_MOIS_PROPRIETAIRE: aggregate reel + comptable
    def _agg_prop(df_v):
        if len(df_v) == 0:
            return pd.DataFrame()
        return df_v.groupby(["mois", "proprietaire_id", "vision"]).agg(
            total_produits=("total_produits", "sum"),
            total_charges =("total_charges",  "sum"),
            resultat      =("resultat",        "sum"),
            nb_flux       =("nb_flux",         "sum"),
        ).reset_index()

    df_par_mp = pd.concat(
        [_agg_prop(df_reel), _agg_prop(df_compt)], ignore_index=True
    )

    # GLOBAL — visions calculees (HC inclus, plus de codage en dur — defaut #2 corrige)
    def _sum(df, col):
        return round(df[col].sum(), 2) if len(df) > 0 and col in df.columns else 0.0

    # df_hc peut contenir un placeholder N/A (resultat 0) -> sommes = 0, coherent
    reel_prod, reel_chg, reel_tot    = _sum(df_reel, "total_produits"),  _sum(df_reel, "total_charges"),  _sum(df_reel, "resultat")
    compt_prod, compt_chg, compt_tot = _sum(df_compt, "total_produits"), _sum(df_compt, "total_charges"), _sum(df_compt, "resultat")
    hc_prod, hc_chg, hc_tot          = _sum(df_hc, "total_produits"),    _sum(df_hc, "total_charges"),    _sum(df_hc, "resultat")

    # Coherence REEL = COMPTABLE + HORS_COMPTA (D035 cumul 1.00 EUR)
    ecart_ident = abs(reel_tot - (compt_tot + hc_tot))
    comment_reel = (
        f"REEL=COMPTABLE+HC verifie (ecart={ecart_ident:.2f} EUR)"
        if ecart_ident <= 1.00 else
        f"!! RUPTURE REEL != COMPTABLE+HC (ecart={ecart_ident:.2f} EUR)"
    )
    comment_hc = "Aucun flux HC" if hc_tot == 0.0 and hc_prod == 0.0 and hc_chg == 0.0 else "Flux HC presents"

    df_global = pd.DataFrame([
        {"vision": "REEL",        "total_produits": reel_prod,  "total_charges": reel_chg,  "resultat": reel_tot,  "commentaire_hc": comment_reel},
        {"vision": "COMPTABLE",   "total_produits": compt_prod, "total_charges": compt_chg, "resultat": compt_tot, "commentaire_hc": "Vision comptable (IC)"},
        {"vision": "HORS_COMPTA", "total_produits": hc_prod,    "total_charges": hc_chg,    "resultat": hc_tot,    "commentaire_hc": comment_hc},
    ])

    wb2 = Workbook()
    wb2.remove(wb2.active)
    _write_sheet(wb2, "PAR_MOIS_LOGEMENT",    df_par_ml)
    _write_sheet(wb2, "PAR_MOIS_PROPRIETAIRE", df_par_mp)
    _write_sheet(wb2, "GLOBAL",                df_global)
    wb2.save(RESULT_FILE)
    log.info(f"  {RESULT_FILE.name} : OK")

    # ── MASTER_CALC_NetProprietaire ──
    wb3 = Workbook()
    wb3.remove(wb3.active)
    _write_sheet(wb3, "EXPLOITATION", df_exploit)
    _write_sheet(wb3, "REGLEMENT",   df_reg)
    _write_sheet(wb3, "VUE_MOIS",    df_vue)
    wb3.save(NET_FILE)
    log.info(f"  {NET_FILE.name} : OK")


# ─────────────────────────────────────────────────────────────────────────────
# 7. Controls report (19 points)
# ─────────────────────────────────────────────────────────────────────────────
def print_controls(
    df_flux, df_comm, df_ac, df_reel, df_compt, df_hc,
    df_exploit, df_reg, df_vue, cfix_controls, hh_controls, df_payout,
):
    sep = "=" * 65
    log.info(sep)
    log.info("RAPPORT DE CONTROLES LOT 10")
    log.info(sep)

    for col in ["payout_calcule", "menage_retenu", "assiette_commission",
                "commission_conciergerie", "net_proprietaire"]:
        if col in df_comm.columns:
            df_comm[col] = pd.to_numeric(df_comm[col], errors="coerce").fillna(0.0)

    total_payout   = df_comm["payout_calcule"].sum()        if len(df_comm) > 0 else 0.0
    total_menage   = df_comm["menage_retenu"].sum()          if len(df_comm) > 0 else 0.0
    total_assiette = df_comm["assiette_commission"].sum()    if len(df_comm) > 0 else 0.0
    total_comm_cci = df_comm["commission_conciergerie"].sum()if len(df_comm) > 0 else 0.0
    total_net_avt  = df_comm["net_proprietaire"].sum()       if len(df_comm) > 0 else 0.0

    if len(df_reg) > 0:
        df_reg["charge_fixe_mensuelle"] = pd.to_numeric(
            df_reg.get("charge_fixe_mensuelle", 0), errors="coerce"
        ).fillna(0.0)
        df_reg["net_proprietaire_apres_charge_mois"] = pd.to_numeric(
            df_reg.get("net_proprietaire_apres_charge_mois", 0), errors="coerce"
        ).fillna(0.0)
        total_cfix    = df_reg["charge_fixe_mensuelle"].sum()
        total_net_apr = df_reg["net_proprietaire_apres_charge_mois"].sum()
    else:
        total_cfix    = 0.0
        total_net_apr = 0.0

    reel_tot  = df_reel["resultat"].sum()  if len(df_reel) > 0 else 0.0
    compt_tot = df_compt["resultat"].sum() if len(df_compt) > 0 else 0.0
    hc_tot    = df_hc["resultat"].sum()    if len(df_hc) > 0 else 0.0
    ecart_ident = abs(reel_tot - (compt_tot + hc_tot))

    # Charges globales / non affectables visibles (D-LOT10C-03)
    n_global_lines = 0
    if len(df_reel) > 0 and "logement_id" in df_reel.columns:
        n_global_lines = int((df_reel["logement_id"] == SENTINEL_GLOBAL).sum())

    # Acomptes injectes (REGLEMENT)
    tot_acomptes = 0.0
    if len(df_reg) > 0 and "autres_acomptes_recus" in df_reg.columns:
        tot_acomptes = pd.to_numeric(df_reg["autres_acomptes_recus"], errors="coerce").fillna(0.0).sum()

    n_hh = int((df_comm["source_type"] == "HH").sum()) if "source_type" in df_comm.columns else 0
    n_hh_ctrl = len(hh_controls)

    n_inco = sum(1 for c in cfix_controls
                 if c["code_anomalie"] == "CHARGE_FIXE_DATE_ENTREE_GESTION_INCOHERENTE")
    n_sans = sum(1 for c in cfix_controls
                 if c["code_anomalie"] == "LOG_SANS_FLUX_017")

    n_reel_log  = len(df_reel[df_reel["vision"] == "REEL"]) if len(df_reel) > 0 else 0
    n_reel_prop = len(
        df_reel.groupby(["mois", "proprietaire_id"]).size()
    ) if len(df_reel) > 0 else 0

    ctrs = [
        ("CTR-LOT10-01", "Flux lus depuis MASTER_CALC_Flux",
         f"{len(df_flux)}"),
        ("CTR-LOT10-02", "Reservations NORMAL integrees aux commissions",
         f"{len(df_comm)}"),
        ("CTR-LOT10-03", "Reservations A_CONTROLER exclues",
         f"{len(df_ac)}"),
        ("CTR-LOT10-04", "Total payout calcule (NORMAL)",
         f"{total_payout:,.2f} EUR"),
        ("CTR-LOT10-05", "Total menage retenu",
         f"{total_menage:,.2f} EUR"),
        ("CTR-LOT10-06", "Total assiette commission",
         f"{total_assiette:,.2f} EUR"),
        ("CTR-LOT10-07", "Total commission conciergerie",
         f"{total_comm_cci:,.2f} EUR"),
        ("CTR-LOT10-08", "Total net proprietaire avant charge fixe",
         f"{total_net_avt:,.2f} EUR"),
        ("CTR-LOT10-09", "Total charge fixe mensuelle generee",
         f"{total_cfix:,.2f} EUR"),
        ("CTR-LOT10-10", "Total net proprietaire apres charge fixe",
         f"{total_net_apr:,.2f} EUR"),
        ("CTR-LOT10-11", "Resultat REEL global (Flux)",
         f"{reel_tot:,.2f} EUR"),
        ("CTR-LOT10-12", "Resultat COMPTABLE global (Flux)",
         f"{compt_tot:,.2f} EUR"),
        ("CTR-LOT10-13", "Resultat HORS_COMPTA global (Flux)",
         f"{hc_tot:,.2f} EUR"),
        ("CTR-LOT10-14", "Lignes resultats PAR_MOIS_LOGEMENT (REEL)",
         f"{n_reel_log}"),
        ("CTR-LOT10-15", "Lignes resultats PAR_MOIS_PROPRIETAIRE (REEL)",
         f"{n_reel_prop}"),
        ("CTR-LOT10-16", "CHARGE_FIXE_DATE_ENTREE_GESTION_INCOHERENTE",
         f"{n_inco} logements"),
        ("CTR-LOT10-17", "LOG_SANS_FLUX_017 (forfait>0 sans reservation Flux)",
         f"{n_sans} logements"),
        ("CTR-LOT10-18", "Controles BLOQUANTS detectes",
         "0 (execution terminee sans sys.exit)"),
        ("CTR-LOT10-19", "Sources amont (lecture seule)",
         "FLUX / RES / PAYOUT / HH / ACOMPTES / REF_SETUP non modifies"),
        ("CTR-LOT10-20", "Identite REEL = COMPTABLE + HORS_COMPTA",
         f"{'OK' if ecart_ident <= 1.00 else 'RUPTURE'} (ecart={ecart_ident:.2f} EUR)"),
        ("CTR-LOT10-21", "Commissions HH integrees (montant saisi)",
         f"{n_hh}"),
        ("CTR-LOT10-22", "Controles HH (ecart commission / sans montant)",
         f"{n_hh_ctrl}"),
        ("CTR-LOT10-23", "Lignes charges GLOBAL_NON_AFFECTE (REEL)",
         f"{n_global_lines}"),
        ("CTR-LOT10-24", "Total acomptes injectes (REGLEMENT seulement)",
         f"{tot_acomptes:,.2f} EUR"),
    ]

    for code, desc, val in ctrs:
        log.info(f"  {code}  {desc:<50s}  {val}")

    # BLOQUANT defensif : rupture identite REEL = COMPTABLE + HC
    if ecart_ident > 1.00:
        log.error(f"BLOQUANT REEL_DIFF_COMPTABLE_PLUS_HC — ecart={ecart_ident:.2f} EUR")
        sys.exit(1)

    if hh_controls:
        log.info("  --- Controles HH ---")
        for c in hh_controls:
            log.info(f"    [{c['niveau']}] {c['code_anomalie']}: {c['message']}")

    log.info(sep)
    log.info("ATTENTE VALIDATION HUMAINE avant commit.")
    log.info(sep)


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────
def main():
    log.info("=" * 65)
    log.info("LOT 10 — Calcul resultats, commissions, net proprietaire")
    log.info(f"Date : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    log.info("=" * 65)

    df_flux, df_res, df_payout, df_hh, df_acc, df_log, df_prop = load_sources()

    df_comm, df_ac, hh_controls     = build_commissions(df_flux, df_res, df_payout, df_hh, df_log, df_prop)
    df_cfix, cfix_controls          = build_charge_fixe(df_flux, df_log)
    df_reel, df_compt, df_hc        = build_resultats(df_flux)
    df_exploit, df_reg, df_vue      = build_net_proprietaire(df_comm, df_cfix, df_acc)

    write_all(df_comm, df_ac, df_reel, df_compt, df_hc, df_exploit, df_reg, df_vue)
    print_controls(
        df_flux, df_comm, df_ac, df_reel, df_compt, df_hc,
        df_exploit, df_reg, df_vue, cfix_controls, hh_controls, df_payout,
    )


if __name__ == "__main__":
    main()
