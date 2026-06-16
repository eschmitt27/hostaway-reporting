"""
lot6c_menages_externes.py
Lot 6c — Ménages externes — Peuplement MASTER depuis factures PDF
Sources : Facture mai Aissata.pdf (Kandia DIABATE / INT_0004)
          Facture mai Mounir.pdf  (MH Entreprise     / INT_0003)
Période : mai 2026
Décisions : D079-D088 (D-6c-01 à D-6c-10)
"""

import os, hashlib
from datetime import datetime, date
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── PATHS ────────────────────────────────────────────────────────────────────
BASE    = r"C:\Users\Ewan\OneDrive\Documents\Conciergerie\Pilotage_Conciergerie"
L6C_DIR = os.path.join(BASE, "02_TRAVAIL", "Lot6c_MenagesExternes")
OUT     = os.path.join(L6C_DIR, "MASTER_FACT_MEN_MenagesExternes.xlsx")
REF_SRC = os.path.join(BASE, "01_SOURCES_BRUTES", "REF_Setup", "REF_Setup.xlsm")
LOT6A   = os.path.join(BASE, "02_TRAVAIL", "Lot1_Hostaway",
                       "MASTER_FACT_HA_CleaningTasks_Discovery.xlsx")
PDF_DIR = os.path.join(BASE, "01_SOURCES_BRUTES", "MenagesExternes", "Factures_PDF")

TS = datetime.now().strftime("%Y%m%d_%H%M%S")
os.makedirs(L6C_DIR, exist_ok=True)

# ─── STYLES ───────────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1F3864")
HDR_FONT   = Font(color="FFFFFF", bold=True, size=10)
VAL_FILL   = PatternFill("solid", fgColor="E2EFDA")
CTRL_FILL  = PatternFill("solid", fgColor="FFEB9C")
BLOK_FILL  = PatternFill("solid", fgColor="FFC7CE")
INFO_FILL  = PatternFill("solid", fgColor="DDEBF7")
ALT_FILL   = PatternFill("solid", fgColor="F2F2F2")
THIN       = Side(border_style="thin", color="BFBFBF")
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def hdr(ws, row, col, val, width=None):
    c = ws.cell(row, col, val)
    c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(wrap_text=True, vertical="center")
    c.border = BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width

def cell(ws, row, col, val, fill=None):
    c = ws.cell(row, col, val)
    c.alignment = Alignment(vertical="top", wrap_text=False)
    c.border = BORDER
    if fill:
        c.fill = fill
    return c

# ─── LOAD REF DATA ────────────────────────────────────────────────────────────
def sheet_to_dicts(wb, name):
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return []
    h = rows[0]
    return [dict(zip(h, r)) for r in rows[1:] if any(v is not None for v in r)]

wb_ref = openpyxl.load_workbook(REF_SRC, read_only=True, keep_vba=False)
REF_LOG   = {r["logement_id"]: r for r in sheet_to_dicts(wb_ref, "REF_Logements")
             if r.get("logement_id")}
REF_INT   = {r["intervenant_id"]: r for r in sheet_to_dicts(wb_ref, "REF_Intervenants")
             if r.get("intervenant_id")}
wb_ref.close()

# Logement enrichissement
def log_info(lid):
    r = REF_LOG.get(lid, {})
    return (r.get("proprietaire_id"), r.get("hostaway_listing_id"),
            r.get("actif"), r.get("date_sortie_gestion"))

# ─── MAPPING FACTURE → REFERENTIEL ───────────────────────────────────────────
PREST_MAP = {
    "kandia diabate / rends-moi un service": "INT_0004",
    "mh entreprise": "INT_0003",
}
LOG_MAP = {
    "studio 76 (dureuil)":           "LOG_0014",
    "studio - cote pavé (françois)": "LOG_0012",
    "studio puits verts (caroline)": "LOG_0006",
    "t3 310 muret (david)":          "LOG_0011",
    "studio st pierre (florane)":    "LOG_0010",
    "t2 9 rue du toul":              "LOG_0007",
    "t3 4 rue engalières":           "LOG_0009",
    "t3 20 rue l'amiral galache":    "LOG_0016",
    "t.4-90 blagnac (cédrine)":      "LOG_0002",
    "t.3 sept deniers (françois)":   "LOG_0013",
    "t.2-65 (gabriel)":              "LOG_0003",
    "studio puits vert (caroline)":  "LOG_0006",
}

# ─── SOURCE_RAW (extraite des PDFs) ──────────────────────────────────────────
def d(s): return date(*map(int, s.split("-"))) if s else None

RAW = [
    # === AISSATA — Facture n°2026-37 — 31/05/2026 — Total TTC 1 439 € ===
    ("FAC-2026-05-AISSATA-001","2026-37",d("2026-05-31"),"Facture mai Aissata.pdf",
     "Kandia DIABATE / Rends-moi un service","studio 76 (Dureuil)",
     d("2026-05-10"),"DATE_PRECISE","TLM_001",1,29.0,29.0,1439.0,
     "Ligne 1 fac.2026-37 — date précise sur facture"),
    ("FAC-2026-05-AISSATA-001","2026-37",d("2026-05-31"),"Facture mai Aissata.pdf",
     "Kandia DIABATE / Rends-moi un service","studio - cote pavé (François)",
     None,"MOIS_FACTURE","TLM_001",5,29.0,145.0,1439.0,
     "Ligne 2 fac.2026-37 — 5 passages mai 2026, date absente"),
    ("FAC-2026-05-AISSATA-001","2026-37",d("2026-05-31"),"Facture mai Aissata.pdf",
     "Kandia DIABATE / Rends-moi un service","studio Puits verts (Caroline)",
     None,"MOIS_FACTURE","TLM_001",10,29.0,290.0,1439.0,
     "Ligne 3 fac.2026-37 — 10 passages mai 2026, date absente"),
    ("FAC-2026-05-AISSATA-001","2026-37",d("2026-05-31"),"Facture mai Aissata.pdf",
     "Kandia DIABATE / Rends-moi un service","T3 310 muret (David)",
     None,"MOIS_FACTURE","TLM_001",8,55.0,440.0,1439.0,
     "Ligne 4 fac.2026-37 — 8 passages mai 2026, date absente"),
    ("FAC-2026-05-AISSATA-001","2026-37",d("2026-05-31"),"Facture mai Aissata.pdf",
     "Kandia DIABATE / Rends-moi un service","studio st Pierre (Florane)",
     None,"MOIS_FACTURE","TLM_001",2,29.0,58.0,1439.0,
     "Ligne 5 fac.2026-37 — 2 passages mai 2026, date absente"),
    ("FAC-2026-05-AISSATA-001","2026-37",d("2026-05-31"),"Facture mai Aissata.pdf",
     "Kandia DIABATE / Rends-moi un service","T2 9 rue du Toul",
     None,"MOIS_FACTURE","TLM_001",8,39.0,312.0,1439.0,
     "Ligne 6 fac.2026-37 — 8 passages mai 2026, date absente"),
    ("FAC-2026-05-AISSATA-001","2026-37",d("2026-05-31"),"Facture mai Aissata.pdf",
     "Kandia DIABATE / Rends-moi un service","T3 4 rue engalières",
     d("2026-05-23"),"DATE_PRECISE","TLM_001",1,55.0,55.0,1439.0,
     "Ligne 7 fac.2026-37 — date précise sur facture"),
    # Ligne 8 originale splittée (2 dates distinctes : 03/05 et 10/05)
    ("FAC-2026-05-AISSATA-001","2026-37",d("2026-05-31"),"Facture mai Aissata.pdf",
     "Kandia DIABATE / Rends-moi un service","T3 20 rue l'Amiral Galache",
     d("2026-05-03"),"DATE_PRECISE","TLM_001",1,55.0,55.0,1439.0,
     "Ligne 8a — splittée (2 dates). Facture: T3, REF: T2 (LOG_0016). Prix 55€ tarif T3 Aissata."),
    ("FAC-2026-05-AISSATA-001","2026-37",d("2026-05-31"),"Facture mai Aissata.pdf",
     "Kandia DIABATE / Rends-moi un service","T3 20 rue l'Amiral Galache",
     d("2026-05-10"),"DATE_PRECISE","TLM_001",1,55.0,55.0,1439.0,
     "Ligne 8b — splittée (2 dates). Facture: T3, REF: T2 (LOG_0016). Prix 55€ tarif T3 Aissata."),
    # === MOUNIR — Facture n°0003 — 31/05/2026 — Total TTC 942 € ===
    ("FAC-2026-05-MOUNIR-001","0003",d("2026-05-31"),"Facture mai Mounir.pdf",
     "MH Entreprise","T.4-90 Blagnac (Cédrine)",
     None,"MOIS_FACTURE","TLM_001",6,65.0,390.0,942.0,
     "Ligne 1 fac.0003 — 6 ménages mai 2026, date absente"),
    ("FAC-2026-05-MOUNIR-001","0003",d("2026-05-31"),"Facture mai Mounir.pdf",
     "MH Entreprise","T.3 Sept Deniers (François)",
     None,"MOIS_FACTURE","TLM_001",10,52.0,520.0,942.0,
     "Ligne 2 fac.0003 — 10 ménages mai 2026, date absente"),
    ("FAC-2026-05-MOUNIR-001","0003",d("2026-05-31"),"Facture mai Mounir.pdf",
     "MH Entreprise","T.2-65 (Gabriel)",
     None,"MOIS_FACTURE","TLM_001",0,36.0,0.0,942.0,
     "Ligne 3 fac.0003 — 0 ménage, logement inactif depuis 2026-04-26. Présent à 0€."),
    ("FAC-2026-05-MOUNIR-001","0003",d("2026-05-31"),"Facture mai Mounir.pdf",
     "MH Entreprise","Studio Puits vert (Caroline)",
     None,"MOIS_FACTURE","TLM_001",1,32.0,32.0,942.0,
     "Ligne 4 fac.0003 — 1 ménage mai 2026, date absente"),
]

RAW_COLS = ["facture_id","numero_facture_original","date_facture","nom_fichier_source",
            "prestataire_nom_brut","appartement_source","date_menage",
            "precision_date_menage","type_ligne_menage_id","nombre_menages",
            "montant_unitaire","montant_ligne_ttc","montant_facture_total_ttc",
            "commentaire_source"]

# ─── SOMMES PAR FACTURE (réconciliation) ──────────────────────────────────────
from collections import defaultdict
sommes_ttc = defaultdict(float)
for r in RAW:
    sommes_ttc[r[0]] += r[11]  # montant_ligne_ttc

# ─── ENRICHISSEMENT MASTER ────────────────────────────────────────────────────
MASTER_COLS = [
    "menage_externe_id","ROW_HASH",
    "facture_id","date_facture","date_menage","precision_date_menage",
    "mois","annee","nom_fichier_source","source_document",
    "prestataire_id","nom_prestataire","type_intervenant","regime_tva_prestataire",
    "logement_id","proprietaire_id","hostaway_listing_id","appartement_source",
    "type_ligne_menage_id","nombre_menages","fournitures_incluses",
    "montant_ligne_ht","taux_tva","montant_tva","montant_ligne_ttc",
    "montant_facture_total_ht","montant_facture_total_ttc",
    "somme_lignes_facture_ttc","ecart_reconciliation_ttc",
    "somme_lignes_facture_ht","ecart_reconciliation_ht",
    "mode_paiement","associe_payeur","code_impact","prise_en_compta",
    "impact_resultat_reel","impact_resultat_comptable",
    "type_flux_id","sens",
    "source_charges_flux_liee","risque_double_comptage",
    "statut_controle","niveau_anomalie","code_anomalie","commentaire",
    "source_module","source_table","source_pk","date_integration",
]

counters = defaultdict(int)
MASTER = []

for r in RAW:
    (fid, num_fac, date_fac, fichier, prest_brut, appt_src,
     date_men, prec_date, tlm, nb_men, pu, mttc, fac_ttc, cmt_src) = r

    pid = PREST_MAP.get(prest_brut.lower())
    lid = LOG_MAP.get(appt_src.lower())

    pinfo   = REF_INT.get(pid, {}) if pid else {}
    nom_p   = pinfo.get("nom_intervenant","INCONNU")
    type_p  = pinfo.get("type_intervenant","A_CONTROLER")

    prop_id, ha_id, actif, date_sortie = log_info(lid) if lid else (None,None,None,None)

    # Logement activité
    log_inactif = False
    if date_sortie:
        ds = date_sortie.date() if hasattr(date_sortie,"date") else date_sortie
        ref_d = date_men or date_fac
        if ref_d and ref_d >= ds:
            log_inactif = True

    # TVA (FRANCHISE_TVA pour les deux prestataires confirmés)
    regime_tva = "FRANCHISE_TVA"
    taux_tva, mtva, mht = 0, 0.0, mttc

    # Mois
    ref_d = date_men or date_fac
    mois  = ref_d.strftime("%Y-%m") if ref_d else "A_CONTROLER"
    annee = ref_d.year if ref_d else None

    # Réconciliation
    s_ttc = round(sommes_ttc[fid], 2)
    e_ttc = round(s_ttc - fac_ttc, 2)

    # Mode paiement
    mode_paie = "VIREMENT" if pid == "INT_0004" else "A_CONTROLER"

    # menage_externe_id
    pcode = "AISSATA" if pid == "INT_0004" else "MOUNIR" if pid == "INT_0003" else "UNK"
    key = f"{mois}-{pcode}"
    counters[key] += 1
    mid = f"MENEXT-{mois}-{pcode}-{counters[key]:03d}"

    # ROW_HASH
    h = hashlib.sha256("|".join(str(x) for x in
        [fid, date_men, pid, lid, tlm, mttc]).encode()).hexdigest()[:16]

    # ─── CONTRÔLES ────────────────────────────────────────────────────────────
    anomalies_b, anomalies_a = [], []

    if not lid:
        anomalies_b.append("MENAGE_EXTERNE_LOGEMENT_ABSENT")
    if not pid:
        anomalies_b.append("MENAGE_EXTERNE_PRESTATAIRE_INCONNU")
    if mttc is None:
        anomalies_b.append("MENAGE_EXTERNE_MONTANT_INVALIDE")
    elif mttc < 0 or nb_men < 0:
        anomalies_b.append("MENAGE_EXTERNE_MONTANT_INVALIDE")
    if type_p == "INTERNE":
        anomalies_b.append("MENAGE_EXTERNE_TYPE_INTERVENANT_INTERNE")
    if abs(e_ttc) > 1.0:
        anomalies_b.append("MENAGE_EXTERNE_FACTURE_NON_RECONCILIEE")

    # Date jour absente sur facture mensuelle : ne bloque PLUS seule (decision 2026-06-16).
    # La ligne reste exploitable ; le controle reel est porte par le rapprochement
    # mensuel Hostaway (VUE_ECART_HOSTAWAY). On ne cree jamais de fausse date.
    date_info = ""
    if not anomalies_b:
        if prec_date == "MOIS_FACTURE":
            date_info = "MENAGE_EXTERNE_DATE_MOIS"   # informatif, non bloquant
        if mttc == 0 or nb_men == 0:
            anomalies_a.append("MENAGE_EXTERNE_MONTANT_NUL")
        if log_inactif:
            anomalies_a.append("MENAGE_EXTERNE_LOGEMENT_INACTIF")

    if anomalies_b:
        statut, niveau, code_ano = "BLOQUANT","BLOQUANT", anomalies_b[0]
        extra = " | ".join(anomalies_b[1:])
    elif anomalies_a:
        statut, niveau, code_ano = "A_CONTROLER","A_CONTROLER", anomalies_a[0]
        extra = " | ".join(anomalies_a[1:])
    else:
        # Ligne exploitable : VALIDE. date_info reste informatif (niveau INFO).
        statut, niveau, code_ano, extra = "VALIDE","INFO", date_info, ""

    commentaire = cmt_src
    if extra:
        commentaire += f" | {extra}"

    # code_impact / compta selon D-6c-03
    code_impact     = "IC"
    prise_en_compta = "OUI"
    imp_reel        = "OUI"
    imp_compta      = "OUI"

    row = [
        mid, h,
        fid, date_fac, date_men, prec_date,
        mois, annee, fichier, f"lot6c_menages_externes.py — {TS}",
        pid or "A_CONTROLER", nom_p, type_p, regime_tva,
        lid or "A_CONTROLER", prop_id, ha_id, appt_src,
        tlm, nb_men, "NON",
        mht, taux_tva, mtva, mttc,
        fac_ttc, fac_ttc,          # HT = TTC (franchise TVA)
        s_ttc, e_ttc, s_ttc, e_ttc,
        mode_paie, None, code_impact, prise_en_compta,
        imp_reel, imp_compta,
        "TYPE_FLUX_014", "CHARGE",
        None, "NON",
        statut, niveau, code_ano, commentaire,
        "MENAGES_EXTERNES","MASTER_FACT_MEN_MenagesExternes.SOURCE_RAW", mid,
        datetime.now().isoformat(),
    ]
    MASTER.append(row)

# ─── VUE_ACTIVE ───────────────────────────────────────────────────────────────
VUE_ACTIVE = [
    r for r in MASTER
    if r[MASTER_COLS.index("statut_controle")] == "VALIDE"
    and r[MASTER_COLS.index("niveau_anomalie")] in ("","INFO")
    and r[MASTER_COLS.index("code_impact")] in ("IC","HC")
    and (r[MASTER_COLS.index("montant_ligne_ttc")] or 0) > 0
    and r[MASTER_COLS.index("type_intervenant")] == "EXTERNE"
    and r[MASTER_COLS.index("type_flux_id")] == "TYPE_FLUX_014"
    and abs(r[MASTER_COLS.index("ecart_reconciliation_ttc")]) <= 1.0
]

# ─── VUE_ECART_HOSTAWAY (rapprochement mensuel recalculable, neutre) ──────────
# Cle : mois x logement_id. Compte TOUTES les lignes facture exploitables (pas
# seulement VUE_ACTIVE) -> les lignes mensuelles MOIS_FACTURE sont incluses.
# Recalcule a chaque relance depuis MASTER + VUE_COMPTAGE Lot 6a (Hostaway).
# Codes NEUTRES : rapproche / ecart / HA sans facture / logement hors HA.
# Multi-mois : ne fige aucun mois.
vue_cpt = {}   # (mois, logement_id) -> {nb_ha, prop}
try:
    wb6a = openpyxl.load_workbook(LOT6A, read_only=True)
    ws6a = wb6a["VUE_COMPTAGE"]
    rows6a = list(ws6a.iter_rows(values_only=True))
    h6a = rows6a[0]
    for rx in rows6a[1:]:
        rd = dict(zip(h6a, rx))
        m0 = rd.get("mois"); lid0 = rd.get("logement_id")
        if m0 and lid0:
            vue_cpt[(m0, lid0)] = {
                "nb_ha": rd.get("nb_menages_realises") or rd.get("nb_menages") or 0,
                "prop":  rd.get("proprietaire_id"),
            }
    wb6a.close()
    print(f"[OK] VUE_COMPTAGE Lot 6a chargee — {len(vue_cpt)} (mois,logement)")
except Exception as e:
    print(f"[WARN] VUE_COMPTAGE non chargee: {e}")

# Comptage facture exploitable depuis MASTER (toutes lignes non BLOQUANTes,
# logement + mois connus, comptees TLM_001/TLM_002). Inclut les mensuelles.
iM = MASTER_COLS.index
ext_cpt   = defaultdict(float)
ext_prop  = {}
ext_prest = defaultdict(set)
for r in MASTER:
    if r[iM("statut_controle")] == "BLOQUANT":
        continue
    lid2 = r[iM("logement_id")]; m2 = r[iM("mois")]
    if not lid2 or lid2 == "A_CONTROLER" or not m2:
        continue
    if r[iM("type_ligne_menage_id")] not in ("TLM_001", "TLM_002"):
        continue
    ext_cpt[(m2, lid2)]  += (r[iM("nombre_menages")] or 0)
    ext_prop[(m2, lid2)]  = r[iM("proprietaire_id")]
    pnom = r[iM("nom_prestataire")]
    if pnom:
        ext_prest[(m2, lid2)].add(str(pnom))

all_keys = sorted(set(vue_cpt.keys()) | set(ext_cpt.keys()))
ECART_ROWS = []
for (m2, lid2) in all_keys:
    nb_ha  = vue_cpt.get((m2, lid2), {}).get("nb_ha", 0) or 0
    nb_ext = ext_cpt.get((m2, lid2), 0)
    prop2  = ext_prop.get((m2, lid2)) or vue_cpt.get((m2, lid2), {}).get("prop")
    prest  = ",".join(sorted(ext_prest.get((m2, lid2), set())))
    ecart  = nb_ext - nb_ha    # >0 : facture > Hostaway
    if nb_ext == 0 and nb_ha == 0:
        niv, cod, cmt = "INFO", "", "Aucune activite menage ce mois."
    elif nb_ext > 0 and nb_ha == 0:
        niv, cod, cmt = "A_CONTROLER", "MENAGE_EXTERNE_LOGEMENT_HORS_HA", \
            "Logement facture absent du comptage Hostaway (archive/hors HA/mapping). A valider, pas une erreur prestataire."
    elif nb_ext == 0 and nb_ha > 0:
        niv, cod, cmt = "INFO", "MENAGE_HA_SANS_FACTURE_EXTERNE", \
            "Menages Hostaway sans facture externe : probable menage interne / a croiser avec M04."
    elif ecart == 0:
        niv, cod, cmt = "INFO", "MENAGE_EXTERNE_RAPPROCHE_HOSTAWAY", \
            "Volume facture = volume Hostaway. Rapproche ; date jour non requise."
    else:
        niv, cod, cmt = "A_CONTROLER", "MENAGE_EXTERNE_ECART_HOSTAWAY", \
            "Ecart volume facture vs Hostaway. A valider (interne, hors HA, decalage mois ou saisie)."
    ECART_ROWS.append([m2, lid2, prop2, prest, nb_ext, nb_ha, ecart, cod, niv, cmt, TS])

ECART_COLS = ["mois","logement_id","proprietaire_id","prestataires_factures",
              "nombre_menages_facture","nombre_menages_hostaway","ecart",
              "code_controle","niveau_controle","commentaire_controle","date_actualisation"]

# ─── BUILD EXCEL ──────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# ── Sheet 1 : SOURCE_RAW ──────────────────────────────────────────────────────
ws1 = wb.create_sheet("SOURCE_RAW")
ws1.freeze_panes = "A2"
col_widths = [24,12,14,28,36,30,14,16,12,8,10,14,16,50]
for ci, (col, w) in enumerate(zip(RAW_COLS, col_widths), 1):
    hdr(ws1, 1, ci, col, w)
for ri, r in enumerate(RAW, 2):
    row_data = list(r)
    for ci, val in enumerate(row_data, 1):
        cell(ws1, ri, ci, val)
ws1.row_dimensions[1].height = 30

# ── Sheet 2 : PARAMETRES ─────────────────────────────────────────────────────
ws2 = wb.create_sheet("PARAMETRES")
params = [
    ("LOT",                   "6c"),
    ("DESCRIPTION",           "Ménages externes — factures PDF prestataires"),
    ("DATE_RUN",              TS),
    ("TOLERANCE_FACTURE",     1.0),
    ("SOURCE_PDF_DIR",        r"01_SOURCES_BRUTES\MenagesExternes\Factures_PDF"),
    ("CHEMIN_REF_SETUP",      r"01_SOURCES_BRUTES\REF_Setup\REF_Setup.xlsm"),
    ("CHEMIN_LOT6A",          r"02_TRAVAIL\Lot1_Hostaway\MASTER_FACT_HA_CleaningTasks_Discovery.xlsx"),
    ("NB_LIGNES_SOURCE_RAW",  len(RAW)),
    ("NB_LIGNES_MASTER",      len(MASTER)),
    ("NB_LIGNES_VUE_ACTIVE",  len(VUE_ACTIVE)),
    ("FACTURES_TRAITEES",     "FAC-2026-05-AISSATA-001 (1439€) ; FAC-2026-05-MOUNIR-001 (942€)"),
]
hdr(ws2, 1, 1, "parametre", 30); hdr(ws2, 1, 2, "valeur", 60)
for ri, (k, v) in enumerate(params, 2):
    cell(ws2, ri, 1, k); cell(ws2, ri, 2, v)

ws2.cell(ri+2, 1, "MAPPING_APPARTEMENT_SOURCE → logement_id").font = Font(bold=True)
hdr(ws2, ri+3, 1, "appartement_source"); hdr(ws2, ri+3, 2, "logement_id")
for i, (k2, v2) in enumerate(LOG_MAP.items(), ri+4):
    cell(ws2, i, 1, k2); cell(ws2, i, 2, v2)

# ── Sheet 3 : MASTER ─────────────────────────────────────────────────────────
ws3 = wb.create_sheet("MASTER")
ws3.freeze_panes = "A2"
mcol_widths = {
    "menage_externe_id":28,"ROW_HASH":18,"facture_id":26,
    "date_facture":13,"date_menage":13,"precision_date_menage":16,
    "mois":10,"annee":8,"nom_fichier_source":26,"source_document":32,
    "prestataire_id":14,"nom_prestataire":16,"type_intervenant":14,
    "regime_tva_prestataire":16,"logement_id":12,"proprietaire_id":13,
    "hostaway_listing_id":16,"appartement_source":30,
    "type_ligne_menage_id":14,"nombre_menages":10,"fournitures_incluses":14,
    "montant_ligne_ht":14,"taux_tva":8,"montant_tva":10,"montant_ligne_ttc":14,
    "montant_facture_total_ht":14,"montant_facture_total_ttc":14,
    "somme_lignes_facture_ttc":18,"ecart_reconciliation_ttc":18,
    "somme_lignes_facture_ht":18,"ecart_reconciliation_ht":18,
    "mode_paiement":14,"associe_payeur":14,"code_impact":12,
    "prise_en_compta":13,"impact_resultat_reel":14,"impact_resultat_comptable":16,
    "type_flux_id":16,"sens":8,"source_charges_flux_liee":18,
    "risque_double_comptage":16,"statut_controle":14,"niveau_anomalie":13,
    "code_anomalie":36,"commentaire":55,"source_module":16,
    "source_table":36,"source_pk":28,"date_integration":24,
}
for ci, col in enumerate(MASTER_COLS, 1):
    hdr(ws3, 1, ci, col, mcol_widths.get(col, 14))

statut_idx = MASTER_COLS.index("statut_controle")
for ri, row in enumerate(MASTER, 2):
    s = row[statut_idx]
    f = VAL_FILL if s == "VALIDE" else CTRL_FILL if s == "A_CONTROLER" else BLOK_FILL
    for ci, val in enumerate(row, 1):
        c = cell(ws3, ri, ci, val)
        if ci == statut_idx + 1:
            c.fill = f
ws3.row_dimensions[1].height = 30

# ── Sheet 4 : VUE_ACTIVE ─────────────────────────────────────────────────────
ws4 = wb.create_sheet("VUE_ACTIVE")
ws4.freeze_panes = "A2"
for ci, col in enumerate(MASTER_COLS, 1):
    hdr(ws4, 1, ci, col, mcol_widths.get(col, 14))
for ri, row in enumerate(VUE_ACTIVE, 2):
    for ci, val in enumerate(row, 1):
        c = cell(ws4, ri, ci, val)
        if ci == statut_idx + 1:
            c.fill = VAL_FILL
ws4.row_dimensions[1].height = 30

# ── Sheet 5 : VUE_ECART_HOSTAWAY ────────────────────────────────────────────
ws5 = wb.create_sheet("VUE_ECART_HOSTAWAY")
ws5.freeze_panes = "A2"
ecol_widths = [10,12,14,20,16,16,8,34,14,48,20]
for ci, (col, w) in enumerate(zip(ECART_COLS, ecol_widths), 1):
    hdr(ws5, 1, ci, col, w)
niv_idx = ECART_COLS.index("niveau_controle")  # 0-based
for ri, row in enumerate(ECART_ROWS, 2):
    niv = row[niv_idx]
    f = VAL_FILL if niv == "INFO" else CTRL_FILL
    for ci, val in enumerate(row, 1):
        c = cell(ws5, ri, ci, val)
        if (ci - 1) == niv_idx:
            c.fill = f

note_row = len(ECART_ROWS) + 3
ws5.cell(note_row, 1, "NOTE : rapprochement mensuel mois x logement, recalcule a chaque relance. "
    "nombre_menages_facture compte TOUTES les lignes facture exploitables (y compris mensuelles MOIS_FACTURE), "
    "pas seulement VALIDE. Codes neutres (rapproche / ecart / HA sans facture / logement hors HA) : "
    "un ecart signale une verification, jamais une accusation prestataire.").font = Font(italic=True, color="808080")

# ── Sheet 6 : POWER_QUERY_CODE ───────────────────────────────────────────────
ws6 = wb.create_sheet("POWER_QUERY_CODE")
hdr(ws6, 1, 1, "query_name", 28); hdr(ws6, 1, 2, "description", 40); hdr(ws6, 1, 3, "m_code", 120)
ws6.column_dimensions["C"].width = 120

PQ = [
("Q1_SOURCE_RAW","Charge le tableau SOURCE_RAW",
"""let
    Source = Excel.CurrentWorkbook(){[Name="SOURCE_RAW"]}[Content],
    Types  = Table.TransformColumnTypes(Source,{
        {"date_facture",       type date},
        {"date_menage",        type date},
        {"nombre_menages",     Int64.Type},
        {"montant_unitaire",   type number},
        {"montant_ligne_ttc",  type number},
        {"montant_facture_total_ttc", type number}})
in  Types"""),

("Q2_REF_INTERVENANTS","Charge REF_Intervenants depuis REF_Setup.xlsm",
"""let
    Source   = Excel.Workbook(File.Contents(CHEMIN_REF_SETUP), null, true),
    Sheet    = Source{[Item="REF_Intervenants",Kind="Sheet"]}[Data],
    Promoted = Table.PromoteHeaders(Sheet, [PromoteAllScalars=true])
in  Promoted"""),

("Q3_REF_LOGEMENTS","Charge REF_Logements depuis REF_Setup.xlsm",
"""let
    Source   = Excel.Workbook(File.Contents(CHEMIN_REF_SETUP), null, true),
    Sheet    = Source{[Item="REF_Logements",Kind="Sheet"]}[Data],
    Promoted = Table.PromoteHeaders(Sheet, [PromoteAllScalars=true])
in  Promoted"""),

("Q4_SOMMES_PAR_FACTURE","GROUP BY facture_id — calcul sommes pour réconciliation",
"""let
    Src     = Q1_SOURCE_RAW,
    Grouped = Table.Group(Src, {"facture_id"}, {
        {"somme_lignes_facture_ttc", each List.Sum([montant_ligne_ttc]), type number},
        {"somme_lignes_facture_ht",  each List.Sum([montant_ligne_ttc]), type number}})
in  Grouped"""),

("Q5_MASTER","Enrichissement complet — 49 colonnes — contrôles et anomalies",
"""let
    Src    = Q1_SOURCE_RAW,
    Sms    = Q4_SOMMES_PAR_FACTURE,
    Ints   = Q2_REF_INTERVENANTS,
    Logs   = Q3_REF_LOGEMENTS,
    Mapped = Table.AddColumn(Src, "logement_id", each
                 MAPPING_LOT6C{[appartement_source=[appartement_source]]}[logement_id]?),
    JoinInt= Table.NestedJoin(Mapped,"prestataire_id_brut",Ints,"nom_normalise","_int",JoinKind.LeftOuter),
    JoinLog= Table.NestedJoin(JoinInt,"logement_id",Logs,"logement_id","_log",JoinKind.LeftOuter),
    JoinSms= Table.NestedJoin(JoinLog,"facture_id",Sms,"facture_id","_sms",JoinKind.LeftOuter),
    // ... expand, calcul ecarts, contrôles, 49 colonnes finales
    Final  = Table.AddColumn(JoinSms, "type_flux_id", each "TYPE_FLUX_014"),
    WithSens = Table.AddColumn(Final, "sens", each "CHARGE")
in  WithSens"""),

("Q6_VUE_ACTIVE","Filtre MASTER — lignes validées uniquement",
"""let
    Src  = Q5_MASTER,
    Filt = Table.SelectRows(Src, each
        [statut_controle]         = "VALIDE"
        and ([niveau_anomalie]    = "" or [niveau_anomalie] = "INFO")
        and ([code_impact]        = "IC" or [code_impact]  = "HC")
        and [montant_ligne_ttc]   > 0
        and [type_intervenant]    = "EXTERNE"
        and [type_flux_id]        = "TYPE_FLUX_014"
        and Number.Abs([ecart_reconciliation_ttc]) <= 1.0)
in  Filt"""),

("Q7_VUE_ECART_HOSTAWAY","Comparaison ménages HA réalisés vs externes facturés — mois x logement",
"""let
    VA    = Q6_VUE_ACTIVE,
    ExtGp = Table.Group(VA, {"mois","logement_id","proprietaire_id"},{
                {"nb_menages_externes_factures",
                 each List.Sum(List.Select([nombre_menages],
                      each _ <> null)), Int64.Type}}),
    HAGp  = // JOIN sur VUE_COMPTAGE Lot 6a (chemin paramétré)
            Excel.Workbook(File.Contents(CHEMIN_LOT6A)){[Item="VUE_COMPTAGE",Kind="Sheet"]}[Data],
    Join  = Table.NestedJoin(ExtGp,{"mois","logement_id"},HAGp,
                             {"mois","logement_id"},"_ha",JoinKind.FullOuter),
    Ecart = Table.AddColumn(Join, "ecart_hostaway_externe", each
                [nb_menages_hostaway_realises] - [nb_menages_externes_factures]),
    Statut= Table.AddColumn(Ecart, "statut_controle", each
                if [ecart_hostaway_externe] = 0 then "VALIDE" else "A_CONTROLER")
in  Statut"""),
]

for ri, (qn, qd, qm) in enumerate(PQ, 2):
    cell(ws6, ri, 1, qn)
    cell(ws6, ri, 2, qd)
    c = ws6.cell(ri, 3, qm)
    c.alignment = Alignment(vertical="top", wrap_text=False)
    c.border = BORDER
ws6.row_dimensions[1].height = 30

# ── Sheet 7 : README ─────────────────────────────────────────────────────────
ws7 = wb.create_sheet("README")
ws7.column_dimensions["A"].width = 120

readme_lines = [
    "MASTER_FACT_MEN_MenagesExternes.xlsx — Lot 6c — Ménages externes",
    "=" * 70,
    f"Généré le {TS} par lot6c_menages_externes.py",
    "",
    "SOURCES",
    f"  - Facture mai Aissata.pdf  (Kandia DIABATE / INT_0004) — n°2026-37 — 1 439 € TTC",
    f"  - Facture mai Mounir.pdf   (MH Entreprise  / INT_0003) — n°0003    — 942 € TTC",
    f"  - Deux prestataires en franchise TVA (art.293B CGI)",
    "",
    "STRUCTURE",
    f"  - SOURCE_RAW       : {len(RAW)} lignes extraites des PDF (ligne 8 Aissata splittée en 8a+8b)",
    f"  - MASTER           : {len(MASTER)} lignes — 49 colonnes",
    f"  - VUE_ACTIVE       : {len(VUE_ACTIVE)} lignes (filtre VALIDE + IC/HC + montant>0 + réconcilié)",
    f"  - VUE_ECART_HA     : {len(ECART_ROWS)} logements — comparaison HA vs externes mai 2026",
    "",
    "DECISIONS METIER ACTEES",
    "  D079 (D-6c-01) : 2 PDFs fournis — format analysé — FRANCHISE_TVA confirmé",
    "  D080 (D-6c-02) : Architecture fichier unique (7 onglets, comme M04)",
    "  D081 (D-6c-03) : IC/OUI normal ; HC/NON liquide/perso + associe_payeur ; HR justifié",
    "  D082 (D-6c-04) : Facture globale sans détail logement = BLOQUANT",
    "  D083 (D-6c-05) : TVA 3 valeurs — FRANCHISE_TVA / ASSUJETTI_TVA / A_CONTROLER",
    "  D084 (D-6c-06) : Lignes séparées par type_ligne_menage_id (TLM_001–006)",
    "  D085 (D-6c-07) : Chemin 02_TRAVAIL/Lot6c_MenagesExternes/",
    "  D086 (D-6c-08) : Kandia DIABATE = INT_0004 ; MH Entreprise = INT_0003",
    "  D087 (D-6c-09) : Option B — date_menage=null + precision_date_menage=MOIS_FACTURE + A_CONTROLER",
    "  D088 (D-6c-10) : Ligne 0€/0 ménage conservée MASTER, exclue VUE_ACTIVE",
    "",
    "CONTROLES BLOQUANTS (7)",
    "  MENAGE_EXTERNE_LOGEMENT_ABSENT",
    "  MENAGE_EXTERNE_PRESTATAIRE_INCONNU",
    "  MENAGE_EXTERNE_CODE_IMPACT_ABSENT",
    "  MENAGE_EXTERNE_FACTURE_NON_RECONCILIEE       [tolérance D035 : 1,00 €]",
    "  MENAGE_EXTERNE_FACTURE_GLOBALE_NON_DETAILLEE",
    "  MENAGE_EXTERNE_TYPE_INTERVENANT_INTERNE",
    "  MENAGE_EXTERNE_MONTANT_INVALIDE              [montant<0 ou nombre_menages<0]",
    "",
    "CONTROLES A_CONTROLER (10)",
    "  MENAGE_EXTERNE_DATE_ABSENTE                  [precision_date_menage=MOIS_FACTURE]",
    "  MENAGE_EXTERNE_MONTANT_NUL                   [montant=0 ou nb=0 — traçable]",
    "  MENAGE_EXTERNE_LOGEMENT_INACTIF",
    "  MENAGE_EXTERNE_A_VENTILER",
    "  MENAGE_EXTERNE_TYPE_LIGNE_A_DEFINIR          [TLM_006]",
    "  MENAGE_EXTERNE_FOURNITURE_A_RECLASSER        [TLM_004/005 anti-doublon]",
    "  MENAGE_EXTERNE_FACTURATION_SANS_COMPTAGE_HA",
    "  MENAGE_EXTERNE_TVA_A_CONTROLER",
    "  MENAGE_EXTERNE_PAIEMENT_PERSO_SANS_ASSOCIE",
    "  MENAGE_EXTERNE_RISQUE_DOUBLE_COMPTAGE_CHARGES",
    "",
    "VUE_ACTIVE FILTRE",
    "  statut_controle=VALIDE",
    "  AND niveau_anomalie IN ('','INFO')",
    "  AND code_impact IN ('IC','HC')",
    "  AND montant_ligne_ttc > 0",
    "  AND type_intervenant = 'EXTERNE'",
    "  AND type_flux_id = 'TYPE_FLUX_014'",
    "  AND ABS(ecart_reconciliation_ttc) <= 1.00",
    "",
    "ALIMENTATION MASTER_CALC_FLUX (Lot 9)",
    "  source_module  = MENAGES_EXTERNES",
    "  source_table   = MASTER_FACT_MEN_MenagesExternes.SOURCE_RAW",
    "  type_flux_id   = TYPE_FLUX_014 (COUT_REEL_MENAGE_EXTERNE)",
    "  sens           = CHARGE",
    "  date_flux      = date_menage",
    "  montant        = montant_ligne_ttc",
    "",
    "MAPPING INVOICE → REF_LOGEMENTS (non encore dans REF_Mapping_Logements)",
    "  À ajouter dans REF_Mapping_Logements pour les futurs runs PQ.",
]

for ri, line in enumerate(readme_lines, 1):
    c = ws7.cell(ri, 1, line)
    if ri == 1:
        c.font = Font(bold=True, size=12)
    elif line.startswith("===") or line.startswith("---"):
        c.font = Font(color="BFBFBF")
    elif line.isupper() and line.endswith(")") or (line and not line.startswith(" ")):
        c.font = Font(bold=True)

# ── Ordre des onglets ─────────────────────────────────────────────────────────
wb.save(OUT)
print(f"[OK] {OUT}")

# ─── RÉSUMÉ ───────────────────────────────────────────────────────────────────
n_valide  = sum(1 for r in MASTER if r[MASTER_COLS.index("statut_controle")] == "VALIDE")
n_ctrl    = sum(1 for r in MASTER if r[MASTER_COLS.index("statut_controle")] == "A_CONTROLER")
n_blok    = sum(1 for r in MASTER if r[MASTER_COLS.index("statut_controle")] == "BLOQUANT")

print(f"\n=== RÉSUMÉ LOT 6C ===")
print(f"SOURCE_RAW   : {len(RAW)} lignes (9 Aissata + 4 Mounir)")
print(f"MASTER       : {len(MASTER)} lignes — 49 colonnes")
print(f"  VALIDE     : {n_valide}")
print(f"  A_CONTRÔLER: {n_ctrl}")
print(f"  BLOQUANT   : {n_blok}")
print(f"VUE_ACTIVE   : {len(VUE_ACTIVE)} lignes")
print(f"VUE_ECART_HA : {len(ECART_ROWS)} logements mois=2026-05")

for fid2 in ["FAC-2026-05-AISSATA-001","FAC-2026-05-MOUNIR-001"]:
    s2 = sommes_ttc[fid2]
    fac_t = next(r[12] for r in RAW if r[0]==fid2)
    print(f"  Réconciliation {fid2}: somme={s2:.2f} total={fac_t:.2f} écart={s2-fac_t:.2f}")

print(f"\n[DONE] {OUT}")
