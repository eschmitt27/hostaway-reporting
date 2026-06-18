"""
lot6e_gainperte_menages.py — PARTIE 2 : écart gain/perte vs coût standard ménage
================================================================================
Vue ANALYTIQUE pure (D105). N'écrit RIEN dans Flux / résultats. Ne touche pas
M04 / SAISIE_Charges_Flux / banque / VRBO / lot9-12.

Calcul de BASE uniquement (validé) :
  ecart_total = cout_standard_total - cout_reel_total
  >0 GAIN | <0 PERTE | =0 EQUILIBRE

Méthodes (D101/D102) :
  EXTERNE_FACTURE            : coût réel = montant facturé (lot6c)
  INTERNE_HEURES_M04         : mois <= 2026-05, coût réel = nb_heures × taux PARAM_004
  INTERNE_STANDARD_PARAMETRE : mois >= 2026-06, coût réel = nb × REF_Couts_Menage_Interne
  NON_CALCULABLE             : donnée manquante

HORS PÉRIMÈTRE de cette version (étape avancée séparée) : cave/local, courses,
consommables, lavage, quote-parts. S'ils existent -> seulement signalés en contrôle,
JAMAIS intégrés au calcul.

DRY-RUN : sortie 02_TRAVAIL/Lot6_DryRun/DRYRUN_GainPerte_Menages.xlsx
Périmètre courant : MOIS = 2026-05.
"""

import sys, os, glob, hashlib, datetime, collections, warnings
warnings.filterwarnings("ignore")
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import openpyxl
from openpyxl.styles import Font, PatternFill

MONTH = "2026-05"
PIVOT = "2026-06"          # >= pivot : méthode interne paramétrée
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REF  = os.path.join(ROOT, "01_SOURCES_BRUTES", "REF_Setup", "REF_Setup.xlsm")
OUTD = os.path.join(ROOT, "02_TRAVAIL", "Lot6e_GainPerte_Menages")
OUT  = os.path.join(OUTD, "MASTER_CALC_GainPerte_Menages.xlsx")
DRY_M04 = os.path.join(ROOT, "02_TRAVAIL", "Lot6b_DeclarationsInternes", "MASTER_NORM_Declarations_Internes.xlsx")

def sh(p, s):
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True); ws = wb[s]
    rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]; wb.close()
    return [dict(zip([str(c) for c in rows[0]], r)) for r in rows[1:]]

def rowhash(*v):
    return hashlib.sha256("|".join("" if x is None else str(x) for x in v).encode()).hexdigest()[:16]

def to_d(v):
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, datetime.date): return v
    try: return datetime.date.fromisoformat(str(v)[:10])
    except (ValueError, TypeError): return None

DREF = datetime.date.fromisoformat(MONTH + "-01")

# ── Référentiels ─────────────────────────────────────────────────────────────
log_info = {d["logement_id"]: d for d in sh(REF, "REF_Logements") if d.get("logement_id") and str(d["logement_id"]) != "logement_id"}
typ_lib  = {d["type_logement_id"]: d.get("type_logement") for d in sh(REF, "REF_Types_Logements")}
int_info = {d["intervenant_id"]: d for d in sh(REF, "REF_Intervenants")}

def cout_standard_unit(type_id):
    best = None
    for d in sh_std:
        if d.get("type_logement_id") != type_id or str(d.get("actif")) != "OUI":
            continue
        deb, fin = to_d(d.get("date_debut_validite")), to_d(d.get("date_fin_validite"))
        if deb and DREF < deb: continue
        if fin and DREF > fin: continue
        best = d.get("cout_standard_menage")
    return best
sh_std = sh(REF, "REF_Couts_Standards_Menage")

def cout_interne_unit(type_id):
    best = None
    for d in sh_int:
        if d.get("type_logement_id") != type_id or str(d.get("actif")) != "OUI":
            continue
        deb, fin = to_d(d.get("date_debut_validite")), to_d(d.get("date_fin_validite"))
        if deb and DREF < deb: continue
        if fin and DREF > fin: continue
        best = d.get("montant_interne_standard")
    return best
sh_int = sh(REF, "REF_Couts_Menage_Interne") if "REF_Couts_Menage_Interne" in openpyxl.load_workbook(REF, read_only=True).sheetnames else []

# taux horaire (PARAM_004), pas en dur
taux_horaire = None
for d in sh(REF, "REF_Parametres_Generaux"):
    if d.get("nom_parametre") == "TAUX_HORAIRE_MENAGE_INTERNE":
        try: taux_horaire = float(d.get("valeur"))
        except (TypeError, ValueError): pass

rows_out = []
controls = []

def add(mois, lg, iid, typ_interv, nb, nb_h, methode, reel_total, code="", comm=""):
    tinfo = log_info.get(lg) or {}
    type_id = tinfo.get("type_logement_id")
    std_u = cout_standard_unit(type_id)
    statut_c, niveau, ctrl = "VALIDE", "INFO", code
    if lg is None: statut_c, niveau, ctrl = "A_CONTROLER", "A_CONTROLER", "LOGEMENT_NON_MAPPE"
    elif type_id is None: statut_c, niveau, ctrl = "A_CONTROLER", "A_CONTROLER", "TYPE_LOGEMENT_ABSENT"
    elif std_u is None: statut_c, niveau, ctrl = "A_CONTROLER", "A_CONTROLER", "COUT_STANDARD_ABSENT"
    std_total = (std_u or 0) * nb if std_u is not None else None
    if reel_total is None or std_u is None:
        ecart, statut_e = None, "NON_CALCULABLE"
        if ctrl == "": ctrl = code or "NON_CALCULABLE"
        statut_c = "A_CONTROLER" if statut_c == "VALIDE" else statut_c
    else:
        ecart = round(std_total - reel_total, 2)
        statut_e = "GAIN" if ecart > 0 else ("PERTE" if ecart < 0 else "EQUILIBRE")
    rows_out.append({
        "mois": mois, "nom_appartement": tinfo.get("nom_logement_officiel"), "logement_id": lg,
        "type_logement_id": type_id, "type_logement_libelle": typ_lib.get(type_id),
        "intervenant_id": iid, "nom_intervenant": (int_info.get(iid) or {}).get("nom_intervenant"),
        "type_intervenant": typ_interv,
        "nb_menages": nb, "nb_heures": nb_h,
        "cout_standard_unitaire": std_u, "cout_standard_total": std_total,
        "methode_cout_reel": methode,
        "cout_reel_unitaire": round(reel_total / nb, 2) if (reel_total is not None and nb) else None,
        "cout_reel_total": reel_total,
        "ecart_total": ecart, "statut_ecart": statut_e,
        "statut_controle": statut_c, "code_controle": ctrl, "commentaire": comm,
        "ROW_HASH": rowhash(mois, lg, iid, methode, nb, reel_total),
    })

# ── A. EXTERNE_FACTURE (lot6c) ───────────────────────────────────────────────
fc = glob.glob(os.path.join(ROOT, "02_TRAVAIL", "**", "MASTER_FACT_MEN_MenagesExternes.xlsx"), recursive=True)[0]
ext_agg = collections.defaultdict(lambda: [0, 0.0])   # (lg, prestataire) -> [nb, montant]
for d in sh(fc, "MASTER"):
    if str(d.get("mois"))[:7] != MONTH: continue
    if str(d.get("type_ligne_menage_id")) not in ("TLM_001", "TLM_002"): continue
    q = d.get("nombre_menages") or 0
    m = d.get("montant_ligne_ttc") or 0
    if (q or 0) == 0 and (m or 0) == 0:
        controls.append(("EXCLU_VOLUME", "INFO", f"facture {d.get('facture_id')} ligne 0€/q0 exclue")); continue
    e = ext_agg[(d.get("logement_id"), d.get("prestataire_id"))]; e[0] += q; e[1] += m
for (lg, iid), (nb, mont) in ext_agg.items():
    add(MONTH, lg, iid, "EXTERNE", nb, None, "EXTERNE_FACTURE", round(mont, 2),
        comm="Coût réel = montant facturé TTC prestataire")

# ── B/C. INTERNE (source dry-run sheet) ──────────────────────────────────────
internal_method = "INTERNE_HEURES_M04" if MONTH < PIVOT else "INTERNE_STANDARD_PARAMETRE"
if os.path.exists(DRY_M04):
    int_agg = collections.defaultdict(lambda: [0, 0.0, False])  # nb, heures, has_h
    for d in sh(DRY_M04, "MASTER_NORMALISE"):
        if str(d.get("mois"))[:7] != MONTH: continue
        k = (d.get("logement_id"), d.get("intervenant_id"))
        int_agg[k][0] += d.get("nb_menages") or 0
        h = d.get("nb_heures")
        if h is not None: int_agg[k][1] += h; int_agg[k][2] = True
    for (lg, iid), (nb, heures, has_h) in int_agg.items():
        type_id = (log_info.get(lg) or {}).get("type_logement_id")
        if internal_method == "INTERNE_HEURES_M04":
            if not has_h or taux_horaire is None:
                add(MONTH, lg, iid, "INTERNE", nb, heures if has_h else None, "NON_CALCULABLE", None,
                    code=("TAUX_HORAIRE_ABSENT" if taux_horaire is None else "HEURES_M04_ABSENTES"),
                    comm="Heures ou taux indisponibles"); continue
            add(MONTH, lg, iid, "INTERNE", nb, heures, "INTERNE_HEURES_M04", round(heures * taux_horaire, 2),
                comm=f"Coût réel = {heures}h × {taux_horaire}€")
        else:
            ci = cout_interne_unit(type_id)
            if ci is None:
                add(MONTH, lg, iid, "INTERNE", nb, None, "NON_CALCULABLE", None,
                    code="COUT_INTERNE_TYPE_LOGEMENT_ABSENT", comm="Coût interne paramétré absent"); continue
            add(MONTH, lg, iid, "INTERNE", nb, None, "INTERNE_STANDARD_PARAMETRE", round(nb * ci, 2),
                comm=f"Coût réel = {nb} × {ci}€ (paramétré)")
else:
    controls.append(("SOURCE_INTERNE_ABSENTE", "A_CONTROLER", "DRYRUN_M04 normalisé absent"))

# ── Résumés ──────────────────────────────────────────────────────────────────
def resume(keyf):
    agg = collections.defaultdict(lambda: [0, 0.0, 0.0, 0.0, True])  # nb, std, reel, ecart, calc
    for r in rows_out:
        k = keyf(r); a = agg[k]; a[0] += r["nb_menages"] or 0
        if r["ecart_total"] is None: a[4] = False; continue
        a[1] += r["cout_standard_total"] or 0; a[2] += r["cout_reel_total"] or 0; a[3] += r["ecart_total"]
    return agg
res_log = resume(lambda r: (r["logement_id"], r["nom_appartement"]))
res_int = resume(lambda r: (r["intervenant_id"], r["nom_intervenant"], r["type_intervenant"]))

# contrôles de synthèse
for r in rows_out:
    if r["statut_ecart"] == "NON_CALCULABLE":
        controls.append((r["code_controle"] or "NON_CALCULABLE", "A_CONTROLER", f"{r['logement_id']}/{r['intervenant_id']}"))
controls.append(("DOUBLE_COMPTAGE_RAPPEL", "INFO", "Vue analytique : aucun montant réinjecté dans Flux/REEL/COMPTABLE/HC"))

# ── Écriture ──────────────────────────────────────────────────────────────────
os.makedirs(OUTD, exist_ok=True)
wb = openpyxl.Workbook()
def wsheet(title, cols, rows):
    ws = wb.create_sheet(title) if title != wb.active.title else wb.active
    ws.append(cols)
    for c in ws[1]: c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="DDDDDD")
    for r in rows: ws.append([r.get(c) if isinstance(r, dict) else r[i] for i, c in enumerate(cols)])
    return ws
wb.active.title = "DETAIL_ECART_COUT"
DCOLS = ["mois","nom_appartement","logement_id","type_logement_id","type_logement_libelle",
    "intervenant_id","nom_intervenant","type_intervenant","nb_menages","nb_heures",
    "cout_standard_unitaire","cout_standard_total","methode_cout_reel","cout_reel_unitaire",
    "cout_reel_total","ecart_total","statut_ecart","statut_controle","code_controle","commentaire"]
wsheet("DETAIL_ECART_COUT", DCOLS, rows_out)
wsheet("RESUME_LOGEMENT", ["mois","logement_id","nom_appartement","nb_menages","cout_standard_total","cout_reel_total","ecart_total","calculable"],
    [{"mois":MONTH,"logement_id":k[0],"nom_appartement":k[1],"nb_menages":v[0],"cout_standard_total":round(v[1],2),"cout_reel_total":round(v[2],2),"ecart_total":round(v[3],2),"calculable":v[4]} for k,v in sorted(res_log.items())])
wsheet("RESUME_INTERVENANT", ["mois","intervenant_id","nom_intervenant","type_intervenant","nb_menages","cout_standard_total","cout_reel_total","ecart_total","calculable"],
    [{"mois":MONTH,"intervenant_id":k[0],"nom_intervenant":k[1],"type_intervenant":k[2],"nb_menages":v[0],"cout_standard_total":round(v[1],2),"cout_reel_total":round(v[2],2),"ecart_total":round(v[3],2),"calculable":v[4]} for k,v in sorted(res_int.items())])
cc = collections.Counter((c[0], c[1]) for c in controls)
wsheet("CONTROLES", ["code_controle","niveau","nb","exemple"],
    [{"code_controle":k[0],"niveau":k[1],"nb":n,"exemple":next(c[2] for c in controls if (c[0],c[1])==k)} for k,n in cc.most_common()])
try:
    wb.save(OUT)
except PermissionError:
    OUT = OUT.replace(".xlsx","_MAJ.xlsx"); wb.save(OUT); print(f"[lot6e] original verrouillé -> {os.path.basename(OUT)}")

# ── Rapport ──────────────────────────────────────────────────────────────────
tot_std = sum(r["cout_standard_total"] or 0 for r in rows_out if r["ecart_total"] is not None)
tot_reel = sum(r["cout_reel_total"] or 0 for r in rows_out if r["ecart_total"] is not None)
tot_ec = round(tot_std - tot_reel, 2)
nc = sum(1 for r in rows_out if r["statut_ecart"] == "NON_CALCULABLE")
print(f"[lot6e] DRY-RUN mois={MONTH} taux_horaire={taux_horaire} -> {OUT}")
print(f"  DETAIL_ECART_COUT : {len(rows_out)} lignes | NON_CALCULABLE : {nc}")
print(f"  GAIN/PERTE GLOBAL : standard={round(tot_std,2)} reel={round(tot_reel,2)} ecart={tot_ec} ({'GAIN' if tot_ec>0 else 'PERTE' if tot_ec<0 else 'EQUILIBRE'})")
print("\n  par intervenant:")
for k,v in sorted(res_int.items()):
    print(f"    {str(k[0]):12} {str(k[1])[:9]:9} {str(k[2]):8} nb={v[0]:3} std={round(v[1],2):8} reel={round(v[2],2):8} ecart={round(v[3],2):8} {'' if v[4] else '(partiel)'}")
print("\n  par logement:")
for k,v in sorted(res_log.items()):
    print(f"    {str(k[0]):8} {str(k[1])[:24]:24} nb={v[0]:3} std={round(v[1],2):8} reel={round(v[2],2):8} ecart={round(v[3],2):8}")
print("\n  CONTROLES:", dict(cc))
