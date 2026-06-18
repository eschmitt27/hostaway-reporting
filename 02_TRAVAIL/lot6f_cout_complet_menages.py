"""
lot6f_cout_complet_menages.py — Coût complet ménage AVANCÉ (vue analytique)
================================================================================
Vue ANALYTIQUE pure (D105). N'écrit RIEN dans Flux / Resultats / Commissions /
NetProprietaire / Factures. Ne modifie pas M04 / SAISIE_Charges_Flux / REF_Setup /
banque / VRBO. Ne relance pas lot9-12.

cout_complet_total = cout_direct_total + Σ quote_parts (local + lavage + courses + conso + autres)
ecart_vs_standard_total = cout_standard_total - cout_complet_total   ( >0 GAIN | <0 PERTE | =0 EQUILIBRE )

Périmètre ACCEPTÉ (§6) :
  - direct externe = montant facture prestataire (lot6c)
  - direct interne <=2026-05 = heures × taux PARAM_004 ; >=2026-06 = forfait REF_Couts_Menage_Interne
  - LOCAL_CAVE = REC_002 (date-aware) ventilé sur tous les ménages du mois
  - LAVAGE interne = Google Sheet / M04 (attribuable par appart + non-attribuable par intervenant)
  - COURSES/CONSO/ACHATS = SAISIE_Charges_Flux (affectable_menage=OUI) -> vide aujourd'hui => 0 POOL_VIDE_NON_SAISI
HORS périmètre (étape ultérieure) : heures cave, heures courses, consommables non saisis.
Pas de logique 'fournitures_incluses'. L'affectation dépend de affectable_menage + intervenant_concerne.

Clé de ventilation (D103) : poids = nb_menages × cout_standard_unitaire.

DRY-RUN : 02_TRAVAIL/Lot6_DryRun/DRYRUN_CoutComplet_Menages.xlsx — MOIS = 2026-05.
"""

import sys, os, io, csv, glob, subprocess, hashlib, datetime, collections, unicodedata, warnings
warnings.filterwarnings("ignore")
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import openpyxl
from openpyxl.styles import Font, PatternFill

MONTH = "2026-05"; PIVOT = "2026-06"
DREF = datetime.date.fromisoformat(MONTH + "-01")
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REF  = os.path.join(ROOT, "01_SOURCES_BRUTES", "REF_Setup", "REF_Setup.xlsm")
SAISIE = os.path.join(ROOT, "01_SOURCES_BRUTES", "Charges", "SAISIE_Charges_Flux.xlsx")
OUTD = os.path.join(ROOT, "02_TRAVAIL", "Lot6f_CoutComplet_Menages")
OUT  = os.path.join(OUTD, "MASTER_CALC_CoutComplet_Menages.xlsx")
def _m04_url():
    """URL Google Sheet M04 lue depuis REF_Setup > REF_Sources_Systeme (pas d'URL en dur)."""
    wb = openpyxl.load_workbook(REF, read_only=True, data_only=True); ws = wb["REF_Sources_Systeme"]
    rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]; wb.close()
    H = [str(c) for c in rows[0]]
    for r in rows[1:]:
        d = dict(zip(H, r))
        if str(d.get("nom_source")) == "GOOGLE_SHEET_M04_DECLARATIONS" and str(d.get("actif")) == "OUI":
            u = str(d.get("dossier_source") or "").strip()
            if u.startswith("http"): return u
    raise SystemExit("[BLOQUANT lot6f] URL GOOGLE_SHEET_M04_DECLARATIONS absente de REF_Sources_Systeme (SRC_011).")
SHEET_URL = _m04_url()

def sh(p, s):
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    if s not in wb.sheetnames: wb.close(); return []
    ws = wb[s]; rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]; wb.close()
    return [dict(zip([str(c) for c in rows[0]], r)) for r in rows[1:]]

def norm(s):
    s = str(s or "").strip().lower()
    return " ".join("".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn").split())

def rh(*v): return hashlib.sha256("|".join("" if x is None else str(x) for x in v).encode()).hexdigest()[:16]
def to_d(v):
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, datetime.date): return v
    try: return datetime.date.fromisoformat(str(v)[:10])
    except (ValueError, TypeError): return None
def f(x):
    try: return float(x)
    except (TypeError, ValueError): return None

# ── Référentiels ─────────────────────────────────────────────────────────────
log_info = {d["logement_id"]: d for d in sh(REF, "REF_Logements") if d.get("logement_id") and str(d["logement_id"]) != "logement_id"}
typ_lib  = {d["type_logement_id"]: d.get("type_logement") for d in sh(REF, "REF_Types_Logements")}
int_info = {d["intervenant_id"]: d for d in sh(REF, "REF_Intervenants")}
std_ref  = sh(REF, "REF_Couts_Standards_Menage")
int_ref  = sh(REF, "REF_Couts_Menage_Interne")
rec_ref  = sh(REF, "REF_Charges_Recurrentes")
taux = next((f(d.get("valeur")) for d in sh(REF, "REF_Parametres_Generaux") if d.get("nom_parametre") == "TAUX_HORAIRE_MENAGE_INTERNE"), None)

def date_aware(rows, type_id, montant_field, type_field="type_logement_id"):
    best = None
    for d in rows:
        if d.get(type_field) != type_id or str(d.get("actif")) != "OUI": continue
        deb, fin = to_d(d.get("date_debut_validite")), to_d(d.get("date_fin_validite"))
        if deb and DREF < deb: continue
        if fin and DREF > fin: continue
        best = f(d.get(montant_field))
    return best
def std_unit(type_id): return date_aware(std_ref, type_id, "cout_standard_menage")
def int_unit(type_id): return date_aware(int_ref, type_id, "montant_interne_standard")

# REC_002 local cave date-aware
def rec_montant(rec_id):
    best = None
    for d in rec_ref:
        if d.get("charge_recurrente_id") != rec_id or str(d.get("actif")) != "OUI": continue
        deb, fin = to_d(d.get("date_debut_validite")), to_d(d.get("date_fin_validite"))
        if deb and DREF < deb: continue
        if fin and DREF > fin: continue
        best = f(d.get("montant_ttc"))
    return best
local_cave_montant = rec_montant("REC_002") or 0.0

# mapping libellé appart -> logement_id
lmap = {}
for d in sh(REF, "REF_Mapping_Logements"):
    if d.get("valeur_source"): lmap[norm(d["valeur_source"])] = d.get("logement_id")

controls = []

# ── Direct EXTERNE (lot6c) ───────────────────────────────────────────────────
fc = glob.glob(os.path.join(ROOT, "02_TRAVAIL", "**", "MASTER_FACT_MEN_MenagesExternes.xlsx"), recursive=True)[0]
ext = collections.defaultdict(lambda: [0, 0.0])
for d in sh(fc, "MASTER"):
    if str(d.get("mois"))[:7] != MONTH: continue
    if str(d.get("type_ligne_menage_id")) not in ("TLM_001", "TLM_002"): continue
    q = d.get("nombre_menages") or 0; m = d.get("montant_ligne_ttc") or 0
    if (q or 0) == 0 and (m or 0) == 0:
        controls.append(("EXCLU_VOLUME", "INFO", f"facture {d.get('facture_id')} 0€/q0")); continue
    e = ext[(d.get("logement_id"), d.get("prestataire_id"))]; e[0] += q; e[1] += m

# ── Interne + LAVAGE (Google Sheet) ──────────────────────────────────────────
INTMAP = {"imene": "INT_0001", "kira": "INT_0002", "kheira": "INT_0002"}
MOIS = {"janvier":"01","fevrier":"02","mars":"03","avril":"04","mai":"05","juin":"06","juillet":"07","aout":"08","septembre":"09","octobre":"10","novembre":"11","decembre":"12"}
txt = subprocess.run(["curl","-sL","-A","Mozilla/5.0",SHEET_URL], capture_output=True, text=True, encoding="utf-8").stdout
srows = list(csv.reader(io.StringIO(txt))); shdr = srows[0]
i_pre = shdr.index("Prénom"); i_mois = shdr.index("Mois des ménages"); i_an = shdr.index("Année des ménages")
appcols = [i for i, h in enumerate(shdr) if h.strip() == "Appartement"]
# colonne lavage non-attribuable (queue)
i_lav_na = next((i for i, h in enumerate(shdr) if h.strip().startswith("Coûts de lavage du linge (hors")), None)

interne = collections.defaultdict(lambda: [0, 0.0, 0.0])   # (lg,iid) -> [nb, heures, lavage_attribuable]
lav_na_by_int = collections.Counter()                       # iid -> lavage non-attribuable du mois
for r in srows[1:]:
    if not any(c.strip() for c in r): continue
    pre = r[i_pre].strip(); mm = MOIS.get(norm(r[i_mois])); yr = r[i_an].strip()
    miso = f"{yr}-{mm}" if (mm and yr) else None
    if miso != MONTH: continue
    iid = INTMAP.get(norm(pre))
    for ci in appcols:
        app = r[ci].strip() if ci < len(r) else ""
        if not app: continue
        lg = lmap.get(norm(app))
        nb = int(f(r[ci+1]) or 0) if (ci+1 < len(r) and r[ci+1].strip()) else 0
        h  = f(r[ci+2]) if (ci+2 < len(r) and r[ci+2].strip()) else 0
        lav = f(r[ci+3]) if (ci+3 < len(r) and r[ci+3].strip()) else 0
        e = interne[(lg, iid)]; e[0] += nb; e[1] += (h or 0); e[2] += (lav or 0)
    if i_lav_na is not None and i_lav_na < len(r) and r[i_lav_na].strip():
        lav_na_by_int[iid] += f(r[i_lav_na]) or 0

# ── Construction lignes de base (direct + standard) ──────────────────────────
lines = []   # dict par (mois,lg,iid)
def base_line(lg, iid, typ_interv, nb, heures, methode, direct, lav_attr=0.0):
    ti = log_info.get(lg) or {}; type_id = ti.get("type_logement_id")
    su = std_unit(type_id)
    poids = (nb * su) if (su is not None) else 0
    lines.append({"mois": MONTH, "logement_id": lg, "nom_appartement": ti.get("nom_logement_officiel"),
        "proprietaire_id": ti.get("proprietaire_id"),
        "type_logement_id": type_id, "type_logement_libelle": typ_lib.get(type_id),
        "intervenant_id": iid, "nom_intervenant": (int_info.get(iid) or {}).get("nom_intervenant"),
        "type_intervenant": typ_interv, "nb_menages": nb, "nb_heures": heures,
        "cout_standard_unitaire": su, "cout_standard_total": (su*nb if su is not None else None),
        "methode": methode, "cout_direct_total": direct, "poids": poids,
        "lavage_attribuable": lav_attr})

for (lg, iid), (nb, mont) in ext.items():
    base_line(lg, iid, "EXTERNE", nb, None, "EXTERNE_FACTURE", round(mont, 2))
internal_method = "INTERNE_HEURES_M04" if MONTH < PIVOT else "INTERNE_STANDARD_PARAMETRE"
for (lg, iid), (nb, heures, lav) in interne.items():
    if internal_method == "INTERNE_HEURES_M04":
        direct = round(heures * taux, 2) if (heures and taux is not None) else 0.0
    else:
        ti = (log_info.get(lg) or {}).get("type_logement_id"); iu = int_unit(ti)
        direct = round(nb * iu, 2) if iu is not None else 0.0
    base_line(lg, iid, "INTERNE", nb, heures, internal_method, direct, lav_attr=round(lav, 2))

# ── POOLS de charges communes ────────────────────────────────────────────────
sum_poids_all = sum(l["poids"] for l in lines) or 1
# La cave/local ne sert QU'AUX ménages internes -> ventilée sur le poids interne uniquement.
sum_poids_interne = sum(l["poids"] for l in lines if l["type_intervenant"] == "INTERNE") or 1
sum_poids_int = collections.Counter()
for l in lines:
    if l["type_intervenant"] == "INTERNE": sum_poids_int[l["intervenant_id"]] += l["poids"]

# SAISIE courses/conso/achats (affectable_menage=OUI) — vide aujourd'hui
saisie_rows = sh(SAISIE, "SAISIE")
saisie_has_col = bool(saisie_rows) and "affectable_menage" in saisie_rows[0]
pool_courses = pool_conso = pool_autres = 0.0
if saisie_has_col:
    for d in saisie_rows:
        if str(d.get("affectable_menage")) != "OUI" or str(d.get("mois"))[:7] != MONTH: continue
        cat = str(d.get("categorie_charge_id")); m = f(d.get("montant")) or 0
        if cat == "CHG_004": pool_conso += m
        elif cat in ("CHG_018",): pool_autres += m
        else: pool_courses += m
if pool_courses == 0 and pool_conso == 0 and pool_autres == 0:
    controls.append(("POOL_VIDE_NON_SAISI", "INFO", "Aucune charge ménage affectable saisie dans SAISIE_Charges_Flux (pools courses/conso=0)"))

# contrôle double source lavage
if saisie_has_col:
    lav_saisie = [d for d in saisie_rows if str(d.get("categorie_charge_id")) == "CHG_003" and str(d.get("affectable_menage")) == "OUI" and str(d.get("mois"))[:7] == MONTH]
    if lav_saisie and sum(l["lavage_attribuable"] for l in lines) > 0:
        controls.append(("DOUBLE_SOURCE_LAVAGE_A_CONTROLER", "A_CONTROLER", f"{len(lav_saisie)} lignes lavage SAISIE affectable=OUI + lavage Google Sheet présent"))

POOLS = {"LOCAL_CAVE": local_cave_montant, "COURSES": pool_courses, "CONSOMMABLES": pool_conso, "AUTRES": pool_autres}
# Règle figée : cave/local REC_002 ventilée UNIQUEMENT sur les ménages internes (jamais les externes).
controls.append(("REC_002_LOCAL_CAVE_INTERNE_ONLY", "INFO",
    f"Cave {local_cave_montant}€ ventilée sur poids internes ({round(sum_poids_interne,2)}) — externes exclus"))

# ── Quote-parts + coût complet par ligne ─────────────────────────────────────
ventil = []
for l in lines:
    w = l["poids"]
    # cave = ménages internes uniquement
    qp_local = round(local_cave_montant * w / sum_poids_interne, 2) if (local_cave_montant and l["type_intervenant"] == "INTERNE") else 0.0
    # lavage = attribuable (direct sheet) + non-attribuable ventilé (interne seulement, par poids intervenant)
    qp_lav_na = 0.0
    if l["type_intervenant"] == "INTERNE" and lav_na_by_int.get(l["intervenant_id"]):
        sp = sum_poids_int.get(l["intervenant_id"]) or 1
        qp_lav_na = round(lav_na_by_int[l["intervenant_id"]] * w / sp, 2)
    qp_lavage = round(l["lavage_attribuable"] + qp_lav_na, 2)
    qp_courses = round(pool_courses * w / sum_poids_all, 2) if pool_courses else 0.0
    qp_conso = round(pool_conso * w / sum_poids_all, 2) if pool_conso else 0.0
    qp_autres = round(pool_autres * w / sum_poids_all, 2) if pool_autres else 0.0
    cc = round((l["cout_direct_total"] or 0) + qp_local + qp_lavage + qp_courses + qp_conso + qp_autres, 2)
    st = l["cout_standard_total"]
    ecart = round(st - cc, 2) if st is not None else None
    statut_e = "NON_CALCULABLE" if ecart is None else ("GAIN" if ecart > 0 else "PERTE" if ecart < 0 else "EQUILIBRE")
    statut_c, code = ("A_CONTROLER", "COUT_STANDARD_ABSENT") if st is None else ("VALIDE", "")
    l.update({"quote_part_local": qp_local, "quote_part_courses": qp_courses, "quote_part_lavage": qp_lavage,
        "quote_part_consommables": qp_conso, "quote_part_autres_charges_menage": qp_autres,
        "cout_complet_total": cc, "cout_complet_unitaire": round(cc/l["nb_menages"], 2) if l["nb_menages"] else None,
        "ecart_vs_standard_total": ecart, "statut_ecart": statut_e, "statut_controle": statut_c,
        "code_controle": code, "commentaire": "", "ROW_HASH": rh(MONTH, l["logement_id"], l["intervenant_id"], cc)})
    for nm, val in [("LOCAL_CAVE", qp_local), ("LAVAGE", qp_lavage), ("COURSES", qp_courses), ("CONSOMMABLES", qp_conso), ("AUTRES", qp_autres)]:
        if val: ventil.append({"mois": MONTH, "pool": nm, "logement_id": l["logement_id"], "intervenant_id": l["intervenant_id"], "poids": round(w,2), "quote_part": val})

# ── Écriture ──────────────────────────────────────────────────────────────────
os.makedirs(OUTD, exist_ok=True)
wb = openpyxl.Workbook()
def wsheet(title, cols, rows, first=False):
    ws = wb.active if first else wb.create_sheet(title)
    if first: ws.title = title
    ws.append(cols)
    for c in ws[1]: c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="DDDDDD")
    for r in rows: ws.append([r.get(c) for c in cols])
DET = ["mois","logement_id","nom_appartement","proprietaire_id","type_logement_id","intervenant_id","nom_intervenant","type_intervenant",
    "nb_menages","cout_standard_total","cout_direct_total","quote_part_local","quote_part_courses","quote_part_lavage",
    "quote_part_consommables","quote_part_autres_charges_menage","cout_complet_total","cout_complet_unitaire",
    "ecart_vs_standard_total","statut_ecart","statut_controle","code_controle","commentaire"]
wsheet("DETAIL_COUT_COMPLET", DET, lines, first=True)
wsheet("POOLS_CHARGES_MENAGE", ["mois","pool","montant_total","source","cle_repartition"],
    [{"mois":MONTH,"pool":k,"montant_total":round(v,2),"source":("REC_002 (date-aware)" if k=="LOCAL_CAVE" else "Google Sheet" if k=="LAVAGE" else "SAISIE_Charges_Flux"),"cle_repartition":"poids=nb×cout_standard"} for k,v in {**POOLS,"LAVAGE":sum(l['lavage_attribuable'] for l in lines)+sum(lav_na_by_int.values())}.items()])
wsheet("VENTILATION_CHARGES", ["mois","pool","logement_id","intervenant_id","poids","quote_part"], ventil)
def resume(keyf):
    agg = collections.defaultdict(lambda: [0,0.0,0.0,0.0])
    for l in lines:
        k = keyf(l); a = agg[k]; a[0]+=l["nb_menages"]; a[1]+=l["cout_standard_total"] or 0; a[2]+=l["cout_complet_total"]; a[3]+=l["ecart_vs_standard_total"] or 0
    return agg
wsheet("RESUME_LOGEMENT", ["mois","logement_id","nom_appartement","nb_menages","cout_standard_total","cout_complet_total","ecart_total"],
    [{"mois":MONTH,"logement_id":k[0],"nom_appartement":k[1],"nb_menages":v[0],"cout_standard_total":round(v[1],2),"cout_complet_total":round(v[2],2),"ecart_total":round(v[3],2)} for k,v in sorted(resume(lambda l:(l["logement_id"],l["nom_appartement"])).items())])
wsheet("RESUME_INTERVENANT", ["mois","intervenant_id","nom_intervenant","type_intervenant","nb_menages","cout_standard_total","cout_complet_total","ecart_total"],
    [{"mois":MONTH,"intervenant_id":k[0],"nom_intervenant":k[1],"type_intervenant":k[2],"nb_menages":v[0],"cout_standard_total":round(v[1],2),"cout_complet_total":round(v[2],2),"ecart_total":round(v[3],2)} for k,v in sorted(resume(lambda l:(l["intervenant_id"],l["nom_intervenant"],l["type_intervenant"])).items())])
wsheet("RESUME_PRESTATAIRE", ["mois","intervenant_id","nom_intervenant","nb_menages","cout_complet_total","ecart_total"],
    [{"mois":MONTH,"intervenant_id":k[0],"nom_intervenant":k[1],"nb_menages":v[0],"cout_complet_total":round(v[2],2),"ecart_total":round(v[3],2)} for k,v in sorted(resume(lambda l:(l["intervenant_id"],l["nom_intervenant"],l["type_intervenant"])).items()) if k[2]=="EXTERNE"])
cc_ctrl = collections.Counter((c[0],c[1]) for c in controls)
# rappel charges déjà en Flux
cc_ctrl[("CHARGE_EXTERNE_DEJA_EN_FLUX_NON_REINJECTEE","INFO")] += 1
wsheet("CONTROLES_DOUBLE_COMPTAGE", ["code_controle","niveau","nb","exemple"],
    [{"code_controle":k[0],"niveau":k[1],"nb":n,"exemple":next((c[2] for c in controls if (c[0],c[1])==k),"ménage externe (TYPE_FLUX_014) déjà compté dans Flux ; coût complet = analytique, non réinjecté")} for k,n in cc_ctrl.most_common()])
try: wb.save(OUT)
except PermissionError: OUT = OUT.replace(".xlsx","_MAJ.xlsx"); wb.save(OUT); print(f"[lot6f] original verrouillé -> {os.path.basename(OUT)}")

# ── Rapport ──────────────────────────────────────────────────────────────────
ts = sum(l["cout_standard_total"] or 0 for l in lines); tc = sum(l["cout_complet_total"] for l in lines)
print(f"[lot6f] DRY-RUN mois={MONTH} | local_cave={local_cave_montant} | taux={taux} -> {OUT}")
print(f"  lignes DETAIL : {len(lines)}")
print(f"  POOLS : LOCAL_CAVE={round(local_cave_montant,2)} LAVAGE={round(sum(l['lavage_attribuable'] for l in lines)+sum(lav_na_by_int.values()),2)} COURSES={pool_courses} CONSO={pool_conso}")
print(f"  COUT GLOBAL : standard={round(ts,2)} complet={round(tc,2)} ecart={round(ts-tc,2)} ({'GAIN' if ts-tc>0 else 'PERTE' if ts-tc<0 else 'EQ'})")
print("\n  par intervenant:")
for k,v in sorted(resume(lambda l:(l["intervenant_id"],l["nom_intervenant"],l["type_intervenant"])).items()):
    print(f"    {str(k[0]):10} {str(k[1])[:9]:9} {str(k[2]):8} nb={v[0]:3} std={round(v[1],2):8} complet={round(v[2],2):8} ecart={round(v[3],2):8}")
print("\n  CONTROLES:", dict(cc_ctrl))
