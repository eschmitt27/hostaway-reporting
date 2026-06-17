"""
lot6_normalise_declarations_internes_dryrun.py — DRY-RUN
================================================================================
Normalise (dépivot) les déclarations internes depuis la Google Sheet publiée
"Suivi ménage" -> table normalisée temporaire.

PARTIE 1 (volumes) uniquement : nb_menages + nb_heures conservés ; lavage/cave/
courses HORS périmètre. Aucun coût, aucun gain/perte.

⚠️ DRY-RUN : n'écrit PAS dans M04 réel. Sortie de test :
   02_TRAVAIL/Lot6_DryRun/DRYRUN_M04_Declarations_Internes_Normalisees.xlsx
"""

import sys, os, io, csv, subprocess, hashlib, datetime, collections, unicodedata, warnings
warnings.filterwarnings("ignore")
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import openpyxl
from openpyxl.styles import Font, PatternFill

URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vTYNAYa4WcY-ask9viQHJT-wQMwyxkKpoDZGpjrnqqtCzXC1xR_6DnZ0oSVTv5LkQTdqPcfaD1PtdAt/pub?output=csv"
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REF  = os.path.join(ROOT, "01_SOURCES_BRUTES", "REF_Setup", "REF_Setup.xlsm")
OUTD = os.path.join(ROOT, "02_TRAVAIL", "Lot6_DryRun")
OUT  = os.path.join(OUTD, "DRYRUN_M04_Declarations_Internes_Normalisees.xlsx")
NOW  = datetime.datetime.now().isoformat(timespec="seconds")

MOIS = {"janvier":"01","fevrier":"02","mars":"03","avril":"04","mai":"05","juin":"06",
        "juillet":"07","aout":"08","septembre":"09","octobre":"10","novembre":"11","decembre":"12"}
INTMAP = {"imene":("INT_0001","Imène"), "kira":("INT_0002","Kheira"), "kheira":("INT_0002","Kheira")}

def norm(s):
    s = str(s or "").strip().lower()
    return " ".join("".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn").split())

def rowhash(*v):
    return hashlib.sha256("|".join("" if x is None else str(x) for x in v).encode()).hexdigest()[:16]

def sh(p, s):
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True); ws = wb[s]
    rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]; wb.close()
    return [dict(zip([str(c) for c in rows[0]], r)) for r in rows[1:]]

# mapping logement (libellé -> logement_id) + nom officiel + type
lmap, lognom, logtype = {}, {}, {}
for d in sh(REF, "REF_Mapping_Logements"):
    if d.get("valeur_source"):
        lmap[norm(d["valeur_source"])] = d.get("logement_id")
for d in sh(REF, "REF_Logements"):
    if d.get("logement_id"):
        lognom[d["logement_id"]] = d.get("nom_logement_officiel")
        logtype[d["logement_id"]] = d.get("type_logement_id")
inttype = {d["intervenant_id"]: d.get("type_intervenant") for d in sh(REF, "REF_Intervenants")}

# CSV
txt = subprocess.run(["curl","-sL","-A","Mozilla/5.0",URL], capture_output=True, text=True, encoding="utf-8").stdout
rows = list(csv.reader(io.StringIO(txt)))
hdr = rows[0]; data = [r for r in rows[1:] if any(c.strip() for c in r)]
i_pre = hdr.index("Prénom"); i_mois = hdr.index("Mois des ménages"); i_an = hdr.index("Année des ménages")
appcols = [i for i, h in enumerate(hdr) if h.strip() == "Appartement"]

out = []
for r in data:
    pre = r[i_pre].strip()
    mo = norm(r[i_mois]); yr = r[i_an].strip()
    mm = MOIS.get(mo); mois_iso = f"{yr}-{mm}" if (mm and yr) else None
    iid, inom = INTMAP.get(norm(pre), (None, pre))
    for ci in appcols:
        app = r[ci].strip() if ci < len(r) else ""
        if not app:
            continue
        def g(off):
            v = r[ci+off].strip() if ci+off < len(r) else ""
            return v
        try: nb = int(float(g(1))) if g(1) else 0
        except ValueError: nb = 0
        try: nh = float(g(2)) if g(2) else None
        except ValueError: nh = None
        lid = lmap.get(norm(app))
        statut, code, comm = "VALIDE", "", ""
        if lid is None:
            statut, code, comm = "A_CONTROLER", "LOGEMENT_NON_MAPPE", f"appartement '{app}' non mappé"
        elif iid is None:
            statut, code, comm = "A_CONTROLER", "INTERVENANT_NON_MAPPE", f"intervenant '{pre}' non mappé"
        out.append({
            "mois": mois_iso, "annee": yr, "mois_saisie": r[i_mois].strip(),
            "appartement_source": app, "nom_appartement": lognom.get(lid), "logement_id": lid,
            "intervenant_source": pre, "intervenant_id": iid, "nom_intervenant": inom,
            "type_intervenant": inttype.get(iid),
            "nb_menages": nb, "nb_heures": nh,
            "statut_controle": statut, "code_controle": code, "commentaire": comm,
            "source_url": URL, "date_extraction": NOW,
            "ROW_HASH": rowhash(mois_iso, lid, iid, nb),
        })

COLS = ["mois","annee","mois_saisie","appartement_source","nom_appartement","logement_id",
        "intervenant_source","intervenant_id","nom_intervenant","type_intervenant",
        "nb_menages","nb_heures","statut_controle","code_controle","commentaire",
        "source_url","date_extraction","ROW_HASH"]
os.makedirs(OUTD, exist_ok=True)
wb = openpyxl.Workbook(); ws = wb.active; ws.title = "MASTER_NORMALISE"
ws.append(COLS)
for c in ws[1]: c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="DDDDDD")
for d in out: ws.append([d.get(c) for c in COLS])
wb.save(OUT)

print(f"[normalise] CSV OK | lignes sheet lues : {len(data)} | lignes normalisées : {len(out)}")
print(f"[normalise] écrit (DRY-RUN, M04 réel NON touché) : {OUT}")
mai = [d for d in out if d["mois"] == "2026-05"]
print(f"\n-- mai 2026 ({len(mai)} lignes) --")
for d in mai:
    print(f"  {d['nom_intervenant']:8} {d['intervenant_id']} {d['logement_id']:8} {str(d['nom_appartement'])[:24]:24} nb={d['nb_menages']}")
ag = collections.Counter()
for d in mai: ag[(d["intervenant_id"], d["nom_intervenant"])] += d["nb_menages"]
print("  agrégat:", dict(ag))
