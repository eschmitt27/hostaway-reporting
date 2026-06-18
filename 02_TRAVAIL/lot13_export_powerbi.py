"""
lot13_export_powerbi.py — Export Power BI (CSV plats, lecture seule)
================================================================================
Génère 03_EXPORTS/PowerBI/PBI_*.csv depuis les sorties MASTER_* du pipeline.
Power BI lit uniquement ces tables plates — AUCUNE règle métier dans Power BI.

Contraintes (AUD-008) :
  - lecture seule des sources ; ne modifie aucun calcul ni fichier source ;
  - CSV UTF-8 (BOM), séparateur ';', un fichier par table ;
  - whitelist de colonnes par export (exclusion stricte des données sensibles) ;
  - filet de sécurité : ABORT si une colonne sensible apparaît dans une sortie ;
  - idempotent : réécrit proprement les CSV ;
  - exports régénérables -> non versionnés (.gitignore 03_EXPORTS/PowerBI/).

EXCLUS (confidentialité) : email, téléphone, IBAN/RIB, adresses, libellés bancaires
bruts, noms voyageurs, .env/secrets/tokens.
"""

import sys, os, csv, re, warnings, datetime
warnings.filterwarnings("ignore")
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTD = os.path.join(ROOT, "03_EXPORTS", "PowerBI")
T = os.path.join(ROOT, "02_TRAVAIL")
REF = os.path.join(ROOT, "01_SOURCES_BRUTES", "REF_Setup", "REF_Setup.xlsm")

# Filet anti-sensible : motifs interdits dans les NOMS de colonnes exportées
SENSIBLE = re.compile(r"(e[-_ ]?mail|^mail|t[ée]l[ée]?phone|^tel$|iban|rib|adresse|secret|token|password|"
                      r"guest|voyageur|libelle_brut|libelle_banque|compte_bancaire|num_compte)", re.I)

def cols_of(p, s):
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    if s not in wb.sheetnames:
        wb.close(); return None, None
    ws = wb[s]
    rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
    wb.close()
    if not rows: return [], []
    h = [str(c) for c in rows[0]]
    return h, [dict(zip(h, r)) for r in rows[1:]]

# (nom_export, fichier, onglet, colonnes whitelist)
EXPORTS = [
 ("PBI_Flux", f"{T}/Lot9_FluxUnifie/MASTER_CALC_Flux.xlsx", "MASTER",
  ["flux_id","mois","date_flux","logement_id","proprietaire_id","associe_id","type_flux_id","sens","montant",
   "code_impact","inclure_resultat_reel","inclure_resultat_comptable","inclure_resultat_hors_compta",
   "statut_controle","niveau_anomalie","code_anomalie"]),
 ("PBI_Resultats_Mensuels", f"{T}/Lot10_Resultats/MASTER_CALC_Resultats.xlsx", "PAR_MOIS_LOGEMENT",
  ["mois","logement_id","proprietaire_id","vision","total_produits","total_charges","resultat","nb_flux"]),
 ("PBI_Reservations_Resolues", f"{T}/Lot4quater_SourceResolue/MASTER_CALC_Reservations_Resolues.xlsx", "MASTER",
  ["reservation_calc_id","reservation_id_hostaway","reservation_hh_id","canal","source","mois","logement_id",
   "proprietaire_id","date_arrivee","date_depart","nuits","montant_retenu","code_impact","etat_mois",
   "origine_initiale","statut_controle","niveau_anomalie","code_anomalie"]),
 ("PBI_Commissions", f"{T}/Lot10_Resultats/MASTER_CALC_Commissions.xlsx", "COMMISSIONS",
  ["reservation_calc_id","reservation_id_hostaway","logement_id","proprietaire_id","mois","date_arrivee","date_depart",
   "nuits","channel_type","source_type","statut_calcul_payout","payout_calcule","menage_retenu","assiette_commission",
   "taux_commission","commission_conciergerie","net_proprietaire"]),
 ("PBI_Net_Proprietaire", f"{T}/Lot10_Resultats/MASTER_CALC_NetProprietaire.xlsx", "VUE_MOIS",
  ["mois","proprietaire_id","total_payout_mois","total_menage_mois","total_commission_mois","charge_fixe_mensuelle",
   "montant_du_conciergerie","reste_a_payer_conciergerie","net_proprietaire_avant_charge_mois",
   "net_proprietaire_apres_charge_mois","nb_reservations"]),
 ("PBI_Prefactures_Proprietaires", f"{T}/Lot12_Factures/MASTER_FACT_Proprietaires.xlsx", "FACT_FACTURE_LIGNES",
  ["facture_id","ligne_num","type_ligne","libelle","montant","bloc"]),
 ("PBI_Menages_Cout_Complet", f"{T}/Lot6f_CoutComplet_Menages/MASTER_CALC_CoutComplet_Menages.xlsx", "DETAIL_COUT_COMPLET",
  ["mois","logement_id","proprietaire_id","type_logement_id","intervenant_id","type_intervenant","nb_menages",
   "cout_standard_total","cout_direct_total","cout_complet_total","ecart_vs_standard_total","ecart_unitaire",
   "statut_ecart","statut_controle"]),
 ("PBI_Menages_Rapprochement", f"{T}/Lot6d_Rapprochement_Menages/MASTER_CTRL_Rapprochement_Menages.xlsx", "TABLEAU_COMPARAISON",
  ["mois","logement_id","proprietaire_id","intervenant_id","type_intervenant","nb_menages_tasks_hostaway_completed",
   "nb_menages_declares_externe","nb_menages_declares_interne_m04","total_menages_declares","ecart","statut_controle","code_controle"]),
 ("PBI_Controles_Ouverts", f"{T}/Lot11_Controles/MASTER_CTRL_Coherence.xlsx", "A_CONTROLER_OUVERTS",
  ["ctrl_pk","code_controle","severity","mois","logement_id","proprietaire_id","message","statut_resolution"]),
 ("PBI_Referentiel_Logements", REF, "REF_Logements",
  ["logement_id","nom_logement_officiel","nom_court","ville","type_logement_id","proprietaire_id",
   "date_entree_gestion","date_sortie_gestion","sur_hostaway","actif","forfait_logiciel_consommables_mensuel"]),
 ("PBI_Referentiel_Proprietaires", REF, "REF_Proprietaires",
  ["proprietaire_id","nom_proprietaire","mode_facturation","taux_commission","actif"]),
]
# Prefact entête : on ajoute aussi quelques champs entête NON sensibles via un export secondaire optionnel — non requis ici.

def val(v):
    if v is None: return ""
    if isinstance(v, (datetime.date, datetime.datetime)): return str(v)[:10]
    return str(v)

def main():
    os.makedirs(OUTD, exist_ok=True)
    dico = []   # (table, colonne)
    rapport = []
    abort_msgs = []
    for name, path, sheet, wl in EXPORTS:
        if not os.path.exists(path):
            rapport.append((name, "SOURCE_ABSENTE", 0)); continue
        h, rows = cols_of(path, sheet)
        if h is None:
            rapport.append((name, "ONGLET_ABSENT", 0)); continue
        keep = [c for c in wl if c in h]            # whitelist ∩ colonnes réelles
        manquantes = [c for c in wl if c not in h]
        # filet anti-sensible sur colonnes exportées
        sens = [c for c in keep if SENSIBLE.search(c)]
        if sens:
            abort_msgs.append(f"{name}: colonne sensible détectée {sens}")
            continue
        out = os.path.join(OUTD, f"{name}.csv")
        with open(out, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f, delimiter=";")
            w.writerow(keep)
            for r in rows:
                w.writerow([val(r.get(c)) for c in keep])
        for c in keep: dico.append((name, c))
        rapport.append((name, "OK" + (f" (cols manquantes ignorées: {manquantes})" if manquantes else ""), len(rows)))

    # Dictionnaire des colonnes
    dpath = os.path.join(OUTD, "PBI_Dictionnaire_Colonnes.csv")
    with open(dpath, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";"); w.writerow(["table","colonne"])
        for t, c in dico: w.writerow([t, c])

    if abort_msgs:
        print("[BLOQUANT lot13] colonnes sensibles dans des exports :")
        for m in abort_msgs: print("  ", m)
        sys.exit(1)

    print(f"[lot13] Export Power BI -> {OUTD}")
    print(f"  exports générés : {sum(1 for _,s,_ in rapport if s.startswith('OK'))}/{len(EXPORTS)} + dictionnaire")
    for name, statut, n in rapport:
        print(f"  {name:32} {n:5} lignes | {statut}")
    print(f"  PBI_Dictionnaire_Colonnes.csv : {len(dico)} entrées")
    print("  EXCLUS confidentialité : email, telephone, adresse(s), IBAN/RIB, libellés bancaires, noms voyageurs, .env/secrets.")

if __name__ == "__main__":
    main()
