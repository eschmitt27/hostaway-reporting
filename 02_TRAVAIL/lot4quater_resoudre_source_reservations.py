"""
lot4quater_resoudre_source_reservations.py
Couche de RÉSOLUTION de la source des réservations (générique, après lot4ter).

Rôle :
  Produire une table unique résolue consommée par lot9/lot10/lot11/lot12, en
  appliquant la bascule mois ouvert / mois clôturé UNE SEULE FOIS ici. lot9-12
  ne décident plus eux-mêmes entre live et historique.

Règle structurante :
  - mois ouvert / non clôturé / absent de REF = MASTER_CALC_Reservations (live lot4bis) ;
  - mois clôturé (REF_Cloture_Mensuelle.statut_mois == CLOTURE) = HIST_Reservations_Cloturees ;
  - sortie unique = MASTER_CALC_Reservations_Resolues (onglets MASTER + VUE_FLUX).
  - upsert sans suppression : une réservation présente en HIST mais disparue du live
    après clôture est conservée (réinjectée).
  - toute différence sur un mois clôturé (live vs HIST) => alerte (HIST prime, pas d'écrasement).

Sorties :
  - 02_TRAVAIL/Lot4quater_SourceResolue/MASTER_CALC_Reservations_Resolues.xlsx
"""

import sys
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.stderr.reconfigure(encoding="utf-8", errors="replace")

import os
import datetime
import collections

import openpyxl
from openpyxl.styles import Font, PatternFill

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PATH_LIVE = os.path.join(ROOT, "02_TRAVAIL", "Lot4bis_TableCommune", "MASTER_CALC_Reservations.xlsx")
PATH_HIST = os.path.join(ROOT, "02_DONNEES_NORMALISEES", "historique_reservations", "HIST_Reservations_Cloturees.xlsx")
PATH_PAY  = os.path.join(ROOT, "02_TRAVAIL", "Lot1_Hostaway", "MASTER_CALC_HA_Payout.xlsx")
PATH_REF  = os.path.join(ROOT, "01_SOURCES_BRUTES", "REF_Setup", "REF_Setup.xlsm")
OUT_DIR   = os.path.join(ROOT, "02_TRAVAIL", "Lot4quater_SourceResolue")
OUT_FILE  = os.path.join(OUT_DIR, "MASTER_CALC_Reservations_Resolues.xlsx")

DATE_INTEGRATION = datetime.datetime.now().isoformat(timespec="seconds")

# 24 colonnes de base (compat lot9/lot10/lot11 — mêmes noms que MASTER_CALC_Reservations)
BASE_COLS = [
    "reservation_calc_id", "ROW_HASH", "source", "reservation_id_hostaway",
    "reservation_hh_id", "mois", "logement_id", "proprietaire_id",
    "date_arrivee", "date_depart", "nuits", "montant_retenu", "source_montant",
    "code_impact", "impact_resultat_reel", "impact_resultat_comptable",
    "statut_controle", "niveau_anomalie", "code_anomalie", "commentaire",
    "source_module", "source_table", "source_pk", "date_integration",
]
# colonnes ajoutées par la résolution (gouvernance + financiers)
EXTRA_COLS = [
    "canal", "etat_mois", "origine_initiale", "source_ligne", "methode",
    "payout_calcule", "menage_retenu", "assiette_commission",
]
RESOLUES_COLS = BASE_COLS + EXTRA_COLS

SRC_HIST = "HIST_RESERVATIONS_CLOTUREES"
METHODE_HIST = "HIST_PRIME_MOIS_CLOTURE"

CANAL_MAP = {
    "HOSTAWAY_AIRBNB": "AIRBNB", "HOSTAWAY_BOOKING": "BOOKING",
    "HOSTAWAY_VRBO_A_CONTROLER": "VRBO", "HOSTAWAY_VRBO_HH": "VRBO",
    "HOSTAWAY_DIRECT_HH": "DIRECT", "MANUEL_HORS_HOSTAWAY": "HH",
    "OWNERSTAY_EXCLU": "OWNERSTAY",
}
# canal -> source classification résolue (pas de *_HIST, pas de *_A_CONTROLER une fois résolu)
SOURCE_FROM_CANAL = {
    "AIRBNB": "HOSTAWAY_AIRBNB", "BOOKING": "HOSTAWAY_BOOKING",
    "VRBO": "HOSTAWAY_VRBO", "DIRECT": "HOSTAWAY_DIRECT_HH",
    "HH": "MANUEL_HORS_HOSTAWAY", "OWNERSTAY": "OWNERSTAY_EXCLU",
}


def load_sheet(path, sheet):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
    wb.close()
    if not rows:
        return []
    h = list(rows[0])
    return [dict(zip(h, r)) for r in rows[1:]]


def closed_months():
    out = set()
    if not os.path.exists(PATH_REF):
        return out
    for d in load_sheet(PATH_REF, "REF_Cloture_Mensuelle"):
        if str(d.get("statut_mois") or "").strip().upper() == "CLOTURE" and d.get("mois"):
            out.add(str(d["mois"])[:7])
    return out


def main():
    cmonths = closed_months()
    print(f"[lot4quater] mois CLOTURE : {sorted(cmonths) or 'AUCUN'}")

    live = load_sheet(PATH_LIVE, "MASTER")
    hist = load_sheet(PATH_HIST, "HIST_Reservations_Cloturees") if os.path.exists(PATH_HIST) else []
    pay = {r["reservation_id"]: r for r in load_sheet(PATH_PAY, "data")} if os.path.exists(PATH_PAY) else {}

    hist_by_mois = collections.defaultdict(list)
    for h in hist:
        hist_by_mois[str(h.get("mois"))[:7]].append(h)
    hist_months = set(hist_by_mois.keys())

    resolved = []
    alertes = []
    n_open = n_closed = n_reinject = n_cloture_sans_hist = 0

    # ── Mois OUVERT (ou non clôturé / absent REF) : live lot4bis ──
    for r in live:
        mois = str(r.get("mois"))[:7]
        if mois in cmonths:
            continue
        canal = CANAL_MAP.get(r.get("source"), "INCONNU")
        rid = r.get("reservation_id_hostaway")
        p = pay.get(rid, {}) if rid else {}
        row = {c: r.get(c) for c in BASE_COLS}
        row.update({
            "canal": canal,
            "etat_mois": "OUVERT",
            "origine_initiale": "API_HOSTAWAY" if rid else "SAISIE_HH",
            "source_ligne": r.get("source"),
            "methode": "LIVE",
            "payout_calcule": p.get("payout_calcule"),
            "menage_retenu": p.get("menage_retenu"),
            "assiette_commission": p.get("assiette_commission"),
        })
        resolved.append(row)
        n_open += 1

    # ── Mois CLOTURE : HIST_Reservations_Cloturees (autorité) ──
    live_by_mois = collections.defaultdict(dict)
    for r in live:
        key = r.get("reservation_id_hostaway") or r.get("reservation_hh_id") or r.get("reservation_calc_id")
        live_by_mois[str(r.get("mois"))[:7]][key] = r

    for mois in sorted(cmonths):
        hrows = hist_by_mois.get(mois, [])
        if not hrows:
            # mois déclaré CLOTURE mais aucune ligne en HIST -> alerte + repli live (non bloquant)
            for r in live_by_mois.get(mois, {}).values():
                canal = CANAL_MAP.get(r.get("source"), "INCONNU")
                row = {c: r.get(c) for c in BASE_COLS}
                row.update({
                    "canal": canal, "etat_mois": "CLOTURE_SANS_HIST",
                    "origine_initiale": "API_HOSTAWAY" if r.get("reservation_id_hostaway") else "SAISIE_HH",
                    "source_ligne": r.get("source"), "methode": "REPLI_LIVE",
                    "payout_calcule": None, "menage_retenu": None, "assiette_commission": None,
                    "statut_controle": "A_CONTROLER",
                    "niveau_anomalie": "A_CONTROLER",
                    "code_anomalie": "MOIS_CLOTURE_SANS_HISTORIQUE",
                    "commentaire": f"Mois {mois} CLOTURE mais absent de HIST — repli live",
                })
                resolved.append(row)
                n_cloture_sans_hist += 1
            alertes.append(("MOIS_CLOTURE_SANS_HISTORIQUE", mois, len(live_by_mois.get(mois, {}))))
            continue

        live_keys_mois = set(live_by_mois.get(mois, {}).keys())
        hist_keys_mois = set()
        for h in hrows:
            canal = h.get("canal") or "INCONNU"
            key = h.get("reservation_id_hostaway") or h.get("reservation_hh_id") or h.get("reservation_calc_id")
            hist_keys_mois.add(key)
            # détection écart live vs HIST (HIST prime)
            lr = live_by_mois.get(mois, {}).get(key)
            if lr is not None:
                lm = lr.get("montant_retenu")
                hm = h.get("montant_retenu")
                try:
                    if lm is not None and hm is not None and round(float(lm), 2) != round(float(hm), 2):
                        alertes.append(("AJUSTEMENT_POST_CLOTURE", mois,
                                        f"key={key} live={lm} HIST={hm}"))
                except (TypeError, ValueError):
                    pass
            else:
                n_reinject += 1   # présent en HIST, disparu du live -> conservé
            row = {
                "reservation_calc_id": h.get("reservation_calc_id"),
                "ROW_HASH": h.get("ROW_HASH"),
                "source": SOURCE_FROM_CANAL.get(canal, "INCONNU"),
                "reservation_id_hostaway": h.get("reservation_id_hostaway"),
                "reservation_hh_id": h.get("reservation_hh_id"),
                "mois": mois,
                "logement_id": h.get("logement_id"),
                "proprietaire_id": h.get("proprietaire_id"),
                "date_arrivee": h.get("date_arrivee"),
                "date_depart": h.get("date_depart"),
                "nuits": h.get("nuits"),
                "montant_retenu": h.get("montant_retenu"),
                "source_montant": SRC_HIST,
                "code_impact": h.get("code_impact"),
                "impact_resultat_reel": h.get("impact_resultat_reel"),
                "impact_resultat_comptable": h.get("impact_resultat_comptable"),
                "statut_controle": h.get("statut_controle"),
                "niveau_anomalie": h.get("niveau_anomalie"),
                "code_anomalie": h.get("code_anomalie"),
                "commentaire": f"Historique clôturé {mois} (origine {h.get('origine_initiale')})",
                "source_module": "lot4quater",
                "source_table": "HIST_Reservations_Cloturees",
                "source_pk": str(h.get("cle_historisation")),
                "date_integration": DATE_INTEGRATION,
                "canal": canal,
                "etat_mois": "CLOTURE",
                "origine_initiale": h.get("origine_initiale"),
                "source_ligne": SRC_HIST,
                "methode": METHODE_HIST,
                "payout_calcule": h.get("payout_calcule"),
                "menage_retenu": h.get("menage_retenu"),
                "assiette_commission": h.get("assiette_commission"),
            }
            resolved.append(row)
            n_closed += 1

    # VUE_FLUX : VALIDE + impact_reel=OUI + montant≠0 (même filtre que lot4bis)
    vue = [r for r in resolved
           if r.get("statut_controle") == "VALIDE"
           and r.get("impact_resultat_reel") == "OUI"
           and (r.get("montant_retenu") or 0) != 0]

    os.makedirs(OUT_DIR, exist_ok=True)
    wb = openpyxl.Workbook()
    ws_m = wb.active
    ws_m.title = "MASTER"
    ws_m.append(RESOLUES_COLS)
    for r in resolved:
        ws_m.append([r.get(c) for c in RESOLUES_COLS])
    ws_v = wb.create_sheet("VUE_FLUX")
    ws_v.append(RESOLUES_COLS)
    for r in vue:
        ws_v.append([r.get(c) for c in RESOLUES_COLS])
    for ws in (ws_m, ws_v):
        for c in ws[1]:
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="DDDDDD")
    wb.save(OUT_FILE)

    by_etat = collections.Counter(r["etat_mois"] for r in resolved)
    print(f"\n[lot4quater] Resolues écrit : {OUT_FILE}")
    print(f"  MASTER total           : {len(resolved)}")
    print(f"  VUE_FLUX               : {len(vue)}")
    print(f"  mois ouverts (live)    : {n_open}")
    print(f"  mois clôturés (HIST)   : {n_closed}")
    print(f"  réinjectées (disparues live) : {n_reinject}")
    print(f"  CLOTURE sans HIST (repli)    : {n_cloture_sans_hist}")
    print(f"  par etat_mois          : {dict(by_etat)}")
    if alertes:
        print(f"\n  ALERTES ({len(alertes)}) :")
        for a in alertes[:30]:
            print(f"    {a}")
    if not cmonths:
        print("  NB : aucun mois CLOTURE -> Resolues = live intégral (HIST non utilisé).")


if __name__ == "__main__":
    main()
