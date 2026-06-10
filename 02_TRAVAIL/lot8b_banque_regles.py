#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
lot8b_banque_regles.py
Lot 8b - REF_Banque_Regles + classification deterministe bancaire

Actions:
  1. Backup REF_Setup.xlsm -> 99_ARCHIVES/LOT8_Banque/
  2. Ajoute TYPE_FLUX_016 (FRAIS_BANCAIRES) dans REF_Types_Flux
  3. Ajoute onglet REF_Banque_Regles (20 cols, 30 regles seed)
  4. Backup BANQUE_LOT8_IMPORT.xlsx -> 99_ARCHIVES/LOT8_Banque/
  5. Applique les regles a NORM_Banque (+ 3 nouvelles colonnes)
  6. Cree onglet IA_Classification (catch-all R_099)
  7. Enrichit CTRL_A_CONTROLER (anomalies classification)
  8. Met a jour LOG_Traitement
  9. Sauvegarde BANQUE_LOT8_IMPORT.xlsx

Decisions D-8b-01 a D-8b-09 integrees.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import shutil
import os
import re
from datetime import datetime

# ─────────────────────────────────────────────────────────────────────────────
# CHEMINS
# ─────────────────────────────────────────────────────────────────────────────

BASE = r"C:\Users\Ewan\OneDrive\Documents\Conciergerie\Pilotage_Conciergerie"
REF_PATH = os.path.join(BASE, "01_SOURCES_BRUTES", "REF_Setup", "REF_Setup.xlsm")
BANQUE_PATH = os.path.join(BASE, "02_TRAVAIL", "Lot8_Banque", "BANQUE_LOT8_IMPORT.xlsx")
ARCHIVE_DIR = os.path.join(BASE, "99_ARCHIVES", "LOT8_Banque")

IMPORT_ID = "IMP-BQ-CM-2026-03-001"
COMPTE_ID = "CM_02211_00021321603"

TS = datetime.now().strftime("%Y%m%d_%H%M%S")
NOW_STR = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# ─────────────────────────────────────────────────────────────────────────────
# STYLES
# ─────────────────────────────────────────────────────────────────────────────

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

SUBHDR_FILL = PatternFill("solid", fgColor="2E75B6")
SUBHDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

FILL_OK = PatternFill("solid", fgColor="C6EFCE")
FILL_ACTRL = PatternFill("solid", fgColor="FFEB9C")
FILL_ELEVE = PatternFill("solid", fgColor="FFC7CE")
FILL_MOYEN = PatternFill("solid", fgColor="FFEB9C")
FILL_FAIBLE = PatternFill("solid", fgColor="C6EFCE")
FILL_GRAY = PatternFill("solid", fgColor="F2F2F2")
FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")
FILL_RAPP = PatternFill("solid", fgColor="E2EFDA")
FILL_IA = PatternFill("solid", fgColor="FCE4D6")

DATA_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", bold=True, size=10)

ALN_CTR = Alignment(horizontal="center", vertical="center", wrap_text=False)
ALN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
ALN_LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# ─────────────────────────────────────────────────────────────────────────────
# 30 REGLES SEED
# Ordre: (regle_id, priorite, actif, compte_id, type_match, champ_cible, motif,
#          tiers_detecte, categorie, type_flux_id, code_impact,
#          source_economique, rapprochement_requis, validation_automatique,
#          niveau_risque, statut_controle_defaut, statut_classification_defaut,
#          date_debut_validite, date_fin_validite, commentaire)
# ─────────────────────────────────────────────────────────────────────────────

REGLES_HDR = [
    "regle_id", "priorite", "actif", "compte_id",
    "type_match", "champ_cible", "motif",
    "tiers_detecte", "categorie", "type_flux_id", "code_impact",
    "source_economique", "rapprochement_requis", "validation_automatique",
    "niveau_risque", "statut_controle_defaut", "statut_classification_defaut",
    "date_debut_validite", "date_fin_validite", "commentaire",
]

SEED_RULES = [
    ("R_001", 10, "OUI", "*", "COMMENCE_PAR", "libelle", "VIR AIRBNB PAYMENTS",
     "AIRBNB", "PAYOUT_PLATEFORME", None, None,
     "HOSTAWAY_PRIORITAIRE", "OUI", "OUI", "FAIBLE", "A_CONTROLER", "RAPPROCHEMENT_REQUIS",
     None, None,
     "B8/DC1 - rapprochement Hostaway prioritaire, jamais nouveau produit"),

    ("R_010", 15, "OUI", "*", "COMMENCE_PAR", "libelle", "IMPAYE",
     "INCONNU", "IMPAYE", None, None,
     "BANQUE_SOURCE", "NON", "NON", "ELEVE", "A_CONTROLER", "CLASSE",
     None, None,
     "Impaye - ELEVE, controle obligatoire, pas de validation auto"),

    ("R_020", 20, "OUI", "*", "CONTIENT", "libelle", "WAFA SOUCI",
     "PERS_WAFA", "VIR_ASSOCIE", None, None,
     "BANQUE_SOURCE", "NON", "NON", "ELEVE", "A_CONTROLER", "CLASSE",
     None, None,
     "Virement associe Wafa Souci - ELEVE"),

    ("R_021", 20, "OUI", "*", "CONTIENT", "libelle", "SOUCI WAFA",
     "PERS_WAFA", "VIR_ASSOCIE", None, None,
     "BANQUE_SOURCE", "NON", "NON", "ELEVE", "A_CONTROLER", "CLASSE",
     None, None,
     "Virement Souci Wafa (ordre inverse) - ELEVE"),

    ("R_030", 30, "OUI", "*", "COMMENCE_PAR", "libelle", "F COMM INTERVENTION",
     "CREDIT_MUTUEL", "FRAIS_BANCAIRES", "TYPE_FLUX_016", "IC",
     "BANQUE_SOURCE", "NON", "OUI", "FAIBLE", "VALIDE", "CLASSE",
     None, None,
     "Frais CM intervention - TYPE_FLUX_016 IC, validation auto"),

    ("R_031", 30, "OUI", "*", "COMMENCE_PAR", "libelle", "FRAIS PAIE CB",
     "CREDIT_MUTUEL", "FRAIS_BANCAIRES", "TYPE_FLUX_016", "IC",
     "BANQUE_SOURCE", "NON", "OUI", "FAIBLE", "VALIDE", "CLASSE",
     None, None,
     "Frais paiement CB - TYPE_FLUX_016 IC, validation auto"),

    ("R_040", 35, "OUI", "*", "CONTIENT", "libelle", "EFFET DOMICILIE",
     "INCONNU", "EFFET_DOMICILIE", None, None,
     "A_CONTROLER", "NON", "NON", "ELEVE", "A_CONTROLER", "CLASSE",
     None, None,
     "Effet domicilie - prestataire non identifie, ELEVE"),

    ("R_050", 40, "OUI", "*", "CONTIENT", "libelle", "VINTED CARTE 8259",
     "VINTED", "ACHAT_PERSO_CB", "TYPE_FLUX_002", "IC",
     "BANQUE_SOURCE", "NON", "OUI", "FAIBLE", "A_CONTROLER", "CLASSE",
     None, None,
     "Vinted - achat personnel CB Wafa (TYPE_FLUX_002)"),

    ("R_051", 40, "OUI", "*", "CONTIENT", "libelle", "UBER",
     "UBER", "ACHAT_PERSO_CB", "TYPE_FLUX_002", "IC",
     "BANQUE_SOURCE", "NON", "OUI", "FAIBLE", "A_CONTROLER", "CLASSE",
     None, None,
     "Uber (Eats/Trip) - achat personnel CB (TYPE_FLUX_002)"),

    ("R_052", 40, "OUI", "*", "CONTIENT", "libelle", "PRICELABSINC",
     "PRICELABS", "LOGICIEL_GESTION", None, "IC",
     "BANQUE_SOURCE", "NON", "OUI", "FAIBLE", "A_CONTROLER", "CLASSE",
     None, None,
     "PriceLabs - logiciel gestion tarifs, IC"),

    ("R_053", 45, "OUI", "*", "CONTIENT", "libelle", "CARTE 8259",
     "PERS_WAFA", "DEPENSE_CB_A_CLASSIFIER", None, None,
     "BANQUE_SOURCE", "NON", "NON", "MOYEN", "A_CONTROLER", "CLASSE",
     None, None,
     "Depense CB Wafa non identifiee par regle specifique - MOYEN"),

    ("R_060", 50, "OUI", "*", "CONTIENT", "libelle", "GIE KLESIA",
     "GIE_KLESIA", "COTISATION_PREVOYANCE", None, "IC",
     "BANQUE_SOURCE", "NON", "OUI", "FAIBLE", "A_CONTROLER", "CLASSE",
     None, None,
     "Cotisation prevoyance retraite GIE Klesia"),

    ("R_061", 50, "OUI", "*", "CONTIENT", "libelle", "CABINET GILLES HOURQU",
     "CAB_HOURQUESSEAUX", "HONORAIRES_COMPTABLES", None, "IC",
     "BANQUE_SOURCE", "NON", "OUI", "FAIBLE", "A_CONTROLER", "CLASSE",
     None, None,
     "Cabinet comptable Gilles Hourquesseaux"),

    ("R_062", 50, "OUI", "*", "CONTIENT", "libelle", "URSSAF",
     "URSSAF", "COTISATIONS_SOCIALES", None, "IC",
     "BANQUE_SOURCE", "NON", "OUI", "FAIBLE", "A_CONTROLER", "CLASSE",
     None, None,
     "Cotisations URSSAF"),

    ("R_063", 50, "OUI", "*", "CONTIENT", "libelle", "PASSPASS",
     "PASSPASS", "ABONNEMENT_MOBILITE", None, None,
     "A_CONTROLER", "NON", "NON", "MOYEN", "A_CONTROLER", "CLASSE",
     None, None,
     "PassPass - IC vs HC non tranche, D-8b-02"),

    ("R_064", 50, "OUI", "*", "CONTIENT", "libelle", "PLAN SANTE ENTREPRISE",
     "INCONNU", "SANTE_A_CONTROLER", None, None,
     "A_CONTROLER", "NON", "NON", "MOYEN", "A_CONTROLER", "CLASSE",
     None, None,
     "Libelle bancaire sante non confirme ; ne pas valider sans justificatif."),

    ("R_065", 65, "OUI", "*", "COMMENCE_PAR", "libelle", "FACT",
     "INCONNU", "FACTURE_PRESTATAIRE", None, None,
     "A_CONTROLER", "NON", "NON", "MOYEN", "A_CONTROLER", "CLASSE",
     None, None,
     "Facture prestataire non identifie (SGT...) - D-8b-04"),

    ("R_071", 71, "OUI", "*", "CONTIENT", "libelle", "DELRIEU",
     "PROP_0002", "VIREMENT_PROPRIETAIRE_A_RAPPROCHER", None, None,
     "BANQUE_RAPPROCHEMENT", "OUI", "NON", "MOYEN", "A_CONTROLER", "RAPPROCHEMENT_REQUIS",
     None, None,
     "Proprietaire Cedrine Delrieu - PROP_0002 - D-8b-07"),

    ("R_072", 72, "OUI", "*", "CONTIENT", "libelle", "DINNEWETH",
     "PROP_0006", "VIREMENT_PROPRIETAIRE_A_RAPPROCHER", None, None,
     "BANQUE_RAPPROCHEMENT", "OUI", "NON", "MOYEN", "A_CONTROLER", "RAPPROCHEMENT_REQUIS",
     None, None,
     "Proprietaire Caroline Pons-Dinneweth - PROP_0006 - D-8b-07"),

    ("R_073", 73, "OUI", "*", "CONTIENT", "libelle", "MAURER",
     "PROP_0009", "VIREMENT_PROPRIETAIRE_A_RAPPROCHER", None, None,
     "BANQUE_RAPPROCHEMENT", "OUI", "NON", "MOYEN", "A_CONTROLER", "RAPPROCHEMENT_REQUIS",
     None, None,
     "Proprietaire Francois Maurer - PROP_0009 (multi-variantes) - D-8b-07"),

    ("R_074", 74, "OUI", "*", "CONTIENT", "libelle", "UZON",
     "FAMILLE_UZON_A_CONTROLER", "VIREMENT_PROPRIETAIRE_A_RAPPROCHER", None, None,
     "BANQUE_RAPPROCHEMENT", "OUI", "NON", "MOYEN", "A_CONTROLER", "RAPPROCHEMENT_REQUIS",
     None, None,
     "Famille Uzon (Didier PROP_0001 ou Maryline PROP_0011) - identifier - D-8b-07"),

    ("R_075", 75, "OUI", "*", "CONTIENT", "libelle", "DUREUIL",
     "PROP_0010", "VIREMENT_PROPRIETAIRE_A_RAPPROCHER", None, None,
     "BANQUE_RAPPROCHEMENT", "OUI", "NON", "MOYEN", "A_CONTROLER", "RAPPROCHEMENT_REQUIS",
     None, None,
     "Proprietaire Noel Dureuil - PROP_0010 - D-8b-07"),

    ("R_076", 76, "OUI", "*", "CONTIENT", "libelle", "VASSAL",
     "PROP_0005", "VIREMENT_PROPRIETAIRE_A_RAPPROCHER", None, None,
     "BANQUE_RAPPROCHEMENT", "OUI", "NON", "MOYEN", "A_CONTROLER", "RAPPROCHEMENT_REQUIS",
     None, None,
     "Proprietaire Florane Vassal - PROP_0005 - D-8b-07"),

    ("R_077", 77, "OUI", "*", "CONTIENT", "libelle", "DAVID TOURE",
     "PROP_0008", "VIREMENT_PROPRIETAIRE_A_RAPPROCHER", None, None,
     "BANQUE_RAPPROCHEMENT", "OUI", "NON", "MOYEN", "A_CONTROLER", "RAPPROCHEMENT_REQUIS",
     None, None,
     "Proprietaire David Toure - PROP_0008 (NB: contient WAFA non declenche R_020) - D-8b-07"),

    ("R_078", 78, "OUI", "*", "CONTIENT", "libelle", "LOYER",
     "INCONNU", "LOYER_LOCAL_A_CONTROLER", None, None,
     "A_CONTROLER", "NON", "NON", "MOYEN", "A_CONTROLER", "CLASSE",
     None, None,
     "Loyer local - TYPE_FLUX_010 non confirme - D-8b-03"),

    ("R_080", 80, "OUI", "*", "COMMENCE_PAR", "libelle", "PRLV SEPA",
     "INCONNU", "PRELEVEMENT_NON_IDENTIFIE", None, None,
     "A_CONTROLER", "NON", "NON", "MOYEN", "A_CONTROLER", "CLASSE",
     None, None,
     "Prelevement SEPA non identifie par regle specifique"),

    ("R_085", 85, "OUI", "*", "COMMENCE_PAR", "libelle", "VIR SEPA",
     "INCONNU", "VIREMENT_SORTANT_A_CONTROLER", None, None,
     "A_CONTROLER", "NON", "NON", "MOYEN", "A_CONTROLER", "CLASSE",
     None, None,
     "Virement SEPA sortant non identifie - D-8b-08"),

    ("R_090", 90, "OUI", "*", "COMMENCE_PAR", "libelle", "VIR INST",
     "INCONNU", "VIR_INST_GENERIQUE", None, None,
     "A_CONTROLER", "NON", "NON", "ELEVE", "A_CONTROLER", "CLASSE",
     None, None,
     "VIR INST non identifie - ELEVE, controle obligatoire"),

    ("R_091", 91, "OUI", "*", "COMMENCE_PAR", "libelle", "VIR",
     "INCONNU", "VIR_GENERIQUE", None, None,
     "A_CONTROLER", "NON", "NON", "ELEVE", "A_CONTROLER", "CLASSE",
     None, None,
     "Virement generique non identifie - ELEVE"),

    ("R_099", 99, "OUI", "*", "CATCH_ALL", "*", "*",
     "INCONNU", "NON_CLASSE", None, None,
     "A_CONTROLER", "NON", "NON", "MOYEN", "A_CONTROLER", "A_ENVOYER_IA",
     None, None,
     "Catch-all technique - vers IA_Classification (stub)"),
]

# Regles triggering CTRL entries
CTRL_TRIGGER_RULES = {
    "R_010": ("IMPAYE_DETECTE",           "A_CONTROLER"),
    "R_020": ("VIR_ASSOCIE_DETECTE",      "A_CONTROLER"),
    "R_021": ("VIR_ASSOCIE_DETECTE",      "A_CONTROLER"),
    "R_040": ("VIREMENT_BANCAIRE_AMBIGU", "A_CONTROLER"),
    "R_085": ("VIREMENT_BANCAIRE_AMBIGU", "A_CONTROLER"),
    "R_090": ("VIREMENT_BANCAIRE_AMBIGU", "A_CONTROLER"),
    "R_091": ("VIREMENT_BANCAIRE_AMBIGU", "A_CONTROLER"),
    "R_099": ("IA_CONFIANCE_INSUFFISANTE","A_CONTROLER"),
}

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def rules_as_dicts(seed):
    """Convert SEED_RULES tuples to dicts, sorted by priority."""
    result = []
    for t in sorted(seed, key=lambda x: (x[1], x[0])):
        result.append({k: v for k, v in zip(REGLES_HDR, t)})
    return result


def apply_rule(libelle_norm, rules):
    """First-match deterministic classifier. libelle_norm = uppercase stripped."""
    for r in rules:
        tm = r["type_match"]
        motif = r["motif"]
        if tm == "CATCH_ALL":
            return r
        elif tm == "COMMENCE_PAR":
            if libelle_norm.startswith(motif):
                return r
        elif tm == "CONTIENT":
            if motif in libelle_norm:
                return r
        elif tm == "REGEX":
            if re.search(motif, libelle_norm):
                return r
    return None  # never reached if CATCH_ALL present


def style_cell(cell, fill=None, font=None, alignment=None, border=None):
    if fill:      cell.fill      = fill
    if font:      cell.font      = font
    if alignment: cell.alignment = alignment
    if border:    cell.border    = border


def write_header_row(ws, row, headers, col_widths=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row, i, h)
        style_cell(c, fill=HDR_FILL, font=HDR_FONT, alignment=ALN_CTR)
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def backup_file(src_path, label):
    if not os.path.exists(src_path):
        print("ERREUR backup: fichier absent - %s" % src_path)
        return
    os.makedirs(ARCHIVE_DIR, exist_ok=True)
    ext = os.path.splitext(src_path)[1]
    dst = os.path.join(ARCHIVE_DIR, "%s_PRE_LOT8B_%s%s" % (label, TS, ext))
    shutil.copy2(src_path, dst)
    print("  Backup: %s" % os.path.basename(dst))


# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 - UPDATE REF_Setup.xlsm
# ─────────────────────────────────────────────────────────────────────────────

def update_ref_setup():
    print("\n=== STEP 1: REF_Setup.xlsm ===")

    backup_file(REF_PATH, "REF_Setup")

    wb = openpyxl.load_workbook(REF_PATH, keep_vba=True)

    # --- A. Ajouter TYPE_FLUX_016 dans REF_Types_Flux ---
    ws_tf = wb["REF_Types_Flux"]
    existing_ids = [ws_tf.cell(r, 1).value for r in range(2, ws_tf.max_row + 1)]
    if "TYPE_FLUX_016" in existing_ids:
        print("  TYPE_FLUX_016 deja present - skip")
    else:
        new_row = ws_tf.max_row + 1
        tf_data = [
            "TYPE_FLUX_016",
            "FRAIS_BANCAIRES",
            "Frais bancaires",
            "IC",
            "NON",
            "NON",
            "OUI",
            "OUI",
            "Source banque - frais bancaires deterministes, hors frais lies a une operation a controler.",
        ]
        for c, val in enumerate(tf_data, 1):
            cell = ws_tf.cell(new_row, c, val)
            style_cell(cell, font=DATA_FONT, alignment=ALN_LEFT, border=THIN_BORDER)
        print("  TYPE_FLUX_016 ajoute (ligne %d)" % new_row)

    # --- B. Ajouter onglet REF_Banque_Regles ---
    if "REF_Banque_Regles" in wb.sheetnames:
        del wb["REF_Banque_Regles"]
        print("  Onglet REF_Banque_Regles existant supprime (recréation)")

    ws_r = wb.create_sheet("REF_Banque_Regles")

    col_widths = [12, 10, 8, 14, 18, 14, 30, 28, 36, 20, 12, 28, 20, 20, 12, 26, 30, 18, 18, 55]
    write_header_row(ws_r, 1, REGLES_HDR, col_widths)
    ws_r.row_dimensions[1].height = 30
    ws_r.freeze_panes = "A2"

    rules_dicts = rules_as_dicts(SEED_RULES)
    for row_idx, r in enumerate(rules_dicts, 2):
        for c_idx, key in enumerate(REGLES_HDR, 1):
            val = r[key]
            cell = ws_r.cell(row_idx, c_idx, val)
            # Fill by niveau_risque
            nr = r["niveau_risque"]
            if nr == "ELEVE":
                fill = FILL_ELEVE
            elif nr == "MOYEN":
                fill = FILL_MOYEN
            elif nr == "FAIBLE":
                fill = FILL_FAIBLE
            else:
                fill = FILL_WHITE
            style_cell(cell, fill=fill, font=DATA_FONT, alignment=ALN_LEFT, border=THIN_BORDER)
        ws_r.row_dimensions[row_idx].height = 16

    print("  REF_Banque_Regles: %d regles seed ecrites" % len(rules_dicts))

    wb.save(REF_PATH)
    print("  REF_Setup.xlsm sauvegarde")
    return rules_dicts


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 - CLASSIFY NORM_Banque + CTRL + IA + LOG
# ─────────────────────────────────────────────────────────────────────────────

def update_banque_import(rules_dicts):
    print("\n=== STEP 2: BANQUE_LOT8_IMPORT.xlsx ===")

    backup_file(BANQUE_PATH, "BANQUE_LOT8_IMPORT")

    wb = openpyxl.load_workbook(BANQUE_PATH)
    ws_norm = wb["NORM_Banque"]
    ws_ctrl = wb["CTRL_A_CONTROLER"]
    ws_log  = wb["LOG_Traitement"]

    # ── Column map for NORM_Banque (1-indexed) ──
    hdr = {ws_norm.cell(1, c).value: c for c in range(1, ws_norm.max_column + 1)}
    COL_LIBELLE          = hdr["libelle"]         # 7
    COL_MOUVEMENT_ID     = hdr["mouvement_id"]    # 1
    COL_LIGNE_SOURCE     = hdr["ligne_source"]    # 4
    COL_DATE_OP          = hdr["date_operation"]  # 5
    COL_DATE_VAL         = hdr["date_valeur"]     # 6
    COL_MONTANT          = hdr["montant"]         # 9
    COL_SENS             = hdr["sens"]            # 10
    COL_TIERS            = hdr["tiers_detecte"]   # 13
    COL_CAT              = hdr["categorie"]       # 14
    COL_TF               = hdr["type_flux_id"]    # 15
    COL_CI               = hdr["code_impact"]     # 16
    COL_SRC_CLASS        = hdr["source_classification"]  # 17
    COL_SRC_ECO          = hdr["source_economique"]      # 18
    COL_STATUT_CTRL      = hdr["statut_controle"]        # 19
    COL_NIV_RISQUE       = hdr["niveau_risque"]          # 20
    COL_CODES_ANOM       = hdr["codes_anomalie"]         # 21
    # New columns (extend header if not already present)
    COL_STATUT_CLASS     = 24
    COL_NIV_ANOM         = 25
    COL_REGLE_ID         = 26

    # Write new headers
    new_headers = ["statut_classification", "niveau_anomalie", "regle_id_appliquee"]
    for i, h in enumerate(new_headers, COL_STATUT_CLASS):
        c = ws_norm.cell(1, i, h)
        style_cell(c, fill=SUBHDR_FILL, font=HDR_FONT, alignment=ALN_CTR)

    # ── Classify each NORM row ──
    stats = {
        "total": 0, "valide": 0, "a_controler": 0,
        "rapprochement": 0, "a_envoyer_ia": 0, "classe": 0,
        "eleve": 0, "moyen": 0, "faible": 0,
    }

    ctrl_rows_new = []   # (mouvement_id, ligne_source, date_op, date_val, lib, montant, sens, code, sev, desc, sc)
    ia_rows = []         # rows for IA_Classification

    for row_idx in range(2, ws_norm.max_row + 1):
        libelle_raw = ws_norm.cell(row_idx, COL_LIBELLE).value or ""
        lib_norm = libelle_raw.upper().strip()
        mouvement_id = ws_norm.cell(row_idx, COL_MOUVEMENT_ID).value
        ligne_source = ws_norm.cell(row_idx, COL_LIGNE_SOURCE).value
        date_op = ws_norm.cell(row_idx, COL_DATE_OP).value
        date_val = ws_norm.cell(row_idx, COL_DATE_VAL).value
        montant = ws_norm.cell(row_idx, COL_MONTANT).value
        sens = ws_norm.cell(row_idx, COL_SENS).value

        rule = apply_rule(lib_norm, rules_dicts)
        if rule is None:
            print("  WARN: aucune regle pour ligne %d - %s" % (row_idx, lib_norm[:60]))
            continue

        stats["total"] += 1
        regle_id = rule["regle_id"]
        statut_ctrl = rule["statut_controle_defaut"]
        statut_class = rule["statut_classification_defaut"]
        niveau_risque = rule["niveau_risque"]

        # Safety: ELEVE -> force A_CONTROLER
        if niveau_risque == "ELEVE":
            statut_ctrl = "A_CONTROLER"

        # niveau_anomalie
        lib_upper = lib_norm
        has_remboursement = "REMBOURSEMENT" in lib_upper or "REMBT" in lib_upper
        if niveau_risque == "ELEVE" or has_remboursement:
            niveau_anom = "A_CONTROLER"
        else:
            niveau_anom = None

        # Source classification
        src_class = "REGLE_DETERMINISTE"

        # Stats
        if statut_ctrl == "VALIDE":
            stats["valide"] += 1
        else:
            stats["a_controler"] += 1
        if statut_class == "RAPPROCHEMENT_REQUIS":
            stats["rapprochement"] += 1
        elif statut_class == "A_ENVOYER_IA":
            stats["a_envoyer_ia"] += 1
        else:
            stats["classe"] += 1
        if niveau_risque == "ELEVE":   stats["eleve"] += 1
        elif niveau_risque == "MOYEN": stats["moyen"] += 1
        elif niveau_risque == "FAIBLE":stats["faible"] += 1

        # ── Write classification to NORM_Banque cells ──
        def set_cell(col, val, fill=None, font=None, aln=None):
            cc = ws_norm.cell(row_idx, col)
            cc.value = val  # explicit assignment — ws.cell(r,c,None) skips None in openpyxl
            style_cell(cc, fill=fill, font=font or DATA_FONT, alignment=aln or ALN_LEFT)

        def _none_or(val):
            return val if val is not None else None

        set_cell(COL_TIERS,       _none_or(rule["tiers_detecte"]))
        set_cell(COL_CAT,         _none_or(rule["categorie"]))
        set_cell(COL_TF,          _none_or(rule["type_flux_id"]))
        set_cell(COL_CI,          _none_or(rule["code_impact"]))
        set_cell(COL_SRC_CLASS,   src_class)
        set_cell(COL_SRC_ECO,     _none_or(rule["source_economique"]))
        set_cell(COL_NIV_RISQUE,  niveau_risque,
                 fill=FILL_ELEVE if niveau_risque=="ELEVE" else (FILL_MOYEN if niveau_risque=="MOYEN" else FILL_FAIBLE))

        # statut_controle: fill + value
        sc_cell = ws_norm.cell(row_idx, COL_STATUT_CTRL, statut_ctrl)
        style_cell(sc_cell,
                   fill=FILL_OK if statut_ctrl == "VALIDE" else FILL_ACTRL,
                   font=DATA_FONT, alignment=ALN_CTR)

        # New cols
        sc2_fill = FILL_RAPP if statut_class == "RAPPROCHEMENT_REQUIS" else (FILL_IA if statut_class == "A_ENVOYER_IA" else FILL_WHITE)
        set_cell(COL_STATUT_CLASS, statut_class, fill=sc2_fill, aln=ALN_CTR)
        set_cell(COL_NIV_ANOM,     niveau_anom,
                 fill=FILL_ELEVE if niveau_anom else FILL_WHITE)
        set_cell(COL_REGLE_ID,     regle_id, aln=ALN_CTR)

        # ── CTRL entries ──
        ctrl_code, ctrl_sev = CTRL_TRIGGER_RULES.get(regle_id, (None, None))
        if ctrl_code:
            desc_map = {
                "IMPAYE_DETECTE":           "Impaye detecte - verifier retour debit et impact (ELEVE)",
                "VIR_ASSOCIE_DETECTE":      "Virement associe Wafa Souci detecte - ELEVE - controle obligatoire",
                "VIREMENT_BANCAIRE_AMBIGU": "Virement/effet non identifie - controle humain requis",
                "IA_CONFIANCE_INSUFFISANTE":"Aucune regle deterministe - envoi IA (stub)",
            }
            ctrl_rows_new.append((
                mouvement_id, ligne_source, date_op, date_val, libelle_raw,
                montant, sens, ctrl_code, ctrl_sev, desc_map.get(ctrl_code, ""),
                "A_CONTROLER",
            ))

        if has_remboursement and ctrl_code not in ("IMPAYE_DETECTE", "VIR_ASSOCIE_DETECTE"):
            ctrl_rows_new.append((
                mouvement_id, ligne_source, date_op, date_val, libelle_raw,
                montant, sens, "REMBOURSEMENT_BANCAIRE_AMBIGU", "A_CONTROLER",
                "Libelle contient REMBOURSEMENT/REMBT - nature et beneficiaire a verifier",
                "A_CONTROLER",
            ))

        # ── IA rows ──
        if statut_class == "A_ENVOYER_IA":
            ia_rows.append((
                mouvement_id, date_op, libelle_raw, montant, sens,
                None, None, "Catch-all R_099 - classification automatique indisponible",
                "EN_ATTENTE_IA",
            ))

    # ── CTRL: purge lignes LOT8B des runs precedents (idempotence) ──
    LOT8B_CTRL_CODES = {
        "IMPAYE_DETECTE", "VIR_ASSOCIE_DETECTE", "VIREMENT_BANCAIRE_AMBIGU",
        "IA_CONFIANCE_INSUFFISANTE", "REMBOURSEMENT_BANCAIRE_AMBIGU",
    }
    for r in range(ws_ctrl.max_row, 1, -1):
        if ws_ctrl.cell(r, 8).value in LOT8B_CTRL_CODES:
            ws_ctrl.delete_rows(r, 1)

    # ── Append CTRL rows ──
    next_ctrl_row = ws_ctrl.max_row + 1
    for r in ctrl_rows_new:
        for c_idx, val in enumerate(r, 1):
            cell = ws_ctrl.cell(next_ctrl_row, c_idx, val)
            style_cell(cell, font=DATA_FONT, alignment=ALN_LEFT, border=THIN_BORDER)
            if c_idx == 9:  # severite
                cell.fill = FILL_ACTRL
        next_ctrl_row += 1

    print("  CTRL: %d nouvelles entrees ajoutees" % len(ctrl_rows_new))

    # ── Create IA_Classification sheet ──
    if "IA_Classification" in wb.sheetnames:
        del wb["IA_Classification"]

    ws_ia = wb.create_sheet("IA_Classification")
    ia_headers = [
        "mouvement_id", "date_operation", "libelle", "montant", "sens",
        "categorie_proposee", "score_confiance", "commentaire_ia", "statut_ia",
    ]
    ia_widths = [30, 14, 60, 12, 10, 30, 16, 60, 18]
    write_header_row(ws_ia, 1, ia_headers, ia_widths)
    ws_ia.freeze_panes = "A2"
    for row_idx, r in enumerate(ia_rows, 2):
        for c_idx, val in enumerate(r, 1):
            cell = ws_ia.cell(row_idx, c_idx, val)
            style_cell(cell, font=DATA_FONT, alignment=ALN_LEFT, border=THIN_BORDER)

    ia_count = len(ia_rows)
    if ia_count == 0:
        info_cell = ws_ia.cell(2, 1, "Aucune ligne orientee IA - toutes les lignes classees par regles deterministes")
        info_cell.font = Font(name="Calibri", italic=True, color="888888", size=10)
    print("  IA_Classification: %d lignes" % ia_count)

    # ── LOG: purge entrees LOT8B_CLASSIF des runs precedents (idempotence) ──
    for r in range(ws_log.max_row, 1, -1):
        if ws_log.cell(r, 3).value == "LOT8B_CLASSIF":
            ws_log.delete_rows(r, 1)

    # ── Update LOG_Traitement ──
    run_id = "RUN-LOT8B-%s" % TS
    nb_eleve = stats["eleve"]
    log_row = [
        run_id, IMPORT_ID, "LOT8B_CLASSIF", NOW_STR,
        stats["total"], None, None, stats["total"],
        nb_eleve, stats["a_controler"],
        0, None,
        "lot8b_banque_regles.py - %d regles, %d VALIDE, %d A_CONTROLER, %d RAPPROCHEMENT_REQUIS, %d IA"
        % (len(rules_dicts), stats["valide"], stats["a_controler"],
           stats["rapprochement"], ia_count),
    ]
    log_row_idx = ws_log.max_row + 1
    for c_idx, val in enumerate(log_row, 1):
        cell = ws_log.cell(log_row_idx, c_idx, val)
        style_cell(cell, font=DATA_FONT, alignment=ALN_LEFT, border=THIN_BORDER)
    print("  LOG_Traitement: ligne LOT8B_CLASSIF ajoutee (row %d)" % log_row_idx)

    wb.save(BANQUE_PATH)
    print("  BANQUE_LOT8_IMPORT.xlsx sauvegarde")

    return stats, len(ctrl_rows_new), ia_count


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("=" * 68)
    print("LOT 8B - REF_Banque_Regles + Classification deterministe")
    print("Timestamp: %s" % NOW_STR)
    print("=" * 68)

    rules_dicts = update_ref_setup()
    stats, nb_ctrl, nb_ia = update_banque_import(rules_dicts)

    print("\n" + "=" * 68)
    print("BILAN LOT 8B")
    print("=" * 68)
    print("  REF_Setup.xlsm:")
    print("    TYPE_FLUX_016 (FRAIS_BANCAIRES) ajoute")
    print("    REF_Banque_Regles: %d regles seed (20 colonnes)" % len(rules_dicts))
    print("")
    print("  NORM_Banque (%d lignes classifiees):" % stats["total"])
    print("    statut_controle VALIDE      : %d" % stats["valide"])
    print("    statut_controle A_CONTROLER : %d" % stats["a_controler"])
    print("    statut_classification RAPPROCHEMENT_REQUIS : %d" % stats["rapprochement"])
    print("    statut_classification CLASSE               : %d" % stats["classe"])
    print("    statut_classification A_ENVOYER_IA         : %d" % stats["a_envoyer_ia"])
    print("    niveau_risque ELEVE  : %d" % stats["eleve"])
    print("    niveau_risque MOYEN  : %d" % stats["moyen"])
    print("    niveau_risque FAIBLE : %d" % stats["faible"])
    print("")
    print("  CTRL_A_CONTROLER: %d nouvelles entrees" % nb_ctrl)
    print("  IA_Classification: %d lignes (catch-all)" % nb_ia)
    print("  LOG_Traitement: entree LOT8B_CLASSIF ajoutee")
    print("")
    print("  Backups:")
    print("    REF_Setup_PRE_LOT8B_%s.xlsm" % TS)
    print("    BANQUE_LOT8_IMPORT_PRE_LOT8B_%s.xlsx" % TS)
    print("")
    print("  Fichiers modifies:")
    print("    01_SOURCES_BRUTES/REF_Setup/REF_Setup.xlsm")
    print("    02_TRAVAIL/Lot8_Banque/BANQUE_LOT8_IMPORT.xlsx")
    print("")
    print("CONTRAINTE: ne pas commiter les fichiers bancaires.")
    print("Seuls lot8b_banque_regles.py et REF_Setup.xlsm seront commites.")
    print("=" * 68)


if __name__ == "__main__":
    main()
