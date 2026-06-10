#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
lot8a_banque_import.py
Lot 8a — Import & normalisation bancaire Crédit Mutuel
Ingère 01_SOURCES_BRUTES/Banque/2026_03_BRUT_Banque_CreditMutuel.xlsx
Produit 02_TRAVAIL/Lot8_Banque/BANQUE_LOT8_IMPORT.xlsx

Onglets :
  BRUT_Banque           — copie brute intégrale
  NORM_Banque           — mouvements normalisés + contrôles structurels
  CTRL_A_CONTROLER      — anomalies bloquantes et à contrôler
  LOG_Traitement        — trace du run import
  REF_Cloture_Mensuelle — structure vide des états de mois (alimentée Lot 8c)
  POWER_QUERY_CODE      — documentation technique

ARCHI §13.2–§13.4 / REGLES §6 B1–B10 / Plan Lot 8a.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import hashlib
import os
import sys
import shutil
import re
from datetime import datetime, date

# ── Chemins ────────────────────────────────────────────────────────────────
BASE        = os.path.dirname(os.path.abspath(__file__))
ROOT        = os.path.dirname(BASE)
BANQUE_DIR  = os.path.join(ROOT, '01_SOURCES_BRUTES', 'Banque')
BRUT_FILE   = os.path.join(BANQUE_DIR, '2026_03_BRUT_Banque_CreditMutuel.xlsx')
OUT_DIR     = os.path.join(BASE, 'Lot8_Banque')
OUT_FILE    = os.path.join(OUT_DIR, 'BANQUE_LOT8_IMPORT.xlsx')
ARCHIVE_DIR = os.path.join(ROOT, '99_ARCHIVES', 'LOT8_Banque')

TS = datetime.now().strftime('%Y%m%d_%H%M%S')

os.makedirs(OUT_DIR,     exist_ok=True)
os.makedirs(ARCHIVE_DIR, exist_ok=True)

# ── Constantes ────────────────────────────────────────────────────────────
COMPTE_ID    = 'CM_02211_00021321603'
IMPORT_ID    = 'IMP-BQ-CM-2026-03-001'
DEVISE_REF   = 'EUR'
SHEET_METIER = 'Cpt 02211 00021321603'
HDR_ROW      = 5
DATA_ROW     = 6
NOM_ANNEE    = 2026
NOM_MOIS     = 3

BLOQUANT_CODES = {
    'BANQUE_DATE_INEXPLOITABLE',
    'BANQUE_LIGNE_SANS_LIBELLE',
    'BANQUE_DEBIT_CREDIT_VIDES',
    'BANQUE_DEBIT_CREDIT_DOUBLES',
    'BANQUE_MONTANT_NON_NUMERIQUE',
}

# ── Styles ─────────────────────────────────────────────────────────────────
FILL_HDR      = PatternFill('solid', fgColor='1F4E79')
FILL_SEC      = PatternFill('solid', fgColor='2E75B6')
FILL_BLOQUANT = PatternFill('solid', fgColor='C00000')
FILL_ACTRL    = PatternFill('solid', fgColor='ED7D31')
FILL_OK       = PatternFill('solid', fgColor='70AD47')
FILL_INFO     = PatternFill('solid', fgColor='DEEAF1')

FONT_HDR  = Font(bold=True, color='FFFFFF', size=10)
FONT_BOLD = Font(bold=True, size=10)
FONT_NRM  = Font(size=10)
FONT_INFO = Font(italic=True, size=9, color='555555')

ALIGN_CTR  = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(vertical='center')


def style_header(ws, row, ncols, fill=None):
    f = fill or FILL_HDR
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.fill = f
        cell.font = FONT_HDR
        cell.alignment = ALIGN_CTR


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Helpers ────────────────────────────────────────────────────────────────

def normalize_libelle(s):
    if s is None:
        return ''
    s = str(s).strip().upper()
    s = re.sub(r'\s+', ' ', s)
    return s


def parse_date(val):
    """Return (date_obj, True) or (None, False)."""
    if val is None:
        return None, False
    if isinstance(val, datetime):
        return val.date(), True
    if isinstance(val, date):
        return val, True
    s = str(val).strip()
    if not s:
        return None, False
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y'):
        try:
            return datetime.strptime(s, fmt).date(), True
        except ValueError:
            pass
    return None, False


def to_float(val):
    """Return (float, True) or (None, False)."""
    if val is None or str(val).strip() == '':
        return None, False
    cleaned = str(val).replace(',', '.').replace('\xa0', '').replace(' ', '')
    try:
        return float(cleaned), True
    except ValueError:
        return None, False


def make_hash(parts):
    raw = '|'.join(str(p) for p in parts)
    return hashlib.sha256(raw.encode('utf-8')).hexdigest()


def is_cm_footer_row(cols):
    """Return True for known CM export footer rows to exclude (ARCHI §13.6).

    Handles three CM footer patterns:
    - 'LISTE DE VOS COMPTES' in any cell
    - 'SOLDE AU' prefix in any cell (ligne de solde final, cols Débit/Crédit = texte)
    - Row with no date, no libellé, no *numeric* amount (blank structural row)
    """
    def nonempty(v):
        return v is not None and str(v).strip() != ''

    def is_numeric(v):
        if v is None:
            return False
        if isinstance(v, (int, float)):
            return True
        try:
            float(str(v).replace(',', '.').replace('\xa0', '').replace(' ', ''))
            return True
        except (ValueError, TypeError):
            return False

    # Pattern check on all cells
    for c in cols:
        val = str(c or '').strip().upper()
        if 'LISTE DE VOS COMPTES' in val:
            return True
        if val.startswith('SOLDE AU'):
            return True

    # Structural blank: no date, no libellé, no numeric amount
    lib = str(cols[2] or '').strip().upper() if len(cols) > 2 else ''
    has_date          = nonempty(cols[0]) or (len(cols) > 1 and nonempty(cols[1]))
    has_lib           = bool(lib)
    has_numeric_amount = (
        (len(cols) > 3 and is_numeric(cols[3])) or
        (len(cols) > 4 and is_numeric(cols[4]))
    )
    return not has_date and not has_lib and not has_numeric_amount


# ==========================================================================
# 1. LECTURE DU FICHIER BRUT (B2 : jamais modifié)
# ==========================================================================

if not os.path.exists(BRUT_FILE):
    print('[ERREUR BLOQUANT] Fichier bancaire brut introuvable :')
    print(f'  {BRUT_FILE}')
    print('Déposez le fichier dans 01_SOURCES_BRUTES/Banque/ puis relancez.')
    sys.exit(1)

print(f'[OK] Source brute : {BRUT_FILE}')
wb_src = openpyxl.load_workbook(BRUT_FILE, data_only=True, read_only=True)

if SHEET_METIER not in wb_src.sheetnames:
    print(f'[ERREUR BLOQUANT] Feuille "{SHEET_METIER}" absente.')
    print(f'  Feuilles disponibles : {wb_src.sheetnames}')
    sys.exit(1)

ws_src = wb_src[SHEET_METIER]
raw_rows       = []
excluded_footer = []
for row in ws_src.iter_rows(min_row=DATA_ROW, values_only=True):
    # Ignorer lignes entièrement vides
    if all(c is None or str(c).strip() == '' for c in row):
        continue
    # Pad à 7 colonnes (Date|Valeur|Libellé|Débit|Crédit|Solde|Dev)
    padded = list(row) + [None] * max(0, 7 - len(row))
    cols7  = padded[:7]
    # Exclure pieds d'export CM (ARCHI §13.6)
    if is_cm_footer_row(cols7):
        excluded_footer.append(cols7)
        continue
    raw_rows.append(cols7)
wb_src.close()
print(f'[OK] {len(raw_rows)} mouvements retenus '
      f'({len(excluded_footer)} pied(s) export exclus) depuis ligne {DATA_ROW}')

# ==========================================================================
# 2. NORMALISATION + CONTRÔLES STRUCTURELS
# ==========================================================================

date_integration = datetime.now().strftime('%Y-%m-%d')
empreintes_vues  = {}
norm_data        = []
ctrl_data        = []
stats = dict(bloquants=0, a_controler=0, doublons=0, ok=0)

for idx, r in enumerate(raw_rows, 1):
    ligne_src = DATA_ROW + idx - 1
    date_brute, val_brute, lib_brut, deb_brut, cre_brut, sol_brut, dev_brut = r

    anomalies = []

    # ── Libellé ──────────────────────────────────────────────────────────
    lib_norm = normalize_libelle(lib_brut)
    if not lib_norm:
        anomalies.append('BANQUE_LIGNE_SANS_LIBELLE')

    # ── Dates (B1 : rattachement par colonnes, jamais par nom du fichier) ─
    date_op,  op_ok  = parse_date(date_brute)
    date_val, val_ok = parse_date(val_brute)

    if not op_ok and not val_ok:
        anomalies.append('BANQUE_DATE_INEXPLOITABLE')
    elif not op_ok and val_ok:
        anomalies.append('BANQUE_LIGNE_SANS_DATE')
        date_op = date_val  # fallback : utiliser date_valeur

    # ── Devise ────────────────────────────────────────────────────────────
    devise = str(dev_brut).strip().upper() if dev_brut else DEVISE_REF
    if not devise:
        devise = DEVISE_REF
    if devise != DEVISE_REF:
        anomalies.append('BANQUE_DEVISE_NON_EUR')

    # ── Montant / Sens ────────────────────────────────────────────────────
    deb_f, deb_ok = to_float(deb_brut)
    cre_f, cre_ok = to_float(cre_brut)
    sens             = None
    montant          = None
    montant_centimes = 0

    deb_renseigne = deb_ok and deb_f is not None
    cre_renseigne = cre_ok and cre_f is not None

    if not deb_renseigne and not cre_renseigne:
        if not deb_ok and not cre_ok:
            anomalies.append('BANQUE_DEBIT_CREDIT_VIDES')
        else:
            anomalies.append('BANQUE_MONTANT_NON_NUMERIQUE')
    elif deb_renseigne and cre_renseigne:
        anomalies.append('BANQUE_DEBIT_CREDIT_DOUBLES')
    elif deb_renseigne:
        sens             = 'DEBIT'
        montant          = round(abs(deb_f), 2)
        montant_centimes = int(round(montant * 100))
    else:
        sens             = 'CREDIT'
        montant          = round(abs(cre_f), 2)
        montant_centimes = int(round(montant * 100))

    # ── Empreinte anti-doublon (B9) ───────────────────────────────────────
    date_op_str  = date_op.strftime('%Y-%m-%d')  if date_op  else 'NODATE'
    date_val_str = date_val.strftime('%Y-%m-%d') if date_val else 'NODATE'

    row_hash   = make_hash([COMPTE_ID, date_op_str, date_val_str,
                             str(sens), str(montant_centimes), lib_norm, devise])
    hash_court = row_hash[:6].upper()

    if row_hash in empreintes_vues:
        anomalies.append('DOUBLON_BANCAIRE_POTENTIEL')
        stats['doublons'] += 1
    else:
        empreintes_vues[row_hash] = ligne_src

    # ── mouvement_id ──────────────────────────────────────────────────────
    date_str     = date_op_str.replace('-', '')
    mouvement_id = (f'MVT-{COMPTE_ID}-{date_str}'
                    f'-{sens or "NOFLOW"}-{montant_centimes}-{hash_court}')

    # ── statut_controle ───────────────────────────────────────────────────
    bloquant = any(c in BLOQUANT_CODES for c in anomalies)
    if bloquant:
        statut = 'BLOQUANT'
        stats['bloquants'] += 1
    elif anomalies:
        statut = 'A_CONTROLER'
        stats['a_controler'] += 1
    else:
        statut = 'EN_ATTENTE_CLASSIFICATION'
        stats['ok'] += 1

    codes_str = ' | '.join(anomalies) if anomalies else ''

    norm_data.append([
        mouvement_id,                # 1  mouvement_id
        row_hash[:16],               # 2  ROW_HASH (16 premiers chars)
        IMPORT_ID,                   # 3  import_id
        ligne_src,                   # 4  ligne_source
        date_op_str,                 # 5  date_operation
        date_val_str,                # 6  date_valeur
        lib_norm,                    # 7  libelle
        str(lib_brut or ''),         # 8  libelle_brut
        montant,                     # 9  montant
        sens,                        # 10 sens
        devise,                      # 11 devise
        COMPTE_ID,                   # 12 compte_id
        None,                        # 13 tiers_detecte
        None,                        # 14 categorie
        None,                        # 15 type_flux_id
        None,                        # 16 code_impact
        'NON_CLASSE',                # 17 source_classification
        None,                        # 18 source_economique
        statut,                      # 19 statut_controle
        None,                        # 20 niveau_risque
        codes_str,                   # 21 codes_anomalie
        date_integration,            # 22 date_integration
        None,                        # 23 commentaire
    ])

    for code in anomalies:
        sev = 'BLOQUANT' if code in BLOQUANT_CODES else 'A_CONTROLER'
        ctrl_data.append([
            mouvement_id,
            ligne_src,
            date_op_str,
            date_val_str,
            lib_norm or str(lib_brut or ''),
            montant,
            sens,
            code,
            sev,
            'Contrôle structurel Lot 8a',
            statut,
        ])

# ── Contrôle global BANQUE_FICHIER_PERIODE_INCOHERENTE ────────────────────
real_dates = []
for row in norm_data:
    d = row[4]  # date_operation
    if d and d != 'NODATE':
        try:
            real_dates.append(datetime.strptime(d, '%Y-%m-%d').date())
        except ValueError:
            pass

periode_incoherente = False
dmin_global = dmax_global = None
if real_dates:
    dmin_global = min(real_dates)
    dmax_global = max(real_dates)
    if not (dmin_global.year == NOM_ANNEE and dmin_global.month == NOM_MOIS
            and dmax_global.year == NOM_ANNEE and dmax_global.month == NOM_MOIS):
        periode_incoherente = True
        ctrl_data.insert(0, [
            f'IMPORT-{IMPORT_ID}',
            'GLOBAL',
            str(dmin_global),
            str(dmax_global),
            (f'Période réelle {dmin_global}→{dmax_global}, '
             f'nominal {NOM_ANNEE}-{NOM_MOIS:02d}'),
            None,
            None,
            'BANQUE_FICHIER_PERIODE_INCOHERENTE',
            'A_CONTROLER',
            (f'Le fichier est nommé {NOM_ANNEE}-{NOM_MOIS:02d} '
             f'mais couvre {dmin_global} → {dmax_global}. '
             f'Non bloquant si au moins une date exploitable (B10).'),
            'A_CONTROLER',
        ])
        print(f'[WARN] BANQUE_FICHIER_PERIODE_INCOHERENTE : {dmin_global} -> {dmax_global}')

print(f'[OK] Normalisation terminée : {len(norm_data)} lignes NORM | '
      f'{stats["bloquants"]} BLOQUANT | {stats["a_controler"]} A_CONTROLER | '
      f'{stats["doublons"]} doublons | {stats["ok"]} EN_ATTENTE_CLASSIFICATION')

# ==========================================================================
# 3. ÉCRITURE EXCEL
# ==========================================================================

if os.path.exists(OUT_FILE):
    bak = os.path.join(ARCHIVE_DIR, f'BANQUE_LOT8_IMPORT_BACKUP_{TS}.xlsx')
    shutil.copy2(OUT_FILE, bak)
    print(f'[OK] Backup : {bak}')

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ==========================================================================
# Onglet 1 — BRUT_Banque
# ==========================================================================
ws_b = wb.create_sheet('BRUT_Banque')
BRUT_HDR = ['import_id', 'ligne_source', 'date_brute', 'valeur_brute',
            'libelle_brut', 'debit_brut', 'credit_brut', 'solde_brut', 'devise_brute']
BRUT_WID = [26, 13, 14, 14, 65, 14, 14, 14, 12]

ws_b.append(BRUT_HDR)
style_header(ws_b, 1, len(BRUT_HDR))
for idx, r in enumerate(raw_rows, 1):
    ws_b.append([IMPORT_ID, DATA_ROW + idx - 1] + r)
for row in ws_b.iter_rows(min_row=2):
    for cell in row:
        cell.font = FONT_NRM
        cell.alignment = ALIGN_LEFT
set_widths(ws_b, BRUT_WID)
ws_b.freeze_panes = 'A2'

# ==========================================================================
# Onglet 2 — NORM_Banque
# ==========================================================================
ws_n = wb.create_sheet('NORM_Banque')
NORM_HDR = [
    'mouvement_id', 'ROW_HASH', 'import_id', 'ligne_source',
    'date_operation', 'date_valeur', 'libelle', 'libelle_brut',
    'montant', 'sens', 'devise', 'compte_id',
    'tiers_detecte', 'categorie', 'type_flux_id', 'code_impact',
    'source_classification', 'source_economique',
    'statut_controle', 'niveau_risque',
    'codes_anomalie', 'date_integration', 'commentaire',
]
NORM_WID = [55, 20, 26, 13, 14, 14, 65, 65, 12, 10, 8, 30,
            25, 25, 20, 14, 28, 30, 28, 14, 55, 14, 35]

ws_n.append(NORM_HDR)
style_header(ws_n, 1, len(NORM_HDR))
for row in norm_data:
    ws_n.append(row)
    r_idx = ws_n.max_row
    statut = row[18]
    if statut == 'BLOQUANT':
        fill = FILL_BLOQUANT
        fnt  = Font(bold=True, color='FFFFFF', size=10)
    elif statut == 'A_CONTROLER':
        fill = FILL_ACTRL
        fnt  = Font(bold=True, color='FFFFFF', size=10)
    else:
        fill = FILL_INFO
        fnt  = FONT_NRM
    ws_n.cell(r_idx, 19).fill = fill
    ws_n.cell(r_idx, 19).font = fnt
    ws_n.cell(r_idx, 19).alignment = ALIGN_CTR
    for c in range(1, len(NORM_HDR) + 1):
        if c != 19:
            ws_n.cell(r_idx, c).font = FONT_NRM
            ws_n.cell(r_idx, c).alignment = ALIGN_LEFT
set_widths(ws_n, NORM_WID)
ws_n.freeze_panes = 'A2'
ws_n.auto_filter.ref = ws_n.dimensions

# ==========================================================================
# Onglet 3 — CTRL_A_CONTROLER
# ==========================================================================
ws_c = wb.create_sheet('CTRL_A_CONTROLER')
CTRL_HDR = [
    'mouvement_id', 'ligne_source', 'date_operation', 'date_valeur',
    'libelle', 'montant', 'sens',
    'code_controle', 'severite', 'description', 'statut_controle',
]
CTRL_WID = [55, 13, 14, 14, 55, 12, 10, 42, 14, 65, 22]

ws_c.append(CTRL_HDR)
style_header(ws_c, 1, len(CTRL_HDR))
for row in ctrl_data:
    ws_c.append(row)
    r_idx = ws_c.max_row
    sev = row[8]
    fill = FILL_BLOQUANT if sev == 'BLOQUANT' else FILL_ACTRL
    for col_idx in [9, 11]:
        cell = ws_c.cell(r_idx, col_idx)
        cell.fill = fill
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.alignment = ALIGN_CTR
    for col_idx in range(1, len(CTRL_HDR) + 1):
        if col_idx not in [9, 11]:
            ws_c.cell(r_idx, col_idx).font = FONT_NRM
            ws_c.cell(r_idx, col_idx).alignment = ALIGN_LEFT
set_widths(ws_c, CTRL_WID)
ws_c.freeze_panes = 'A2'
ws_c.auto_filter.ref = ws_c.dimensions

# ==========================================================================
# Onglet 4 — LOG_Traitement
# ==========================================================================
ws_l = wb.create_sheet('LOG_Traitement')
LOG_HDR = [
    'run_id', 'import_id', 'etape', 'timestamp',
    'nb_lignes_lues', 'nb_lignes_exclues_pied', 'nb_lignes_brut', 'nb_lignes_norm',
    'nb_bloquants', 'nb_a_controler', 'nb_doublons',
    'periode_incoherente', 'commentaire',
]
LOG_WID = [30, 26, 18, 22, 15, 22, 15, 15, 14, 16, 13, 22, 55]

ws_l.append(LOG_HDR)
style_header(ws_l, 1, len(LOG_HDR))
run_id = f'RUN-LOT8A-{TS}'
nb_lues_total = len(raw_rows) + len(excluded_footer)
ws_l.append([
    run_id,
    IMPORT_ID,
    'LOT8A_IMPORT',
    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    nb_lues_total,
    len(excluded_footer),
    len(raw_rows),
    len(norm_data),
    stats['bloquants'],
    stats['a_controler'],
    stats['doublons'],
    'OUI' if periode_incoherente else 'NON',
    f'lot8a_banque_import.py — {TS}',
])
for row in ws_l.iter_rows(min_row=2):
    for cell in row:
        cell.font = FONT_NRM
        cell.alignment = ALIGN_LEFT
set_widths(ws_l, LOG_WID)
ws_l.freeze_panes = 'A2'

# ==========================================================================
# Onglet 5 — REF_Cloture_Mensuelle (structure, alimentée Lot 8c)
# ==========================================================================
ws_cl = wb.create_sheet('REF_Cloture_Mensuelle')
CLOTURE_HDR = [
    'mois', 'statut_mois', 'date_passage_controle',
    'date_cloture', 'nb_lignes_bancaires_non_classees',
    'nb_controles_bloquants_ouverts', 'commentaire',
]
CLOTURE_WID = [12, 16, 24, 16, 34, 32, 55]

ws_cl.append(CLOTURE_HDR)
style_header(ws_cl, 1, len(CLOTURE_HDR))

# Initialiser les mois observés dans le fichier
months_obs = sorted(set((d.year, d.month) for d in real_dates)) if real_dates else []
for y, m in months_obs:
    ws_cl.append([
        f'{y}-{m:02d}', 'OUVERT', None, None, None, None,
        f'Auto-créé Lot 8a — à compléter Lot 8c',
    ])
for row in ws_cl.iter_rows(min_row=2):
    for cell in row:
        cell.font = FONT_NRM
        cell.alignment = ALIGN_LEFT
set_widths(ws_cl, CLOTURE_WID)
ws_cl.freeze_panes = 'A2'

# ==========================================================================
# Onglet 6 — POWER_QUERY_CODE (documentation)
# ==========================================================================
ws_pq = wb.create_sheet('POWER_QUERY_CODE')
ws_pq.column_dimensions['A'].width = 120
title_cell = ws_pq.cell(1, 1, 'POWER_QUERY_CODE — Lot 8a — Import Bancaire (Python)')
title_cell.fill = FILL_HDR
title_cell.font = FONT_HDR
title_cell.alignment = ALIGN_LEFT

doc_lines = [
    '',
    'Ce fichier est généré par lot8a_banque_import.py.',
    'Pas de transformation Power Query dans cet onglet — pipeline Python uniquement.',
    '',
    'Onglets produits :',
    '  BRUT_Banque           — Copie brute intégrale (import_id, ligne_source, 7 cols CM)',
    '  NORM_Banque           — Mouvements normalisés, empreinte anti-doublon, 23 colonnes',
    '  CTRL_A_CONTROLER      — Anomalies structurelles (BLOQUANT + A_CONTROLER)',
    '  LOG_Traitement        — Trace du run import',
    '  REF_Cloture_Mensuelle — États de mois (OUVERT/EN_CONTROLE/CLOTURE) — alimentée Lot 8c',
    '',
    'Règles appliquées :',
    '  B1  — Rattachement par colonnes Date/Valeur, jamais par nom de fichier',
    '  B2  — Fichier brut jamais modifié',
    '  B9  — Empreinte : compte_id|date_op|date_val|sens|montant_centimes|libellé_norm|devise',
    '  B10 — BANQUE_FICHIER_PERIODE_INCOHERENTE : A_CONTROLER (non bloquant)',
    '',
    'Source brute (jamais committée, .gitignore) :',
    '  01_SOURCES_BRUTES/Banque/2026_03_BRUT_Banque_CreditMutuel.xlsx',
    '',
    'Script : 02_TRAVAIL/lot8a_banque_import.py',
    'ARCHI §13.2–§13.4 / REGLES §6 / Plan Lot 8a',
]
for i, line in enumerate(doc_lines, 2):
    ws_pq.cell(i, 1, line).font = FONT_NRM

# ==========================================================================
# 4. SAUVEGARDE
# ==========================================================================

wb.save(OUT_FILE)
print(f'[OK] Fichier produit : {OUT_FILE}')

print()
print('=' * 65)
print('BILAN LOT 8a — Import & normalisation bancaire')
print('=' * 65)
nb_lues_total = len(raw_rows) + len(excluded_footer)
print(f'  Source brute       : 2026_03_BRUT_Banque_CreditMutuel.xlsx')
print(f'  Lignes lues total  : {nb_lues_total}')
print(f'  Pieds export exclus: {len(excluded_footer)}')
print(f'  Lignes BRUT/NORM   : {len(raw_rows)}')
print(f'  BLOQUANT           : {stats["bloquants"]}')
print(f'  A_CONTROLER        : {stats["a_controler"]}')
print(f'  Doublons potentiels: {stats["doublons"]}')
print(f'  EN_ATTENTE_CLASS.  : {stats["ok"]}')
if periode_incoherente and dmin_global and dmax_global:
    print(f'  Periode reelle     : {dmin_global} -> {dmax_global}')
    print(f'  [WARN A_CONTROLER] BANQUE_FICHIER_PERIODE_INCOHERENTE')
print(f'  Mois REF_Cloture   : {[f"{y}-{m:02d}" for y, m in months_obs]}')
print(f'  Sortie             : {OUT_FILE}')
print('=' * 65)
print()
print('Prochain lot : 8b — REF_Banque_Regles + classification déterministe')
print('Ne pas lancer avant validation humaine de ce bilan.')
