#!/usr/bin/env python3
"""
lot12_generer_factures.py
Lot 12 — Préfactures propriétaires (Module 10)

Produit (lecture seule des sorties Lot 10/11) :
  02_TRAVAIL/Lot12_Factures/MASTER_FACT_Proprietaires.xlsx
  Onglets : FACT_FACTURE_ENTETE / FACT_FACTURE_LIGNES / CONTROLE_MENSUEL /
            DASHBOARD_FACTURATION / A_CONTROLER

RÈGLE FONDAMENTALE : PRÉFACTURES UNIQUEMENT.
  Aucune facture finale tant que facturation_lot12_ok != OUI (DASHBOARD_MOIS, Lot 11).

Décisions (validées 2026-06-15) :
  D-LOT12-01 granularité prop × logement × mois (+ récap prop × mois en contrôle)
  D-LOT12-02 fichier unique multi-onglets
  D-LOT12-03 Excel seul (structure PDF-ready)
  D-LOT12-04 période mensuelle
  D-LOT12-05 balises {{...}} si coords société / adresse absentes
  D-LOT12-06 préfactures seulement
  D-LOT12-07 GLOBAL_NON_AFFECTE hors factures, en contrôle uniquement
  D-LOT12-08 numérotation PREF-AAAA-MM-PROP-LOG-NNN

Sources amont : lecture seule, aucune modification.
Dépendances : pip install pandas openpyxl
"""

import sys
import logging
from datetime import date
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

BASE       = Path(__file__).resolve().parent.parent
NET_FILE   = BASE / "02_TRAVAIL/Lot10_Resultats/MASTER_CALC_NetProprietaire.xlsx"
COMM_FILE  = BASE / "02_TRAVAIL/Lot10_Resultats/MASTER_CALC_Commissions.xlsx"
RSLT_FILE  = BASE / "02_TRAVAIL/Lot10_Resultats/MASTER_CALC_Resultats.xlsx"
CTRL_FILE  = BASE / "02_TRAVAIL/Lot11_Controles/MASTER_CTRL_Coherence.xlsx"
REF_FILE   = BASE / "01_SOURCES_BRUTES/REF_Setup/REF_Setup.xlsm"
OUT_DIR    = BASE / "02_TRAVAIL/Lot12_Factures"
OUT_FILE   = OUT_DIR / "MASTER_FACT_Proprietaires.xlsx"

SENTINEL_GLOBAL = "GLOBAL_NON_AFFECTE"
TOL = 0.10  # D035
TODAY = date.today().isoformat()

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)-8s %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger("lot12")

FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FONT_HEADER = Font(bold=True, color="FFFFFF", size=10)
FONT_DATA   = Font(size=10)


def _read(path, sheet=None, keep_vba=False):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=keep_vba)
    ws = wb[sheet] if sheet else wb.active
    it = ws.iter_rows(values_only=True)
    headers = list(next(it))
    rows = [dict(zip(headers, r)) for r in it]
    wb.close()
    return pd.DataFrame(rows)


def _n(v):
    try:
        return round(float(v), 2) if v is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


def _write_sheet(wb, name, df):
    ws = wb.create_sheet(name)
    cols = list(df.columns)
    for ci, col in enumerate(cols, 1):
        c = ws.cell(1, ci, col)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center")
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws.cell(ri, ci, val).font = FONT_DATA
    ws.freeze_panes = "A2"
    for ci, col in enumerate(cols, 1):
        m = len(str(col))
        if len(df) > 0:
            ln = df[col].astype(str).str.len().max()
            m = max(m, int(ln) if pd.notna(ln) else m)
        ws.column_dimensions[get_column_letter(ci)].width = min(m + 2, 50)


# ──────────────────────────────────────────────────────────────────────────
# 1. Chargement sources
# ──────────────────────────────────────────────────────────────────────────
log.info("=== Lot 12 — Préfactures propriétaires ===")
log.info("Chargement sources (lecture seule)...")

df_reg   = _read(NET_FILE,  "REGLEMENT")
df_ac    = _read(COMM_FILE, "A_CONTROLER")
df_glob  = _read(RSLT_FILE, "GLOBAL")
df_dash  = _read(CTRL_FILE, "DASHBOARD_MOIS")
df_prop  = _read(REF_FILE,  "REF_Proprietaires", keep_vba=True)
df_log   = _read(REF_FILE,  "REF_Logements",     keep_vba=True)

df_prop = df_prop[df_prop["proprietaire_id"].astype(str) != "proprietaire_id"].reset_index(drop=True)
df_log  = df_log[df_log["logement_id"].astype(str) != "logement_id"].reset_index(drop=True)

prop_idx = {r["proprietaire_id"]: r for _, r in df_prop.iterrows()}
log_idx  = {r["logement_id"]: r for _, r in df_log.iterrows()}

# Données partielles : HORS_COMPTA global = 0 -> sources HC réelles vides
hc_global = 0.0
for _, r in df_glob.iterrows():
    if r.get("vision") == "HORS_COMPTA":
        hc_global = _n(r.get("resultat"))
donnees_partielles = (hc_global == 0.0)

# Gate facturation par mois (DASHBOARD_MOIS Lot 11)
dash_idx = {}
for _, r in df_dash.iterrows():
    dash_idx[str(r["mois"])] = {
        "bloquants":  int(_n(r.get("nb_bloquants_ouverts"))),
        "a_controler": int(_n(r.get("nb_a_controler_ouverts"))),
        "ok":         str(r.get("facturation_lot12_ok")),
        "statut_banque": str(r.get("statut_mois_banque")),
    }
# A_CONTROLER transverses (non rattachés à un mois) -> impactent tous les mois
transverse = dash_idx.get("TRANSVERSE", {"bloquants": 0, "a_controler": 0, "ok": "NON", "statut_banque": "OUVERT"})

# Réservations A_CONTROLER par (mois) — best effort (A_CONTROLER n'a pas toujours mois/prop)
n_ac_reservations = len(df_ac)

log.info(f"  REGLEMENT       : {len(df_reg)} lignes")
log.info(f"  A_CONTROLER res : {n_ac_reservations}")
log.info(f"  DASHBOARD_MOIS  : {len(df_dash)} mois")
log.info(f"  Proprietaires   : {len(df_prop)} / Logements : {len(df_log)}")
log.info(f"  HORS_COMPTA global = {hc_global} -> donnees_partielles={donnees_partielles}")


# ──────────────────────────────────────────────────────────────────────────
# 2. Construction ENTETE + LIGNES (préfactures)
# ──────────────────────────────────────────────────────────────────────────
log.info("Construction préfactures...")

entetes = []
lignes  = []
controle = []
compteur = {}  # (mois) -> n

# GLOBAL_NON_AFFECTE -> contrôle uniquement (D-LOT12-07)
for _, r in df_reg.iterrows():
    mois   = r.get("mois")
    log_id = r.get("logement_id")
    prop_id = r.get("proprietaire_id")

    rec = {
        "mois": mois, "proprietaire_id": prop_id, "logement_id": log_id,
        "total_payout": _n(r.get("total_payout_mois")),
        "total_menage": _n(r.get("total_menage_mois")),
        "total_commission": _n(r.get("total_commission_mois")),
        "charge_fixe": _n(r.get("charge_fixe_mensuelle")),
        "revenu_net_exploitation": _n(r.get("net_proprietaire_apres_charge_mois")),
        "montant_du": _n(r.get("montant_du_conciergerie")),
        "acomptes": _n(r.get("autres_acomptes_recus")),
        "reste_a_payer": _n(r.get("reste_a_payer_conciergerie")),
        "nb_reservations": int(_n(r.get("nb_reservations"))),
        "statut": "CONTROLE_GLOBAL_NON_AFFECTE" if log_id == SENTINEL_GLOBAL else "RATTACHE_PROPRIETAIRE",
    }
    controle.append(rec)

    # D-LOT12-07 : charges globales non affectées -> jamais en facture
    if log_id == SENTINEL_GLOBAL or not prop_id or prop_id == SENTINEL_GLOBAL:
        continue

    # Numérotation PREF-AAAA-MM-PROP-LOG-NNN
    compteur[mois] = compteur.get(mois, 0) + 1
    facture_id = f"PREF-{mois}-{prop_id}-{log_id}-{compteur[mois]:03d}"

    pr = prop_idx.get(prop_id, {})
    lg = log_idx.get(log_id, {})
    nom_prop = f"{pr.get('prenom_proprietaire','') or ''} {pr.get('nom_proprietaire','') or ''}".strip()
    adresse_prop = pr.get("adresse_facturation")
    mode_fact = str(pr.get("mode_facturation") or "A_DEFINIR")

    # ── Gate statut (par mois) ──
    g = dash_idx.get(str(mois))
    bloquants = (g["bloquants"] if g else 0) + transverse["bloquants"]
    a_controler = (g["a_controler"] if g else 0) + transverse["a_controler"]
    statut_banque = g["statut_banque"] if g else "OUVERT"
    facturable_flag = (g["ok"] if g else "NON")

    if bloquants > 0:
        statut_facture = "NON_FACTURABLE_BLOQUANT"
    elif a_controler > 0 or facturable_flag != "OUI" or mode_fact == "A_DEFINIR":
        statut_facture = "NON_FACTURABLE_A_CONTROLER"
    else:
        statut_facture = "FACTURABLE"

    # D-LOT12-06 : préfactures uniquement -> generation toujours PREFACTURE
    statut_generation = "PREFACTURE_CONTROLE"

    # ── Balises ──
    balises = ["{{LOGO_A_INSERER}}", "{{SIRET_A_COMPLETER}}",
               "{{ADRESSE_SOCIETE_A_COMPLETER}}", "{{FACTURE_NON_FINALE}}"]
    if not adresse_prop or str(adresse_prop).strip() == "":
        balises.append("{{ADRESSE_PROPRIETAIRE_A_COMPLETER}}")
    if mode_fact == "A_DEFINIR":
        balises.append("{{MODE_FACTURATION_A_DEFINIR}}")
    if statut_banque != "CLOTURE":
        balises.append("{{BANQUE_NON_CLOTUREE}}")
    if a_controler > 0:
        balises.append("{{RESERVATIONS_A_CONTROLER}}")
    if rec["acomptes"] == 0.0:
        balises.append("{{ACOMPTES_NON_ALIMENTES}}")
    if donnees_partielles:
        balises.append("{{DONNEES_PARTIELLES}}")
    balises_str = " ".join(balises)

    # période mensuelle
    periode_debut = f"{mois}-01"
    periode_fin   = str(pd.Period(mois, freq="M").end_time.date())

    entetes.append({
        "facture_id": facture_id, "mois": mois,
        "proprietaire_id": prop_id, "nom_proprietaire": nom_prop or "{{NOM_PROPRIETAIRE}}",
        "adresse_proprietaire": adresse_prop or "{{ADRESSE_PROPRIETAIRE_A_COMPLETER}}",
        "logement_id": log_id, "nom_logement": lg.get("nom_logement_officiel") or lg.get("nom_court") or log_id,
        "periode_debut": periode_debut, "periode_fin": periode_fin,
        "nb_reservations": rec["nb_reservations"],
        "total_exploitation_net": rec["revenu_net_exploitation"],
        "total_reglement_du": rec["montant_du"],
        "reste_a_payer": rec["reste_a_payer"],
        "mode_facturation": mode_fact,
        "statut_facture": statut_facture,
        "statut_generation": statut_generation,
        "balises": balises_str,
        "date_generation": TODAY,
    })

    # ── 12 lignes §17.3 ──
    L = [
        (1,  "TOTAL_PAYOUT",            "Total payout",                                     rec["total_payout"],            "EXPLOITATION"),
        (2,  "MENAGE_FACTURE",          "Ménage facturé",                                   rec["total_menage"],            "EXPLOITATION"),
        (3,  "COMMISSION_CONCIERGERIE", "Commission conciergerie",                          rec["total_commission"],        "EXPLOITATION"),
        (4,  "CHARGE_FIXE",             "Charge fixe mensuelle",                            rec["charge_fixe"],             "EXPLOITATION"),
        (5,  "REVENU_NET_EXPLOITATION", "Revenu net d'exploitation propriétaire",           rec["revenu_net_exploitation"], "EXPLOITATION"),
        (6,  "MONTANT_DU",              "Montant total dû à la conciergerie",               rec["montant_du"],              "REGLEMENT"),
        (7,  "ACOMPTE_AIRBNB",          "Acompte reçu via Airbnb",                          0.0,                            "REGLEMENT"),
        (8,  "PAIEMENT_DEJA_RECU",      "Autres paiements déjà reçus",                      0.0,                            "REGLEMENT"),
        (9,  "RESTE_A_PAYER",           "Reste à payer à la conciergerie",                  rec["reste_a_payer"],           "REGLEMENT"),
        (10, "CHARGES_EXCEPT_REFAC",    "Charges / achats exceptionnels refacturés",        0.0,                            "REGLEMENT"),
        (11, "ACOMPTES_PROPRIETAIRES",  "Acomptes propriétaires (réservations hors HA)",    rec["acomptes"],                "REGLEMENT"),
    ]
    for num, t, lib, mt, bloc in L:
        lignes.append({
            "facture_id": facture_id, "ligne_num": num, "type_ligne": t,
            "libelle": lib, "montant": mt, "bloc": bloc, "commentaire": "",
        })
    # Ligne 12 : statut règlement (texte)
    lignes.append({
        "facture_id": facture_id, "ligne_num": 12, "type_ligne": "STATUT_REGLEMENT",
        "libelle": "Statut règlement", "montant": None, "bloc": "REGLEMENT",
        "commentaire": statut_facture,
    })

df_entete   = pd.DataFrame(entetes)
df_lignes   = pd.DataFrame(lignes)
df_controle = pd.DataFrame(controle)


# ──────────────────────────────────────────────────────────────────────────
# 3. DASHBOARD_FACTURATION (prop × mois)
# ──────────────────────────────────────────────────────────────────────────
dash_rows = []
if len(df_entete) > 0:
    for (mois, prop), grp in df_entete.groupby(["mois", "proprietaire_id"]):
        g = dash_idx.get(str(mois))
        bloquants = (g["bloquants"] if g else 0) + transverse["bloquants"]
        a_controler = (g["a_controler"] if g else 0) + transverse["a_controler"]
        balises_all = set()
        for b in grp["balises"]:
            balises_all.update(b.split())
        crit = [b for b in balises_all if b in (
            "{{MODE_FACTURATION_A_DEFINIR}}", "{{BANQUE_NON_CLOTUREE}}",
            "{{ADRESSE_PROPRIETAIRE_A_COMPLETER}}", "{{RESERVATIONS_A_CONTROLER}}")]
        dash_rows.append({
            "mois": mois, "proprietaire_id": prop,
            "nb_logements": grp["logement_id"].nunique(),
            "nb_bloquants_mois": bloquants, "nb_a_controler_mois": a_controler,
            "facturation_lot12_ok": (g["ok"] if g else "NON"),
            "mode_facturation": grp["mode_facturation"].iloc[0],
            "statut_facture": grp["statut_facture"].iloc[0],
            "balises_non_resolues": " ".join(sorted(crit)),
        })
df_dashf = pd.DataFrame(dash_rows)


# ──────────────────────────────────────────────────────────────────────────
# 4. A_CONTROLER (réservations exclues + anomalies facturation)
# ──────────────────────────────────────────────────────────────────────────
ac_rows = []
for _, r in df_ac.iterrows():
    ac_rows.append({
        "mois": None, "proprietaire_id": None,
        "logement_id": r.get("logement_id_snapshot"),
        "reservation": r.get("reservation_id"),
        "code_anomalie": r.get("code_anomalie_lot10") or "RESERVATION_A_CONTROLER",
        "severite": "A_CONTROLER",
        "impact_facturation": "EXCLUE_DE_FACTURE",
        "message": f"Réservation {r.get('source')} exclue — saisie/contrôle requis",
    })
# Anomalie transverse mode_facturation
n_adef = int((df_prop["mode_facturation"].astype(str) == "A_DEFINIR").sum())
if n_adef > 0:
    ac_rows.append({
        "mois": None, "proprietaire_id": None, "logement_id": None, "reservation": None,
        "code_anomalie": "MODE_FACTURATION_A_DEFINIR", "severite": "A_CONTROLER",
        "impact_facturation": "BLOQUE_FACTURE_FINALE",
        "message": f"{n_adef} proprietaires sans mode_facturation",
    })
df_acout = pd.DataFrame(ac_rows)


# ──────────────────────────────────────────────────────────────────────────
# 5. Contrôles BLOQUANTS Lot 12
# ──────────────────────────────────────────────────────────────────────────
log.info("Contrôles Lot 12...")
bloquants_l12 = []

# Aucune facture finale ne doit exister (préfactures only)
if len(df_entete) > 0:
    non_pref = df_entete[df_entete["statut_generation"] != "PREFACTURE_CONTROLE"]
    if len(non_pref) > 0:
        bloquants_l12.append(f"FACTURE_FINALE_GENEREE: {len(non_pref)}")
    # 12 lignes par facture
    cnt = df_lignes.groupby("facture_id").size()
    bad = cnt[cnt != 12]
    if len(bad) > 0:
        bloquants_l12.append(f"LIGNE_FACTURE_INCOMPLETE: {len(bad)} factures != 12 lignes")
    # Confusion net vs reste_a_payer (L5 == L9 alors que montants non nuls)
    for fid, grp in df_lignes.groupby("facture_id"):
        l5 = grp[grp["ligne_num"] == 5]["montant"].iloc[0]
        l9 = grp[grp["ligne_num"] == 9]["montant"].iloc[0]
        if _n(l5) != 0 and _n(l5) == _n(l9) and _n(grp[grp["ligne_num"]==6]["montant"].iloc[0]) != 0:
            bloquants_l12.append(f"CONFUSION_NET_VS_RESTE_A_PAYER: {fid}")
            break

if bloquants_l12:
    for b in bloquants_l12:
        log.error(f"BLOQUANT Lot12 — {b}")
    sys.exit(1)
log.info("  0 BLOQUANT Lot 12")


# ──────────────────────────────────────────────────────────────────────────
# 6. Écriture
# ──────────────────────────────────────────────────────────────────────────
OUT_DIR.mkdir(parents=True, exist_ok=True)
wb = Workbook()
wb.remove(wb.active)
_write_sheet(wb, "FACT_FACTURE_ENTETE",   df_entete)
_write_sheet(wb, "FACT_FACTURE_LIGNES",   df_lignes)
_write_sheet(wb, "CONTROLE_MENSUEL",      df_controle)
_write_sheet(wb, "DASHBOARD_FACTURATION", df_dashf)
_write_sheet(wb, "A_CONTROLER",           df_acout)
wb.save(OUT_FILE)


# ──────────────────────────────────────────────────────────────────────────
# 7. Rapport
# ──────────────────────────────────────────────────────────────────────────
from collections import Counter
sep = "=" * 60
log.info(sep)
log.info("RAPPORT LOT 12 — PRÉFACTURES")
log.info(sep)
log.info(f"  Fichier            : {OUT_FILE}")
log.info(f"  Préfactures (ENTETE): {len(df_entete)}")
log.info(f"  Lignes facture      : {len(df_lignes)}  (attendu {len(df_entete)*12})")
log.info(f"  CONTROLE_MENSUEL    : {len(df_controle)} lignes")
log.info(f"    dont GLOBAL_NON_AFFECTE : {int((df_controle['logement_id']==SENTINEL_GLOBAL).sum()) if len(df_controle)>0 else 0}")
log.info(f"  A_CONTROLER         : {len(df_acout)}")
if len(df_entete) > 0:
    log.info(f"  Statuts facture     : {dict(Counter(df_entete['statut_facture']))}")
    log.info(f"  Statuts generation  : {dict(Counter(df_entete['statut_generation']))}")
    finales = int((df_entete['statut_generation'] != 'PREFACTURE_CONTROLE').sum())
    log.info(f"  Factures finales    : {finales} (doit etre 0)")
log.info(sep)
log.info("Sources amont : AUCUNE modification (lecture seule).")
log.info("ATTENTE VALIDATION HUMAINE avant commit.")
